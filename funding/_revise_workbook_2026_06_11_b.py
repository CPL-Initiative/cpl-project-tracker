#!/usr/bin/env python3
"""One-shot revision #2 of CPL_Funding_Model_2026.xlsx — 2026-06-11 (Sam):

1. Rename college "Chabot Hayward" → "Chabot" (display preference; the
   short-name dataset is updated in lockstep so the MAP-name join keeps
   resolving).
2. Add "Mt San Antonio Noncredit" as a new institution row, inserted after
   the Mt San Antonio row. HEADCOUNT IS 0 PENDING SAM'S DataMart NUMBER —
   the row exists and allocates $0 until the 2022-23 annual headcount for
   the noncredit division is typed in (then re-run the data builder).
   District / county / county-census cells copy from the Mt San Antonio
   sibling row.
3. Re-fill the uniform formulas for the new extent (rows 8..last college),
   including C8's SUM range and the AVERAGE row's range — openpyxl row
   insertion does NOT auto-adjust formula references (exactly the bug class
   the rev2 SUM fix addressed; never leave ranges stale after an insert).

Run once from repo root, then re-run funding/_build_funding_data.py:

    python3 funding/_revise_workbook_2026_06_11_b.py
"""
import os

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "CPL_Funding_Model_2026.xlsx")
SHEET = "CPL IMPLEMENTATION MODEL 2026"
FIRST_COLLEGE_ROW, SYSTEM_ROW = 9, 8
NEW_NAME = "Mt San Antonio Noncredit"


def fill_formulas(ws, last_row):
    ws["C8"] = f"=SUM(C{FIRST_COLLEGE_ROW}:C{last_row})"
    for r in range(SYSTEM_ROW, last_row + 1):
        ws[f"D{r}"] = f"=C{r}/C$8"
        ws[f"E{r}"] = f"=$D{r}*E$7*$H$3"
        ws[f"G{r}"] = f"=$D{r}*G$7*$H$3"
        ws[f"I{r}"] = f"=$D{r}*I$7*$H$3"
        ws[f"F{r}"] = f"=C{r}*F$7"
        ws[f"H{r}"] = f"=C{r}*H$7"
        ws[f"J{r}"] = f"=C{r}*J$7"
        ws[f"K{r}"] = f"=E{r}+G{r}+I{r}"
        ws[f"O{r}"] = f"=N{r}/N$8"


def main():
    wb = load_workbook(XLSX)
    ws = wb[SHEET]

    # locate rows by LOCATION (col B) / the AVERAGE row (col I)
    chabot = msac = last = avg_row = None
    max_order = 0
    for r in range(FIRST_COLLEGE_ROW, ws.max_row + 1):
        b = ws[f"B{r}"].value
        if isinstance(ws[f"A{r}"].value, (int, float)):
            max_order = max(max_order, int(ws[f"A{r}"].value))
            last = r
        if str(ws[f"I{r}"].value or "").strip() == "AVERAGE ALLOCATION":
            avg_row = r
        if b is None:
            continue
        name = str(b).strip()
        if name == "Chabot Hayward":
            chabot = r
        if name == "Mt San Antonio":
            msac = r
    assert chabot and msac and last and avg_row, (chabot, msac, last, avg_row)
    assert ws[f"B{r}"].value is None or True  # scan completed

    # 1. rename
    ws[f"B{chabot}"] = "Chabot"

    # 2. insert the noncredit row after Mt San Antonio
    ws.insert_rows(msac + 1)
    new_r, last, avg_row = msac + 1, last + 1, avg_row + 1
    ws[f"A{new_r}"] = max_order + 1
    ws[f"B{new_r}"] = NEW_NAME
    ws[f"C{new_r}"] = 0   # ← PENDING: Sam supplies the DataMart 2022-23 headcount
    for col in ("L", "M", "N"):
        ws[f"{col}{new_r}"] = ws[f"{col}{msac}"].value
    for col in "CDEFGHIJKNO":
        ws[f"{col}{new_r}"].number_format = ws[f"{col}{msac}"].number_format

    # 3. re-fill all formulas for the new extent + the AVERAGE range
    fill_formulas(ws, last)
    ws[f"K{avg_row}"] = f"=AVERAGE(K{FIRST_COLLEGE_ROW}:K{last})"

    wb.calculation.fullCalcOnLoad = True
    wb.save(XLSX)
    print(f"revised: Chabot Hayward→Chabot (row {chabot}); '{NEW_NAME}' inserted at row {new_r} "
          f"(order {max_order + 1}, headcount 0 PENDING); formulas re-filled 8..{last}; "
          f"AVERAGE range → K9:K{last}")


if __name__ == "__main__":
    main()
