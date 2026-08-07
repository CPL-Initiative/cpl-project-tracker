"""
_student_detail_local.py — hash + aggregate the student-detail export, LOCALLY.

RUN THIS ON YOUR OWN MACHINE. Nothing about it belongs on a runner or in a
session: the whole point is that the per-student rows never leave the computer
they are already on. What comes out the other end is a small aggregate file you
can paste or upload anywhere.

    python funding/_student_detail_local.py "G:\\My Drive\\AI Documentation\\Copy of StudentDetailsCredit_080626.xlsx"

Needs openpyxl for .xlsx  ->  pip install openpyxl
(.csv needs nothing; if the install is a nuisance, save-as CSV in Excel instead.)

WHY LOCAL, AND WHY THIS SHAPE
-----------------------------
Sam supplied the 2026-08-06 export and stripped the masked name/SSN columns
himself, then asked for InternalMAPStudentID to be hashed as a further
precaution. Both are the right instincts. But the export is ~51 MB and the Drive
connector caps at 10 MB with no range requests, so a session cannot read the file
to do the hashing — and it should not want to. This project already settled that
question: funding/_build_cr_backlog.py states "the student grain never leaves the
runner", and Session 26 / #227 was a PII forward-stop that needed a git history
rewrite to undo. A session's context is a worse place than a runner, not a
better one.

So the grain stays with you, and only counts travel.

ABOUT THE SALT — PLEASE READ, IT IS THE WHOLE PROTECTION
-------------------------------------------------------
A hash of an internal ID is NOT anonymous by itself. The ID space is small and
enumerable, so anyone holding the salt (or a salt-free hash) can hash every
plausible ID and invert the whole column in seconds. The hash is only worth
anything while the salt is secret.

Therefore this script:
  * reads the salt from the MAP_HASH_SALT environment variable, or from
    ~/.map_hash_salt (your home directory — deliberately OUTSIDE any git repo,
    so it cannot be committed by accident);
  * generates a strong random salt and saves it there on first run;
  * never prints the salt, and never writes it into the output.

Keep that file. The same salt gives the same hashes, which is what lets two
exports be compared over time. Lose it and past hashes stop lining up — that is
the only cost, and it is preferable to the alternative.

WHAT THE HASH BUYS
------------------
A real DISTINCT-STUDENT count. Without a per-student key you can only count
credit recommendations, and one person holding several CRs under one exhibit —
or the same credential at two colleges — inflates the number. Every headcount
below is a count of distinct hashes; every "recommendations" figure is rows. The
output labels which is which, because conflating them is exactly how a metric
starts overstating itself.

WHAT COMES OUT
--------------
`student_detail_aggregate.json` next to the input, plus a console summary you can
paste straight back into a session. Aggregates only: no ID, no hash, no free-text
that could carry a name. Cells below 5 are suppressed to null with a `_suppressed`
flag, matching docs/kb-notes/adr-funding-priority-metrics-privacy.md.
"""
import hashlib
import json
import os
import re
import secrets
import sys
from collections import Counter, defaultdict

SUPPRESS_BELOW = 5
NEEDS_ACTION = "Needs Action"

# Anything other than Needs Action means a human looked at it. Not Applicable
# counts as work done — reviewing a recommendation and correctly ruling it out is
# the job, and an applied-only measure penalises a college for doing it properly
# (Cabrillo: 844 N/A vs 320 Applied).
ACTED = ("Applied to CPL Plan", "Not Applicable", "In Process")

# ACE says outright that no credit is recommended for these, yet they sit in
# plans as Needs Action. Unarticulable by construction — carved out and counted
# separately rather than silently dropped, because the count is itself a finding.
NOT_RECOMMENDED = re.compile(r"credit\s+is\s+not\s+recommended", re.I)

TEST_COLLEGES = {"RivTest City College", "MorTest City College", "Nortest City College",
                 "CA MAP INITIATIVE COLLEGE", "RivTest", "MorTest", "Nortest"}

# Same family the chat layer expands, so this answers the REAL question rather
# than a tidier one: "how many students statewide are eligible for CPL based on
# CPR, AED, or similar certs?"
FAMILIES = {
    "CPR / AED / first aid": re.compile(
        r"\b(cpr|aed|first\s*aid|basic\s*life\s*support|\bbls\b|cardiopulmonary|"
        r"defibrillat|lifesaving|life\s*saving|heartsaver)", re.I),
}

# Column names are matched loosely — exports rename things between versions, and
# a hard-coded header that silently matches nothing is worse than an error.
#
# ORDERED, and that ordering is load-bearing. Each entry is a list of patterns
# tried MOST-SPECIFIC FIRST, across all columns, before falling back. The first
# version used one regex with alternation, which takes the first COLUMN matching
# any branch — so on the 2026-08-06 export a column plainly named `Status` won
# over `CPLStatusPlan` purely by sitting earlier in the header row. `Status`
# holds the workflow STAGE (Needs Action / Implementation / Faculty / Initiator /
# Articulation Officer), not the CPL disposition, and the run reported a
# statewide disposition rate of 0.0% across 525,362 rows. A wrong column is far
# more dangerous than a missing one: it produces a confident, plausible-looking,
# completely false number.
WANTED = {
    "location":   [r"^location$", r"^college$"],
    "status":     [r"^cpl\s*status\s*plan$", r"cpl\s*status\s*plan",
                   r"cpl.*status", r"disposition", r"^status$"],
    "cr":         [r"^credit\s*recommendation$", r"credit\s*recommendation"],
    "course":     [r"^college\s*course$", r"college\s*course", r"^course$"],
    "exhibit":    [r"^exhibit\s*id$", r"exhibit\s*id"],
    "student_id": [r"internal.*map.*student.*id", r"^student\s*map\s*id$",
                   r"student\s*map\s*id", r"student.*id"],
}

# The four values MAP writes into CPLStatusPlan. Used only to CHECK that the
# column we matched is the one we meant — see _check_status_column.
DISPOSITION_VOCAB = {"applied to cpl plan", "not applicable", "in process",
                     "needs action"}


def load_salt():
    """Env var, else ~/.map_hash_salt, else generate and save one there."""
    env = os.environ.get("MAP_HASH_SALT", "").strip()
    if env:
        print("  salt: from MAP_HASH_SALT")
        return env.encode()
    path = os.path.join(os.path.expanduser("~"), ".map_hash_salt")
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            s = f.read().strip()
        if s:
            print(f"  salt: reusing {path} (same hashes as last run)")
            return s.encode()
    s = secrets.token_hex(32)
    with open(path, "w", encoding="utf-8") as f:
        f.write(s)
    try:
        os.chmod(path, 0o600)
    except OSError:
        pass
    print(f"  salt: NEW salt generated and saved to {path}")
    print("        Keep this file. Do NOT commit it or paste it anywhere — it is")
    print("        the only thing making the hashes non-reversible.")
    return s.encode()


def hasher(salt):
    cache = {}

    def h(v):
        v = str(v or "").strip()
        if not v:
            return ""
        if v not in cache:
            cache[v] = hashlib.blake2b(salt + v.encode(), digest_size=16).hexdigest()
        return cache[v]
    return h


def iter_rows(path):
    """Yield dicts, streaming. Never materialises the file.

    Takes whichever export is to hand: .xlsx/.xlsm, .csv, or .json — the JSON
    being either a bare array of row objects or the CustomReport envelope
    ({viewName, columnName:[...], columnValue:[[...]]}), which is column-oriented
    and NOT a list of dicts. Same two shapes funding/_build_cr_backlog.py already
    accepts, so whichever file MAP hands over works without a conversion step.
    """
    if path.lower().endswith(".json"):
        with open(path, encoding="utf-8", errors="replace") as f:
            text = f.read()
        start = text.find("[")
        if start == -1:
            return
        if '"viewName"' in text[:4000]:
            for report in json.loads(text) or []:
                cols = report.get("columnName", [])
                for row in report.get("columnValue") or []:
                    yield dict(zip(cols, row))
            return
        dec = json.JSONDecoder()
        i, n = start + 1, len(text)
        while True:
            while i < n and text[i] in " \n\r\t,":
                i += 1
            if i >= n or text[i] == "]":
                return
            obj, i = dec.raw_decode(text, i)
            yield obj
        return
    if path.lower().endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Needs openpyxl for .xlsx —  pip install openpyxl\n"
                     "(or save the sheet as .csv in Excel and re-run on that)")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip() if c is not None else "" for c in row]
                continue
            yield dict(zip(header, row))
        wb.close()
        return
    import csv
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        for rec in csv.DictReader(f):
            yield rec


def resolve_columns(sample_keys):
    """Most-specific pattern first, ACROSS ALL COLUMNS, before falling back."""
    got = {}
    cols = [str(c) for c in sample_keys if c]
    for key, patterns in WANTED.items():
        for pat in patterns:
            hit = next((c for c in cols if re.search(pat, c, re.I)), None)
            if hit:
                got[key] = hit
                break
    return got


def check_status_column(observed, col_name):
    """Did we match the DISPOSITION column, or something that merely looks like one?

    Returns True when the values look like CPLStatusPlan. A statement of the
    problem beats a silent wrong answer: the numbers downstream are only
    meaningful if this column is the disposition, and every other column in the
    export is plausible enough to pass unnoticed.
    """
    vals = {v.lower() for v in observed if v}
    known = vals & DISPOSITION_VOCAB
    # "Needs Action" alone is not enough — it appears in the workflow-stage
    # column too, which is exactly how the wrong column passed the first time.
    if known - {"needs action"}:
        return True
    print("\n" + "!" * 74)
    print(f"  WRONG COLUMN: '{col_name}' does not hold CPL dispositions.")
    print(f"  values seen: {sorted(vals)[:8]}")
    print("  expected some of: Applied to CPL Plan / Not Applicable / In Process")
    print("  This is the WORKFLOW STAGE, not the disposition. Every disposition")
    print("  rate below is meaningless and has been withheld rather than printed")
    print("  as 0%. Student and exhibit COUNTS are unaffected and still valid.")
    print("  Fix: confirm the export carries a CPLStatusPlan column; if it was")
    print("  dropped when the PII columns were stripped, re-export with it.")
    print("!" * 74)
    return False


def suppress(n):
    return (None, True) if 0 < n < SUPPRESS_BELOW else (n, False)


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__.strip().split("\n\n")[2])
    path = sys.argv[1]
    if not os.path.exists(path):
        sys.exit(f"not found: {path}")

    print(f"Reading {path} ({os.path.getsize(path) / 1e6:.1f} MB)")
    salt = load_salt()
    h = hasher(salt)

    rows = iter_rows(path)
    try:
        first = next(rows)
    except StopIteration:
        sys.exit("no rows")
    cols = resolve_columns(first.keys())
    # ALWAYS print every header, not only on failure. The first run matched the
    # wrong status column and there was no way to see what it could have picked.
    print(f"  headers in file ({len(first)}): {list(first.keys())}")
    print(f"  columns matched: { {k: v for k, v in cols.items()} }")
    missing = [k for k in ("location", "status", "cr") if k not in cols]
    if missing:
        sys.exit(f"could not find columns for {missing}. Headers seen:\n"
                 f"  {list(first.keys())}")
    if "student_id" not in cols:
        print("  ⚠ no InternalMAPStudentID column found — distinct-student counts")
        print("    will be omitted; recommendation counts still work.")

    def g(r, key):
        c = cols.get(key)
        return str(r.get(c) or "").strip() if c else ""

    state = Counter()
    students_state = set()
    per_college = defaultdict(Counter)
    college_students = defaultdict(set)
    # PER-EXHIBIT is the rollup Sierra actually needs. ExhibitID is the join key
    # to chatbox_exhibits.exhibit_id, so this is what would let a topic question
    # ("CPR/AED") total real dispositions across the exhibits it matched — the
    # gap that made Sierra hedge. No college x exhibit cross-tab: that is where
    # cells get thin enough to point at a person.
    per_exhibit = defaultdict(Counter)
    exhibit_students = defaultdict(set)
    exhibit_colleges = defaultdict(set)
    fam_status = {name: Counter() for name in FAMILIES}
    fam_students = {name: set() for name in FAMILIES}
    fam_colleges = {name: set() for name in FAMILIES}
    fam_by_college = {name: defaultdict(set) for name in FAMILIES}
    not_rec = 0
    seen = 0

    # `first` was consumed to sniff the headers — put it back without ever
    # materialising the rest (the point of streaming a 51 MB export).
    for r in _chain(first, rows):
        loc = g(r, "location")
        if not loc or loc in TEST_COLLEGES:
            continue
        seen += 1
        status = g(r, "status") or NEEDS_ACTION
        cr = g(r, "cr")
        if status == NEEDS_ACTION and NOT_RECOMMENDED.search(cr):
            not_rec += 1
            continue
        sid = h(g(r, "student_id")) if "student_id" in cols else ""
        state[status] += 1
        per_college[loc][status] += 1
        if sid:
            students_state.add(sid)
            college_students[loc].add(sid)
        ex = g(r, "exhibit")
        if ex:
            per_exhibit[ex][status] += 1
            exhibit_colleges[ex].add(loc)
            if sid:
                exhibit_students[ex].add(sid)
        blob = cr + " " + g(r, "course")
        for name, pat in FAMILIES.items():
            if pat.search(blob):
                fam_status[name][status] += 1
                fam_colleges[name].add(loc)
                if sid:
                    fam_students[name].add(sid)
                    fam_by_college[name][loc].add(sid)
        if seen % 100000 == 0:
            print(f"    …{seen:,} rows")

    total = sum(state.values())
    acted = sum(state[d] for d in ACTED)
    ok = check_status_column(state.keys(), cols.get("status", "?"))

    def rate(a, t):
        """None — never 0.0 — when the status column is not the disposition."""
        return round(a / t, 4) if (ok and t) else None

    print(f"\nROWS (non-test): {seen:,}  |  scored: {total:,}  |  "
          f"'credit is not recommended' carved out: {not_rec:,}")
    print(f"DISTINCT STUDENTS statewide: {len(students_state):,}" if students_state
          else "DISTINCT STUDENTS: unavailable (no id column)")
    if ok and total:
        print(f"DISPOSITION RATE statewide: {100 * acted / total:.1f}% "
              f"({acted:,} acted / {total:,})")
    else:
        print("DISPOSITION RATE statewide: WITHHELD (see the warning above)")
    print("\nby disposition:")
    for d, n in state.most_common():
        print(f"  {d or '(blank)':<26} {n:>9,}  ({100 * n / total:.1f}%)")

    payload = {
        "source": os.path.basename(path),
        "note": ("Aggregates only. Student ids salted-hashed locally and never "
                 "written here. 'students' = DISTINCT hashed ids; "
                 "'recommendations' = rows. Cells <5 suppressed."),
        "suppress_below": SUPPRESS_BELOW,
        "status_column": cols.get("status"),
        "disposition_rates_valid": ok,
        "statewide": {
            "recommendations": total,
            "students": len(students_state) or None,
            "acted": acted,
            "disposition_rate": rate(acted, total),
            "not_recommended_carved": not_rec,
            "by_disposition": dict(state),
        },
        "families": {},
        "colleges": {},
        "exhibits": {},
    }

    # Keyed by ExhibitID, which joins to chatbox_exhibits.exhibit_id — the payload
    # that would let Sierra say "how many students, and how far along" for the
    # exhibits a topic search matched.
    ex_dropped = ex_dropped_rows = 0
    for ex, c in per_exhibit.items():
        t = sum(c.values())
        if t < SUPPRESS_BELOW:
            # SAY SO rather than just shrinking the map. A quietly short list
            # reads as "that is all the data there is", and a scan returning
            # nothing looks like good news, which nobody investigates.
            ex_dropped += 1
            ex_dropped_rows += t
            continue
        a = sum(c[d] for d in ACTED)
        payload["exhibits"][ex] = {
            "recommendations": t,
            "students": suppress(len(exhibit_students[ex]))[0],
            "colleges": len(exhibit_colleges[ex]),
            "acted": a,
            "disposition_rate": rate(a, t),
            "by_disposition": dict(c),
        }
    payload["exhibits_suppressed"] = {
        "exhibits": ex_dropped, "recommendations": ex_dropped_rows,
        "why": f"fewer than {SUPPRESS_BELOW} credit recommendations statewide",
    }

    for name in FAMILIES:
        fs = fam_status[name]
        ftotal = sum(fs.values())
        facted = sum(fs[d] for d in ACTED)
        print(f"\n=== {name} ===")
        print(f"  credit recommendations : {ftotal:,}")
        print(f"  DISTINCT STUDENTS      : {len(fam_students[name]):,}"
              if fam_students[name] else "  DISTINCT STUDENTS      : n/a")
        print(f"  colleges involved      : {len(fam_colleges[name])}")
        for d, n in fs.most_common():
            print(f"    {d or '(blank)':<26} {n:>8,}  ({100 * n / ftotal:.1f}%)"
                  if ftotal else "")
        top = sorted(fam_by_college[name].items(), key=lambda kv: -len(kv[1]))[:15]
        print("  top colleges by distinct students (<5 suppressed):")
        for loc, sids in top:
            v, sup = suppress(len(sids))
            print(f"    {loc[:44]:<46} {'<5' if sup else v:>6}")
        payload["families"][name] = {
            "recommendations": ftotal,
            "students": len(fam_students[name]) or None,
            "colleges": len(fam_colleges[name]),
            "acted": facted,
            "disposition_rate": rate(facted, ftotal),
            "by_disposition": dict(fs),
            "top_colleges": [
                {"college": loc, "students": suppress(len(s))[0],
                 "_suppressed": suppress(len(s))[1]} for loc, s in top],
        }

    for loc, c in per_college.items():
        t = sum(c.values())
        if t < SUPPRESS_BELOW:
            continue
        a = sum(c[d] for d in ACTED)
        payload["colleges"][loc] = {
            "recommendations": t,
            "students": suppress(len(college_students[loc]))[0],
            "acted": a,
            "disposition_rate": rate(a, t),
        }

    print(f"\nexhibits reported: {len(payload['exhibits']):,}  |  "
          f"suppressed as too thin: {ex_dropped:,} "
          f"({ex_dropped_rows:,} recommendations)")

    out = os.path.join(os.path.dirname(os.path.abspath(path)),
                       "student_detail_aggregate.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=1, ensure_ascii=False)
    print(f"\nWrote {out}")
    print("That file is safe to share — no ids, no hashes, no free text.")


def _chain(first, rest):
    yield first
    for r in rest:
        yield r


if __name__ == "__main__":
    main()
