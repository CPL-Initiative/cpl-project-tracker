#!/usr/bin/env python3
"""One-shot revision of CPL_Funding_Model_2026.xlsx — 2026-06-11 rev2.

Applies the formula-review recommendations Sam approved ("Great
recommendations! Make the changes"):

1. SUM-range fix: C8 =SUM(C9:C125) → =SUM(C9:C126) (the list ends at C126/
   Yuba; the old range excluded the last row, so the pool block under-counted
   statewide headcount by 8,417 and the model over-allocated its own tranche
   by ~$44.6K/yr).
2. Shares-first model: the three PRIORITY SHARES (30% / 42% / 28%) become the
   typed inputs (row 7, E7/G7/I7). College dollars are now
       E9 = $D9 * E$7 * $H$3      (headcount share × priority share × tranche)
   The old FUNDING FACTOR is gone; the projection percents (F7/H7/J7) remain
   as PROJECTED HEADCOUNT **TARGETS** that no longer move money. Balance is
   exact by construction whenever the shares sum to 100%.
3. One formula per column, filled down: D8:D126 = C/C$8 (D126 was a typed
   literal — the mixed-denominator artifact); E/F/G/H/I/J/K rewritten
   uniformly for every row 8-126. Column P ("...PCT CPUNTY TOTAL") is REMOVED
   — internally inconsistent literals with no recoverable denominator; the
   dashboard never rendered it.
4. Cosmetics: J2 header 26-27 → 28-29; input cells are plain values (C3 was
   =9040307; E3 =400000*3 → 1200000 with the "(2 FTE*3 YRS)" label carrying
   the derivation).

Run once from repo root (kept for provenance, like the kb/ one-shots):

    python3 funding/_revise_workbook_2026_06_11.py

The original edition is preserved in git history (committed 2026-06-11,
PR #353). After running, re-run funding/_build_funding_data.py.

NOTE: openpyxl cannot evaluate formulas, so edited cells carry no cached
values until the workbook is next opened+saved in Excel; fullCalcOnLoad is
set so Excel recalculates everything on open. The data builder computes the
chain from the typed inputs and never reads cached derived values.
"""
import os

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "CPL_Funding_Model_2026.xlsx")
SHEET = "CPL IMPLEMENTATION MODEL 2026"
FIRST_COLLEGE_ROW, LAST_COLLEGE_ROW = 9, 126
SYSTEM_ROW = 8


def main():
    wb = load_workbook(XLSX)
    ws = wb[SHEET]

    # ── 4. input-cell hygiene + header fixes ────────────────────────────
    ws["C3"] = 9040307            # was the formula constant =9040307
    ws["E3"] = 1200000            # was =400000*3; the label already documents it
    ws["E2"] = "ADMIN COST (2 FTE*3 YRS = 400000*3)"
    ws["J2"] = "28-29 AVAILABLE COLLEGE FUNDING"   # was a repeated "26-27"

    # ── 2. shares-first inputs (row 6 labels + row 7 values) ────────────
    ws["E6"] = "PRIORITY SHARE"
    ws["G6"] = "PRIORITY SHARE"
    ws["I6"] = "PRIORITY SHARE"
    ws["F6"] = "PROJECTED HEADCOUNT (TARGET)"
    ws["H6"] = "PROJECTED HEADCOUNT (TARGET)"
    ws["J6"] = "PROJECTED HEADCOUNT (TARGET)"
    for col, share in (("E", 0.30), ("G", 0.42), ("I", 0.28)):
        ws[f"{col}7"] = share
        ws[f"{col}7"].number_format = "0%"
    for col in ("F", "H", "J"):   # projection targets keep their values
        ws[f"{col}7"].number_format = "0.00%"

    # ── 1 + 3. uniform formulas, filled down (SYSTEM row 8 + colleges) ──
    ws["C8"] = f"=SUM(C{FIRST_COLLEGE_ROW}:C{LAST_COLLEGE_ROW})"   # was C9:C125
    for r in range(SYSTEM_ROW, LAST_COLLEGE_ROW + 1):
        ws[f"D{r}"] = f"=C{r}/C$8"                 # D126 was a typed literal
        ws[f"E{r}"] = f"=$D{r}*E$7*$H$3"           # headcount share × share × tranche
        ws[f"G{r}"] = f"=$D{r}*G$7*$H$3"
        ws[f"I{r}"] = f"=$D{r}*I$7*$H$3"
        ws[f"F{r}"] = f"=C{r}*F$7"                 # projection TARGET (no $ effect)
        ws[f"H{r}"] = f"=C{r}*H$7"
        ws[f"J{r}"] = f"=C{r}*J$7"
        ws[f"K{r}"] = f"=E{r}+G{r}+I{r}"
        ws[f"O{r}"] = f"=N{r}/N$8"                 # was =SUM(N9/$N$8) — same, plainly

    # ── 3. drop column P (inconsistent literals, never rendered) ────────
    ws["P5"] = None
    for r in range(SYSTEM_ROW, LAST_COLLEGE_ROW + 1):
        ws[f"P{r}"] = None

    # Excel recalculates everything on open (edited cells carry no caches).
    wb.calculation.fullCalcOnLoad = True
    wb.save(XLSX)
    print(f"revised {os.path.basename(XLSX)}: shares-first model, "
          f"C8=SUM(C9:C{LAST_COLLEGE_ROW}), formulas filled {SYSTEM_ROW}-{LAST_COLLEGE_ROW}, "
          "column P removed, headers fixed")


if __name__ == "__main__":
    main()
