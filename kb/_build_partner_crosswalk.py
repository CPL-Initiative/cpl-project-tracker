#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build a partner occupation → CPL crosswalk workbook.

"An employment training center sent us a list of the occupations they train
for. Which of those can our students already get college credit for, and at
which colleges?"  (SkyWalker, 2026-08-05 — first run: San Joaquin County
Office of Education, 160 occupation rows.)

This is the REUSABLE engine behind that question. Point it at any partner's
occupation/program list and it emits a distributable workbook: every occupation
matched to the CPL credit recommendations California Community Colleges have
articulated on the MAP platform, plus the colleges that adopted each one and
the local courses that grant the credit.

    python3 kb/_build_partner_crosswalk.py \
        --input <list.xlsx|.csv|.txt> --partner "San Joaquin COE" --slug sjcoe \
        [--column 2] [--sheet Sheet1] [--region-preset san-joaquin]

WHY A SHARED MAP (the scaling bit)
----------------------------------
The occupation→credential match is subject-matter JUDGMENT, not a mechanical
join: partner lists carry job and apprenticeship titles ("RESIDENTIAL WIREMAN",
"Hydro Plant Operator"), never credential names. That judgment is expensive, so
it is NOT thrown away per run — it accumulates in

    kb/occupation_credential_map.json

keyed by a normalized occupation string, so the next partner's "Plumber" reuses
the "PLUMBER" ruling curated for the last one. Each run reports what it could
NOT map (`unmapped.json`) — that file IS the curator worklist for growing the
map. Coverage compounds across partners instead of restarting.

TWO SOURCES, AND WHY BOTH
-------------------------
  1. statewide_data.js        — the adoption view. Carries `collaborative_type`,
     `adopter_names`, `potential_names`, and exhibit-level `credit_recs`.
     STATEWIDE IS DEFINED HERE: `collaborative_type == "CCC Collaborative"`
     (138 unified titles as of 2026-08-05).
  2. credential_reference_data.js — the Common Exhibit Reference. Carries the
     per-college articulation detail (C-ID/M-ID → local subj/num → colleges),
     which is the only place the "which COURSE at which COLLEGE" answer lives.

⚠ The two disagree on `statewide`, and the disagreement is not symmetric: the
credential reference flags only **84** titles, a strict SUBSET of the adoption
file's 138. The 54-title delta is the newer statewide cohort — CSLB contractor
licences, Carpenters Apprenticeship, NCCER, child development — i.e. exactly the
building-trades rows a workforce partner's list is full of. So this tool takes
statewide from the ADOPTION file. Re-check that subset relationship if either
generator changes (`tests/partner_crosswalk_test.py` guards it).

ADOPTER SEMANTICS
-----------------
`adopters` is the UNION of `adopter_names` across every exhibit record sharing a
unified title — statewide adoptions AND local articulations of the same
credential. That is deliberate: the partner's question is "where can my student
get credit for this?", and a local articulation answers it just as well as a
statewide one. It means a credential's adopter count here can EXCEED the count
on the statewide tab, which shows the CCC-Collaborative record alone.

OUTPUT (kb/partner_crosswalk_out/<date>-<slug>/)
------------------------------------------------
  <date>_<Slug>_Occupation_CPL_Crosswalk.xlsx   the distributable deliverable
  summary.json                                  per-occupation status + counts
  unmapped.json                                 occupations with no map entry
                                                (the curator worklist)

The workbook is a REGENERABLE ARTIFACT — per the repo artifact policy it is not
committed; re-run the tool. `summary.json`/`unmapped.json` are small and are
the run receipt.
"""
import argparse
import csv
import datetime
import json
import os
import re
import sys
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MAP_PATH = os.path.join(ROOT, "kb", "occupation_credential_map.json")
REGION_PATH = os.path.join(ROOT, "kb", "partner_crosswalk_regions.json")
OUT_ROOT = os.path.join(ROOT, "kb", "partner_crosswalk_out")

TIER_LABEL = {"D": "Direct", "R": "Related"}

# AP/CLEP/DSST-style credit-by-exam. Used only to split a college's portfolio
# into "academic exam" vs "career/technical" on the Regional Capacity sheet —
# the distinction a workforce partner actually cares about.
ACADEMIC_EXAM = re.compile(r"^(AP |CLEP|DSST|IB |UExcel|Excelsior|DLPT|Cambridge|A-Level)", re.I)


# ---------------------------------------------------------------- loading ---
def load_window_json(rel_path):
    """Parse a `window.X = {...};` data file into a dict."""
    path = os.path.join(ROOT, rel_path)
    with open(path, encoding="utf-8") as fh:
        s = fh.read()
    s = s[s.index("{"):].rstrip().rstrip(";")
    return json.loads(s)


def build_credential_index():
    """Join the adoption view + the Common Exhibit Reference, keyed by unified title.

    Returns {title: {...}} — see the module docstring for adopter semantics.
    """
    exhibits = load_window_json("statewide_data.js")["exhibits"]
    cref = load_window_json("credential_reference_data.js")["unified_titles"]

    idx = {}

    def blank(title, issuer=""):
        return {
            "title": title, "statewide": False, "cref_statewide": False,
            "issuer": issuer, "cpl_types": set(), "disciplines": set(),
            "adopters": set(), "potential": set(), "credit_recs": [],
            "raw_variants": set(), "college_courses": defaultdict(set),
            "ccc_rec": "", "discipline": "",
        }

    for e in exhibits:
        t = e["unified_title"]
        r = idx.setdefault(t, blank(t, e.get("issuing_agency") or ""))
        if e.get("collaborative_type") == "CCC Collaborative":
            r["statewide"] = True
        if e.get("cpl_type"):
            r["cpl_types"].add(e["cpl_type"])
        if e.get("discipline"):
            r["disciplines"].add(e["discipline"])
        r["adopters"].update(e.get("adopter_names") or [])
        r["potential"].update(e.get("potential_names") or [])
        r["raw_variants"].update(x for x in (e.get("raw_titles") or []) if x)
        for cr in (e.get("credit_recs") or []):
            r["credit_recs"].append((cr.get("course", ""), cr.get("credit", "")))
        if not r["issuer"]:
            r["issuer"] = e.get("issuing_agency") or ""

    for u in cref:
        t = u["ut"]
        r = idx.setdefault(t, blank(t, u.get("issuer") or ""))
        r["cref_statewide"] = bool(u.get("statewide"))
        r["ccc_rec"] = u.get("ccc_rec") or ""
        r["discipline"] = u.get("disc_modal") or ""
        if not r["issuer"]:
            r["issuer"] = u.get("issuer") or ""
        r["potential"].update(u.get("potential_colleges") or [])
        r["cpl_types"].update(u.get("cpl_types") or [])
        r["raw_variants"].update(v.get("r", "") for v in (u.get("raw_variants") or []) if v.get("r"))
        for a in (u.get("articulations") or []):
            for loc in (a.get("local") or []):
                label = ("%s %s — %s" % (loc.get("subj", ""), loc.get("num", ""),
                                         loc.get("t", ""))).strip(" —")
                for col in (loc.get("colleges") or []):
                    r["college_courses"][col].add(label)

    for r in idx.values():
        r["adopters"] = sorted(r["adopters"])
        r["potential"] = sorted(r["potential"])
        r["cpl_types"] = sorted(r["cpl_types"])
        r["raw_variants"] = sorted(r["raw_variants"])
        r["college_courses"] = {k: sorted(v) for k, v in r["college_courses"].items()}
        if not r["discipline"] and r["disciplines"]:
            r["discipline"] = sorted(r["disciplines"])[0]
        del r["disciplines"]
    return idx


# ------------------------------------------------------------ occupations ---
def norm_key(s):
    """Normalized occupation key — the join key for the shared map.

    Collapses case, punctuation and whitespace so a partner's "Plumber",
    "PLUMBER" and "plumber " all reuse one curated ruling.
    """
    return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()


def read_occupations(path, column, sheet):
    """Read a partner list from .xlsx / .csv / .tsv / .txt.

    `column` is 1-indexed to match what a person reads off a spreadsheet.
    Returns [(display_label, times_listed)] in first-seen order.
    """
    ext = os.path.splitext(path)[1].lower()
    values = []
    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            sys.exit("openpyxl required to read %s — pip install openpyxl" % ext)
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            if column <= len(row) and row[column - 1]:
                values.append(str(row[column - 1]).strip())
    elif ext in (".csv", ".tsv"):
        delim = "\t" if ext == ".tsv" else ","
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for row in csv.reader(fh, delimiter=delim):
                if column <= len(row) and row[column - 1].strip():
                    values.append(row[column - 1].strip())
    else:
        with open(path, encoding="utf-8") as fh:
            values = [ln.strip() for ln in fh if ln.strip()]

    # Drop a header row if it looks like one.
    if values and norm_key(values[0]) in (
            "occupation", "occupations", "program name", "program", "job title",
            "title", "occupation title", "program title"):
        values = values[1:]

    order, counts = [], defaultdict(int)
    labels = {}
    for v in values:
        k = norm_key(v)
        if not k:
            continue
        if k not in labels:
            labels[k] = v
            order.append(k)
        counts[k] += 1
    return [(labels[k], counts[k], k) for k in order]


def load_occupation_map():
    if not os.path.exists(MAP_PATH):
        return {"_note": "", "occupations": {}}
    with open(MAP_PATH, encoding="utf-8") as fh:
        return json.load(fh)


def load_region(preset, explicit):
    if explicit:
        return [c.strip() for c in explicit.split(",") if c.strip()]
    if not preset:
        return []
    if not os.path.exists(REGION_PATH):
        sys.exit("region preset %r requested but %s is missing" % (preset, REGION_PATH))
    with open(REGION_PATH, encoding="utf-8") as fh:
        presets = json.load(fh).get("presets", {})
    if preset not in presets:
        sys.exit("unknown region preset %r — have: %s" % (preset, ", ".join(sorted(presets))))
    return presets[preset]["colleges"]


# ----------------------------------------------------------------- resolve ---
def resolve(occupations, occ_map, idx):
    """Match occupations to credentials. Returns (rows, summary, unmapped, bad_titles).

    `bad_titles` are map entries naming a credential that no longer exists in the
    index — a re-mint or a title consolidation upstream. Surfaced, never silently
    dropped, because a vanished title means the map needs re-curating.
    """
    entries = occ_map.get("occupations", {})
    rows, summary, unmapped, bad_titles = [], [], [], []

    for label, times, key in occupations:
        entry = entries.get(key)
        if entry is None:
            unmapped.append({"occupation": label, "key": key, "times_listed": times})
            summary.append({"occupation": label, "key": key, "times_listed": times,
                            "status": "Not yet mapped", "credentials": 0, "direct": 0,
                            "statewide": 0, "colleges": 0, "region_colleges": 0})
            continue

        creds = entry.get("credentials", [])
        matched, colleges = [], set()
        n_direct = n_sw = 0
        for c in creds:
            title, tier = c["title"], c.get("tier", "R")
            r = idx.get(title)
            if r is None:
                bad_titles.append({"occupation": label, "title": title})
                continue
            matched.append((title, tier, r))
            colleges |= set(r["adopters"])
            if tier == "D":
                n_direct += 1
            if r["statewide"]:
                n_sw += 1

        if not creds:
            status = "No CPL found"
        elif not matched:
            status = "Mapping stale — credential titles missing"
        elif n_sw:
            status = "Statewide CPL available"
        else:
            status = "Local CPL only"

        rows.append({"occupation": label, "times_listed": times, "matched": matched})
        summary.append({"occupation": label, "key": key, "times_listed": times,
                        "status": status, "credentials": len(matched), "direct": n_direct,
                        "statewide": n_sw, "colleges": len(colleges),
                        "region_colleges": 0})
    return rows, summary, unmapped, bad_titles


# ------------------------------------------------------------------ output ---
def write_workbook(path, partner, rows, summary, idx, region, as_of, source_name):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    FONT = "Arial"
    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    SW_FILL = PatternFill("solid", fgColor="E2EFDA")
    GAP_FILL = PatternFill("solid", fgColor="FCE4E4")
    BODY = Font(name=FONT, size=10)
    TITLE = Font(name=FONT, size=14, bold=True, color="1F3864")
    HEAD = Font(name=FONT, size=11, bold=True, color="1F3864")
    WRAP = Alignment(vertical="top", wrap_text=True)
    CTR = Alignment(vertical="top", horizontal="center")
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def header(ws, hdr, row=1):
        ws.append(hdr) if row == 1 else None
        for i, _ in enumerate(hdr, 1):
            c = ws.cell(row=row, column=i)
            c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, WRAP, BORDER

    def style_body(ws, first_row, ncols, center_cols=()):
        for r in range(first_row, ws.max_row + 1):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.font, cell.border = BODY, BORDER
                cell.alignment = CTR if c in center_cols else WRAP

    def widths(ws, ws_widths):
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(ws_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = openpyxl.Workbook()

    # ---- Read Me ----
    ws = wb.active
    ws.title = "Read Me"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 112
    lines = [
        ("t", "Occupation → CPL Crosswalk for %s" % partner),
        ("", ""),
        ("h", "What this is"),
        ("p", "Every occupation on the %s list, matched to the Credit for Prior Learning (CPL) credit "
              "recommendations California Community Colleges have articulated on the Mapping Articulated "
              "Pathways (MAP) platform — and the colleges that have adopted each one." % partner),
        ("p", "A student holding one of these certifications, licences, or apprenticeship completions can "
              "request college credit for it at any college listed, without repeating the coursework."),
        ("", ""),
        ("h", "How to read it"),
        ("b", "Occupation Summary — one row per occupation. Start here."),
        ("b", "Crosswalk — one row per occupation × credential. The credential is what the student holds."),
        ("b", "College Detail — one row per occupation × credential × college, with the course(s) that "
              "college awards credit for. This is the row you hand a student."),
        ("b", "Regional Capacity — what the nearest colleges actually offer CPL for."),
        ("b", "Gaps — occupations with no CPL anywhere. These are the build-it opportunities."),
        ("", ""),
        ("h", "Key terms"),
        ("b", "Statewide credit recommendation (\"CCC Collaborative\") — vetted and published at the state "
              "level for ANY college to adopt. Portable and already reviewed."),
        ("b", "Local articulation — one college's own credit recommendation. Just as real for the student, "
              "but it applies only at that college."),
        ("b", "Match type — \"Direct\" means the credential is the occupation's own certification or "
              "licence. \"Related\" means it is commonly held in that occupation, or covers part of the "
              "role. Direct matches are the safe ones to advertise; Related ones are worth a conversation."),
        ("b", "Colleges — colleges with an articulation on record for that credential (statewide adopters "
              "and local articulations combined)."),
        ("", ""),
        ("h", "Source and currency"),
        ("b", "MAP platform statewide exhibit adoption data and the Common Exhibit Reference, as of %s." % as_of),
        ("b", "Live dashboard: https://cpl-initiative.github.io/cpl-project-tracker/"),
        ("b", "Generated %s by the CPL Initiative, California Community Colleges Chancellor's Office, "
              "from %s." % (datetime.date.today().isoformat(), source_name)),
        ("", ""),
        ("h", "Important caveats — please read before distributing"),
        ("b", "The occupation-to-credential match is subject-matter judgment, not an official crosswalk. "
              "Source lists carry job and apprenticeship titles, not credential names. Verify a Direct "
              "match before promising credit to a student."),
        ("b", "A college listed here has ARTICULATED the credential — it has not guaranteed an award. Each "
              "student still applies through that college's CPL process, and unit awards vary."),
        ("b", "\"No CPL found\" means no articulation exists in MAP today. It does not mean a college would "
              "refuse credit — it means nobody has published a recommendation yet."),
    ]
    r = 1
    for kind, txt in lines:
        c = ws.cell(row=r, column=2)
        if kind == "t":
            c.value, c.font = txt, TITLE
        elif kind == "h":
            c.value, c.font = txt, HEAD
        elif kind in ("b", "p"):
            c.value = ("•  " + txt) if kind == "b" else txt
            c.font, c.alignment = BODY, WRAP
            ws.row_dimensions[r].height = max(14, 13 * (1 + len(txt) // 105))
        r += 1

    # ---- Occupation Summary ----
    ws = wb.create_sheet("Occupation Summary")
    hdr = ["Occupation (as submitted)", "Times listed", "CPL status", "Credentials matched",
           "  of which Direct", "  of which Statewide", "Colleges offering"]
    if region:
        hdr.append("…in partner region")
    header(ws, hdr)
    for s in summary:
        row = [s["occupation"], s["times_listed"], s["status"], s["credentials"],
               s["direct"], s["statewide"], s["colleges"]]
        if region:
            row.append(s["region_colleges"])
        ws.append(row)
    style_body(ws, 2, len(hdr), center_cols=set(range(2, len(hdr) + 1)) - {3})
    for r in range(2, ws.max_row + 1):
        st = ws.cell(row=r, column=3)
        st.alignment = WRAP
        if st.value == "Statewide CPL available":
            st.fill = SW_FILL
        elif st.value in ("No CPL found", "Not yet mapped"):
            st.fill = GAP_FILL
    widths(ws, [46, 11, 22, 13, 12, 13, 12] + ([14] if region else []))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (chr(64 + len(hdr)), ws.max_row)

    # ---- Crosswalk ----
    ws = wb.create_sheet("Crosswalk")
    hdr = ["Occupation", "Match", "CPL opportunity (credential)", "Statewide?", "Issuing agency",
           "CPL type", "# colleges", "Colleges that have adopted it", "Typical credit awarded"]
    header(ws, hdr)
    for row in rows:
        for title, tier, r_ in row["matched"]:
            rec = r_["ccc_rec"] or (r_["credit_recs"][0][1] if r_["credit_recs"] else "")
            ws.append([row["occupation"], TIER_LABEL[tier], title,
                       "Yes" if r_["statewide"] else "No", r_["issuer"] or "—",
                       ", ".join(r_["cpl_types"]) or "—", len(r_["adopters"]),
                       "; ".join(r_["adopters"]) or "— (published, not yet adopted)",
                       rec or "—"])
    style_body(ws, 2, len(hdr), center_cols={7})
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=4).value == "Yes":
            ws.cell(row=r, column=4).fill = SW_FILL
    widths(ws, [34, 9, 44, 10, 34, 20, 9, 60, 40])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:I%d" % ws.max_row

    # ---- College Detail ----
    ws = wb.create_sheet("College Detail")
    hdr = ["Occupation", "CPL opportunity (credential)", "Statewide?", "Match", "College",
           "Course(s) that grant the credit at this college"]
    header(ws, hdr)
    region_rows = []
    for row in rows:
        for title, tier, r_ in row["matched"]:
            for col in r_["adopters"]:
                courses = r_["college_courses"].get(col, [])
                course_txt = "; ".join(courses) if courses else "(course detail not published)"
                ws.append([row["occupation"], title, "Yes" if r_["statewide"] else "No",
                           TIER_LABEL[tier], col, course_txt])
                if col in region:
                    region_rows.append([col, row["occupation"], title,
                                        "Yes" if r_["statewide"] else "No",
                                        TIER_LABEL[tier], course_txt])
    style_body(ws, 2, len(hdr))
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=3).value == "Yes":
            ws.cell(row=r, column=3).fill = SW_FILL
    widths(ws, [32, 42, 10, 9, 32, 66])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F%d" % ws.max_row

    # ---- Partner Region (only when a region is supplied) ----
    if region:
        ws = wb.create_sheet("Partner Region")
        ws["A1"] = ("CPL available at the %d colleges nearest %s — the realistic first stop for a "
                    "local student." % (len(region), partner))
        ws["A1"].font = Font(name=FONT, size=10, italic=True, color="1F3864")
        ws.merge_cells("A1:F1")
        hdr = ["College", "Occupation", "CPL opportunity (credential)", "Statewide?", "Match",
               "Course(s) that grant the credit"]
        ws.append(hdr)
        header(ws, hdr, row=2)
        region_rows.sort(key=lambda x: (x[0], x[1], x[2]))
        for row in region_rows:
            ws.append(row)
        style_body(ws, 3, len(hdr))
        for r in range(3, ws.max_row + 1):
            if ws.cell(row=r, column=4).value == "Yes":
                ws.cell(row=r, column=4).fill = SW_FILL
        widths(ws, [28, 32, 42, 10, 9, 62])
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = "A2:F%d" % ws.max_row

        # ---- Regional Capacity ----
        ws = wb.create_sheet("Regional Capacity")
        ws["A1"] = ("What the nearest colleges actually offer CPL for. \"Academic exam\" = AP / CLEP / "
                    "DSST credit-by-exam. \"Career/technical\" = industry certifications, licences and "
                    "apprenticeships — the kind your students hold.")
        ws["A1"].font = Font(name=FONT, size=10, italic=True, color="1F3864")
        ws.merge_cells("A1:E1")
        ws.row_dimensions[1].height = 28
        hdr = ["College", "Total credentials", "Academic exam (AP/CLEP)", "Career / technical",
               "  of which statewide"]
        ws.append(hdr)
        header(ws, hdr, row=2)
        cap = []
        for col in region:
            ts = [t for t, r_ in idx.items() if col in r_["adopters"]]
            acad = [t for t in ts if ACADEMIC_EXAM.match(t)]
            career = [t for t in ts if not ACADEMIC_EXAM.match(t)]
            sw = [t for t in career if idx[t]["statewide"]]
            cap.append([col, len(ts), len(acad), len(career), len(sw)])
        cap.sort(key=lambda x: -x[3])
        for row in cap:
            ws.append(row)
        style_body(ws, 3, 5, center_cols={2, 3, 4, 5})
        for r in range(3, ws.max_row + 1):
            if ws.cell(row=r, column=4).value <= 1:
                ws.cell(row=r, column=1).fill = GAP_FILL
                ws.cell(row=r, column=4).fill = GAP_FILL
        widths(ws, [30, 17, 22, 18, 19])
        ws.freeze_panes = "A3"
        thin_ = [c for c in cap if c[3] <= 1]
        if thin_:
            note = ws.max_row + 2
            cell = ws.cell(row=note, column=1)
            cell.value = ("Read before planning referrals: %s %s effectively no career/technical CPL — "
                          "their portfolios are almost entirely AP and CLEP. The nearest real career-CPL "
                          "capacity is %s." % (
                              ", ".join(c[0] for c in thin_),
                              "has" if len(thin_) == 1 else "have",
                              cap[0][0] if cap else "—"))
            cell.font = Font(name=FONT, size=10, bold=True, color="C00000")
            cell.alignment = WRAP
            ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=5)
            ws.row_dimensions[note].height = 60

    # ---- Gaps ----
    ws = wb.create_sheet("Gaps")
    ws["A1"] = ("Occupations with no CPL credit recommendation anywhere in the California Community "
                "Colleges system. Each is an opportunity to build one.")
    ws["A1"].font = Font(name=FONT, size=10, italic=True, color="C00000")
    ws.merge_cells("A1:C1")
    hdr = ["Occupation", "Times listed", "Finding"]
    ws.append(hdr)
    header(ws, hdr, row=2)
    for s in summary:
        if s["status"] in ("No CPL found", "Not yet mapped"):
            ws.append([s["occupation"], s["times_listed"],
                       "No credential in MAP matches this occupation"
                       if s["status"] == "No CPL found"
                       else "Not yet curated into kb/occupation_credential_map.json"])
    style_body(ws, 3, 3, center_cols={2})
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=1).fill = GAP_FILL
    widths(ws, [52, 13, 56])
    ws.freeze_panes = "A3"

    wb.save(path)


# -------------------------------------------------------------------- main ---
def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--input", required=True, help="partner occupation list (.xlsx/.csv/.tsv/.txt)")
    ap.add_argument("--partner", required=True, help='display name, e.g. "San Joaquin COE"')
    ap.add_argument("--slug", required=True, help="short slug for the output directory, e.g. sjcoe")
    ap.add_argument("--column", type=int, default=1, help="1-indexed column holding the occupation (default 1)")
    ap.add_argument("--sheet", default=None, help="worksheet name (xlsx only; default first sheet)")
    ap.add_argument("--region-preset", default=None, help="named college region from kb/partner_crosswalk_regions.json")
    ap.add_argument("--region", default=None, help="comma-separated college names (overrides --region-preset)")
    ap.add_argument("--out", default=None, help="output directory (default kb/partner_crosswalk_out/<date>-<slug>)")
    args = ap.parse_args()

    occupations = read_occupations(args.input, args.column, args.sheet)
    if not occupations:
        sys.exit("no occupations read from %s (column %d) — check --column/--sheet"
                 % (args.input, args.column))

    idx = build_credential_index()
    occ_map = load_occupation_map()
    region = load_region(args.region_preset, args.region)

    rows, summary, unmapped, bad_titles = resolve(occupations, occ_map, idx)

    # region college counts, now that we know each occupation's colleges
    if region:
        region_set = set(region)
        by_occ = {r["occupation"]: r for r in rows}
        for s in summary:
            row = by_occ.get(s["occupation"])
            if not row:
                continue
            cols = set()
            for _t, _tier, r_ in row["matched"]:
                cols |= set(r_["adopters"])
            s["region_colleges"] = len(cols & region_set)

    today = datetime.date.today().isoformat()
    out_dir = args.out or os.path.join(OUT_ROOT, "%s-%s" % (today, args.slug))
    os.makedirs(out_dir, exist_ok=True)

    as_of = load_window_json("statewide_data.js").get("generated_at", "unknown")
    xlsx = os.path.join(out_dir, "%s_%s_Occupation_CPL_Crosswalk.xlsx"
                        % (today.replace("-", ""), args.slug.upper()))
    write_workbook(xlsx, args.partner, rows, summary, idx, region, as_of,
                   os.path.basename(args.input))

    status_counts = defaultdict(int)
    for s in summary:
        status_counts[s["status"]] += 1

    with open(os.path.join(out_dir, "summary.json"), "w", encoding="utf-8") as fh:
        json.dump({
            "_partner": args.partner, "_slug": args.slug, "_generated": today,
            "_source_list": os.path.basename(args.input),
            "_data_as_of": as_of,
            "_region_preset": args.region_preset,
            "_statewide_definition": "statewide_data.js collaborative_type == 'CCC Collaborative'",
            "counts": dict(status_counts),
            "occupations_unique": len(summary),
            "occupation_credential_matches": sum(len(r["matched"]) for r in rows),
            "stale_credential_titles": bad_titles,
            "occupations": summary,
        }, fh, indent=1)

    with open(os.path.join(out_dir, "unmapped.json"), "w", encoding="utf-8") as fh:
        json.dump({
            "_note": "Occupations with no entry in kb/occupation_credential_map.json. "
                     "THIS IS THE CURATOR WORKLIST — add rulings here and coverage compounds "
                     "for every future partner.",
            "_partner": args.partner,
            "count": len(unmapped),
            "occupations": unmapped,
        }, fh, indent=1)

    print("partner            : %s" % args.partner)
    print("occupations (uniq) : %d" % len(summary))
    for k in sorted(status_counts):
        print("  %-30s %d" % (k, status_counts[k]))
    print("credential matches : %d" % sum(len(r["matched"]) for r in rows))
    if bad_titles:
        print("⚠ stale map titles : %d (credential no longer in the index — re-curate)"
              % len(bad_titles))
    if unmapped:
        print("⚠ unmapped         : %d → %s/unmapped.json" % (len(unmapped), out_dir))
    print("workbook           : %s" % xlsx)


if __name__ == "__main__":
    main()
