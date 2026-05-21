"""
One-shot enrichment pass: join credit status + synthetic units + TOP code +
noncredit program category onto the minted M-IDs (STAGING).

WHY
The minted M-ID layer was derived from a per-college course list that lacked
funding fields. MAP later provided a richer (deduplicated) "Unique Course List"
with UnitValue, CreditType, Non_Credit_Category, and TopCode. This pass JOINS
those fields onto the existing M-IDs by (subject, course_number) — it does NOT
regenerate the M-IDs, because the new file is deduplicated and would collapse
the per-college corroboration counts the M-IDs are built on.

CREDIT STATUS — derived from the CreditType column (the funding type):
  'Credit Course'                          -> Credit
  'Other Noncredit Enhanced Funding'       -> Noncredit Enhanced
  'Workforce Preparation Enhanced Funding' -> Noncredit Enhanced
  'Non-Enhanced Funding'                   -> Noncredit
  blank / unrecognized                     -> by UnitValue: >0 Credit, else Noncredit
(Non_Credit_Category is the CDCP *program* type — Short-term Vocational, ESL,
Older Adults, … — carried separately as `noncredit_category`, not the funding
signal.)

AGGREGATION — an M-ID groups many member courses (and the new file may carry
several rows per course). We gather every matched row for all of an M-ID's
members and take the MODAL value: credit_status (+ a `credit_status_mixed` flag
when members disagree), typical_units (synthetic unit value — best-judgment
representative), top_code, noncredit_category. Unmatched M-IDs keep null.

Updates in place (additive fields): kb/coci_minted_courses.json and
kb/coci_minted_singletons.json. memberships, the unified crosswalk, and the
curated common_courses.json / course_crosswalk.json are untouched.

Run from repo root:
  python3 kb/_join_credit_status.py /path/to/Unique_Course_List_from_MAP.xlsx
"""
import json
import os
import re
import sys
from collections import Counter, defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
CATALOG = os.path.join(HERE, "coci_minted_courses.json")
SINGLETONS = os.path.join(HERE, "coci_minted_singletons.json")
MEMBERSHIPS = os.path.join(HERE, "coci_minted_memberships.json")

JOINED_AT = "2026-05-21"
JOINED_BY = "claude-opus-4-7 (credit-status / units / topcode join)"
SOURCE_DESC = ('MAP "20260521 Unique Course List from MAP.xlsx" sheet Sheet2 '
               "(adds UnitValue, CreditType, Non_Credit_Category, TopCode)")

ENHANCED = {"Other Noncredit Enhanced Funding", "Workforce Preparation Enhanced Funding"}
BLANKS = {"", "(blank)", "not applicable", "non applicable", "null", "n/a"}


def jkey(subj, num):
    return (re.sub(r"\s+", " ", str(subj)).strip().upper(),
            re.sub(r"\s+", " ", str(num)).strip().upper())


def isblank(v):
    return v is None or str(v).strip().lower() in BLANKS


def parse_units(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def derive_status(credit_type, units):
    ct = "" if credit_type is None else str(credit_type).strip()
    if ct == "Credit Course":
        return "Credit"
    if ct in ENHANCED:
        return "Noncredit Enhanced"
    if ct == "Non-Enhanced Funding":
        return "Noncredit"
    u = parse_units(units)  # blank / unrecognized CreditType -> units rule
    return "Credit" if (u is not None and u > 0) else "Noncredit"


def build_lookup(xlsx_path):
    import openpyxl
    ws = openpyxl.load_workbook(xlsx_path, read_only=True)["Sheet2"]
    rows = ws.iter_rows(values_only=True)
    H = next(rows)
    I = {h: i for i, h in enumerate(H)}
    lut = defaultdict(lambda: {"status": Counter(), "units": Counter(),
                               "topcode": Counter(), "ncc": Counter()})
    n = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        n += 1
        key = jkey(r[I["Subject"]], r[I["Course_Number"]])
        ct, uv = r[I["CreditType"]], r[I["UnitValue"]]
        rec = lut[key]
        rec["status"][derive_status(ct, uv)] += 1
        u = parse_units(uv)
        if u is not None:
            rec["units"][u] += 1
        tc = r[I["TopCode"]]
        if not isblank(tc):
            # store the numeric TOP code only; the label is in TOP_Code_Lookup.xlsx
            rec["topcode"][str(tc).split(":")[0].strip()] += 1
        ncc = r[I["Non_Credit_Category"]]
        if not isblank(ncc) and str(ncc).strip() != "Credit Course":
            rec["ncc"][str(ncc).strip()] += 1
    return lut, n


def aggregate(member_keys, lut):
    """Merge modal attributes across all of an M-ID's matched member rows."""
    status, units, tc, ncc = Counter(), Counter(), Counter(), Counter()
    matched = 0
    for k in member_keys:
        if k in lut:
            matched += 1
            r = lut[k]
            status += r["status"]; units += r["units"]
            tc += r["topcode"]; ncc += r["ncc"]
    if not status:
        return None
    return {
        "credit_status": status.most_common(1)[0][0],
        "credit_status_mixed": len(status) > 1,
        "typical_units": units.most_common(1)[0][0] if units else None,
        "top_code": tc.most_common(1)[0][0] if tc else None,
        # TOP codes vary often for the same course (colleges assign them with
        # discretion + no definitive guidance), so surface the spread rather
        # than silently trusting the modal pick.
        "top_code_mixed": len(tc) > 1,
        "top_code_distribution": dict(tc.most_common()) if len(tc) > 1 else None,
        "noncredit_category": ncc.most_common(1)[0][0] if ncc else None,
        "_matched": matched,
    }


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: python3 kb/_join_credit_status.py /path/to/Unique_Course_List.xlsx")
    lut, n_rows = build_lookup(sys.argv[1])

    dist = Counter()
    n_matched_cat = n_mixed = 0

    # --- corroborated catalog (members live in the memberships file) ---
    cat_doc = json.load(open(CATALOG))
    members = json.load(open(MEMBERSHIPS))["memberships"]
    for cid, v in cat_doc["courses"].items():
        keys = [jkey(m["subject"], m["course_number"]) for m in members.get(cid, [])]
        agg = aggregate(keys, lut)
        if agg:
            n_matched_cat += 1
            n_mixed += agg["credit_status_mixed"]
            dist[agg["credit_status"]] += 1
            v["credit_status"] = agg["credit_status"]
            v["credit_status_mixed"] = agg["credit_status_mixed"]
            v["typical_units"] = agg["typical_units"]
            v["top_code"] = agg["top_code"]
            v["top_code_mixed"] = agg["top_code_mixed"]
            v["top_code_distribution"] = agg["top_code_distribution"]
            v["noncredit_category"] = agg["noncredit_category"]
        else:
            dist["(unmatched)"] += 1
            v["credit_status"] = None
            v["credit_status_mixed"] = False
            v["top_code"] = None
            v["top_code_mixed"] = False
            v["top_code_distribution"] = None
            v["noncredit_category"] = None

    # --- singletons (single embedded member; lean schema -> omit null fields) ---
    sg_doc = json.load(open(SINGLETONS))
    sdist = Counter()
    n_matched_sg = 0
    for cid, v in sg_doc["courses"].items():
        agg = aggregate([jkey(v["subject"], v["course_number"])], lut)
        for f in ("credit_status", "credit_status_mixed", "typical_units",
                  "top_code", "top_code_mixed", "top_code_distribution",
                  "noncredit_category"):
            v.pop(f, None)
        if agg:
            n_matched_sg += 1
            sdist[agg["credit_status"]] += 1
            v["credit_status"] = agg["credit_status"]
            if agg["credit_status_mixed"]:
                v["credit_status_mixed"] = True
            if agg["typical_units"] is not None:
                v["typical_units"] = agg["typical_units"]
            if agg["top_code"]:
                v["top_code"] = agg["top_code"]
            if agg["top_code_mixed"]:
                v["top_code_mixed"] = True
            if agg["noncredit_category"]:
                v["noncredit_category"] = agg["noncredit_category"]
        else:
            sdist["(unmatched)"] += 1

    join_note = {
        "_credit_join_source": SOURCE_DESC,
        "_credit_join_at": JOINED_AT,
        "_credit_join_by": JOINED_BY,
        "_credit_status_rule": ("CreditType: 'Credit Course'->Credit; 'Other Noncredit "
                                "Enhanced Funding'/'Workforce Preparation Enhanced Funding'->"
                                "Noncredit Enhanced; 'Non-Enhanced Funding'->Noncredit; blank->"
                                "by UnitValue (>0 Credit else Noncredit). typical_units/top_code/"
                                "noncredit_category are the modal value across matched member "
                                "rows; credit_status_mixed flags members that disagree."),
        "_topcode_note": ("TOP codes vary often for the same course — colleges assign them in "
                          "COCI with discretion and no definitive guidance for ambiguous cases. "
                          "top_code is the modal (plurality) pick; top_code_mixed flags spread; "
                          "the catalog also carries top_code_distribution (code->count) when mixed."),
        "catalog_matched": n_matched_cat,
        "catalog_credit_status_distribution": dict(dist),
        "catalog_mixed_status": n_mixed,
        "singletons_matched": n_matched_sg,
        "singletons_credit_status_distribution": dict(sdist),
    }
    cat_doc["_credit_join"] = join_note
    sg_doc["_credit_join"] = join_note
    # document the new singleton defaults
    sg_doc.setdefault("_record_defaults", {})
    sg_doc["_record_defaults"]["credit_status"] = None
    sg_doc["_record_defaults"]["credit_status_mixed"] = False
    sg_doc["_record_defaults"]["top_code"] = None
    sg_doc["_record_defaults"]["top_code_mixed"] = False
    sg_doc["_record_defaults"]["noncredit_category"] = None

    for path, doc in ((CATALOG, cat_doc), (SINGLETONS, sg_doc)):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(doc, f, indent=2, ensure_ascii=False)
            f.write("\n")

    print(f"new-file rows scanned: {n_rows}; lookup keys: {len(lut)}")
    print(f"catalog M-IDs matched: {n_matched_cat}/{len(cat_doc['courses'])} "
          f"(mixed-status: {n_mixed})")
    print(f"  catalog credit_status: {dict(dist)}")
    print(f"singleton M-IDs matched: {n_matched_sg}/{len(sg_doc['courses'])}")
    print(f"  singleton credit_status: {dict(sdist)}")


if __name__ == "__main__":
    main()
