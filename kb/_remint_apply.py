"""
Re-mint generator — STEP 1a: re-key the minted identity layer + rebuild
memberships at College/CourseControlNumber granularity.

Writes a PREVIEW under kb/remint_out/ (does NOT overwrite the live coci_*.json
and does NOT touch Supabase). Outputs:
  coci_minted_courses.json      — re-keyed corroborated catalog (new <SUBJ4> M#### ids)
  coci_minted_memberships.json  — new id -> member college courses, each carrying its
                                  OWN College/CourseControlNumber + official C-ID/CCN
  coci_minted_singletons.json   — re-keyed single-college identities (<SUBJ4> M<band><d><LL>)
  alias_map.json                — authoritative old M-ID -> new id (supersedes the dry-run map)
  promotions.json               — split manifest: title rows that carry an official
                                  C-ID/CCN (the remnant stays minted; these promote in 1b)

Keying mirrors CCN's 4-digit SUBJ C#### (leading digit = band): corroborated ->
"<SUBJ4> M<band><seq:03d>" (band 9 noncredit / 1 credit), stand-alone ->
"<SUBJ4> M<band><d><LL>" (band + 1 digit + 2 letters, cap 6,760/bucket). Banding
is credit_status-only. SUBJ4 = synthetic 4-letter subject (MAP surrogate, not the
official CCN list). All other per-identity metadata (discipline + inference
provenance, descriptions, confidence, *_mixed flags) is carried 1:1 from the
same-title old identity so the inference/curation work is preserved.
"""
import json
import os
import re
from collections import Counter, defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(HERE, "reference", "coci_course_list.xlsx")
OUT_DIR = os.path.join(HERE, "remint_out")
GENERATED_AT = "2026-05-22"
GENERATED_BY = "kb/_remint_apply.py (CourseControlNumber re-mint, step 1a)"

SENT = {None, "", "null", "n/a", "na", "not applicable", "(blank)", "blank", "none"}
STOP_RE = re.compile("|".join([
    r"independent stud", r"directed stud", r"dir stud", r"special stud",
    r"special project", r"special topic", r"selected topic", r"special problem",
    r"work experience", r"cooperative (work )?(education|experience)", r"coop ",
    r"internship", r"\bintern\b", r"supervised tutoring",
    r"student instructional assistant", r"service learning", r"occupational work",
    r"tutoring", r"practicum", r"fieldwork", r"field work", r"field experience",
    r"field study", r"directed practice", r"clinical practice", r"cooperative work",
    r"work based learning", r"work-based", r"on the job", r"apprenticeship",
    r"seminar$", r"^seminar", r"special assignment", r"volunteer", r"community service",
]))
CODE_RE = re.compile(r"^[a-z]{1,6} ?\d{1,4}[a-z]?$")
CID_RE = re.compile(r"^[A-Z]+(?:-[A-Z]+)? \d{2,4}[A-Z]?(?: [A-Z]{1,2})*$")
LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s.lower() in SENT else s


def ntitle(t):
    if t is None:
        return ""
    t = re.sub(r"\s+", " ", str(t)).strip().lower()
    t = re.sub(r"[^a-z0-9 ]", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def subj4(subject):
    return (re.sub(r"[^A-Za-z]", "", str(subject or "")).upper()[:4] or "MISC")


def credit_status(ct, units):
    c = (ct or "").strip().lower()
    if c == "credit course":
        return "Credit"
    if c in ("other noncredit enhanced funding", "workforce preparation enhanced funding"):
        return "Noncredit Enhanced"
    if c == "non-enhanced funding":
        return "Noncredit"
    try:
        u = float(units)
    except (TypeError, ValueError):
        u = 0
    return "Credit" if u > 0 else "Noncredit"


def parse_cids(raw):
    raw = clean(raw)
    if not raw:
        return []
    out = []
    for part in raw.split(","):
        ded = []
        for t in part.split():
            if t == "000" or (ded and t == ded[-1]):
                continue
            ded.append(t)
        cid = " ".join(ded)
        if cid and CID_RE.match(cid):
            out.append(cid)
        elif cid:
            out.append(cid)  # keep; flagged by caller if it fails CID_RE
    return out


def sing_code(n):  # 0-based seq -> "<d><L><L>" (base 10*26*26 = 6,760)
    d, r = divmod(n, 26 * 26)
    l1, l2 = divmod(r, 26)
    return f"{d}{LETTERS[l1]}{LETTERS[l2]}"


def num_units(u):
    try:
        return float(u)
    except (TypeError, ValueError):
        return None


def build_from_raw():
    """Stream raw list once. ntitle -> {minted:[memberdict], official:{oid:[memberdict]}}."""
    import openpyxl
    wb = openpyxl.load_workbook(RAW, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it)
    idx = defaultdict(lambda: {"minted": [], "official": defaultdict(list)})
    for (College, CCNum, Subj, Cnum, Title, Units, CreditType,
         NCcat, Top, CID, Desc, CCNcommon) in it:
        nt = ntitle(Title)
        if not nt or len(nt) < 4 or STOP_RE.search(nt) or CODE_RE.match(nt):
            continue
        ccn = clean(CCNcommon)
        cids = parse_cids(CID)
        m = {
            "college": clean(College), "control_number": clean(CCNum),
            "subject": (str(Subj).strip() if Subj is not None else ""),
            "course_number": (str(Cnum).strip() if Cnum is not None else ""),
            "title": clean(Title), "units": num_units(Units),
            "credit_status": credit_status(CreditType, Units),
            "top_code": clean(Top), "noncredit_category": clean(NCcat),
        }
        e = idx[nt]
        if ccn:
            e["official"][f"CCN:{ccn}"].append(m)
        elif cids:
            for c in cids:
                e["official"][f"C-ID:{c}"].append(m)
        else:
            e["minted"].append(m)
    wb.close()
    return idx


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    old_cat = json.load(open(os.path.join(HERE, "coci_minted_courses.json")))["courses"]
    old_sing = json.load(open(os.path.join(HERE, "coci_minted_singletons.json")))["courses"]
    old_by_title = {}
    for mid, r in old_cat.items():
        old_by_title[ntitle(r.get("common_title"))] = ("corr", mid, r)
    for mid, r in old_sing.items():
        old_by_title[ntitle(r.get("common_title"))] = ("sing", mid, r)

    idx = build_from_raw()

    # ── Pass 1: assign new keys to every title that retains a minted remnant ──
    meta = {}  # nt -> (s4, band, corroborated?, credit_status)
    for nt, e in idx.items():
        if not e["minted"]:
            continue
        cs = Counter(m["credit_status"] for m in e["minted"]).most_common(1)[0][0]
        s4 = subj4(Counter(m["subject"] for m in e["minted"]).most_common(1)[0][0])
        band = "9" if cs in ("Noncredit", "Noncredit Enhanced") else "1"
        meta[nt] = (s4, band, len(e["minted"]) >= 2, cs)
    seq_corr, seq_sing = defaultdict(int), defaultdict(int)
    new_key = {}
    for nt in sorted(meta):
        s4, band, corr, cs = meta[nt]
        if corr:
            seq_corr[(s4, band)] += 1
            new_key[nt] = f"{s4} M{band}{seq_corr[(s4, band)]:03d}"
        else:
            new_key[nt] = f"{s4} M{band}{sing_code(seq_sing[(s4, band)])}"
            seq_sing[(s4, band)] += 1

    # ── Pass 2: build catalog / singletons / memberships, carrying metadata ──
    courses, memberships, singletons, alias, promotions = {}, {}, {}, {}, {}
    tier_flips = Counter()
    for nt in sorted(meta):
        s4, band, corr, cs = meta[nt]
        e = idx[nt]
        members = e["minted"]
        nid = new_key[nt]
        old = old_by_title.get(nt)
        old_rec = old[2] if old else {}
        if old:
            alias[old[1]] = nid
            tier_flips[(old[0], "corr" if corr else "sing")] += 1

        subj_counts = Counter(m["subject"] for m in members)
        title_counts = Counter(m["title"] for m in members if m["title"])
        units_counts = Counter(m["units"] for m in members if m["units"] is not None)
        common_title = old_rec.get("common_title") or (title_counts.most_common(1)[0][0] if title_counts else nt)

        if corr:
            courses[nid] = {
                "course_id": nid, "id_system": "M-ID", "ccn_id": None, "c_id": None,
                "common_title": common_title,
                "common_title_source": old_rec.get("common_title_source", "local catalog (representative/modal)"),
                "description": old_rec.get("description"),
                "description_source": old_rec.get("description_source"),
                "subject": subj_counts.most_common(1)[0][0],
                "subject_4letter": s4,
                "discipline": old_rec.get("discipline"),
                "discipline_provisional": old_rec.get("discipline_provisional", subj_counts.most_common(1)[0][0]),
                "typical_units": (units_counts.most_common(1)[0][0] if units_counts else old_rec.get("typical_units")),
                "confidence": old_rec.get("confidence", 0.68),
                "corroboration_members": len(members),
                "subject_spread": len(subj_counts),
                "source_college_count": len({m["college"] for m in members}),
                "classified_at": old_rec.get("classified_at", GENERATED_AT),
                "classified_by": old_rec.get("classified_by", GENERATED_BY),
                "reviewed_at": old_rec.get("reviewed_at"),
                "reviewed_by": old_rec.get("reviewed_by"),
                "_notes": old_rec.get("_notes"),
                "credit_status": cs,
                "credit_status_mixed": len({m["credit_status"] for m in members}) > 1,
                "top_code": old_rec.get("top_code") or (Counter(m["top_code"] for m in members if m["top_code"]).most_common(1)[0][0] if any(m["top_code"] for m in members) else None),
                "noncredit_category": old_rec.get("noncredit_category"),
                "top_code_mixed": len({m["top_code"] for m in members if m["top_code"]}) > 1,
                "top_code_distribution": old_rec.get("top_code_distribution"),
                "noncredit_category_mixed": old_rec.get("noncredit_category_mixed"),
                "noncredit_category_distribution": old_rec.get("noncredit_category_distribution"),
                "discipline_source": old_rec.get("discipline_source"),
                "discipline_confidence": old_rec.get("discipline_confidence"),
                "discipline_inferred_at": old_rec.get("discipline_inferred_at"),
                "_remint_from": old[1] if old else None,
            }
            memberships[nid] = [
                {"college": m["college"], "control_number": m["control_number"],
                 "subject": m["subject"], "course_number": m["course_number"],
                 "units": m["units"], "credit_status": m["credit_status"],
                 "top_code": m["top_code"]}
                for m in sorted(members, key=lambda x: (x["college"] or "", x["control_number"] or ""))
            ]
        else:
            m0 = members[0]
            singletons[nid] = {
                "course_id": nid, "common_title": common_title,
                "subject": m0["subject"], "subject_4letter": s4,
                "course_number": m0["course_number"],
                "college": m0["college"], "control_number": m0["control_number"],
                "credit_status": cs, "typical_units": m0["units"],
                "top_code": old_rec.get("top_code") or m0["top_code"],
                "discipline": old_rec.get("discipline"),
                "discipline_source": old_rec.get("discipline_source"),
                "discipline_confidence": old_rec.get("discipline_confidence"),
                "discipline_inferred_at": old_rec.get("discipline_inferred_at"),
                "_remint_from": old[1] if old else None,
            }

        if e["official"]:  # split: capture promotions for step 1b
            promotions[nid] = {
                "minted_remnant_members": len(members),
                "official_targets": {
                    oid: {"members": len(ms),
                          "colleges": sorted({m["college"] for m in ms})}
                    for oid, ms in sorted(e["official"].items())
                },
            }

    # ── Validate + write ──
    fresh = [nt for nt in meta if nt not in old_by_title]
    assert not fresh, f"unexpected fresh (no-old) minted titles: {len(fresh)}"
    payloads = {
        "coci_minted_courses.json": {
            "_status": "PREVIEW — CourseControlNumber re-mint (step 1a). Not yet landed.",
            "_generated_by": GENERATED_BY, "_generated_at": GENERATED_AT,
            "_memberships_file": "coci_minted_memberships.json",
            "count": len(courses), "courses": dict(sorted(courses.items()))},
        "coci_minted_memberships.json": {
            "_status": "PREVIEW — members carry College + CourseControlNumber (the re-key fix).",
            "_generated_at": GENERATED_AT,
            "count": len(memberships),
            "member_courses_total": sum(len(v) for v in memberships.values()),
            "memberships": dict(sorted(memberships.items()))},
        "coci_minted_singletons.json": {
            "_status": "PREVIEW — re-keyed single-college identities.",
            "_generated_at": GENERATED_AT,
            "count": len(singletons), "courses": dict(sorted(singletons.items()))},
        "alias_map.json": {
            "_about": "AUTHORITATIVE old M-ID -> new id (supersedes the dry-run map). PREVIEW.",
            "count": len(alias), "alias": dict(sorted(alias.items()))},
        "promotions.json": {
            "_about": "Split manifest: minted titles whose other members carry an official "
                      "C-ID/CCN. Remnant stays minted (above); these promote in step 1b.",
            "count": len(promotions), "promotions": dict(sorted(promotions.items()))},
    }
    for name, doc in payloads.items():
        with open(os.path.join(OUT_DIR, name), "w", encoding="utf-8") as f:
            json.dump(doc, f, indent=2, ensure_ascii=False)
            f.write("\n")

    print(f"corroborated catalog : {len(courses):,}")
    print(f"singletons           : {len(singletons):,}")
    print(f"memberships rows      : {sum(len(v) for v in memberships.values()):,}")
    print(f"alias entries         : {len(alias):,} (old M-IDs covered)")
    print(f"split (promotions)    : {len(promotions):,}")
    print(f"tier flips (old->new) : {dict(tier_flips)}")
    print(f"wrote preview to {OUT_DIR}/")


if __name__ == "__main__":
    main()
