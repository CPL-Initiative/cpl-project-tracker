"""
Validate parity between the Excel-derived workplan-goals ladder and the Supabase
public.workplan_goals table — the Phase 1 read-only diff that has to land green
before we cut excel_to_dashboard.py over to Supabase as source-of-truth.

Excel-side derivation ("A+" rules; Sam's call 2026-05-28):
  * Auto-discover every project in the Project List with at least one non-zero
    KPI cell (kpi_goal_* or kpi_stretch_*)
  * Exclude `D.*` dashboard-metric rows (they carry KPI ladders but are not
    workplan-goal-class entries)
  * No parent/child aggregation — each project is its own row. `4.1` renders as
    itself, `4.1.1`-`4.1.4` each render as themselves. The renderer's old
    hardcoded `core_ids` list retires.

What it checks:
  * Coverage      — every A+-derived activity has both a GOAL and a STRETCH row in Supabase
  * Value parity  — for overlapping (activity_id, row_type), every year column matches
  * Orphans       — Supabase rows whose activity_id is not in A+-derived set
  * Associations  — PR-A: every workplan_activity_associations row resolves to
                    a real kind='activity' AND a real kind='project' row, and
                    every project has at least one association

PR-A note: the Excel-vs-Supabase comparison is scoped to `kind='project'`
rows. The 5 `kind='activity'` rows in workplan_goals are curator-managed and
do not appear in Excel A+ derivation — they're excluded from the diff.

Auth:
  * SUPABASE_URL          (default https://hvuwhnbuahrtptokpqfh.supabase.co)
  * SUPABASE_SERVICE_KEY  (required for REST fetch)
  * --supabase-json PATH  (alt: read rows from a JSON file dumped via the MCP
    or a prior REST fetch; lets the script run without the service key)

Outputs:
  * kb/workplan_goals_validation.md  (human-reviewable report)
  * exit 0 if perfectly in sync; exit 1 if any mismatch / missing / orphan

Run from repo root:
  SUPABASE_SERVICE_KEY=... python3 kb/_validate_workplan_goals.py
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = REPO_ROOT / "CPL_Initiative_Project_List_v3.xlsx"
OUT_PATH = REPO_ROOT / "kb" / "workplan_goals_validation.md"

SUPABASE_URL = os.environ.get(
    "SUPABASE_URL", "https://hvuwhnbuahrtptokpqfh.supabase.co"
).rstrip("/")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")

# Excel column indices — mirror excel_to_dashboard.py constants
COL_ID, COL_NAME = 1, 2
COL_KPI_G2526, COL_KPI_S2526 = 20, 21
COL_KPI_G2627, COL_KPI_S2627 = 22, 23
COL_KPI_G2728, COL_KPI_S2728 = 24, 25
COL_KPI_G2829, COL_KPI_S2829 = 26, 27
COL_KPI_G2930, COL_KPI_S2930 = 28, 29

YEAR_COLS = [
    ("yr_2025_26", "2025-26", COL_KPI_G2526, COL_KPI_S2526),
    ("yr_2026_27", "2026-27", COL_KPI_G2627, COL_KPI_S2627),
    ("yr_2027_28", "2027-28", COL_KPI_G2728, COL_KPI_S2728),
    ("yr_2028_29", "2028-29", COL_KPI_G2829, COL_KPI_S2829),
    ("yr_2029_30", "2029-30", COL_KPI_G2930, COL_KPI_S2930),
]

# Treat |a - b| < EPS as equal for numeric comparison
EPS = 1e-6


def parse_num(val) -> float:
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("+", "")
    if not s:
        return 0.0
    if s.lower().endswith("k"):
        try:
            return float(s[:-1]) * 1000
        except ValueError:
            return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _fetch_path(path: str) -> list[dict]:
    if not SUPABASE_KEY:
        sys.exit(
            "Set SUPABASE_SERVICE_KEY in the env, or pass --supabase-json PATH"
            " to read rows from a file."
        )
    endpoint = f"{SUPABASE_URL}/rest/v1/{path}"
    req = urllib.request.Request(
        endpoint,
        headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.load(r)


def fetch_supabase_rows() -> list[dict]:
    return _fetch_path(
        "workplan_goals"
        "?select=activity_id,name,row_type,kind,"
        "yr_2025_26,yr_2026_27,yr_2027_28,yr_2028_29,yr_2029_30,total"
    )


def fetch_associations() -> list[dict]:
    return _fetch_path(
        "workplan_activity_associations?select=project_id,activity_id"
    )


def read_supabase_json(path: Path) -> list[dict]:
    """
    Accept either a flat list[dict] (legacy validator input) or a snapshot-
    shaped dict {"rows": [...], "associations": [...]}. Returns the rows.
    Use `read_associations_json()` for the associations sidecar.
    """
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        return data.get("rows") or []
    return data


def read_associations_json(path: Path) -> list[dict]:
    """Read associations from a snapshot file (returns [] if not present)."""
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        return data.get("associations") or []
    return []


def derive_excel_workplan_goals() -> dict[str, dict[str, dict[str, float]]]:
    """
    A+ derivation of the workplan-goals set from the canonical Excel.

    Returns {pid: {"name": str, "GOAL": {yr_key: val}, "STRETCH": {yr_key: val}}}

    Rules:
      - Read every row in Project List (starting row 4)
      - Skip rows whose id starts with "D." (dashboard KPI metric rows)
      - Include the row only if at least one of its 10 KPI cells is non-zero
      - No aggregation, no parent suppression — each project is its own row
    """
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Project List"]

    out: dict[str, dict] = {}
    for r in range(4, ws.max_row + 1):
        pid = ws.cell(r, COL_ID).value
        if not pid:
            break
        pid_s = str(pid).strip()
        if pid_s.startswith("D."):
            continue
        name = ws.cell(r, COL_NAME).value or ""
        goal_row = {}
        stretch_row = {}
        for yr_key, _yr_label, g_col, s_col in YEAR_COLS:
            goal_row[yr_key] = parse_num(ws.cell(r, g_col).value)
            stretch_row[yr_key] = parse_num(ws.cell(r, s_col).value)
        if not any(goal_row.values()) and not any(stretch_row.values()):
            continue
        out[pid_s] = {"name": str(name), "GOAL": goal_row, "STRETCH": stretch_row}
    return out


def reshape_supabase(rows: list[dict]) -> dict[str, dict[str, dict[str, float]]]:
    """
    {activity_id: {"name": ..., "GOAL": {yr_key: val}, "STRETCH": {yr_key: val}}}

    Scopes to `kind='project'` rows (the Excel-A+ comparison is project-vs-project).
    Activity rows (kind='activity') are skipped — they're curator-managed and
    don't appear in Excel A+. Rows without a kind field default to 'project'
    so pre-PR-A snapshots still validate correctly.
    """
    out: dict[str, dict] = {}
    for row in rows:
        if (row.get("kind") or "project") != "project":
            continue
        aid = row["activity_id"]
        rt = row["row_type"]
        bucket = out.setdefault(aid, {"name": row.get("name", "")})
        bucket[rt] = {k: parse_num(row.get(k)) for k, *_ in YEAR_COLS}
        if not bucket.get("name"):
            bucket["name"] = row.get("name", "")
    return out


def validate_associations(
    rows: list[dict], assocs: list[dict]
) -> tuple[list[dict], list[dict], list[dict]]:
    """
    Returns (orphan_activity, orphan_project, projects_without_assoc).
    Application-enforced FK shape (no DB FK per the PR-A design): every
    association row must resolve to an existing activity-kind row AND a
    project-kind row in workplan_goals.
    """
    activity_ids = {
        r["activity_id"] for r in rows
        if (r.get("kind") or "project") == "activity"
    }
    project_ids = {
        r["activity_id"] for r in rows
        if (r.get("kind") or "project") == "project"
    }
    orphan_activity = [
        a for a in assocs if a.get("activity_id") not in activity_ids
    ]
    orphan_project = [
        a for a in assocs if a.get("project_id") not in project_ids
    ]
    projects_with_assoc = {a.get("project_id") for a in assocs}
    projects_without_assoc = [
        {"project_id": pid} for pid in sorted(project_ids - projects_with_assoc)
    ]
    return orphan_activity, orphan_project, projects_without_assoc


def compare_rows(
    excel: dict[str, dict], supabase: dict[str, dict]
) -> tuple[list, list, list, list]:
    """
    Returns (matches, mismatches, missing_in_supabase, orphan_in_supabase).
    Each match/mismatch entry is a dict carrying the diagnostic detail.
    """
    matches, mismatches = [], []
    missing = []

    for pid, ex_data in excel.items():
        sb_data = supabase.get(pid)
        if not sb_data:
            missing.append(
                {"activity_id": pid, "name": ex_data["name"], "reason": "no_supabase_row"}
            )
            continue
        for row_type in ("GOAL", "STRETCH"):
            ex_row = ex_data[row_type]
            sb_row = sb_data.get(row_type)
            if not sb_row:
                missing.append(
                    {
                        "activity_id": pid,
                        "row_type": row_type,
                        "name": ex_data["name"],
                        "reason": f"missing_{row_type.lower()}_row",
                    }
                )
                continue
            diffs = []
            for yr_key, yr_label, *_ in YEAR_COLS:
                ev = ex_row.get(yr_key, 0.0)
                sv = sb_row.get(yr_key, 0.0)
                if abs(ev - sv) >= EPS:
                    diffs.append({"year": yr_label, "excel": ev, "supabase": sv})
            entry = {
                "activity_id": pid,
                "name": ex_data["name"],
                "row_type": row_type,
                "excel": ex_row,
                "supabase": sb_row,
                "diffs": diffs,
            }
            if diffs:
                mismatches.append(entry)
            else:
                matches.append(entry)

    # Orphans: Supabase activity_ids not in Excel CORE_IDS
    orphans = []
    for aid, sb_data in supabase.items():
        if aid not in excel:
            orphans.append(
                {"activity_id": aid, "name": sb_data.get("name", "")}
            )

    return matches, mismatches, missing, orphans


def fmt_num(v: float) -> str:
    if v == int(v):
        return f"{int(v):,}"
    return f"{v:,.2f}"


def render_report(
    excel: dict,
    supabase: dict,
    matches,
    mismatches,
    missing,
    orphans,
    assoc_orphan_activity=None,
    assoc_orphan_project=None,
    assoc_projects_without=None,
    assoc_count: int = 0,
) -> str:
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    assoc_orphan_activity = assoc_orphan_activity or []
    assoc_orphan_project = assoc_orphan_project or []
    assoc_projects_without = assoc_projects_without or []
    in_sync = not (
        mismatches
        or missing
        or orphans
        or assoc_orphan_activity
        or assoc_orphan_project
        or assoc_projects_without
    )

    lines = []
    lines.append("---")
    lines.append("title: Workplan Goals — Excel vs Supabase Validation")
    lines.append(f"date: {today}")
    lines.append("tags: [workplan-goals, supabase, excel-migration, validation, phase-1]")
    lines.append("related:")
    lines.append("  - kb/_validate_workplan_goals.py")
    lines.append("  - CPL_Initiative_Project_List_v3.xlsx (Project List sheet)")
    lines.append("  - public.workplan_goals (Supabase)")
    lines.append("---")
    lines.append("")
    lines.append("# Workplan Goals — Excel vs Supabase Validation")
    lines.append("")
    status = "✅ in sync" if in_sync else "⚠ drift detected"
    lines.append(f"**Status:** {status}")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- Excel A+-derived sub-activities (non-zero ladder, excl. D.*): **{len(excel)}**")
    lines.append(f"- Supabase rows: **{len(supabase)}** distinct activity_ids")
    lines.append(f"- Matches (GOAL+STRETCH × 5 years agree): **{len(matches)}**")
    lines.append(f"- Mismatches (overlapping rows that disagree): **{len(mismatches)}**")
    lines.append(f"- Missing in Supabase (A+ sub-activities without a Supabase row): **{len(missing)}**")
    lines.append(f"- Orphans in Supabase (rows whose activity_id isn't in the A+ set): **{len(orphans)}**")
    lines.append(f"- Activity↔Project associations: **{assoc_count}** "
                 f"(orphan-activity: {len(assoc_orphan_activity)}, "
                 f"orphan-project: {len(assoc_orphan_project)}, "
                 f"projects-without-assoc: {len(assoc_projects_without)})")
    lines.append("")

    if assoc_orphan_activity or assoc_orphan_project or assoc_projects_without:
        lines.append("## Association integrity")
        lines.append("")
        if assoc_orphan_activity:
            lines.append("**Orphan-activity associations** (point at activity_id with no `kind='activity'` row):")
            lines.append("")
            for a in assoc_orphan_activity:
                lines.append(f"- project_id=`{a.get('project_id')}` → activity_id=`{a.get('activity_id')}`")
            lines.append("")
        if assoc_orphan_project:
            lines.append("**Orphan-project associations** (point at project_id with no `kind='project'` row):")
            lines.append("")
            for a in assoc_orphan_project:
                lines.append(f"- project_id=`{a.get('project_id')}` → activity_id=`{a.get('activity_id')}`")
            lines.append("")
        if assoc_projects_without:
            lines.append(
                "**Projects with no association** (project rows that don't link "
                "to any Activity):"
            )
            lines.append("")
            for p in assoc_projects_without:
                lines.append(f"- `{p['project_id']}`")
            lines.append("")

    # Missing
    if missing:
        lines.append("## Missing in Supabase")
        lines.append("")
        lines.append("A+-derived sub-activities that have no row in `workplan_goals`. "
                     "These will be INSERTed by the seed step before Supabase can become source of truth.")
        lines.append("")
        lines.append("| activity_id | name | reason |")
        lines.append("|---|---|---|")
        for m in missing:
            name = m.get("name", "")
            rt = m.get("row_type", "")
            reason = m["reason"] + (f" ({rt})" if rt else "")
            lines.append(f"| `{m['activity_id']}` | {name} | {reason} |")
        lines.append("")

    # Orphans
    if orphans:
        lines.append("## Orphans in Supabase")
        lines.append("")
        lines.append("Supabase `workplan_goals` rows whose `activity_id` is NOT in "
                     "the A+-derived set. The seed step will DELETE these so "
                     "Supabase matches Excel exactly.")
        lines.append("")
        lines.append("| activity_id | name |")
        lines.append("|---|---|")
        for o in orphans:
            lines.append(f"| `{o['activity_id']}` | {o['name']} |")
        lines.append("")

    # Mismatches
    if mismatches:
        lines.append("## Value Mismatches")
        lines.append("")
        lines.append("Overlapping (activity_id, row_type) pairs where the year-by-year "
                     "values disagree between Excel and Supabase. Each table shows the "
                     "drift; you decide which side wins per row.")
        lines.append("")
        for mm in mismatches:
            lines.append(
                f"### `{mm['activity_id']}` — {mm['name']} — {mm['row_type']}"
            )
            lines.append("")
            lines.append("| year | Excel | Supabase | Δ |")
            lines.append("|---|---:|---:|---:|")
            for yr_key, yr_label, *_ in YEAR_COLS:
                ev = mm["excel"].get(yr_key, 0.0)
                sv = mm["supabase"].get(yr_key, 0.0)
                delta = sv - ev
                marker = "" if abs(delta) < EPS else " ⚠"
                lines.append(
                    f"| {yr_label} | {fmt_num(ev)} | {fmt_num(sv)} | {fmt_num(delta)}{marker} |"
                )
            lines.append("")

    # Matches summary (one-line per row, no table — keeps the report short)
    if matches:
        lines.append("## In-sync rows (no action needed)")
        lines.append("")
        for m in matches:
            lines.append(f"- `{m['activity_id']}` {m['row_type']} — {m['name']}")
        lines.append("")

    lines.append("## How to read this")
    lines.append("")
    lines.append(
        "- **Missing** rows will be INSERTed by `kb/_seed_workplan_goals.py`."
    )
    lines.append(
        "- **Mismatches** will be UPDATEd by the seed step (Excel A+ wins by construction)."
    )
    lines.append(
        "- **Orphans** will be DELETEd by the seed step (Excel A+ is the source of truth)."
    )
    lines.append("")
    lines.append("Re-run: `python3 kb/_validate_workplan_goals.py`")
    lines.append("")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--supabase-json",
        type=Path,
        default=None,
        help="Read Supabase rows from this JSON file instead of fetching via REST.",
    )
    args = parser.parse_args()

    if args.supabase_json:
        rows = read_supabase_json(args.supabase_json)
        assocs = read_associations_json(args.supabase_json)
    else:
        rows = fetch_supabase_rows()
        assocs = fetch_associations()

    excel = derive_excel_workplan_goals()
    supabase = reshape_supabase(rows)
    matches, mismatches, missing, orphans = compare_rows(excel, supabase)
    assoc_orphan_activity, assoc_orphan_project, assoc_projects_without = (
        validate_associations(rows, assocs)
    )

    report = render_report(
        excel,
        supabase,
        matches,
        mismatches,
        missing,
        orphans,
        assoc_orphan_activity=assoc_orphan_activity,
        assoc_orphan_project=assoc_orphan_project,
        assoc_projects_without=assoc_projects_without,
        assoc_count=len(assocs),
    )
    OUT_PATH.write_text(report, encoding="utf-8")

    print(f"wrote {OUT_PATH.relative_to(REPO_ROOT)}")
    print(
        f"  excel A+ activities: {len(excel)}; "
        f"supabase activity_ids: {len(supabase)}"
    )
    print(
        f"  matches={len(matches)} mismatches={len(mismatches)} "
        f"missing={len(missing)} orphans={len(orphans)}"
    )
    print(
        f"  associations={len(assocs)} "
        f"orphan_activity={len(assoc_orphan_activity)} "
        f"orphan_project={len(assoc_orphan_project)} "
        f"projects_without_assoc={len(assoc_projects_without)}"
    )

    in_sync = not (
        mismatches
        or missing
        or orphans
        or assoc_orphan_activity
        or assoc_orphan_project
        or assoc_projects_without
    )
    if not in_sync:
        sys.exit(1)


if __name__ == "__main__":
    main()
