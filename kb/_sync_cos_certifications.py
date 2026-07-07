"""
CareerOneStop Certification Finder → kb/reference/cos_certifications.json

The COS authority-anchoring lane (Sam, 2026-07-07: "Build the CareerOneStop"):
pull the national certification registry — certification name, certifying
ORGANIZATION (the issuing-agency vocabulary the CER's credentials.json is
missing), acronym, type, in-demand flag — into a slim committed reference file
that `kb/_match_cos_authority.py` joins against the credential KB and the CER
renders as "✓ COS" provenance badges. Credential Engine/CTDL integration comes
AFTER (the MAP↔CE partnership will code MAP CPL data with CTDL tags; the
`cos_cert_id` anchors minted here become join keys then).

RUNS ON A GITHUB ACTIONS RUNNER ONLY (.github/workflows/cos-authority-sync.yml)
— every careeronestop.org host is egress-blocked from the agent sandbox
(runner-as-proxy pattern, docs/kb-notes/playbook-runner-as-external-api-proxy.md).

Two lanes, tried in order:
  1. BULK — scrape the public data-download page for the current
     `ZIP_Certification_Finder_Data_<MMDDYYYY>.zip` link, download, parse the
     delimited/XLSX member inside. No account needed (probe confirms).
  2. API — page `GET /v1/certificationfinder/{userId}/{keyword}/...` with the
     registered COS credentials (repo secrets COS_USER_ID + COS_API_TOKEN).
     Keyword enumeration a–z/0–9 with de-dupe by cert Id (the API has no
     confirmed list-all call), so it's the fallback, not the default.

Modes:
  python3 kb/_sync_cos_certifications.py --probe   # report reachability, the
      bulk link found, archive members + headers + 3 sample rows, API status.
      WRITES NOTHING. Run this first from a workflow_dispatch and read the log.
  python3 kb/_sync_cos_certifications.py --apply   # write the reference file
      (idempotent: unchanged certification set → file left untouched).

Terms: free/royalty-free with REQUIRED attribution to USDOL ETA + Minnesota
DEED wherever the data displays (the CER badge/footer carries it); the synced
file stays tracker-internal under kb/reference/ per
docs/kb-notes/reference-authority-anchored-credential-naming.md.
"""
import csv
import io
import json
import os
import re
import ssl
import sys
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime, timezone

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "reference", "cos_certifications.json")

DATA_PAGE = "https://www.careeronestop.org/Developers/Data/certifications.aspx"
API_BASE = "https://api.careeronestop.org"
UA = ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
      "Chrome/126.0 Safari/537.36 cpl-project-tracker-sync")

ATTRIBUTION = ("Source: CareerOneStop (www.careeronestop.org), sponsored by the "
               "U.S. Department of Labor, Employment and Training Administration; "
               "data maintained by the Minnesota Department of Employment and "
               "Economic Development (DEED).")

COS_USER_ID = os.environ.get("COS_USER_ID")
COS_API_TOKEN = os.environ.get("COS_API_TOKEN")


def _get(url, headers=None, timeout=60, attempts=3):
    # www.careeronestop.org's WAF INTERMITTENTLY 403s runner IPs (probe run 1
    # got a 200, run 2 a 403, same UA, 2 minutes apart) — retry with backoff
    # before declaring the lane down. api.careeronestop.org (Bearer token) is
    # the reliable leg; the bulk lane is opportunistic.
    import time
    req = urllib.request.Request(url, headers={"User-Agent": UA, **(headers or {})})
    ctx = ssl.create_default_context()
    last = None
    for i in range(attempts):
        try:
            with urllib.request.urlopen(req, timeout=timeout, context=ctx) as r:
                return r.read(), dict(r.headers)
        except urllib.error.HTTPError as e:
            last = e
            if e.code not in (403, 429, 500, 502, 503):
                raise
            if i < attempts - 1:
                wait = 5 * (i + 1)
                print(f"  HTTP {e.code} on {url.split('?')[0]} — retry in {wait}s")
                time.sleep(wait)
    raise last


# ── Lane 1: bulk download ────────────────────────────────────────────────────

def find_bulk_link(html, probe=False):
    """The download page links ZIP_Certification_Finder_Data_<MMDDYYYY>.zip —
    return the newest absolute URL, or None. Accepts single/double-quoted or
    JS-embedded hrefs; in probe mode dumps every candidate href so a pattern
    drift is diagnosable straight from the run log."""
    if probe:
        print(f"  page bytes: {len(html):,}")
        for marker in ("just a moment", "cf-challenge", "captcha", "access denied",
                       "request unsuccessful", "incapsula"):
            if marker in html.lower():
                print(f"  ⚠ bot-challenge marker in page: {marker!r}")
        cands = re.findall(r"""["']([^"']*(?:\.zip|\.xlsx|certif[^"']*))["']""",
                           html, flags=re.I)
        seen = []
        for c in cands:
            if c not in seen and re.search(r"\.(zip|xlsx)$|download|data", c, re.I):
                seen.append(c)
        print(f"  candidate hrefs ({len(seen)}):")
        for c in seen[:40]:
            print(f"    {c}")
    links = re.findall(
        r"""["']([^"']*Certification[^"']*\.zip)["']""", html, flags=re.I)
    if not links:
        return None
    def date_key(u):
        m = re.search(r'(\d{2})(\d{2})(\d{4})\.zip', u)
        return (m.group(3), m.group(1), m.group(2)) if m else ("0", "0", "0")
    best = sorted(links, key=date_key)[-1]
    if best.startswith("http"):
        return best
    if best.startswith("/"):
        return "https://www.careeronestop.org" + best
    return "https://www.careeronestop.org/Developers/Data/" + best


# COS name convention embeds the acronym as a trailing " - ACR" (the first
# authenticated probe, 2026-07-07: "50001 Certified Professional - 50001 CP")
# while the matcher (kb/_match_cos_authority.py) expects a CLEAN name plus a
# separate acronym field — the shape the bulk file's dedicated column gives.
# Split the suffix out so both lanes emit the same record shape. Guards: every
# suffix token must be caps/digits (never "…- AutoCAD"), and a pure roman
# numeral / pure digit suffix is a LEVEL token, not an acronym ("Firefighter
# - II" stays intact — the level-collapse lesson).
_ACR_SUFFIX = re.compile(
    r"^(?P<base>.*\S)\s+-\s+(?P<acr>[A-Z0-9+\-\./]{1,12}(?: [A-Z0-9+\-\./]{1,12}){0,2})$")


def split_name_acronym(name):
    m = _ACR_SUFFIX.match(name or "")
    if not m:
        return name, None
    acr = m.group("acr")
    if not re.search(r"[A-Z]", acr) or re.fullmatch(r"[IVX]+|\d+", acr):
        return name, None
    return m.group("base"), acr


# Tolerant header → canonical field mapping (bulk-file headers are unverified
# until the first probe run; extend here if the probe shows different names).
HEADER_MAP = {
    "name": ["certification name", "certification", "cert name", "name", "title"],
    "org": ["certifying organization", "organization", "certifying org",
            "organization name", "sponsoring organization"],
    "acronym": ["acronym", "certification acronym", "abbreviation"],
    "id": ["id", "certification id", "cert id", "element id", "element_id"],
    "type": ["type", "certification type", "cert type"],
    "in_demand": ["in demand", "in-demand", "indemand", "in demand flag"],
    "url": ["url", "certification url", "website", "web address"],
    "org_url": ["organization url", "org url", "organization website"],
}


def map_headers(headers):
    low = [h.strip().lower() for h in headers]
    out = {}
    for field, names in HEADER_MAP.items():
        for n in names:
            if n in low:
                out[field] = low.index(n)
                break
    return out


def parse_rows(headers, rows):
    idx = map_headers(headers)
    if "name" not in idx or "org" not in idx:
        raise SystemExit(f"Bulk file headers unrecognized: {headers} — extend HEADER_MAP.")
    certs = []
    for row in rows:
        def cell(f):
            i = idx.get(f)
            return (str(row[i]).strip() if i is not None and i < len(row)
                    and row[i] is not None else "")
        name = cell("name")
        if not name:
            continue
        rec = {"name": name, "org": cell("org")}
        for f in ("id", "acronym", "type", "url", "org_url"):
            v = cell(f)
            if v:
                rec[f] = v
        if "acronym" not in rec:
            base, acr = split_name_acronym(name)
            if acr:
                rec["name"], rec["acronym"] = base, acr
        d = cell("in_demand").lower()
        if d in ("y", "yes", "true", "1"):
            rec["in_demand"] = True
        certs.append(rec)
    return certs


def read_archive(blob, probe):
    """Yield (member_name, headers, rows) for the first parseable member."""
    zf = zipfile.ZipFile(io.BytesIO(blob))
    names = zf.namelist()
    if probe:
        print(f"  archive members: {names}")
    for n in names:
        low = n.lower()
        data = zf.read(n)
        if low.endswith((".csv", ".txt", ".tsv")):
            text = data.decode("utf-8-sig", errors="replace")
            delim = "\t" if low.endswith((".tsv",)) or text.count("\t") > text.count(",") else ","
            rows = list(csv.reader(io.StringIO(text), delimiter=delim))
            if len(rows) > 1:
                return n, rows[0], rows[1:]
        elif low.endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True)
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            headers = [str(c) if c is not None else "" for c in next(it)]
            return n, headers, [list(r) for r in it]
    raise SystemExit(f"No parseable member in the bulk zip: {names}")


def lane_bulk(probe):
    print(f"BULK lane: GET {DATA_PAGE}")
    html, _ = _get(DATA_PAGE)
    html = html.decode("utf-8", errors="replace")
    link = find_bulk_link(html, probe)
    print(f"  bulk link found: {link}")
    if not link:
        return None
    blob, hdrs = _get(link, timeout=300)
    print(f"  downloaded {len(blob):,} bytes ({hdrs.get('Content-Type')})")
    member, headers, rows = read_archive(blob, probe)
    print(f"  member: {member} | headers: {headers} | data rows: {len(rows)}")
    if probe:
        for r in rows[:3]:
            print(f"  sample: {r}")
        return "PROBED"
    return parse_rows(headers, rows)


# ── Lane 2: API fallback ─────────────────────────────────────────────────────

def api_get(path):
    blob, _ = _get(API_BASE + path, headers={
        "Authorization": "Bearer " + COS_API_TOKEN, "Accept": "application/json"})
    return json.loads(blob)


def _api_rec(c):
    """API record → the registry shape the matcher expects (clean name +
    separate acronym; the API embeds acronyms as a 'Name - ACR' suffix and
    the first authenticated probe showed no Acronym field)."""
    rec = {"name": (c.get("Name") or "").strip(),
           "org": (c.get("Organization") or "").strip(),
           "id": c.get("Id")}
    for src, dst in (("Acronym", "acronym"), ("Type", "type"),
                     ("Url", "url"), ("OrganizationUrl", "org_url")):
        if c.get(src):
            rec[dst] = str(c[src]).strip()
    if "acronym" not in rec:
        base, acr = split_name_acronym(rec["name"])
        if acr:
            rec["name"], rec["acronym"] = base, acr
    if str(c.get("InDemand", "")).strip().lower() in ("y", "yes", "true", "1"):
        rec["in_demand"] = True
    return rec


def lane_api(probe):
    if not (COS_USER_ID and COS_API_TOKEN):
        print("API lane: COS_USER_ID / COS_API_TOKEN secrets not set — skipped.")
        return None
    def q(keyword, start=1, limit=500):
        kw = urllib.parse.quote(keyword) if keyword else "%20"
        return (f"/v1/certificationfinder/{COS_USER_ID}/{kw}/false/0/0/0/0/0/"
                f"Name/ASC/{start}/{limit}")
    if probe:
        d = api_get(q("building inspector", 1, 5))
        n = d.get("RecordCount") or len(d.get("CertList") or [])
        print(f"API lane probe: 'building inspector' → RecordCount={n}; "
              f"first: {json.dumps((d.get('CertList') or [{}])[0])[:300]}")
        return "PROBED"

    seen, certs = {}, []

    def fetch_keyword(kw):
        """Paginate one keyword search. Defensive on an unverified contract:
        advance by ACTUAL batch size (a server page cap below our limit must
        not end enumeration early), stop at RecordCount, and bail if two
        pages open with the same Id (server ignoring startRecord) so a
        >500-record keyword can never spin the runner / hammer the API."""
        start, total, prev_first, added = 1, None, None, 0
        while True:
            d = api_get(q(kw, start, 500))
            batch = d.get("CertList") or []
            if total is None:
                total = d.get("RecordCount") or 0
            if not batch:
                break
            first = batch[0].get("Id")
            if first is not None and first == prev_first:
                print(f"  keyword {kw!r}: startRecord={start} not advancing — "
                      f"stopping this keyword")
                break
            prev_first = first
            for c in batch:
                cid = c.get("Id")
                if cid and cid not in seen:
                    seen[cid] = 1
                    certs.append(_api_rec(c))
                    added += 1
            if start - 1 + len(batch) >= total:
                break
            start += len(batch)
        return total or 0, added

    # The empty keyword ("%20") may enumerate the whole registry in one pass —
    # the probe showed keyword search matches broadly ('building inspector' →
    # 1,088 records). Try it first; the a-z/0-9 fan-out stays as the fallback
    # (de-dupe by Id makes overlap free). Each keyword fails soft so one bad
    # response can't zero the whole lane.
    for kw in [""] + list("abcdefghijklmnopqrstuvwxyz0123456789"):
        try:
            total, added = fetch_keyword(kw)
        except Exception as e:
            print(f"  keyword {kw!r} failed: {type(e).__name__}: {e}")
            continue
        print(f"  keyword {kw or '<all>'}: RecordCount={total}, new={added}, "
              f"cumulative {len(certs)}")
        if kw == "" and total and len(certs) >= total:
            print("  empty-keyword pass covered the full registry — "
                  "skipping the a-z/0-9 fan-out")
            break
    return certs


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    probe = "--probe" in sys.argv or "--apply" not in sys.argv
    mode = "probe" if probe else "apply"
    print(f"mode: {mode}")

    certs = None
    lane = None
    for lane_name, fn in (("bulk", lane_bulk), ("api", lane_api)):
        try:
            res = fn(probe)
        except Exception as e:  # each lane fails soft; the other still runs
            print(f"{lane_name} lane failed: {type(e).__name__}: {e}")
            continue
        if res == "PROBED":
            lane = lane or lane_name
            continue
        if res:
            certs, lane = res, lane_name
            break

    if probe:
        print("PROBE complete — read the findings above, then dispatch with mode=apply.")
        return
    if not certs:
        sys.exit("Both lanes failed — no certification data retrieved.")

    certs.sort(key=lambda c: (c["name"].lower(), c.get("org", "").lower()))
    # Idempotent: unchanged certification set → no rewrite (no timestamp churn).
    try:
        prior = json.load(open(OUT, encoding="utf-8")).get("certifications")
    except (FileNotFoundError, ValueError):
        prior = None
    if prior == certs:
        print(f"no change: {len(certs)} certifications — file left untouched")
        return

    out = {
        "_about": ("CareerOneStop Certification Finder registry, slimmed for the "
                   "CER authority-anchoring lane (kb/_match_cos_authority.py). "
                   "Synced by kb/_sync_cos_certifications.py on a GitHub Actions "
                   "runner (cos-authority-sync.yml) — the agent sandbox cannot "
                   "reach careeronestop.org. Tracker-internal: display requires "
                   "the attribution below; do not republish in the public KB "
                   "without a CareerOneStop permission check."),
        "_attribution": ATTRIBUTION,
        "_lane": lane,
        "_synced_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "count": len(certs),
        "certifications": certs,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=1, ensure_ascii=False)
        f.write("\n")
    print(f"wrote {OUT}: {len(certs)} certifications via the {lane} lane")


if __name__ == "__main__":
    main()
