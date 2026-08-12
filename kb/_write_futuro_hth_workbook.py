#!/usr/bin/env python3
"""Write the Futuro Health HTH -> CCC CNA crosswalk workbook (.xlsx).

Consumes kb/futuro_hth_out/crosswalk.json (from kb/_build_futuro_hth_crosswalk.py).

Run:  python3 kb/_build_futuro_hth_crosswalk.py && python3 kb/_write_futuro_hth_workbook.py
"""
import argparse
import datetime as dt
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "futuro_hth_out", "crosswalk.json")


def col_of(headers, name):
    """Column letter for a header, BY NAME.

    The two lenses ship different column sets, so hard-coded letters in the summary
    formulas would silently point at the wrong column on one of them. Deriving the
    letter from the header removes that failure mode entirely.
    """
    return get_column_letter(headers.index(name) + 1)

FONT = "Arial"
INK = "1F2937"
BRAND = "1D4ED8"
MUTED = "6B7280"
HDR_FILL = PatternFill("solid", fgColor="1D4ED8")
BAND = PatternFill("solid", fgColor="EFF6FF")
WARN = PatternFill("solid", fgColor="FEF3C7")
THIN = Side(style="thin", color="D1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 34
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def body(ws, first=2, wrap_cols=()):
    for r in range(first, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.border = BOX
            cell.alignment = Alignment(
                vertical="top", wrap_text=get_column_letter(c) in wrap_cols)
        if r % 2 == 0:
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=r, column=c).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=r, column=c).fill = BAND


def next_step(r):
    """Plain-language recommendation. Absence never renders as a checkmark."""
    ready_a = r["readiness"].startswith("A")
    ready_b = r["readiness"].startswith("B")
    strong = r["score"] >= 4
    if not r["n_candidates"]:
        return ("Ask the college directly - no receiving course matched the state file. "
                "The course may exist under a title the file abbreviates, or be too new.")
    if ready_a and strong:
        return ("START HERE. CPL is already operating in MAP and a clear receiving course "
                "exists - the conversation is about one course, not about building CPL.")
    if ready_a:
        return ("CPL is operating in MAP. Receiving course is a weaker match - ask the "
                "health faculty which course carries the interpersonal-skills outcomes.")
    if ready_b:
        return ("Exhibits are loaded in MAP but nothing is converting to credit yet. "
                "Good technical readiness; ask what is blocking credit recommendations.")
    if strong:
        return ("Strong curricular fit but no CPL activity in MAP. Needs a CPL conversation "
                "with the named contact before a course-level ask.")
    return ("Lower priority. No CPL activity in MAP and the receiving-course match is weak - "
            "revisit if a regional or student-demand reason brings this college forward.")


def build(outpath, src=SRC):
    with open(src, encoding="utf-8") as f:
        d = json.load(f)
    lens = d.get("lens", "all")
    nc = lens == "noncredit"
    rows = sorted(d["rows"], key=lambda r: (-r["score"], -r["map_credit_recs"], r["college"]))
    n = len(rows)
    t = d["totals"]
    today = dt.date.today().isoformat()

    wb = Workbook()

    # ------------------------------------------------------------ Read me ---
    ws = wb.active
    ws.title = "Read me"
    lines = [
        ("Futuro Health - Human Touch Healthcare (HTH)", "title"),
        ("Statewide crosswalk to California Community College CNA programs"
         + (" - NONCREDIT lens" if nc else ""), "sub"),
        (f"Prepared for Ashley  |  MAP / CPL Initiative  |  {today}", "meta"),
        ("", ""),
        ("What this answers", "h"),
        ("Which California Community Colleges offer a CNA program, and where the Human Touch "
         "Healthcare course could plausibly earn a student college credit at each one.", "p"),
        ("", ""),
        ("The key judgment behind the sheet - please read this one", "h"),
        ("HTH does not map onto the CNA course itself. A California CNA program is a "
         "CDPH-approved 160-hour course (60 theory + 100 clinical) whose hours and content are "
         "fixed in regulation, so a college cannot award CPL against it for an 80-hour online "
         "soft-skills course. What HTH maps onto is the course sitting NEXT TO the CNA course "
         "in the same catalog - interpersonal communication, intercultural communication, "
         "healthcare ethics, communication for allied health. That is what the "
         "'Best-fit receiving course' column names.", "p"),
        ("", ""),
        ("Where the numbers come from", "h"),
        ("Colleges offering CNA: the statewide MIS course file (fall 2025, 109,898 courses "
         "across 118 colleges), TOP code 1230.30 Certified Nurse Assistant.", "p"),
        ("CNA award level: the statewide COCI program export (2026-06-17, 29,147 programs).", "p"),
        ("CPL readiness, contacts, landing pages: MAP itself, via the CPL Initiative's "
         "Supabase mirror, synced 2026-08-12.", "p"),
        ("", ""),
        ("How a receiving course was chosen", "h"),
        ("A course had to pass two independent tests: its TITLE matches one of HTH's six "
         "modules, AND its TOP code sits in a related discipline family. Title-only matching "
         "produced 'Library Teamwork Supervision Skills' and 'Compassion Training for Yoga "
         "Teachers', so both signals are required. Courses teaching another occupation's law "
         "and ethics (funeral service, dental, veterinary) are scored down - they are health "
         "TOP codes but they are not this course.", "p"),
        ("", ""),
        ("What this sheet does NOT tell you", "h"),
        ("It does not confirm that any college will grant credit. Nothing here has been "
         "validated with discipline faculty, and no college website was reviewed - these "
         "sessions cannot reach college domains, so every recommendation rests on state data "
         "files and MAP, not on catalog pages. Treat the receiving courses as the best "
         "available starting point for a conversation, not as an approval.", "p"),
        ("An empty-looking result is always written out as a phrase, never left blank, so you "
         "can tell 'we looked and found nothing' apart from 'this column was not filled in'.", "p"),
        ("", ""),
        ("The finding worth acting on first", "h"),
        ("Futuro Health already exists in MAP as a partner entity (ID 133) with a live landing "
         "page - but it holds ZERO exhibits and ZERO credit recommendations. No CPL credit "
         "recommendation for HTH, or for any Futuro Health course, exists in MAP today. "
         "Getting HTH loaded as an exhibit is the step that turns all 61 of these colleges "
         "from prospects into reachable targets.", "p"),
        ("", ""),
        ("Sheets in this workbook", "h"),
        ("College crosswalk - one row per college, sorted strongest first.", "p"),
        ("Receiving course detail - every candidate course behind the summary column.", "p"),
        ("HTH course profile - the six modules and six SLOs, from the syllabus.", "p"),
        ("Statewide summary - the counts, as live formulas.", "p"),
        ("", ""),
    ] + ([
        ("What the noncredit lens changes, and the finding in it", "h"),
        ("This version keeps only the colleges that teach CNA as a NONCREDIT course - "
         f"{n} of the 61. Noncredit CNA is typically free or near-free to the student, "
         "which is why it is worth looking at on its own for Futuro Health's scholars.", "p"),
        ("The finding is not the college list, it is this: even at these colleges, the "
         "course HTH would articulate into is almost always on the CREDIT side. Only "
         f"{t['with_nc_receiver']} of the {n} have any noncredit receiving course at all - "
         f"and of those, only {t['with_nc_real_match']} matches one of HTH's six modules "
         "(Saddleback's Interprofessional Communication in Healthcare). The other "
         f"{t['with_nc_receiver'] - t['with_nc_real_match']} are career-survey courses "
         "- 'Survey of Health Careers', 'Pathways to Health Careers' - which matched the "
         "broad allied-health professionalism lens, not an HTH module, and are flagged as "
         f"such in the sheet. For {t['nc_only_route_is_credit']} colleges the credit "
         "course is the only route. Community "
         "college noncredit catalogs in health are dominated by clinical and technical "
         "training - Home Health Aide, CPR and basic life support, nursing skills labs, "
         "phlebotomy - not by interpersonal-skills courses. Statewide there are 372 "
         "noncredit courses at these colleges in the relevant discipline families, and "
         "only 12 of them are a soft-skills course of the kind HTH maps to.", "p"),
        ("That matters for how you ask. A noncredit-to-noncredit articulation is the "
         "cleanest ask - no unit cost, no transcript friction - but it is available at "
         "very few colleges. Everywhere else the honest ask is against a credit course, "
         "which is still worth doing but is a different conversation.", "p"),
        ("", ""),
    ] if nc else []) + [
        ("MAP college ID (last column)", "h"),
        ("Each college is matched to MAP through its state MIS college code rather than by "
         "name, using the shared college identity crosswalk the MAP team maintains. A code "
         "cannot be defeated by a spelling, so this is more reliable than matching "
         "\"MOUNT SAN ANTONIO COLLEGE\" to \"Mt. San Antonio College\" by hand. 60 of the 61 "
         "colleges matched that way. The exception is San Diego College of Continuing "
         "Education, which is a real MAP entity but sits outside that crosswalk\'s scope "
         "(it covers the 116 credit colleges), so it was matched by name instead and its "
         "ID cell says so rather than sitting empty.", "p"),
        ("", ""),
        ("One note on the Statewide summary tab", "h"),
        ("Its figures are live formulas over the College crosswalk tab, so they stay correct if you "
         "filter or edit rows. They compute the moment Excel opens the file. If you preview the "
         "workbook in something that does not calculate (a browser preview, a file-preview pane), "
         "those cells can look empty - open it in Excel and they fill in.", "p"),
    ]
    styles = {
        "title": Font(name=FONT, size=17, bold=True, color=BRAND),
        "sub": Font(name=FONT, size=12, color=INK),
        "meta": Font(name=FONT, size=10, italic=True, color=MUTED),
        "h": Font(name=FONT, size=11, bold=True, color=INK),
        "p": Font(name=FONT, size=10, color=INK),
    }
    for i, (text, kind) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if kind:
            c.font = styles[kind]
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 118
    for i, (text, kind) in enumerate(lines, start=1):
        if kind == "p" and len(text) > 110:
            ws.row_dimensions[i].height = 15 * (len(text) // 110 + 1)

    # -------------------------------------------------- College crosswalk ---
    ws = wb.create_sheet("College crosswalk")
    headers = [
        "College", "Region", "County",
        "CNA courses", "CNA delivery", "CNA award (COCI)",
        "Best-fit receiving course(s) for HTH", "Best-fit course type",
    ] + ([
        "Noncredit receiving course", "Credit receiving course",
    ] if nc else []) + [
        "Candidate courses", "HTH modules covered", "Which HTH modules",
        "Alignment score (1-5)", "MAP CPL readiness",
        "MAP exhibits", "MAP credit recs",
        "MAP primary contact", "Contact email", "MAP landing page",
        "Suggested next step", "MAP college ID",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([
            r["college"], r["region"], r["county"],
            r["cna_courses"], r["cna_delivery"], r["cna_award"],
            r["best_fit"], r["best_fit_tier"],
        ] + ([
            " | ".join(
                f"{c['course']} {c['title']}"
                + ("" if c["real_module_match"]
                   else "  [career-survey course - matches allied-health professionalism "
                        "broadly, NOT one of HTH's six modules]")
                for c in r["nc_candidates"][:2])
            or "No noncredit receiving course - the credit course is the only route",
            " | ".join(f"{c['course']} {c['title']}" for c in r["cr_candidates"][:2])
            or "No credit receiving course found",
        ] if nc else []) + [
            r["n_candidates"], r["n_modules"], r["modules"],
            r["score"], r["readiness"],
            r["map_exhibits"], r["map_credit_recs"],
            r["contact"], r["email"], r["landing"],
            next_step(r),
            # Appended LAST on purpose: it is a join key, not reading matter, and
            # adding it here leaves columns A-S (and the summary formulas that
            # reference them) untouched.
            r.get("map_college_id") or "Not in the identity crosswalk (continuing ed)",
        ])
    style_header(ws)
    W = {"College": 30, "Region": 19, "County": 15, "CNA courses": 8,
         "CNA delivery": 22, "CNA award (COCI)": 26,
         "Best-fit receiving course(s) for HTH": 52, "Best-fit course type": 26,
         "Noncredit receiving course": 46, "Credit receiving course": 46,
         "Candidate courses": 10, "HTH modules covered": 10, "Which HTH modules": 34,
         "Alignment score (1-5)": 9, "MAP CPL readiness": 30, "MAP exhibits": 9,
         "MAP credit recs": 9, "MAP primary contact": 22, "Contact email": 30,
         "MAP landing page": 44, "Suggested next step": 60, "MAP college ID": 16}
    widths(ws, {col_of(headers, k): v for k, v in W.items() if k in headers})
    WRAP = ["CNA delivery", "CNA award (COCI)", "Best-fit receiving course(s) for HTH",
            "Best-fit course type", "Noncredit receiving course", "Credit receiving course",
            "Which HTH modules", "MAP CPL readiness", "MAP primary contact",
            "Contact email", "Suggested next step", "MAP college ID"]
    body(ws, wrap_cols=tuple(col_of(headers, k) for k in WRAP if k in headers))
    CENTER = ["CNA courses", "Candidate courses", "HTH modules covered",
              "Alignment score (1-5)", "MAP exhibits", "MAP credit recs"]
    c_ready = headers.index("MAP CPL readiness") + 1
    c_best = headers.index("Best-fit receiving course(s) for HTH") + 1
    c_ncrec = headers.index("Noncredit receiving course") + 1 if nc else None
    for r in range(2, ws.max_row + 1):
        for k in CENTER:
            ws.cell(row=r, column=headers.index(k) + 1).alignment = Alignment(
                horizontal="center", vertical="top")
        if ws.cell(row=r, column=c_ready).value.startswith("A"):
            ws.cell(row=r, column=c_ready).fill = PatternFill("solid", fgColor="DCFCE7")
        # Flag the "we looked and found nothing" rows so they read as a finding.
        if str(ws.cell(row=r, column=c_best).value).startswith("No matching course"):
            ws.cell(row=r, column=c_best).fill = WARN
        if c_ncrec and str(ws.cell(row=r, column=c_ncrec).value).startswith("No noncredit"):
            ws.cell(row=r, column=c_ncrec).fill = WARN
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # ------------------------------------------------ Receiving detail ------
    ws = wb.create_sheet("Receiving course detail")
    ws.append(["College", "Region", "Course", "Course title", "Type", "TOP code",
               "Credit status", "Max units", "HTH modules matched", "Fit score"])
    for r in rows:
        for c in r["candidates"]:
            ws.append([r["college"], r["region"], c["course"], c["title"],
                       {1: "Health discipline", 2: "Communication studies",
                        3: "Human services / psychology"}[c["tier"]],
                       c["top"], c["credit"], c["units"],
                       ", ".join(c["modules_all"]), c["fit"]])
    style_header(ws)
    widths(ws, {"A": 30, "B": 19, "C": 15, "D": 50, "E": 24, "F": 11, "G": 24,
                "H": 10, "I": 42, "J": 9})
    body(ws, wrap_cols=("D", "G", "I"))
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    # ------------------------------------------------- HTH course profile ---
    ws = wb.create_sheet("HTH course profile")
    ws.append(["Field", "Value"])
    profile = [
        ("Course title", "Human Touch Healthcare (HTH)"),
        ("Provider", "Futuro Health"),
        ("Format", "Fully online, instructor-supported, Canvas LMS, with weekly live sessions"),
        ("Length", "Six weeks"),
        ("Total learning hours", "80 hours (approx. 12-13 hours per week)"),
        ("Prerequisites", "None"),
        ("Assessment", "Final assessment, minimum 73% to pass; minimum 730 total course points"),
        ("Graded work", "Rubric-scored assignments, discussions, scenario 'Resolves', "
                        "knowledge checks, weekly live sessions"),
        ("Credential on completion", "Digital completion badge"),
        ("Textbook", "None required"),
        ("", ""),
        ("MODULE 1", "Emotional Intelligence"),
        ("MODULE 2", "Empathy and Compassion"),
        ("MODULE 3", "Effective Communication"),
        ("MODULE 4", "Cultural Competence"),
        ("MODULE 5", "Teamwork and Collaboration"),
        ("MODULE 6", "Ethics and Integrity"),
        ("", ""),
        ("SLO 1", "Apply ethical guidelines to respond to healthcare scenarios in ways that "
                  "show empathy, cultural sensitivity, and professional accountability."),
        ("SLO 2", "Identify personal emotional responses and interpersonal dynamics to "
                  "cultivate ethical, empathetic, and compassionate relationships that honor "
                  "the cultural context and lived experiences of others."),
        ("SLO 3", "Apply clear, empathetic, and culturally responsive communication strategies "
                  "that promote trust, understanding, and inclusive, patient-centered care "
                  "across diverse populations."),
        ("SLO 4", "Recognize personal cultural awareness and implicit biases to develop "
                  "equitable, culturally responsive practices that respect the diverse values "
                  "and beliefs of patients and colleagues."),
        ("SLO 5", "Apply teamwork and conflict-resolution strategies to foster inclusive "
                  "collaboration, clarify roles, and contribute to high-quality, "
                  "coordinated care."),
        ("SLO 6", "Engage in critical self-reflection and incorporate feedback to enhance "
                  "communication, strengthen ethical practice, promote empathy and cultural "
                  "responsiveness, and support professional development."),
        ("", ""),
        ("Source", "HTH General Syllabus 3.0 v2.2, supplied by Futuro Health"),
        ("Note on unit value", "80 hours is broadly consistent with a 3-unit lecture course "
                               "(a 3-unit CCC lecture course is ~48-54 contact hours plus "
                               "homework). Unit award is the receiving college's decision."),
    ]
    for k, v in profile:
        ws.append([k, v])
    style_header(ws)
    widths(ws, {"A": 24, "B": 96})
    body(ws, wrap_cols=("B",))
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value).startswith(("MODULE", "SLO")):
            ws.cell(row=r, column=1).font = Font(name=FONT, size=10, bold=True, color=BRAND)

    # --------------------------------------------------------- Summary -----
    ws = wb.create_sheet("Statewide summary")
    ws.append(["Measure", "Value", "How it is counted"])
    last = n + 1
    C = lambda name: col_of(headers, name)
    facts = ([
        ("Colleges teaching CNA as a NONCREDIT course", n,
         "CB_CREDIT_STATUS = N in the fall-2025 state course file - this sheet's "
         "whole universe", None),
        ("...of those, with a NONCREDIT receiving course for HTH", t["with_nc_receiver"],
         "A noncredit course HTH could articulate into", None),
        ("...where the CREDIT course is the only route", t["nc_only_route_is_credit"],
         "A receiving course exists but only on the credit side - the student would "
         "enrol in a credit course to use the credit", None),
    ] if nc else []) + [
        ("California Community Colleges teaching a CNA course", None,
         "TOP 1230.30 in the fall-2025 statewide MIS course file",
         f"=COUNTA('College crosswalk'!{C('College')}2:{C('College')}{last})"),
        ("...out of colleges in the state course file", t["mis_colleges"],
         "Distinct colleges present in the same file", None),
        ("CNA courses statewide", t["cna_courses"],
         "All TOP 1230.30 course records", None),
        ("CNA programs in COCI (award level)", t["coci_programs"],
         "TOP 1230.30 or CNA-titled programs in the COCI export", None),
        ("Colleges with at least one plausible receiving course", None,
         "Passed both the title lens and the TOP-family test",
         f"=COUNTIF('College crosswalk'!{C('Candidate courses')}2:{C('Candidate courses')}{last},\">0\")"),
        ("...of which cover at least one of HTH's six modules", None,
         "A candidate can match the allied-health professionalism lens without "
         "matching any of the six named modules - one college does",
         f"=COUNTIF('College crosswalk'!{C('HTH modules covered')}2:{C('HTH modules covered')}{last},\">0\")"),
        ("Colleges with a health-discipline receiving course", t["with_tier1"],
         "Receiving course sits in the same division as CNA", None),
        ("Colleges where CPL is already operating in MAP", None,
         "MAP holds both exhibits and credit recommendations",
         f"=COUNTIF('College crosswalk'!{C('MAP CPL readiness')}2:{C('MAP CPL readiness')}{last},\"A*\")"),
        ("Colleges scoring 4 or 5 for HTH alignment", None,
         "Strongest curricular fit",
         f"=COUNTIF('College crosswalk'!{C('Alignment score (1-5)')}2:{C('Alignment score (1-5)')}{last},\">=4\")"),
        ("Colleges with no MAP primary contact on file", None,
         "MAP holds no primary contact - a person must be found first",
         f"=COUNTIF('College crosswalk'!{C('MAP primary contact')}2:{C('MAP primary contact')}{last},\"MAP holds no*\")"),
        ("Futuro Health exhibits in MAP today", 0,
         "Futuro Health is a MAP partner entity (ID 133) with a live landing "
         "page but no exhibits and no credit recommendations", None),
    ]
    for label, value, how, formula in facts:
        ws.append([label, formula if formula else value, how])
    style_header(ws)
    widths(ws, {"A": 52, "B": 14, "C": 62})
    body(ws, wrap_cols=("A", "C"))
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=r, column=2).font = Font(name=FONT, size=11, bold=True, color=BRAND)
    ws.cell(row=ws.max_row, column=2).fill = WARN

    wb.save(outpath)
    return outpath, n


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--lens", choices=("all", "noncredit"), default="all")
    ap.add_argument("--out")
    a = ap.parse_args()
    sub = "" if a.lens == "all" else a.lens
    src = os.path.join(HERE, "futuro_hth_out", sub, "crosswalk.json")
    if not a.out:
        tag = "" if a.lens == "all" else "_Noncredit"
        a.out = os.path.join(
            HERE, "futuro_hth_out", sub,
            f"{dt.date.today():%Y%m%d}_Futuro_Health_HTH_CNA_Statewide_Crosswalk{tag}.xlsx")
    os.makedirs(os.path.dirname(a.out), exist_ok=True)
    p, n = build(a.out, src)
    print(f"wrote {p}  ({n} colleges)")
