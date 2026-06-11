#!/usr/bin/env python3
"""Measure the WITNESS-KINSHIP gate before changing the official-ID fold rule.

Context (Session 41, 2026-06-11 — the AUTO 120X/150X mis-folds): the Phase A/B
official-ID fold consumes kb/promotions.json receipts written at the 2026-05-22
re-mint. A receipt ties a DEPARTED witness (a college course whose COCI row
claims a C-ID/CCN) to the M-ID remnant of its old pre-re-mint family. When that
old family was a lossy-key chimera (the original (subject, number) pair-minting),
the witness is UNRELATED to the remnant — e.g. MiraCosta's "Automatic
Transmissions and Transaxles" (claims AUTO 120 X) vouching for remnant
AUTO M1017 "Advanced Automotive Engine Performance". The fold rule then merges
the engine remnant under the transmissions C-ID.

The gate this script measures: a witness is KIN-VALID for a remnant iff
  max( J(remnant_title, witness_course_title), J(remnant_title, official_title) )
  >= 0.5
where J = token-set Jaccard over the level-safe normalization (parentheticals
stripped, articles dropped, roman numerals -> digits). The official-title branch
keeps folds where the remnant IS the official course by name (no witness needed);
the witness branch keeps the evidence-over-lexical wins ("Spanish 3" -> SPAN 200,
because the witnesses' own courses are titled "Spanish 3").

Read-only. Prints the would-block list for eyeballing; writes nothing.
"""
import json
import os
import re
import sys
from collections import Counter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

_DROP = {"the", "of", "to", "and", "for", "with", "in", "a", "an", "on", "at", "as"}
_ROMAN = {"i": "1", "ii": "2", "iii": "3", "iv": "4", "v": "5",
          "vi": "6", "vii": "7", "viii": "8", "ix": "9", "x": "10"}


def toks(t):
    t = re.sub(r"\([^)]*\)", " ", str(t or "").lower())
    t = re.sub(r"[^a-z0-9 ]+", " ", t)
    return {_ROMAN.get(w, w) for w in t.split() if w not in _DROP}


def jac(a, b):
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def slug(s):
    return re.sub(r"[^A-Z0-9]+", "", str(s or "").upper())


def main():
    promos = json.load(open(os.path.join(HERE, "promotions.json")))["promotions"]
    cat = json.load(open(os.path.join(HERE, "coci_minted_courses.json")))["courses"]
    sgd = json.load(open(os.path.join(HERE, "coci_minted_singletons.json")))
    sg = sgd.get("singletons", sgd)

    # official catalog titles
    cid_titles = {}
    for d in json.load(open(os.path.join(HERE, "reference", "cid_descriptors.json")))["descriptors"]:
        cid_titles[" ".join(str(d.get("descriptor") or "").split()).upper()] = d.get("title") or ""
    ccn_titles = {}
    for c in json.load(open(os.path.join(HERE, "reference", "ccn_courses.json")))["courses"]:
        if c.get("ccn"):
            ccn_titles[" ".join(c["ccn"].split()).upper()] = c.get("title") or ""

    # witness-resolution index: (official id, college slug) -> [course titles]
    from openpyxl import load_workbook
    NULLISH = {"", "NULL", "N/A", "NA", "NONE", "NOT APPLICABLE", "NOT APPLICABLE.", "TBD", "-"}
    claims = {}
    wb = load_workbook(os.path.join(HERE, "reference", "coci_course_list.xlsx"), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rit = ws.iter_rows(values_only=True)
    next(rit)
    for row in rit:
        college, title = row[0], row[4]
        cid = " ".join(str(row[9] or "").split()).upper()
        ccn = " ".join(str(row[11] or "").split()).upper()
        if not college:
            continue
        cs = slug(college)
        if cid and cid not in NULLISH:
            claims.setdefault(("C-ID:" + cid, cs), []).append(str(title or ""))
        if ccn and ccn not in NULLISH:
            claims.setdefault(("CCN:" + ccn, cs), []).append(str(title or ""))
    wb.close()

    def leaf_title(i):
        v = cat.get(i) or sg.get(i) or {}
        return v.get("common_title") or ""

    def official_title(oid):
        ns, _, bare = oid.partition(":")
        return (cid_titles if ns == "C-ID" else ccn_titles).get(bare, "")

    # ---- per-record kinship ----
    n_rec = n_tgt = 0
    tgt_all_kin = tgt_mixed = tgt_none_kin = 0
    unresolvable = 0
    blocked = []   # (leaf, leaf_title, oid, official_title, n_wit, kin_wit, witness titles)
    kept_by_official = 0
    for leaf, rec in promos.items():
        lt = leaf_title(leaf)
        if not lt:
            continue  # leaf id not live (R1 dead ends) — skip
        ltk = toks(lt)
        n_rec += 1
        for oid, w in (rec.get("official_targets") or {}).items():
            n_tgt += 1
            cols = (w or {}).get("colleges") or []
            n = (w or {}).get("members", 0) or 0
            ot = official_title(oid)
            otk = toks(ot)
            off_kin = jac(ltk, otk) >= 0.5
            kin = unkin = 0
            wt_samples = []
            for c in cols:
                titles = claims.get((oid, slug(c)), [])
                if not titles:
                    unresolvable += 1
                    best = 0.0
                else:
                    best = max(jac(ltk, toks(t)) for t in titles)
                    wt_samples.append(max(titles, key=lambda t: jac(ltk, toks(t))))
                if best >= 0.5 or off_kin:
                    kin += 1
                else:
                    unkin += 1
            if unkin == 0 and kin > 0:
                tgt_all_kin += 1
                if off_kin and all(jac(ltk, toks(t)) < 0.5 for t in wt_samples):
                    kept_by_official += 1
            elif kin > 0:
                tgt_mixed += 1
            else:
                tgt_none_kin += 1
                blocked.append((leaf, lt, oid, ot, n, kin, wt_samples[:3]))

    print(f"promotions records with a live leaf: {n_rec}; target edges: {n_tgt}")
    print(f"  all witnesses kin-valid : {tgt_all_kin}  (kept purely by official-title match: {kept_by_official})")
    print(f"  mixed kin               : {tgt_mixed}")
    print(f"  ZERO kin (gate blocks)  : {tgt_none_kin}")
    print(f"  unresolvable witness courses: {unresolvable}")

    # ---- impact on the CURRENT live folds ----
    txt = open(os.path.join(ROOT, "unified_courses_data.js"), encoding="utf-8").read()
    data = json.loads(txt[txt.index("{"): txt.rindex("}") + 1])
    rows = data["rows"]
    folded_rows = [r for r in rows if r.get("consolidated_from")]
    blocked_leaves = {b[0] for b in blocked}
    lost_folds = []
    for r in folded_rows:
        lost = [m for m in r["consolidated_from"] if m in blocked_leaves]
        if lost:
            lost_folds.append((r["id"], r.get("title"), lost))
    n_mids_folded = sum(len(r["consolidated_from"]) for r in folded_rows)
    n_mids_lost = sum(len(l[2]) for l in lost_folds)
    print(f"\ncurrent live folds: {len(folded_rows)} official rows / {n_mids_folded} folded M-IDs")
    print(f"gate would UNFOLD: {n_mids_lost} M-IDs across {len(lost_folds)} official rows")

    # SPAN sanity — the Session-40 wins must survive
    print("\nSPAN sanity (must all be KEPT):")
    for mid in ("FLSP M1342", "FLSP M1043", "FLSP M1362", "FLSP M1352",
                "FLSP M1045", "FLSP M1036", "FLSP M1337"):
        state = "UNFOLDED ✗" if mid in blocked_leaves else "kept ✓"
        print(f"  {mid:12s} {leaf_title(mid)[:42]:42s} {state}")

    print("\n--- the would-block list (leaf -> official), eyeball every line ---")
    for leaf, lt, oid, ot, n, kin, wts in sorted(blocked, key=lambda b: b[2]):
        live = " [LIVE FOLD]" if any(leaf in l[2] for l in lost_folds) else ""
        print(f"  {leaf:14s} {lt[:38]:38s} -> {oid:18s} {ot[:34]:34s} wit={n}{live}")
        for w in wts:
            print(f"      witness course: {w[:70]}")

    if "--json" in sys.argv:
        out = {"blocked": [{"leaf": b[0], "leaf_title": b[1], "target": b[2],
                            "official_title": b[3], "witnesses": b[4]} for b in blocked]}
        path = os.path.join(HERE, "kinship_analysis.json")
        json.dump(out, open(path, "w"), indent=1)
        print(f"\nwrote {path}")


if __name__ == "__main__":
    main()
