#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build a COLLEGE-SCOPED occupation -> CPL crosswalk (workbook + HTML).

Companion to kb/_build_partner_crosswalk.py. That tool answers a partner's
question -- "which of the occupations we train for can our students get college
credit for, and WHERE in California?". This one answers the college's half of
the same conversation:

    "Of those occupations, which can THIS college already carry, using the
     programs and courses it already runs -- and which of them already have a
     credit-recommendation exhibit built in MAP?"

First run: San Joaquin County Office of Education <-> San Joaquin Delta College
(2026-08-19), following the statewide SJCOE run of 2026-08-05.

WHY IT IS A SEPARATE TOOL
-------------------------
The partner tool ranks the whole state and deliberately does not privilege the
partner's in-county college. That is right for a referral question and wrong for
a partnership meeting, where the entire question is what ONE college can do. The
two also disagree about what counts as a good answer: the partner tool's best
outcome is "some college already offers this"; this tool's best outcome is "this
college teaches the content AND the exhibit already exists AND nobody has joined
the two up" -- which is a task, not a fact.

THREE SEPARATE JUDGMENTS, KEPT SEPARATE
---------------------------------------
  1. occupation -> credential      REUSED unchanged from
                                   kb/occupation_credential_map.json.
  2. occupation -> college offering  NEW per college, persisted in
                                   kb/<slug>_offering_map.json. Subject-matter
                                   judgment; a discussion agenda, never a
                                   determination -- faculty decide.
  3. college -> exhibit status     MECHANICAL, read from statewide_data.js
                                   adopter_names / potential_names. No judgment.

Keeping (2) and (3) in different columns is the point of the deliverable. A
college can teach something and not be listed; it can be listed and not teach
it. Collapsing them into one "match" score destroys exactly the distinction the
meeting needs.

ABSENCE HAS TWO MEANINGS AND THEY MUST NOT COLLAPSE (inherited from the partner
tool): "no exhibit exists anywhere in California" is a build opportunity and
often a chance to be first; "an exhibit exists and this college is not on it" is
an adoption task. Same empty cell, opposite next step.

    python3 kb/_build_college_offering_crosswalk.py \
        --college "San Joaquin Delta College" --slug delta \
        --partner "San Joaquin County Office of Education" \
        --partner-run kb/partner_crosswalk_out/2026-08-06-sjcoe/summary.json

Outputs to kb/college_crosswalk_out/<date>-<slug>/ :
    <date>_<College>_CPL_Opportunity_Crosswalk.xlsx   the distributable workbook
    <slug>_cpl_crosswalk.html                         the shareable visual
    crosswalk.json                                    the run receipt
Per the repo artifact policy the workbook and HTML are REGENERABLE and are not
committed; the offering map and this script are.
"""
import argparse
import datetime
import html
import json
import os
import re
import sys
from collections import defaultdict, Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Colleges a San Joaquin student could realistically reach. Mirrors the
# san-joaquin preset in kb/partner_crosswalk_regions.json.
REGION = ["Modesto Junior College", "Columbia College", "Merced College",
          "Sacramento City College", "American River College", "Folsom Lake College",
          "Cosumnes River College", "Las Positas College", "Fresno City College",
          "Sierra College", "Chabot College", "Los Rios"]

FIT_LABEL = {"C": "Confirmed alignment",
             "P": "Potential \u2014 validate with college",
             "N": "No apparent alignment"}

# Priority is derived, never authored: it is a pure function of (does the college
# teach it) x (does an exhibit exist, and is the college flagged on it). Authoring
# it by hand would let the ranking drift away from the two columns it summarises.
TIER = {1: ("1 \u2014 Adopt now",
            "The college teaches it and the MAP exhibit already exists, listing it as a "
            "potential adopter. Nothing to build."),
        2: ("2 \u2014 Build first-in-state",
            "The college teaches it and no credit recommendation exists anywhere in California."),
        3: ("3 \u2014 Adopt (needs MAP listing)",
            "The college teaches it and an exhibit exists, but the college is not listed against it."),
        4: ("4 \u2014 Validate, then adopt",
            "Partial coverage; an exhibit exists with the college flagged as a potential adopter."),
        5: ("5 \u2014 Validate",
            "Partial coverage at the college; no exhibit is teed up for it."),
        6: ("6 \u2014 Refer out / consider building",
            "No pathway at the college, but the credential exists elsewhere in California."),
        7: ("7 \u2014 Statewide gap",
            "No pathway at the college and no credit recommendation anywhere in California.")}


def load_window_json(rel_path):
    """Parse a `window.X = {...};` data file into a dict."""
    with open(os.path.join(ROOT, rel_path), encoding="utf-8") as fh:
        s = fh.read()
    return json.loads(s[s.index("{"):].rstrip().rstrip(";"))


def priority(fit, map_status):
    building = "build new" in map_status.lower()
    flagged = "potential adopter" in map_status
    if fit == "C":
        return 1 if flagged else (2 if building else 3)
    if fit == "P":
        return 4 if flagged else (5 if not building else 5)
    return 7 if building else 6


def build_rows(college, offering_map, partner_run):
    """Join the three judgments into one row per occupation."""
    ex = load_window_json("statewide_data.js")["exhibits"]
    info = defaultdict(lambda: dict(adopters=set(), potential=set(), statewide=False))
    for e in ex:
        ut = e.get("unified_title") or e.get("title")
        i = info[ut]
        i["adopters"].update(e.get("adopter_names") or [])
        i["potential"].update(e.get("potential_names") or [])
        if e.get("collaborative_type") == "CCC Collaborative":
            i["statewide"] = True

    occ_map = json.load(open(os.path.join(ROOT, "kb", "occupation_credential_map.json"),
                             encoding="utf-8"))["occupations"]
    summary = json.load(open(partner_run, encoding="utf-8"))
    status_by_key = {o["key"]: o for o in summary["occupations"]}
    rulings = offering_map["occupations"]

    missing = sorted(set(status_by_key) - set(rulings))
    if missing:
        # A silently unruled occupation would render as a blank row that reads like
        # "we looked and found nothing" -- the exact collapse this tool exists to avoid.
        sys.exit("ERROR: %d occupation(s) have no offering ruling: %s"
                 % (len(missing), ", ".join(missing[:8])))

    rows = []
    for key, meta in status_by_key.items():
        rul = rulings[key]
        creds = []
        for c in occ_map[key]["credentials"]:
            t = c["title"]
            i = info.get(t)
            if not i:
                creds.append(dict(title=t, tier=c["tier"], state="not in MAP index",
                                  adopters=0, statewide=False, region=[]))
                continue
            state = ("ADOPTED by " + college if college in i["adopters"]
                     else ("%s listed as potential adopter" % college if college in i["potential"]
                           else "%s not listed" % college))
            creds.append(dict(title=t, tier=c["tier"], state=state,
                              adopters=len(i["adopters"]), statewide=i["statewide"],
                              region=sorted(x for x in i["adopters"]
                                            if any(r in x for r in REGION))))
        states = {c["state"] for c in creds}
        if any(s.startswith("ADOPTED") for s in states):
            map_status = "Already articulated at %s" % college
        elif any("potential adopter" in s for s in states):
            map_status = "Exhibit exists \u2014 %s flagged as potential adopter" % college
        elif creds:
            map_status = "Exhibit exists \u2014 %s not listed" % college
        else:
            map_status = "No CPL exhibit anywhere (build new)"
        p = priority(rul["fit"], map_status)
        rows.append(dict(occupation=occ_map[key]["label"], key=key,
                         times=meta.get("times_listed", 0),
                         statewide_status=meta.get("status", ""),
                         fit=rul["fit"], programs=rul["programs"], courses=rul["courses"],
                         why=rul["why"], cpl=rul["cpl"], creds=creds,
                         map_status=map_status, prio=p, prio_label=TIER[p][0]))
    return rows


# Phrases that assert CPL does not exist ANYWHERE. Authored prose is often shared across
# a cluster of related occupations, and a sentence true of five rows can be false on the
# sixth -- which is how "no CPL anywhere" ended up on two rows that do have local
# wastewater exhibits (2026-08-19). One over-claim discredits the other 138 rows in a
# room with a college, so this is a hard failure, not a warning.
ABSENCE_CLAIMS = ("no cpl exists anywhere", "zero cpl anywhere", "no cpl anywhere",
                  "nothing exists anywhere in california", "no credit recommendation exists anywhere")


def check_absence_claims(rows):
    """Fail if a row's prose claims statewide absence while its exhibit list is non-empty."""
    bad = []
    for r in rows:
        if not r["creds"]:
            continue  # genuinely nothing matched; the claim is true
        prose = ("%s %s" % (r["why"], r["cpl"])).lower()
        hit = next((c for c in ABSENCE_CLAIMS if c in prose), None)
        if hit:
            bad.append((r["occupation"], hit, len(r["creds"])))
    if bad:
        lines = ["ERROR: %d row(s) claim no CPL exists anywhere, but matched exhibits:" % len(bad)]
        lines += ["  %-46s claims %-32r (%d exhibit(s) matched)" % b for b in bad]
        lines.append("Fix the ruling's prose in the offering map, or the claim ships to a college.")
        sys.exit("\n".join(lines))


def build_workbook(rows, college, partner, programs, out_path, DATE, NARR, V):
    SHORT = NARR.get('college_short', college)
    F='Arial'
    H_FILL=PatternFill('solid', fgColor='1F3864'); H_FONT=Font(name=F,size=10,bold=True,color='FFFFFF')
    TITLE=Font(name=F,size=15,bold=True,color='1F3864')
    SUB=Font(name=F,size=10,italic=True,color='555555')
    BODY=Font(name=F,size=10)
    BOLD=Font(name=F,size=10,bold=True)
    WRAP=Alignment(wrap_text=True, vertical='top')
    TOP=Alignment(vertical='top')
    CTR=Alignment(horizontal='center', vertical='top')
    GREEN=PatternFill('solid', fgColor='C6E0B4')   # confirmed
    AMBER=PatternFill('solid', fgColor='FFE699')   # potential
    GREY =PatternFill('solid', fgColor='E7E6E6')   # none
    BLUE =PatternFill('solid', fgColor='DDEBF7')
    BAND =PatternFill('solid', fgColor='F2F2F2')
    THIN=Side(style='thin', color='BFBFBF'); BOX=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

    FIT_FILL={'C':GREEN,'P':AMBER,'N':GREY}

    def hdr(ws, headers, row=1):
        for c,h in enumerate(headers,1):
            cell=ws.cell(row=row,column=c,value=h); cell.fill=H_FILL; cell.font=H_FONT
            cell.alignment=Alignment(wrap_text=True,vertical='center'); cell.border=BOX
        ws.row_dimensions[row].height=30
        ws.freeze_panes=ws.cell(row=row+1,column=1)

    def widths(ws, w):
        for i,x in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=x

    wb=Workbook()

    # ---------------------------------------------------------------- 1 READ ME
    ws=wb.active; ws.title='Read Me'
    widths(ws,[3,104,52])
    ws['B2']=f'{college} — CPL Opportunity Crosswalk'; ws['B2'].font=TITLE
    ws['B3']=f'Prepared for the {partner} ↔ {college} partnership conversation · {DATE}'
    ws['B3'].font=SUB
    r=5
    def block(title, lines):
        nonlocal r
        ws.cell(row=r,column=2,value=title).font=Font(name=F,size=11,bold=True,color='1F3864'); r+=1
        for ln in lines:
            c=ws.cell(row=r,column=2,value=ln); c.font=BODY; c.alignment=WRAP
            ws.row_dimensions[r].height=max(14, 13*(1+len(ln)//95)); r+=1
        r+=1

    for _title, _lines in NARR['readme']:
        block(_title, [ln.format(**V) for ln in _lines])

    def creds_txt(x, only_state=None):
        out=[]
        for c in x['creds']:
            if only_state and only_state not in c['state']: continue
            tag='statewide' if c['statewide'] else 'local'
            out.append(f"{c['title']} [{tag}; {c['adopters']} adopting colleges; {c['state']}]")
        return '\n'.join(out) or '— no credit-recommendation exhibit matched —'

    def region_txt(x):
        seen=[]
        for c in x['creds']:
            for col in c['region']:
                if col not in seen: seen.append(col)
        return ', '.join(seen) if seen else '—'

    # --------------------------------------------------- 2 PRIORITY OPPORTUNITIES
    ws=wb.create_sheet('Priority Opportunities')
    prio=[x for x in rows if x['prio']<=2]
    prio.sort(key=lambda x:(x['prio'], -x['times'], x['occupation']))
    ws['A1']=f'Priority opportunities to discuss with {college}'; ws['A1'].font=TITLE
    ws['A2']=(f'{college} already teaches the content AND (Priority 1) a MAP exhibit already exists listing it as a potential adopter, '
              f'or (Priority 2) nothing exists anywhere in California and {college} appears uniquely placed to build it.')
    ws['A2'].font=SUB; ws['A2'].alignment=WRAP; ws.row_dimensions[2].height=28
    ws.merge_cells('A2:H2')
    H=['Priority','Occupation',f'{college} program / certificate',f'{college} courses that carry the content',
       'Why it aligns','CPL opportunity','Credit-recommendation exhibit(s) in MAP','Regional colleges already adopting']
    hdr(ws,H,row=4)
    widths(ws,[19,32,38,42,60,60,52,30])
    rr=5
    for x in prio:
        vals=[x['prio_label'], x['occupation'], '\n'.join(x['programs']) or '—',
              '\n'.join(x['courses']) or '—', x['why'], x['cpl'], creds_txt(x), region_txt(x)]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row=rr,column=c,value=v); cell.font=BODY; cell.alignment=WRAP; cell.border=BOX
            if c==1: cell.fill=GREEN if x['prio']==1 else BLUE; cell.font=BOLD
            if c==2: cell.font=BOLD
        ws.row_dimensions[rr].height=max(46, 12*(max(len(x['why']),len(x['cpl']))//58+1))
        rr+=1

    # ------------------------------------------------------------ 3 FULL CROSSWALK
    ws=wb.create_sheet('Full Crosswalk')
    H=['Occupation','Times on partner list','Alignment status','Priority',f'{college} program / certificate',
       f'{college} course(s)','Why it may align','Potential CPL connection / opportunity',
       'Existing MAP exhibit?','Credit-recommendation exhibit(s) matched','Regional colleges already adopting',
       'Statewide CPL status (all CA colleges)']
    hdr(ws,H)
    widths(ws,[34,9,24,19,36,40,58,58,30,50,28,20])
    srt=sorted(rows,key=lambda x:(x['prio'], x['occupation']))
    rr=2
    for x in srt:
        vals=[x['occupation'], x['times'], FIT_LABEL[x['fit']], x['prio_label'],
              '\n'.join(x['programs']) or '—', '\n'.join(x['courses']) or '—',
              x['why'], x['cpl'], x['map_status'], creds_txt(x), region_txt(x), x['statewide_status']]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row=rr,column=c,value=v); cell.font=BODY; cell.alignment=WRAP if c not in (2,) else CTR
            cell.border=BOX
            if c==3: cell.fill=FIT_FILL[x['fit']]; cell.font=BOLD
            if c==1: cell.font=BOLD
        ws.row_dimensions[rr].height=max(44, 12*(max(len(x['why']),len(x['cpl']))//56+1))
        rr+=1
    ws.auto_filter.ref=f'A1:L{rr-1}'
    LAST=rr-1

    # --------------------------------------------------- summary formulas on Read Me
    ws2=wb['Read Me']
    sr=r+1
    ws2.cell(row=sr,column=2,value='Counts (live — recalculate if you edit the Full Crosswalk tab)').font=Font(name=F,size=11,bold=True,color='1F3864')
    sr+=1
    counts=[(f'Occupations on the {partner} list', f'=COUNTA(\'Full Crosswalk\'!A2:A{LAST})'),
            (f'Confirmed alignment with a {SHORT} offering', f'=COUNTIF(\'Full Crosswalk\'!C2:C{LAST},"{FIT_LABEL["C"]}")'),
            ('Potential alignment — validate with college', f'=COUNTIF(\'Full Crosswalk\'!C2:C{LAST},"{FIT_LABEL["P"]}")'),
            (f'No apparent alignment at {SHORT}', f'=COUNTIF(\'Full Crosswalk\'!C2:C{LAST},"{FIT_LABEL["N"]}")'),
            (f'Priority 1 — {SHORT} teaches it AND the exhibit already exists', f'=COUNTIF(\'Full Crosswalk\'!D2:D{LAST},"1 — Adopt now")'),
            (f'Priority 2 — {SHORT} teaches it, nothing exists statewide (build first)', f'=COUNTIF(\'Full Crosswalk\'!D2:D{LAST},"2 — Build first-in-state")'),
            (f'Occupations where MAP flags {SHORT} as a potential adopter', f'=COUNTIF(\'Full Crosswalk\'!I2:I{LAST},"Exhibit exists — {college} flagged as potential adopter")'),
            (f'Occupations already articulated at {SHORT} today', f'=COUNTIF(\'Full Crosswalk\'!I2:I{LAST},"Already articulated at {college}")')]
    for label,f_ in counts:
        ws2.cell(row=sr,column=2,value=label).font=BODY
        c=ws2.cell(row=sr,column=3,value=f_); c.font=BOLD; c.alignment=Alignment(horizontal='center')
        c.fill=BLUE; c.border=BOX
        sr+=1

    # ----------------------------------------------------------- 4 DELTA CPL TODAY
    ex=load_window_json('statewide_data.js')['exhibits']
    D=college
    adopted={}
    for e in ex:
        if D in (e.get('adopter_names') or []):
            ut=e.get('unified_title') or e.get('title')
            adopted.setdefault(ut, dict(ut=ut, cpl=e.get('cpl_type',''), disc=e.get('discipline',''),
                                        sector=e.get('sector',''), sw=e.get('collaborative_type')=='CCC Collaborative'))
    ws=wb.create_sheet(f'{SHORT} CPL Today'[:31])
    ws['A1']=f'Every CPL exhibit {college} has articulated today'; ws['A1'].font=TITLE
    ws['A2']=(f'{len(adopted)} exhibits. Exactly ONE is career/technical (POST Basic Academy); the rest are AP and CLEP '
              'credit-by-exam. This is the starting point the conversation is trying to change.')
    ws['A2'].font=SUB; ws['A2'].alignment=WRAP; ws.merge_cells('A2:E2'); ws.row_dimensions[2].height=28
    hdr(ws,['Credential / exhibit','Type','CPL type','Discipline','Sector'],row=4)
    widths(ws,[52,20,26,34,38])
    rr=5
    for k in sorted(adopted, key=lambda k:(not (not adopted[k]['ut'].upper().startswith(('AP ','CLEP'))), k)):
        a=adopted[k]
        career = not a['ut'].upper().startswith(('AP ','CLEP','DSST','IB '))
        vals=[a['ut'], 'CAREER / TECHNICAL' if career else 'Academic exam', a['cpl'], a['disc'], a['sector']]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row=rr,column=c,value=v); cell.font=BOLD if career else BODY
            cell.alignment=WRAP; cell.border=BOX
            if career: cell.fill=GREEN
        rr+=1
    ws.auto_filter.ref=f'A4:E{rr-1}'

    # ------------------------------------------------------ 5 NO ALIGNMENT / REFER
    ws=wb.create_sheet('No Alignment — Refer Out')
    none=[x for x in rows if x['fit']=='N']
    none.sort(key=lambda x:('build new' in x['map_status'].lower(), x['occupation']))
    ws['A1']=f'Occupations with no apparent alignment at {college}'; ws['A1'].font=TITLE
    ws['A2']=('Split into two very different cases: (a) a credit recommendation EXISTS elsewhere in California, so the student should be '
              f'referred — or {college} should consider building the program; and (b) nothing exists anywhere, which is a statewide gap the '
              'partnership could help make the case for.')
    ws['A2'].font=SUB; ws['A2'].alignment=WRAP; ws.merge_cells('A2:F2'); ws.row_dimensions[2].height=32
    hdr(ws,['Occupation','Case',f'Why nothing at {college}','What to do instead','Exhibit(s) that exist elsewhere','Regional colleges already adopting'],row=4)
    widths(ws,[34,32,54,58,48,30])
    rr=5
    for x in none:
        case=('CPL exists elsewhere — refer or build' if 'build new' not in x['map_status'].lower()
              else 'No CPL anywhere in California')
        vals=[x['occupation'], case, x['why'], x['cpl'], creds_txt(x), region_txt(x)]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row=rr,column=c,value=v); cell.font=BODY; cell.alignment=WRAP; cell.border=BOX
            if c==2: cell.fill=BLUE if 'elsewhere' in case else GREY; cell.font=BOLD
            if c==1: cell.font=BOLD
        ws.row_dimensions[rr].height=max(40, 12*(max(len(x['why']),len(x['cpl']))//54+1))
        rr+=1
    ws.auto_filter.ref=f'A4:F{rr-1}'

    # ------------------------------------------------- 6 DELTA OFFERINGS REFERENCED
    progs=programs
    act=[p for p in progs if p['status'] in ('Active','Approved')]
    ws=wb.create_sheet(f'{SHORT} Offerings (source)'[:31])
    ws['A1']=f'{college} — active program inventory used for this crosswalk'; ws['A1'].font=TITLE
    ws['A2']=(f'{len(act)} active/approved programs from the committed CCCCO COCI program export, '
              f'{sum(1 for p in act if p["cte"])} of them flagged CTE. This is the evidence base behind the "program / certificate" column.')
    ws['A2'].font=SUB; ws['A2'].alignment=WRAP; ws.merge_cells('A2:G2'); ws.row_dimensions[2].height=28
    hdr(ws,['Program title','Award','CTE?','Transfer (ADT)?','TOP code','CIP code','Units','COCI control #'],row=4)
    widths(ws,[54,52,8,14,11,11,16,14])
    rr=5
    for p in sorted(act,key=lambda p:(not p['cte'], p['title'])):
        vals=[p['title'],p['award'],'CTE' if p['cte'] else '','ADT' if p['xfer'] else '',
              p['top'],p['cip'],p['units'],p['ctrl']]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row=rr,column=c,value=v); cell.font=BODY; cell.alignment=WRAP if c in(1,2) else TOP
            cell.border=BOX
            if c==3 and v: cell.fill=GREEN; cell.alignment=CTR
        rr+=1
    ws.auto_filter.ref=f'A4:H{rr-1}'
    wb.save(out_path)
    return out_path


def build_html_page(rows, college, partner, out_path, DATE, NARR, V):
    FIT={'C':('confirmed','Confirmed'),'P':('validate','Validate'),'N':('none','No match')}
    e=html.escape
    def credlist(x):
        out=[]
        for c in x['creds']:
            badge='sw' if c['statewide'] else 'loc'
            st=c['state']
            cls='yes' if 'potential' in st else ('adopted' if 'ADOPTED' in st else 'no')
            out.append(f'<li><span class="cred-name">{e(c["title"])}</span>'
                       f'<span class="tag {badge}">{"statewide" if c["statewide"] else "local"}</span>'
                       f'<span class="tag ct">{c["adopters"]} colleges</span>'
                       f'<span class="tag {cls}">{e(st)}</span></li>')
        return '<ul class="creds">'+''.join(out)+'</ul>' if out else '<p class="nocred">No credit-recommendation exhibit exists anywhere in California for this occupation.</p>'
    def chips(items, cls):
        return ''.join(f'<span class="chip {cls}">{e(i)}</span>' for i in items) or '<span class="chip empty">none</span>'

    cards=[]
    for x in sorted(rows,key=lambda r:(r['prio'], r['occupation'])):
        f_cls,f_lab=FIT[x['fit']]
        cards.append(f'''<article class="row" data-fit="{f_cls}" data-tier="{x['prio']}" data-q="{e((x['occupation']+' '+' '.join(x['programs'])+' '+' '.join(x['courses'])).lower())}">
    <header class="row-head">
      <div class="row-id"><span class="tier t{x['prio']}">P{x['prio']}</span><h3>{e(x['occupation'])}</h3></div>
      <span class="fit {f_cls}">{f_lab}</span>
    </header>
    <div class="row-body">
      <div class="col">
        <p class="lbl">Program / certificate</p>{chips(x['programs'],'prog')}
        <p class="lbl">Courses that carry the content</p>{chips(x['courses'],'course')}
      </div>
      <div class="col">
        <p class="lbl">Why it may align</p><p class="prose">{e(x['why'])}</p>
        <p class="lbl">CPL opportunity</p><p class="prose op">{e(x['cpl'])}</p>
      </div>
      <div class="col">
        <p class="lbl">MAP exhibit status</p><p class="mapst">{e(x['map_status'])}</p>
        <p class="lbl">Credit recommendations matched</p>{credlist(x)}
      </div>
    </div></article>''')

    n=len(rows)
    nC=sum(1 for x in rows if x['fit']=='C'); nP=sum(1 for x in rows if x['fit']=='P'); nN=sum(1 for x in rows if x['fit']=='N')
    n1=sum(1 for x in rows if x['prio']==1); n2=sum(1 for x in rows if x['prio']==2)
    npot=sum(1 for x in rows if 'potential adopter' in x['map_status'])

    tierlegend=''.join(f'<div class="tl"><span class="tier t{k}">P{k}</span><div><strong>{e(v[0])}</strong><p>{e(v[1])}</p></div></div>'
                       for k,v in sorted(TIER.items()) if any(x['prio']==k for x in rows))

    _fmt = lambda t: t.format(**V)
    hero_title = e(NARR['headline_title'].format(**V))
    hero_dek = e(NARR['headline_dek'].format(**V))
    finding_heading = e(NARR['finding_heading'])
    finding_paras = ''.join('<p>%s</p>' % _fmt(x) for x in NARR['finding_paras'])
    footer_paras = ''.join('<p>%s</p>' % _fmt(x) for x in NARR['footer'])
    partner_name = e(partner)
    source_note = e(NARR.get('source_note', 'See the workbook Read Me tab for sources'))
    page_title = e(NARR.get('page_title', '%s CPL Crosswalk' % college))

    DOC=f'''<title>{page_title}</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Archivo:wght@500;600;700&family=Source+Sans+3:ital,wght@0,400;0,600;1,400&family=IBM+Plex+Mono:wght@400;500&display=swap">
    <style>
    :root{{
      --ground:#F2F4F5; --surface:#FFFFFF; --surface-2:#E9EDEF;
      --ink:#16202B; --ink-2:#4A5A68; --ink-3:#75858F;
      --line:#D3DBDF; --line-2:#BCC8CE;
      --accent:#14505E; --accent-soft:#DCEAED;
      --ok:#2F6B3F; --ok-soft:#DEEBE0;
      --warn:#8C5E12; --warn-soft:#F6EBD5;
      --mute:#5A6B76; --mute-soft:#E4E9EB;
      --shadow:0 1px 2px rgba(22,32,43,.06),0 4px 14px rgba(22,32,43,.05);
    }}
    @media (prefers-color-scheme:dark){{ :root:not([data-theme="light"]){{
      --ground:#0F161B; --surface:#171F26; --surface-2:#1E2831;
      --ink:#E4ECF1; --ink-2:#A7B7C1; --ink-3:#7F919C;
      --line:#2A353E; --line-2:#3A4854;
      --accent:#5FB4C4; --accent-soft:#12313A;
      --ok:#7FC08D; --ok-soft:#16301D;
      --warn:#DDAE5C; --warn-soft:#332612;
      --mute:#93A3AD; --mute-soft:#212B33;
      --shadow:0 1px 2px rgba(0,0,0,.4),0 4px 16px rgba(0,0,0,.3);
    }}}}
    :root[data-theme="dark"]{{
      --ground:#0F161B; --surface:#171F26; --surface-2:#1E2831;
      --ink:#E4ECF1; --ink-2:#A7B7C1; --ink-3:#7F919C;
      --line:#2A353E; --line-2:#3A4854;
      --accent:#5FB4C4; --accent-soft:#12313A;
      --ok:#7FC08D; --ok-soft:#16301D;
      --warn:#DDAE5C; --warn-soft:#332612;
      --mute:#93A3AD; --mute-soft:#212B33;
      --shadow:0 1px 2px rgba(0,0,0,.4),0 4px 16px rgba(0,0,0,.3);
    }}
    *{{box-sizing:border-box}}
    body{{margin:0;background:var(--ground);color:var(--ink);
      font-family:"Source Sans 3",ui-sans-serif,system-ui,sans-serif;font-size:16px;line-height:1.55;
      -webkit-font-smoothing:antialiased}}
    .wrap{{max-width:1180px;margin:0 auto;padding:0 24px 80px}}
    h1,h2,h3{{font-family:Archivo,ui-sans-serif,system-ui,sans-serif;text-wrap:balance;margin:0}}
    .mono{{font-family:"IBM Plex Mono",ui-monospace,monospace}}

    header.top{{padding:52px 0 30px;border-bottom:2px solid var(--ink)}}
    .eyebrow{{font-size:12px;letter-spacing:.13em;text-transform:uppercase;color:var(--accent);
      font-weight:600;margin:0 0 14px;font-family:Archivo,sans-serif}}
    h1{{font-size:clamp(30px,4.4vw,46px);font-weight:700;letter-spacing:-.02em;line-height:1.08}}
    .dek{{margin:16px 0 0;max-width:66ch;font-size:17.5px;color:var(--ink-2)}}
    .meta{{margin-top:20px;display:flex;flex-wrap:wrap;gap:8px 22px;font-size:13px;color:var(--ink-3)}}
    .meta span::before{{content:"";display:inline-block;width:5px;height:5px;border-radius:50%;
      background:var(--line-2);margin-right:8px;vertical-align:middle}}

    .finding{{margin:34px 0 0;background:var(--surface);border:1px solid var(--line);
      border-left:4px solid var(--accent);border-radius:3px;padding:22px 26px;box-shadow:var(--shadow)}}
    .finding h2{{font-size:19px;font-weight:600;margin-bottom:8px}}
    .finding p{{margin:0 0 10px;max-width:78ch;color:var(--ink-2)}}
    .finding p:last-child{{margin-bottom:0}}
    .finding strong{{color:var(--ink);font-weight:600}}

    .stats{{display:grid;grid-template-columns:repeat(auto-fit,minmax(168px,1fr));gap:1px;
      background:var(--line);border:1px solid var(--line);margin:34px 0 0;border-radius:3px;overflow:hidden}}
    .stat{{background:var(--surface);padding:20px 18px}}
    .stat b{{display:block;font-family:Archivo,sans-serif;font-size:38px;font-weight:700;line-height:1;
      letter-spacing:-.02em;font-variant-numeric:tabular-nums;color:var(--accent)}}
    .stat.zero b{{color:var(--ink)}}
    .stat span{{display:block;margin-top:9px;font-size:13px;color:var(--ink-2);line-height:1.4}}

    h2.sec{{font-size:13px;letter-spacing:.13em;text-transform:uppercase;font-weight:600;
      color:var(--ink-3);margin:52px 0 16px;padding-bottom:9px;border-bottom:1px solid var(--line)}}

    .legend{{display:grid;grid-template-columns:repeat(auto-fit,minmax(290px,1fr));gap:12px}}
    .tl{{display:flex;gap:11px;align-items:flex-start;background:var(--surface);border:1px solid var(--line);
      border-radius:3px;padding:13px 15px}}
    .tl strong{{font-family:Archivo,sans-serif;font-size:14px;display:block}}
    .tl p{{margin:3px 0 0;font-size:13.5px;color:var(--ink-2);line-height:1.42}}

    .controls{{position:sticky;top:0;z-index:20;background:var(--ground);padding:16px 0 13px;
      margin-bottom:4px;border-bottom:1px solid var(--line);display:flex;flex-wrap:wrap;gap:10px;align-items:center}}
    .filters{{display:flex;flex-wrap:wrap;gap:7px}}
    button.f{{font:inherit;font-size:13.5px;font-weight:600;font-family:Archivo,sans-serif;cursor:pointer;
      border:1px solid var(--line-2);background:var(--surface);color:var(--ink-2);
      padding:6px 13px;border-radius:100px;transition:background .12s,color .12s,border-color .12s}}
    button.f:hover{{border-color:var(--accent);color:var(--ink)}}
    button.f[aria-pressed="true"]{{background:var(--accent);border-color:var(--accent);color:var(--ground)}}
    button.f:focus-visible,input:focus-visible{{outline:2px solid var(--accent);outline-offset:2px}}
    input.search{{font:inherit;font-size:14px;padding:6px 12px;border:1px solid var(--line-2);border-radius:100px;
      background:var(--surface);color:var(--ink);min-width:210px;flex:1}}
    .count{{font-size:13px;color:var(--ink-3);font-variant-numeric:tabular-nums;white-space:nowrap}}

    .row{{background:var(--surface);border:1px solid var(--line);border-radius:3px;margin-top:11px;
      box-shadow:var(--shadow);overflow:hidden}}
    .row-head{{display:flex;justify-content:space-between;align-items:center;gap:14px;flex-wrap:wrap;
      padding:13px 18px;background:var(--surface-2);border-bottom:1px solid var(--line)}}
    .row-id{{display:flex;align-items:center;gap:11px;min-width:0}}
    .row-head h3{{font-size:16.5px;font-weight:600;letter-spacing:-.01em}}
    .tier{{font-family:"IBM Plex Mono",monospace;font-size:11.5px;font-weight:500;padding:3px 7px;border-radius:2px;
      background:var(--mute-soft);color:var(--mute);border:1px solid var(--line-2);white-space:nowrap}}
    .tier.t1{{background:var(--accent);color:var(--surface);border-color:var(--accent)}}
    .tier.t2{{background:var(--warn-soft);color:var(--warn);border-color:var(--warn)}}
    .fit{{font-family:Archivo,sans-serif;font-size:11.5px;font-weight:600;letter-spacing:.05em;
      text-transform:uppercase;padding:4px 10px;border-radius:100px;white-space:nowrap}}
    .fit.confirmed{{background:var(--ok-soft);color:var(--ok)}}
    .fit.validate{{background:var(--warn-soft);color:var(--warn)}}
    .fit.none{{background:var(--mute-soft);color:var(--mute)}}

    .row-body{{display:grid;grid-template-columns:0.9fr 1.3fr 1fr;gap:0}}
    .col{{padding:16px 18px;border-right:1px solid var(--line)}}
    .col:last-child{{border-right:0}}
    .lbl{{font-family:Archivo,sans-serif;font-size:10.5px;letter-spacing:.1em;text-transform:uppercase;
      color:var(--ink-3);font-weight:600;margin:0 0 7px}}
    .col .lbl:not(:first-child){{margin-top:15px}}
    .prose{{margin:0;font-size:14.5px;line-height:1.5;color:var(--ink-2)}}
    .prose.op{{color:var(--ink)}}
    .chip{{display:inline-block;font-size:12.5px;padding:2.5px 8px;border-radius:2px;margin:0 4px 4px 0;
      border:1px solid var(--line-2);background:var(--ground);color:var(--ink-2);line-height:1.4}}
    .chip.course{{font-family:"IBM Plex Mono",monospace;font-size:11.5px}}
    .chip.empty{{opacity:.55;font-style:italic}}
    .mapst{{margin:0;font-size:14px;font-weight:600;color:var(--ink)}}
    .creds{{list-style:none;margin:0;padding:0;display:flex;flex-direction:column;gap:8px}}
    .creds li{{font-size:13px;line-height:1.45;padding-bottom:8px;border-bottom:1px dotted var(--line)}}
    .creds li:last-child{{border-bottom:0;padding-bottom:0}}
    .cred-name{{display:block;font-weight:600;color:var(--ink);margin-bottom:3px}}
    .tag{{display:inline-block;font-size:10.5px;padding:1.5px 6px;border-radius:2px;margin:0 4px 2px 0;
      font-family:"IBM Plex Mono",monospace;border:1px solid var(--line-2);color:var(--ink-3)}}
    .tag.sw{{background:var(--accent-soft);color:var(--accent);border-color:var(--accent)}}
    .tag.yes{{background:var(--ok-soft);color:var(--ok);border-color:var(--ok)}}
    .tag.adopted{{background:var(--ok);color:var(--surface);border-color:var(--ok)}}
    .nocred{{margin:0;font-size:13.5px;font-style:italic;color:var(--ink-3)}}
    .empty-state{{padding:40px;text-align:center;color:var(--ink-3);font-size:15px}}

    footer{{margin-top:56px;padding-top:22px;border-top:1px solid var(--line);
      font-size:13px;color:var(--ink-3);max-width:80ch}}
    footer p{{margin:0 0 8px}}
    @media (max-width:900px){{ .row-body{{grid-template-columns:1fr}}
      .col{{border-right:0;border-bottom:1px solid var(--line)}} .col:last-child{{border-bottom:0}} }}
    @media (prefers-reduced-motion:reduce){{*{{transition:none!important;animation:none!important}}}}
    </style>

    <div class="wrap">
    <header class="top">
      <p class="eyebrow">{partner_name} &nbsp;&middot;&nbsp; Partnership working document</p>
      <h1>{hero_title}</h1>
      <p class="dek">{hero_dek}</p>
      <div class="meta"><span>Prepared {DATE}</span><span>{source_note}</span></div>
    </header>

    <div class="finding">
  <h2>{finding_heading}</h2>
  {finding_paras}
</div>

    <div class="stats">
      <div class="stat"><b>{n}</b><span>Occupations on the partner list</span></div>
      <div class="stat"><b>{nC}</b><span>Confirmed matching offering at the college</span></div>
      <div class="stat"><b>{n1}</b><span>Exhibit exists <em>and</em> the college is flagged as a potential adopter</span></div>
      <div class="stat"><b>{n2}</b><span>College teaches it, nothing exists statewide</span></div>
      <div class="stat zero"><b>0</b><span>Articulated at the college today</span></div>
    </div>

    <h2 class="sec">How the priority tiers work</h2>
    <div class="legend">{tierlegend}</div>

    <h2 class="sec">The register &mdash; all {n} occupations</h2>
    <div class="controls">
      <div class="filters">
        <button class="f" data-filter="all" aria-pressed="true">All</button>
        <button class="f" data-filter="t1" aria-pressed="false">Adopt now</button>
        <button class="f" data-filter="t2" aria-pressed="false">Build first-in-state</button>
        <button class="f" data-filter="confirmed" aria-pressed="false">Confirmed</button>
        <button class="f" data-filter="validate" aria-pressed="false">Validate</button>
        <button class="f" data-filter="none" aria-pressed="false">No match</button>
      </div>
      <input class="search" type="search" placeholder="Search occupation, program or course&hellip;" aria-label="Search">
      <span class="count"></span>
    </div>
    <div id="list">{''.join(cards)}</div>
    <p class="empty-state" id="empty" hidden>Nothing matches that filter.</p>

    <footer>
  {footer_paras}
</footer>
    </div>
    '''

    DOC += """
    <script>
    (function(){
      var rows=[].slice.call(document.querySelectorAll('.row'));
      var btns=[].slice.call(document.querySelectorAll('button.f'));
      var search=document.querySelector('.search');
      var count=document.querySelector('.count');
      var empty=document.getElementById('empty');
      var active='all';
      function apply(){
        var q=(search.value||'').trim().toLowerCase(), shown=0;
        rows.forEach(function(r){
          var okF = active==='all' ||
            (active.charAt(0)==='t' ? r.dataset.tier===active.slice(1) : r.dataset.fit===active);
          var okQ = !q || r.dataset.q.indexOf(q)>-1;
          var vis = okF && okQ;
          r.hidden = !vis; if(vis) shown++;
        });
        count.textContent = shown + ' of ' + rows.length + ' shown';
        empty.hidden = shown>0;
      }
      btns.forEach(function(b){ b.addEventListener('click',function(){
        active=b.dataset.filter;
        btns.forEach(function(o){ o.setAttribute('aria-pressed', String(o===b)); });
        apply();
      }); });
      search.addEventListener('input',apply);
      apply();
    })();
    </script>"""
    with open(out_path, 'w', encoding='utf-8') as fh:
        fh.write(DOC)
    return out_path


def load_programs(college):
    """Active/approved COCI programs for one college, from the committed export."""
    d = load_window_json("coci_programs_data.js")
    # COCI names colleges in short upper case ("SAN JOAQUIN DELTA"); match on the
    # longest college entry contained in the MAP name, so "College"/"Junior College"
    # suffixes do not defeat the join.
    up = college.upper()
    cands = [c for c in d["colleges"] if c in up]
    if not cands:
        sys.exit("ERROR: no COCI college entry matches %r" % college)
    name = max(cands, key=len)
    idx = d["colleges"].index(name)
    progs = [dict(ctrl=r[1], title=r[2], top=r[3], cip=r[4], award=d["awards"][r[5]],
                  status=d["statuses"][r[6]], units=r[7], xfer=r[8], cte=r[9])
             for r in d["rows"] if r[0] == idx]
    return [p for p in progs if p["status"] in ("Active", "Approved")]


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--college", required=True, help="Exact college name as it appears in MAP data")
    ap.add_argument("--slug", required=True, help="Short slug; picks kb/<slug>_offering_map.json")
    ap.add_argument("--partner", required=True, help="Partner org name, for the cover")
    ap.add_argument("--partner-run", required=True,
                    help="summary.json from the matching kb/_build_partner_crosswalk.py run")
    ap.add_argument("--date", default=datetime.date.today().isoformat())
    a = ap.parse_args()

    om_path = os.path.join(ROOT, "kb", "%s_offering_map.json" % a.slug)
    offering_map = json.load(open(om_path, encoding="utf-8"))
    rows = build_rows(a.college, offering_map, a.partner_run)
    check_absence_claims(rows)
    programs = load_programs(a.college)

    NARR = offering_map.get("_narrative")
    if not NARR:
        sys.exit("ERROR: %s carries no _narrative block. The workbook and page copy are FINDINGS "
                 "about one college and live with its rulings, not in this script." % om_path)
    fits = Counter(r["fit"] for r in rows)
    tiers = Counter(r["prio"] for r in rows)
    V = dict(college=a.college, partner=a.partner, n=len(rows),
             nC=fits["C"], nP=fits["P"], nN=fits["N"], n1=tiers[1], n2=tiers[2])

    out_dir = os.path.join(ROOT, "kb", "college_crosswalk_out", "%s-%s" % (a.date, a.slug))
    os.makedirs(out_dir, exist_ok=True)
    safe = re.sub(r"[^A-Za-z0-9]+", "_", a.college).strip("_")
    xlsx_path = os.path.join(out_dir, "%s_%s_CPL_Opportunity_Crosswalk.xlsx"
                             % (a.date.replace("-", ""), safe))
    html_path = os.path.join(out_dir, "%s_cpl_crosswalk.html" % a.slug)

    build_workbook(rows, a.college, a.partner, programs, xlsx_path, a.date, NARR, V)
    build_html_page(rows, a.college, a.partner, html_path, a.date, NARR, V)
    with open(os.path.join(out_dir, "crosswalk.json"), "w", encoding="utf-8") as fh:
        json.dump(rows, fh, indent=1, ensure_ascii=False)

    print("occupations: %d | confirmed %d / potential %d / none %d"
          % (len(rows), fits["C"], fits["P"], fits["N"]))
    print("P1 adopt now: %d | P2 build first-in-state: %d" % (tiers[1], tiers[2]))
    print("workbook: %s" % xlsx_path)
    print("html:     %s" % html_path)


if __name__ == "__main__":
    main()
