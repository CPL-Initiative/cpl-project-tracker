"""
Audit kb/discipline_inference.json's subject_map for COLLEGE-HOMONYM subject
codes — the CRIM M1003 failure class (Session 45, 2026-06-11).

Subject codes are college-local vocabulary: the same 2–5 letters can name
different departments at different colleges. The lexicon's authoring rule has
always been "subject codes whose member titles are unambiguously ONE
discipline" — but nothing enforced it. The motivating case: subject_map had
`CADM → Administration of Justice`, correct for Bakersfield's CADM
(Corrections ADMinistration, 47 courses, TOP 21xx) and wrong for Merced's
CADM (Computer-Aided Drafting & Manufacturing, TOP 09xx). The bad fill then
cascaded: "Introduction to 3D" got discipline Administration of Justice at
0.8 confidence, and the canonical-SUBJ4 apply re-keyed the row to CRIM M1003
— laundering a lexicon error into an authoritative-looking identity.

Detection (per subject_map entry, from the raw COCI list):
  * Group the subject's raw courses by college; a college "votes" with the
    plurality 2-digit TOP division of its own courses, counted only when the
    college is INTERNALLY consistent (plurality share >= MIN_SHARE) and has
    >= MIN_COURSES courses — internal consistency is what distinguishes "a
    different department uses this code" from ordinary TOP-assignment noise.
  * A subject is a HOMONYM CANDIDATE when its consistent colleges disagree:
    >= 1 college's division differs from the subject's global majority
    division.
  * TITLE EVIDENCE then grades each minority college, because division
    disagreement alone conflates two classes: per minority college, the
    share of its course titles containing >= 1 token of the mapped
    discipline's name. HIGH overlap = "consistent" (the college only
    TOP-codes the same content under a different division — Mt. SAC files
    art history under Humanities; the fill is CORRECT, keep the entry).
    LOW overlap = "foreign" (a different department owns the code there —
    Modesto's AP is Anatomy & Physiology, not Photography; the entry must
    be removed/scoped). Thresholds: >= 0.5 consistent, < 0.2 foreign,
    else mixed. A hint for the curator, not a verdict.

Suspect-fill report (which staging rows the entry poisoned):
  * singletons: discipline_source == "subject_map", subject == S, college on
    the minority side of a flagged S.
  * minted M-IDs: discipline_source == "subject_map", discipline == map[S],
    S among member subjects, and the row's own member TOP divisions'
    plurality differs from S's global majority division.

READ-ONLY over the staging files; writes the receipt
kb/subject_map_audit.json (+ console report). Re-runnable. The repair path:
convert the entry to a COLLEGE-SCOPED form ({"discipline": ...,
"colleges": [majority side]}) — or remove it outright when no side is
clean — then re-run the inference chain. kb/_infer_disciplines.py RETRACTS
its own fills that no longer re-derive (see that script's header), and the
TOP/description/division passes re-fill the blanks from per-row evidence.
Scoping beats removal when the majority side is a real department: flat
removal of GSS would have degraded Lassen's 118 correct Gunsmithing fills
to a coarse division umbrella just to fix Saddleback's 6.

Run from repo root:  python3 kb/_audit_subject_map.py
"""
import json
import os
from collections import Counter, defaultdict
from datetime import date

HERE = os.path.dirname(os.path.abspath(__file__))
RAW_XLSX = os.path.join(HERE, "reference", "coci_course_list.xlsx")
OUT = os.path.join(HERE, "subject_map_audit.json")

MIN_COURSES = 3     # a college needs this many courses on the code to vote
MIN_SHARE = 0.6     # …and this internal plurality share (consistency gate)


def load(name):
    with open(os.path.join(HERE, name), encoding="utf-8") as f:
        return json.load(f)


def division(top_code):
    ts = str(top_code or "").strip()
    return ts[:2] if len(ts) >= 2 and ts[:2].isdigit() else None


_STOP = {"of", "and", "the", "in", "for", "to", "a", "an", "studies", "study",
         "technology", "technologies", "general", "related"}


def disc_tokens(disc):
    """Distinctive lowercase tokens of an MQ discipline name (stopword- and
    generic-word-free, so 'Culinary Arts/Food Technology' doesn't match every
    title containing 'technology')."""
    import re as _re
    toks = {t for t in _re.split(r"[^a-z]+", str(disc).lower()) if len(t) >= 3}
    toks |= {t[:-1] for t in toks if t.endswith("s")}   # arts -> art too
    return toks - _STOP


def title_overlap(titles, tokens):
    """Share of titles containing >= 1 discipline token (whole-word-ish)."""
    import re as _re
    if not titles or not tokens:
        return 0.0
    hit = 0
    for t in titles:
        words = set(_re.split(r"[^a-z]+", str(t).lower()))
        # crude singular/plural bridge: 'arts' matches 'art' and vice versa
        words |= {w[:-1] for w in words if w.endswith("s")}
        if words & tokens:
            hit += 1
    return round(hit / len(titles), 2)


def main():
    lex = load("discipline_inference.json")
    subject_map = lex["subject_map"]
    # Curator clearances (lexicon _subject_map_notes.audit_cleared): entries a
    # human verified by reading the minority college's actual titles — the
    # title heuristic under-credits abstract discipline names ('Multimedia',
    # 'Computer Information Systems'), so e.g. Solano's MS-Office BSOT courses
    # grade "foreign" while being correctly disciplined. Cleared entries stay
    # in the report (verdict "cleared (curator)") but produce NO suspects.
    cleared = set((lex.get("_subject_map_notes") or {}).get("audit_cleared") or {})
    div_names = {k: v for k, v in load("top_division_discipline_map.json")
                 .get("map", {}).items()}

    # ---- raw COCI sweep: per (subject-in-map, college) division counters ----
    from openpyxl import load_workbook
    wb = load_workbook(RAW_XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rit = ws.iter_rows(values_only=True)
    next(rit)
    by_subj = defaultdict(lambda: defaultdict(Counter))  # subj -> college -> div counter
    titles = defaultdict(lambda: defaultdict(list))      # subj -> college -> [titles]
    for row in rit:
        college, subj, title, topc = row[0], row[2], row[4], row[8]
        if not college or subj is None:
            continue
        s = str(subj).strip().upper()
        if s not in subject_map:
            continue
        d = division(str(topc).split(":")[0] if topc else "")
        if d:
            by_subj[s][str(college)][d] += 1
            if title:
                titles[s][str(college)].append(str(title))
    wb.close()

    # ---- homonym detection ---------------------------------------------------
    # Scoped entries ({"discipline": ..., "colleges": [...]}) only ever fill
    # allowlisted colleges, so only those vote — the audit checks the
    # allowlist's INTERNAL consistency. Global (string) entries: all colleges.
    homonyms = {}
    for s, colleges in sorted(by_subj.items()):
        entry = subject_map[s]
        scoped = not isinstance(entry, str)
        mapped_disc = entry["discipline"] if scoped else entry
        allow = set(entry["colleges"]) if scoped else None
        votes = {}   # college -> (division, n_courses, share)
        for col, divs in colleges.items():
            if allow is not None and col not in allow:
                continue
            n = sum(divs.values())
            if n < MIN_COURSES:
                continue
            d, dn = divs.most_common(1)[0]
            share = dn / n
            if share >= MIN_SHARE:
                votes[col] = (d, n, round(share, 2))
        if len(votes) < 2:
            continue
        tally = Counter()
        for col, (d, n, _) in votes.items():
            tally[d] += n   # weight by course volume, not college count
        majority = tally.most_common(1)[0][0]
        minority = {col: v for col, v in votes.items() if v[0] != majority}
        if minority:
            toks = disc_tokens(mapped_disc)
            evidence, grades = {}, []
            for col in sorted(minority):
                ov = title_overlap(titles[s].get(col, []), toks)
                grade = ("consistent" if ov >= 0.5
                         else "foreign" if ov < 0.2 else "mixed")
                evidence[col] = {"title_overlap": ov, "grade": grade}
                grades.append(grade)
            homonyms[s] = {
                "mapped_discipline": mapped_disc,
                "scoped": scoped,
                "majority_division": majority,
                "majority_division_umbrella": div_names.get(majority),
                "colleges": {col: {"division": d, "umbrella": div_names.get(d),
                                   "courses": n, "share": sh}
                             for col, (d, n, sh) in sorted(votes.items())},
                "minority_colleges": sorted(minority),
                "minority_title_evidence": evidence,
                # worst grade wins the entry verdict: any foreign minority
                # college means the global entry mis-fills real courses —
                # unless a curator cleared the entry by title inspection
                "verdict": ("cleared (curator)" if s in cleared
                            else "foreign" if "foreign" in grades
                            else "mixed" if "mixed" in grades else "consistent"),
            }

    # ---- suspect fills over the staging files --------------------------------
    suspects = []
    minted = load("coci_minted_courses.json")["courses"]
    singles = load("coci_minted_singletons.json")["courses"]
    memships = load("coci_minted_memberships.json")["memberships"]

    def _actionable(h):
        return h and h["verdict"] in ("foreign", "mixed")

    for cid, v in singles.items():
        s = str(v.get("subject") or "").strip().upper()
        h = homonyms.get(s)
        if (_actionable(h) and v.get("discipline_source") == "subject_map"
                and v.get("college") in h["minority_colleges"]
                and h["minority_title_evidence"][v["college"]]["grade"] != "consistent"):
            suspects.append({"course_id": cid, "kind": "singleton",
                             "subject": s, "college": v.get("college"),
                             "title": v.get("common_title"),
                             "filled_discipline": v.get("discipline")})

    for cid, v in minted.items():
        if v.get("discipline_source") != "subject_map":
            continue
        ms = memships.get(cid) or []
        msubs = {str(m.get("subject") or "").strip().upper() for m in ms}
        hits = [s for s in msubs if _actionable(homonyms.get(s))
                and homonyms[s]["mapped_discipline"] == v.get("discipline")]
        if not hits:
            continue
        mdivs = Counter(d for m in ms
                        if (d := division(str(m.get("top_code") or "").split(":")[0])))
        if not mdivs:
            continue
        row_div = mdivs.most_common(1)[0][0]
        if any(row_div != homonyms[s]["majority_division"] for s in hits):
            suspects.append({"course_id": cid, "kind": "minted",
                             "subject": "/".join(sorted(hits)),
                             "member_plurality_division": row_div,
                             "title": v.get("common_title"),
                             "filled_discipline": v.get("discipline")})

    receipt = {
        "_generated_by": "kb/_audit_subject_map.py",
        "_generated_at": date.today().isoformat(),
        "_params": {"min_courses": MIN_COURSES, "min_share": MIN_SHARE},
        "homonym_candidates": homonyms,
        "suspect_fills": suspects,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(receipt, f, indent=1, ensure_ascii=False)
        f.write("\n")

    print(f"subject_map audit — {len(subject_map)} entries, "
          f"{len(homonyms)} homonym candidate(s), {len(suspects)} suspect fill(s)")
    for s, h in homonyms.items():
        print(f"\n  {s} -> {h['mapped_discipline']!r}  "
              f"(majority div {h['majority_division']} {h['majority_division_umbrella']})"
              f"  VERDICT: {h['verdict'].upper()}")
        for col, info in h["colleges"].items():
            ev = h["minority_title_evidence"].get(col)
            mark = (f" <-- MINORITY title-overlap {ev['title_overlap']} ({ev['grade']})"
                    if ev else "")
            print(f"     {col[:34]:34} div {info['division']} ({info['umbrella']}) "
                  f"x{info['courses']} share {info['share']}{mark}")
    if suspects:
        print("\n  suspect fills:")
        for sp in suspects[:25]:
            print(f"     {sp['course_id']:14} [{sp['kind']:9}] {sp['subject']:10} "
                  f"{(sp.get('college') or '')[:24]:24} {str(sp['title'])[:42]!r} "
                  f"-> {sp['filled_discipline']!r}")
    print(f"\n  receipt: kb/subject_map_audit.json")


if __name__ == "__main__":
    main()
