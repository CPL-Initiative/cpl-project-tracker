"""
Identity-anchored title suggestions for the CER unclassified-triage worklist.

Phase 1 of Sam's Cx title-precedence integration (2026-07-07): for each raw MAP
exhibit title in the unclassified queue, parse the embedded local course code
("ADM JUS 003 - Legal Aspects of Evidence" → ADM JUS 003), join it to the COCI
course universe, and — when the identity is clean — suggest the OFFICIAL title
per skill Rule 5c precedence:

  1. CCN   — the AB-1111 Common Course Number's student-facing statewide title
  2. C-ID  — the descriptor title
  3. (M-ID — gated OFF until Sam declares the M-ID layer stable: --with-mids)
  4. course — the modal COCI course title (the Rule-5c plain-content fallback)

Plus a COS tier when kb/reference/cos_certifications.json exists (CareerOneStop
authority match on the raw title — industry-cert exhibits).

The worklist renders these as one-click 💡 fill chips; the CURATOR always
confirms — suggestions are never auto-applied.

Join guards (the documented CCR membership-join hazard — (subject, number) is
globally ambiguous):
  - a parsed course ref must EXIST in COCI under a normalized (SUBJ, NUM) key;
  - identity tiers need a UNANIMOUS C-ID/CCN across that key's rows;
  - when the raw title carries descriptive text beyond the code, the COCI
    title must share tokens with it (a zero-overlap join is dropped — protects
    "AP 2-D Art…" from parsing as subject "AP" number "2");
  - exam-brand prefixes (AP/IB/CLEP) never parse as course refs.

Runs in the daily cron after the exhibit auditor (fresh queue); idempotent
(unchanged suggestion set → file untouched). Output:
kb/unclassified_suggestions.json — fetched lazily by credential_reference.js.

Run from repo root:  python3 kb/_suggest_unclassified.py [--with-mids]
"""
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime, timezone

HERE = os.path.dirname(os.path.abspath(__file__))
AUDIT = os.path.join(HERE, "exhibit_audit", "latest.json")
COCI = os.path.join(HERE, "reference", "coci_course_list.xlsx")
CID = os.path.join(HERE, "reference", "cid_descriptors.json")
CCN = os.path.join(HERE, "reference", "ccn_courses.json")
COS = os.path.join(HERE, "reference", "cos_certifications.json")
MINTED = os.path.join(HERE, "coci_minted_courses.json")
OUT = os.path.join(HERE, "unclassified_suggestions.json")

EXAM_BRANDS = {"AP", "IB", "CLEP", "AICE", "DSST"}
STOP = {"the", "of", "and", "in", "for", "a", "an", "to", "with", "on", "or",
        "credit", "by", "exam", "cbe", "portfolio", "review", "completion",
        "certification", "certificate", "course", "training"}

# Mechanism-phrase strip (Session 104): 'Credit by Exam ACCTG 022 …'-style Cx
# titles hide their course code behind the mechanism phrase, so the leading-
# code parser never fired on them (19/451 coverage before this). Reuses the
# pre-seed lane's patterns — one definition, no drift between the two scripts.
sys.path.insert(0, HERE)
try:
    import _preseed_unclassified as _ps

    def strip_mechanism(t):
        for pat in (_ps._CX_LEAD, _ps._CX_TRAIL, _ps._CX_CBE,
                    _ps._PORTFOLIO_LEAD, _ps._PORTFOLIO_TRAIL):
            t = pat.sub(" ", t)
        return re.sub(r"\s+", " ", t).strip(" -–,;:")
except ImportError:  # standalone use — the parser simply sees fewer refs
    strip_mechanism = None


def load(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def norm_subj(s):
    return re.sub(r"[^A-Z0-9]", "", (s or "").upper())


def norm_college(c):
    return re.sub(r"[^A-Z0-9]", "", (c or "").upper())


def norm_num(n):
    n = (n or "").strip().upper()
    m = re.match(r"0*(\d+[A-Z]*)$", n)
    return m.group(1) if m else n


def tokens(s):
    return {t for t in re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).split()
            if t not in STOP and len(t) > 1}


def parse_course_refs(raw):
    """Course refs embedded in a raw exhibit title: a LEADING 'SUBJ [-] NUM'
    (subject = 1-2 short words) and/or parenthetical '(SUBJ NUM)' anywhere.
    Returns [(subject, number, remainder_text)]."""
    refs = []
    lead = re.match(
        r"^\s*([A-Za-z][A-Za-z&/\.]{0,9}(?:\s+[A-Za-z]{1,6})?)\s*[-–—:,]?\s+"
        r"[-–—:,]?\s*0*(\d{1,4}[A-Za-z]{0,2})\b[\s\-–—:,\.]*(.*)$", raw or "")
    if not lead:
        # Tight-hyphen form "CD-005" (no whitespace at all — the Lemoore HS
        # articulation pattern, 2026-07-08). Subject needs ≥2 letters so a
        # stray initial never reads as a subject; the COCI join + the
        # title-sanity guard gate anything this over-matches.
        lead = re.match(
            r"^\s*([A-Za-z]{2,10})[-–—]0*(\d{1,4}[A-Za-z]{0,2})\b"
            r"[\s\-–—:,\.]*(.*)$", raw or "")
    if lead:
        subj, num, rest = lead.group(1), lead.group(2), lead.group(3)
        if norm_subj(subj) not in EXAM_BRANDS:
            refs.append((subj, num, rest))
    # Parenthetical refs — but NEVER on an exam-brand exhibit ("AP Biology
    # (BIOSCI 100)"): there the parenthetical is the local TARGET course, and
    # the credential is the exam — suggesting the course title would steer the
    # curator wrong.
    first = re.split(r"[\s:,-]+", (raw or "").strip(), 1)[0]
    if norm_subj(first) not in EXAM_BRANDS:
        for m in re.finditer(r"\(([A-Za-z][A-Za-z&/\. ]{0,11})\s+0*(\d{1,4}[A-Za-z]{0,2})\)", raw or ""):
            subj = m.group(1)
            if norm_subj(subj) not in EXAM_BRANDS:
                refs.append((subj, m.group(2), (raw or "").replace(m.group(0), " ")))
    return refs


def build_coci_index():
    """Two indexes over one streaming read:
      idx:  (SUBJ, NUM) → {'cids': Counter, 'ccns': Counter, 'titles': Counter, 'n': int}
      cidx: (COLLEGE, SUBJ, NUM) → {'cids': set, 'ccns': set, 'titles': Counter}
    The college-scoped index (Session 104) exists because the global
    (subject, number) key is ambiguous across colleges — the documented CCR
    membership-join hazard — while (College, SUBJ, NUM) is essentially unique;
    the exhibit auditor now stamps originating colleges on each queue card."""
    from openpyxl import load_workbook
    wb = load_workbook(COCI, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h) for h in next(it)]
    col = {name: hdr.index(name) for name in
           ("College", "Subject", "Course_Number", "CourseTitle", "CIDNumber",
            "CommonCourseNumber")}
    idx, cidx = {}, {}
    for row in it:
        subj = norm_subj(str(row[col["Subject"]] or ""))
        num = norm_num(str(row[col["Course_Number"]] or ""))
        if not subj or not num:
            continue
        e = idx.setdefault((subj, num), {"cids": Counter(), "ccns": Counter(),
                                         "titles": Counter(), "n": 0})
        college = norm_college(str(row[col["College"]] or ""))
        ce = cidx.setdefault((college, subj, num),
                             {"cids": set(), "ccns": set(), "titles": Counter()}) \
            if college else None
        e["n"] += 1
        title = str(row[col["CourseTitle"]] or "").strip()
        if title and title.lower() not in ("none", "n/a"):
            e["titles"][title] += 1
            if ce is not None:
                ce["titles"][title] += 1
        cid = str(row[col["CIDNumber"]] or "").strip()
        if cid and cid.lower() not in ("none", "n/a", "null", "not applicable"):
            for c in re.split(r"[,;]", cid):  # comma-joined values occur (tmc #642)
                c = c.strip()
                if c:
                    e["cids"][c] += 1
                    if ce is not None:
                        ce["cids"].add(c)
        ccn = str(row[col["CommonCourseNumber"]] or "").strip()
        if ccn and ccn.lower() not in ("none", "n/a", "null", "not applicable"):
            e["ccns"][ccn] += 1
            if ce is not None:
                ce["ccns"].add(ccn)
    return idx, cidx


def main():
    with_mids = "--with-mids" in sys.argv

    audit = load(AUDIT)
    queue = [(c["raw_title"], c.get("colleges") or [])
             for c in audit.get("title_cards", [])
             if not c.get("unified_title")
             and "unclassified_in_map" in (c.get("tags") or [])]
    print(f"unclassified queue: {len(queue)}")
    if not queue:
        print("Queue empty — nothing to suggest.")
        return

    cid_titles = {d["descriptor"].strip().upper(): d["title"]
                  for d in load(CID).get("descriptors", []) if d.get("descriptor")}
    ccn_titles = {c["ccn"].strip().upper(): c["title"]
                  for c in load(CCN).get("courses", []) if c.get("ccn")}
    mid_titles = {}
    if with_mids and os.path.exists(MINTED):
        # Gated on Sam declaring the M-ID layer stable (Rule 5c tier 3).
        for cid_, rec in load(MINTED).get("courses", {}).items():
            t = rec.get("title") or rec.get("representative_title")
            if t:
                mid_titles[cid_] = t

    cos = None
    if os.path.exists(COS):
        sys.path.insert(0, HERE)
        import importlib
        m = importlib.import_module("_match_cos_authority")
        certs = load(COS).get("certifications", [])
        cos = (m, m.build_index(certs))

    idx, cidx = build_coci_index()
    print(f"coci (SUBJ,NUM) keys: {len(idx):,} | college-scoped keys: {len(cidx):,}")

    suggestions = {}
    stats = Counter()
    for raw, colleges in queue:
        out = []
        refs = parse_course_refs(raw)
        if not refs and strip_mechanism:
            # 'Credit by Exam ACCTG 022 …' — the code only leads once the
            # mechanism phrase is gone (Session 104 coverage fix).
            stripped_raw = strip_mechanism(raw)
            if stripped_raw != raw:
                refs = parse_course_refs(stripped_raw)
                if refs:
                    stats["mechanism_stripped"] += 1
        for subj, num, rest in refs:
            key = (norm_subj(subj), norm_num(num))
            e = idx.get(key)
            if not e:
                continue
            # College-scoped narrowing (Session 104): the exhibit auditor stamps
            # originating colleges on each queue card; (College, SUBJ, NUM) is
            # essentially unique where the global key is ambiguous, so prefer
            # the originating colleges' own rows and fall back to the global
            # view when they don't resolve.
            scoped = None
            if colleges:
                agg = {"cids": set(), "ccns": set(), "titles": Counter()}
                for cname in colleges:
                    sub = cidx.get((norm_college(cname), key[0], key[1]))
                    if sub:
                        agg["cids"] |= sub["cids"]
                        agg["ccns"] |= sub["ccns"]
                        agg["titles"].update(sub["titles"])
                if agg["titles"] or agg["cids"] or agg["ccns"]:
                    scoped = agg
                    stats["college_scoped"] += 1
            ccn_set = set(scoped["ccns"]) if scoped else set(e["ccns"])
            cid_set = set(scoped["cids"]) if scoped else set(e["cids"])
            title_pool = scoped["titles"] if scoped and scoped["titles"] else e["titles"]
            modal_title, modal_n = (title_pool.most_common(1) or [("", 0)])[0]
            # Title-sanity guard: a descriptive remainder must overlap the COCI
            # course title — else this (SUBJ, NUM) is a different course and the
            # join is noise (the CCR membership-join hazard).
            rt = tokens(rest)
            if rt and modal_title and not (rt & tokens(modal_title)):
                stats["dropped_title_mismatch"] += 1
                continue
            code = f"{subj.upper().strip()} {num.upper()}"
            # "Unanimous" = a single DISTINCT value among rows that carry one
            # (most rows carry none — only 58 CCNs exist); >1 distinct = ambiguous.
            if len(ccn_set) == 1:
                ccn = next(iter(ccn_set))
                t = ccn_titles.get(ccn.upper()) or modal_title
                if t:
                    out.append({"kind": "ccn", "id": ccn, "title": t, "code": code})
            if len(cid_set) == 1:
                cid_ = next(iter(cid_set))
                t = cid_titles.get(cid_.upper())
                if t:
                    out.append({"kind": "cid", "id": cid_, "title": t, "code": code})
                elif modal_title:
                    out.append({"kind": "cid", "id": cid_, "title": modal_title,
                                "code": code, "unverified": True})
            if with_mids and not out:
                pass  # M-ID tier lands here when unlocked; mid_titles is ready.
            if not any(s["kind"] in ("ccn", "cid") for s in out) and modal_title:
                share = modal_n / max(1, sum(title_pool.values()))
                rec = {"kind": "course", "title": modal_title, "code": code,
                       "share": round(share, 2), "colleges": e["n"]}
                if scoped:
                    rec["scoped"] = True  # resolved from the originating college's own catalog
                out.append(rec)
        if cos:
            m, (prepared, by_norm, by_acr) = cos
            tier, cands = m.match_title(raw, None, prepared, by_norm, by_acr)
            if tier in ("exact", "acronym"):
                uniq = {(m.norm(c["name"]), m.norm(c.get("org", ""))): c for c in cands}
                if len(uniq) == 1:
                    c = list(uniq.values())[0]
                    out.append({"kind": "cos", "id": c.get("id"), "title": c["name"],
                                "org": c.get("org") or None})
        # De-dupe by (kind, title); CCN > C-ID > COS > course ordering.
        rank = {"ccn": 0, "cid": 1, "cos": 2, "course": 3}
        seen, final = set(), []
        for s in sorted(out, key=lambda s: rank[s["kind"]]):
            k = (s["kind"], s["title"].lower())
            if k not in seen:
                seen.add(k)
                final.append(s)
        if final:
            suggestions[raw] = final
            stats["with_" + final[0]["kind"]] += 1

    print(f"suggestions: {len(suggestions)}/{len(queue)} raws | {dict(stats)}")
    for raw in list(suggestions)[:8]:
        s = suggestions[raw][0]
        print(f"  💡 {raw!r} → [{s['kind']}] {s['title']!r}"
              + (f" ({s.get('id')})" if s.get("id") else ""))

    # Idempotent write (the file is a daily-cron artifact — no timestamp churn).
    try:
        prior = load(OUT).get("suggestions")
    except (FileNotFoundError, ValueError):
        prior = None
    if prior == suggestions:
        print("no change — file left untouched")
        return
    payload = {
        "_about": ("Identity-anchored title suggestions for the CER unclassified-"
                   "triage worklist (kb/_suggest_unclassified.py). Precedence per "
                   "skill Rule 5c: CCN > C-ID > (M-ID once stable) > modal local "
                   "course title; plus CareerOneStop authority matches when the "
                   "registry is synced. One-click FILL chips — the curator always "
                   "confirms; never auto-applied."),
        "_generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "count": len(suggestions),
        "suggestions": suggestions,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=1, ensure_ascii=False)
        f.write("\n")
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
