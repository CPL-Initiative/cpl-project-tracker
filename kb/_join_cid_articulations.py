#!/usr/bin/env python3
"""Join the ASCCC C-ID approved-articulation table to the raw COCI course list
— the per-(college, course) official-authority layer for M-ID cleanup.

Input:  kb/reference/cid_articulations.json (curator-supplied c-id.net export;
        see its _schema/_semantics headers)
Output: kb/cid_articulation_joins.json (--write) — control_number -> cid map +
        per-row dispositions; always prints the dry-run analysis.

Dispositions per articulation row (SET-aware — a local course can hold
APPROVALS UNDER MULTIPLE DESCRIPTORS, 1,820 statewide, dominated by
series∧component pairs like CHEM 110 + CHEM 120 S; and COCI's CIDNumber
column can itself be a comma-list):
  already_claimed — this row's descriptor is among the COCI row's own
                    CIDNumber claims (the re-mint already split this member).
  compatible_multi— COCI claims a DIFFERENT descriptor, but one that is also
                    in this course's OWN table approval set (e.g. COCI says
                    PHYS 100 S where the table row says PHYS 105 — the course
                    holds both; the series/component pattern). NOT a conflict.
  new_authority   — COCI carries NO usable CIDNumber but the table approves
                    it. THE PAYOFF CLASS (statewide ~10k courses — colleges
                    under-report C-ID alignment in COCI roughly half the time).
  coci_conflict   — COCI claims a descriptor the table does NOT hold for this
                    course. Surfaced loudly, auto-trusted NEVER.
  unmatched       — no current COCI row at (college, subject, number); the
                    college likely renumbered since the approval term. Goes to
                    curation; NEVER fuzzy-matched (titles are not consulted —
                    that is the entire point of this authority tier).
Rows flagged `sequence: true` by the ingester (multi-course articulations —
neither course alone IS the descriptor course) are SKIPPED for joining and
counted; curation owns them. Series descriptors ("… S") are tagged
`series: true`; a course's other approvals ride along as `also_approved` so
the Phase-1 router can prefer the component descriptor as the display home.

Join rules (docs/cid_articulation_authority_scope.md §3):
  * college matched on an ASCII-alnum slug (mojibake-safe);
  * (subject, number) exact first; a leading-zero-normalized retry ("003A" ==
    "3A") is allowed but FLAGGED (zero_normalized: true) for audit;
  * titles NEVER join; approval_term is carried for audit only.

Re-runnable + read-only by default; --write emits the join artifact the
generator will consume (display-level member routing — scope §4).

Usage (from repo root):
  python3 kb/_join_cid_articulations.py            # dry-run analysis
  python3 kb/_join_cid_articulations.py --write    # + write the join artifact
"""
import json
import os
import re
import sys
from collections import defaultdict
from datetime import date

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
ARTICULATIONS = os.path.join(HERE, "reference", "cid_articulations.json")
RAW_XLSX = os.path.join(HERE, "reference", "coci_course_list.xlsx")
OUT = os.path.join(HERE, "cid_articulation_joins.json")

CID_SENTINELS = {"", "NULL", "N/A", "NA", "NONE", "PENDING"}


def _slug(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _num_norm(n):
    return re.sub(r"\s+", "", str(n or "").upper())


def _num_zero_norm(n):
    # "003A" -> "3A"; keeps a pure-zero number ("0") intact
    n = _num_norm(n)
    stripped = n.lstrip("0")
    return stripped if stripped else n


def load_raw_index():
    """(college_slug, subject, number) -> [raw row dicts]. Streamed once."""
    from openpyxl import load_workbook
    wb = load_workbook(RAW_XLSX, read_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}
    exact, zeroed = defaultdict(list), defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {
            "college": row[ix["College"]],
            "control_number": row[ix["CourseControlNumber"]],
            "subject": str(row[ix["Subject"]] or ""),
            "number": str(row[ix["Course_Number"]] or ""),
            "title": row[ix["CourseTitle"]],
            "cid": str(row[ix["CIDNumber"]] or "").strip(),
            "ccn": str(row[ix["CommonCourseNumber"]] or "").strip(),
        }
        cs = _slug(rec["college"])
        exact[(cs, rec["subject"], _num_norm(rec["number"]))].append(rec)
        zeroed[(cs, rec["subject"], _num_zero_norm(rec["number"]))].append(rec)
    wb.close()
    return exact, zeroed


def main():
    write = "--write" in sys.argv
    doc = json.load(open(ARTICULATIONS, encoding="utf-8"))
    arts = doc["articulations"]
    exact, zeroed = load_raw_index()

    # per-course table approval sets — a course can hold approvals under
    # several descriptors (series∧component etc.). SEQUENCE rows count here
    # (a "PHYS 2A + PHYS 2B" series approval proves both courses participate
    # in that series) even though routing skips them — without this, a COCI
    # series claim against a table component row reads as a false conflict.
    course_cids = defaultdict(set)
    for a in arts:
        course_cids[(_slug(a["college"]), a["subject"], _num_norm(a["number"]))].add(a["cid"].upper())

    joins = []          # the authority map rows
    by_disp = defaultdict(list)
    n_sequence = 0
    for a in arts:
        if a.get("sequence"):
            n_sequence += 1
            continue
        key = (_slug(a["college"]), a["subject"], _num_norm(a["number"]))
        hits, zero_normalized = exact.get(key), False
        if not hits:
            hits = zeroed.get((key[0], key[1], _num_zero_norm(a["number"])))
            zero_normalized = bool(hits)
        if not hits:
            by_disp["unmatched"].append(a)
            continue
        table_set = course_cids[key]
        for h in hits:
            coci_set = {x.strip().upper() for x in h["cid"].split(",")
                        if x.strip().upper() not in CID_SENTINELS}
            cid_u = a["cid"].upper()
            if cid_u in coci_set:
                disp = "already_claimed"
            elif not coci_set:
                disp = "new_authority"
            elif coci_set & table_set:
                disp = "compatible_multi"
            else:
                disp = "coci_conflict"
            row = {"control_number": h["control_number"], "cid": a["cid"],
                   "college": h["college"], "subject": h["subject"],
                   "number": h["number"], "local_title": h["title"],
                   "coci_cid": h["cid"] or None, "coci_ccn": h["ccn"] or None,
                   "effective_term": a.get("effective_term"),
                   "disposition": disp}
            if cid_u.endswith(" S"):
                row["series"] = True
            also = sorted(table_set - {cid_u})
            if also:
                row["also_approved"] = also
            if zero_normalized:
                row["zero_normalized"] = True
            joins.append(row)
            by_disp[disp].append(row)

    # current home of each joined course (where the member sits in staging today)
    mem = json.load(open(os.path.join(HERE, "coci_minted_memberships.json")))["memberships"]
    singles = json.load(open(os.path.join(HERE, "coci_minted_singletons.json")))["courses"]
    cn2home = {}
    for mid, members in mem.items():
        for m in members:
            cn2home[m.get("control_number")] = mid
    for sid, s in singles.items():
        cn2home[s.get("control_number")] = sid
    for j in joins:
        j["current_home"] = cn2home.get(j["control_number"])  # None = departed at re-mint

    print(f"articulation rows: {len(arts)} | sequence-skipped: {n_sequence} | joined COCI courses: {len(joins)}")
    for d in ("already_claimed", "compatible_multi", "new_authority", "coci_conflict"):
        print(f"  {d:<16} {len(by_disp[d])}")
    print(f"  {'unmatched':<16} {len(by_disp['unmatched'])}")
    if by_disp["coci_conflict"]:
        print(f"\nCOCI-conflict rows (true disagreements — curation decides), first 20 of {len(by_disp['coci_conflict'])}:")
        for r in by_disp["coci_conflict"][:20]:
            print(f"  {r['college']}: {r['subject']} {r['number']} — table says {r['cid']}, COCI says {r['coci_cid']}")
    if by_disp["unmatched"]:
        ucol = defaultdict(int)
        for a in by_disp["unmatched"]:
            ucol[a["college"]] += 1
        print(f"\nunmatched rows (no current COCI course — likely renumbered since approval): {len(by_disp['unmatched'])}")
        print("  by college (top 12):", dict(sorted(ucol.items(), key=lambda x: -x[1])[:12]))
    homes = defaultdict(int)
    for j in joins:
        homes[j["current_home"] or "(departed to official at re-mint)"] += 1
    print("\ncurrent homes of joined members (what the authority would move/confirm):")
    for h, n in sorted(homes.items(), key=lambda x: -x[1]):
        print(f"  {h}: {n}")

    if write:
        out = {"_generated_by": "kb/_join_cid_articulations.py",
               "_generated_at": date.today().isoformat(),
               "_source": "kb/reference/cid_articulations.json",
               "_note": ("control_number -> C-ID authority at the (college, course) grain. "
                         "Same trust tier as COCI CIDNumber; consumed by the generator's "
                         "member routing (scope §4). coci_conflict rows are INCLUDED with "
                         "their disposition so the consumer can exclude them — never "
                         "auto-trust a conflict."),
               "count": len(joins),
               "joins": joins,
               "unmatched": by_disp["unmatched"]}
        with open(OUT, "w", encoding="utf-8") as f:
            json.dump(out, f, indent=1, ensure_ascii=False)
        print(f"\nWROTE {OUT}")
    else:
        print("\nDRY-RUN — no files written. Re-run with --write.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
