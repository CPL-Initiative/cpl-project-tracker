#!/usr/bin/env python3
"""Build cip_crosswalk_data.js — the CIP Code Taxonomy reference dataset.

Source: kb/reference/cip_searchable_260715.xlsx — the Chancellor's Office
"CIP Searchable" workbook (2026-07-15 cut) for the TOP-to-CIP transition
(ESS 26-06, effective fall 2026). The CIP Codes tab (cip_crosswalk.js) is the
faculty-facing REFERENCE MANUAL — the successor to the CCC TOP Code Manual —
that replaces the multi-tab spreadsheet ESS was going to email to the field.

The tool used to also recreate the TOP↔CIP crosswalk, but the CO (Jenni Abbott,
AA division) reframed it: the crosswalk is one-to-many and COE already hosts it,
so ours is the AUTHORITATIVE CIP LIST — "clear, comprehensive, user-friendly,"
an easy button. So this dataset is now LEAN: the full federal CIP-2020 list with
each code's certified CTE category, definition, examples, family, 2020-CIP
action, and a course-level C-ID/CCN floor flag. No TOP codes, no crosswalk pairs.

Emits window.CIP_CROSSWALK:
  {
    _built_at, _built_by, _source, _note, _category, _transfer, _topcip,
    fams:   { "<fam2>": "<family title>" },
    rows:   [ {code, t, cat, fam, def, ex, act, x} ]   # 2,325 CIP-2020 codes
    topcip: { "<TOP NNNN.NN>": {t:"<TOP title>", c:[["<cip>","<tier>"], ...]} }
    boiler: [ "<cip>", ... ]   # generic-noncredit CIPs mapped from nearly every TOP
    sub4:   { "<fam4 NN.NN>": "<series title>" }   # OPTIONAL — only when an
            # authoritative NCES all-levels export is present in kb/reference/
            # (the CCCCO workbook is 6-digit-only). See load_sub4.
  }
    cat = certified CTE category: CTE | Both | Non-CTE | Noncredit | Retired | Reserved
    act = 2020-CIP action (New / Deleted / Moved from|to / No substantive changes)
    x   = 1 if any TOP that maps to this CIP has C-ID or CCN coursework (a floor)

The topcip map powers the "Find my course's code" easy button: every COCI course
carries a current TOP code, and the CO's official TOP->CIP crosswalk lists the
candidate CIPs for that TOP. The tab ranks those candidates by description-fit —
the two-signals-agree gate from the repo's TOP doctrine (the crosswalk PROPOSES,
description-fit RANKS, faculty CONFIRMS; TOP never decides). Each candidate carries
a provenance tier so field-submitted mappings read softer than CO-authoritative
ones:
    o = official   (CCCCO TOP-CIP / COCI / COE TOP-CIP)
    f = field-submitted
    n = noncredit   (Noncredit TOP-CIP)
    g = generic     (CIP Code File, no real TOP relationship)
`boiler` lists the two universal noncredit CIPs (32.0107 Career Exploration,
32.0111 Workforce Development) that map from ~280 TOPs each — pure boilerplate the
tab collapses out of the way rather than ranking.

Two data rules baked in here (full story: docs/cip_crosswalk_lessons.md):
  1. CATEGORY comes from the CO consultant's CERTIFIED designations, not either
     workbook tab. The *CIP Descriptions* and *TOP-CIP Data* tabs disagree on 244
     codes in BOTH directions (Jenni's 45.0702 catch), so neither is reliable.
     Authority: kb/reference/cip_cte_certified_260715.json (the 244 disagreements);
     agreed value elsewhere; single-tab value for orphans.
  2. C-ID/CCN is a course-level FLOOR, never "transferable." A CIP is flagged if
     any TOP that maps to it has courses carrying a C-ID (transfer-model
     articulation) or a CCN (AB 1111 common course number). TOP→CIP is
     one-to-many, so it's an association, not a guarantee.

Rebuild: python kb/_build_cip_crosswalk.py   (writes cip_crosswalk_data.js at repo root)
"""
import json
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
SRC = os.path.join(HERE, "reference", "cip_searchable_260715.xlsx")
# The CO consultant's certified CTE designations — the authority for the 244
# codes where the workbook's two tabs disagree (see rule 1 above).
CERT = os.path.join(HERE, "reference", "cip_cte_certified_260715.json")
# COCI course inventory — used for the C-ID/CCN course-level floor flag.
COCI_SRC = os.path.join(HERE, "reference", "coci_course_list.xlsx")
OUT = os.path.join(REPO, "cip_crosswalk_data.js")
# Optional AUTHORITATIVE NCES 4-digit SERIES titles (sub4). The CCCCO workbook we
# build from was exported filtered to "6 Digit - Specific" rows, so it carries the
# 2-digit family titles + 6-digit code titles but NOT the 4-digit series titles.
# Those come ONLY from an authoritative file dropped into kb/reference/ — never
# inferred, paraphrased, or invented (the tool is grounded). Supported inputs, in
# priority order (see load_sub4): a simple {"51.38": "<series title>"} JSON, or the
# NCES all-levels CIPCode2020 export (CSV or XLSX) which includes the 4-digit rows.
SUB4_JSON = os.path.join(HERE, "reference", "cip_series4_titles.json")
SUB4_NCES_CSV = os.path.join(HERE, "reference", "CIPCode2020.csv")
SUB4_NCES_XLSX = os.path.join(HERE, "reference", "CIPCode2020.xlsx")
BUILT_AT = "2026-07-16"
SRC_LABEL = "kb/reference/cip_searchable_260715.xlsx (CCCCO CIP Searchable Workbook, 2026-07-15 cut) + kb/reference/cip_cte_certified_260715.json"

# raw workbook/certified flag -> the category shown in the reference
CATMAP = {
    "CTE": "CTE", "Both": "Both", "Not CTE": "Non-CTE", "Noncredit CIP": "Noncredit",
    "Remove": "Retired", "Canadian CIP": "Reserved", "CIP No Longer in Use": "Retired", "": "",
}


def clean(v):
    return "" if v is None else str(v).strip()


def _norm_cip(v):
    """Normalize a CIP code cell to a plain string like '51.38' / '51.3801'.
    NCES exports sometimes wrap codes as '="01.01"' or store a bare number; keep
    only what looks like a dotted CIP code."""
    s = clean(v).lstrip("=").strip('"').strip("'").strip()
    return s


def load_sub4(codes):
    """AUTHORITATIVE NCES 4-digit SERIES titles, keyed 'NN.NN'. Sourced ONLY from a
    file an operator dropped into kb/reference/ — never inferred (the tool is
    grounded, so an unknown series simply shows its code + count, as before).

    Priority:
      1. cip_series4_titles.json  — {"51.38": "Registered Nursing, ...", ...}
      2. CIPCode2020.csv / .xlsx  — the NCES all-levels export; we keep only rows
         whose CIP code is a 4-digit series (matches ^\\d\\d\\.\\d\\d$).

    Titles are retained ONLY for 4-digit prefixes that actually occur among the
    built 6-digit codes, so the map can never carry a series the taxonomy lacks.
    Returns {} when no authoritative source is present.
    """
    import re
    prefixes = {c[:5] for c in codes if len(c) >= 5}   # "51.3801" -> "51.38"
    out = {}

    # 1) explicit JSON map
    if os.path.exists(SUB4_JSON):
        try:
            with open(SUB4_JSON, encoding="utf-8") as f:
                m = json.load(f)
            for k, v in (m or {}).items():
                k = _norm_cip(k)
                if re.match(r"^\d\d\.\d\d$", k) and clean(v):
                    out[k] = clean(v).rstrip(".")
        except Exception as e:   # noqa: BLE001 — a bad optional file must not break the build
            print(f"  sub4: could not read {SUB4_JSON}: {e}")

    # 2) NCES all-levels export (CSV or XLSX) — pick the 4-digit series rows
    def _add(code, title):
        code = _norm_cip(code)
        if re.match(r"^\d\d\.\d\d$", code) and clean(title):
            out.setdefault(code, clean(title).rstrip("."))

    if os.path.exists(SUB4_NCES_CSV):
        try:
            import csv
            with open(SUB4_NCES_CSV, encoding="utf-8-sig", newline="") as f:
                rd = csv.DictReader(f)
                cols = {c.lower().replace(" ", ""): c for c in (rd.fieldnames or [])}
                cc = cols.get("cipcode"); ct = cols.get("ciptitle")
                if cc and ct:
                    for r in rd:
                        _add(r.get(cc), r.get(ct))
        except Exception as e:   # noqa: BLE001
            print(f"  sub4: could not read {SUB4_NCES_CSV}: {e}")
    elif os.path.exists(SUB4_NCES_XLSX):
        try:
            w = openpyxl.load_workbook(SUB4_NCES_XLSX, read_only=True, data_only=True)
            ws = w.worksheets[0]
            rows_iter = ws.iter_rows(values_only=True)
            hdr = [clean(c).lower().replace(" ", "") for c in next(rows_iter)]
            ci = hdr.index("cipcode") if "cipcode" in hdr else None
            ti = hdr.index("ciptitle") if "ciptitle" in hdr else None
            if ci is not None and ti is not None:
                for row in rows_iter:
                    _add(row[ci], row[ti])
        except Exception as e:   # noqa: BLE001
            print(f"  sub4: could not read {SUB4_NCES_XLSX}: {e}")

    return {k: v for k, v in out.items() if k in prefixes}


def split_code_title(s):
    """'0101.00 - Agriculture Technology...' -> ('0101.00', 'Agriculture Technology...')."""
    s = clean(s)
    if " - " in s:
        code, title = s.split(" - ", 1)
        return code.strip(), title.strip().rstrip(".")
    return s, ""


def canon(code):
    """Normalize a CIP code to LL.RRRR so it matches the certified-JSON keys."""
    code = clean(code)
    left, right = code.split(".", 1) if "." in code else (code, "")
    return left.zfill(2) + "." + right.ljust(4, "0")[:4]


def coci_top_flags():
    """Per-TOP-code {cid, ccn} presence from the COCI course inventory.

    Returns {top_code: {"cid": bool, "ccn": bool}}. The COCI TopCode is
    'CODE: Title' (e.g. '0101.00: Agriculture...'), split on ':' to match the
    clean '0101.00' TOP keys on the TOP-CIP Data sheet. C-ID = the statewide
    transfer-model articulation system; CCN = AB 1111 common course numbering.
    Presence is a FLOOR (a course-level identifier exists), not transferability.
    """
    if not os.path.exists(COCI_SRC):
        print(f"  (COCI source {COCI_SRC} absent — C-ID/CCN flags will all be 0)")
        return {}
    wb = openpyxl.load_workbook(COCI_SRC, read_only=True, data_only=True)
    it = wb.active.iter_rows(values_only=True)
    idx = {h: i for i, h in enumerate(next(it))}
    ti, ci, ni = idx.get("TopCode"), idx.get("CIDNumber"), idx.get("CommonCourseNumber")
    out = {}
    for r in it:
        raw = clean(r[ti]) if ti is not None else ""
        if not raw:
            continue
        code = raw.split(":", 1)[0].strip() if ":" in raw else raw
        rec = out.setdefault(code, {"cid": False, "ccn": False})
        if ci is not None and clean(r[ci]):
            rec["cid"] = True
        if ni is not None and clean(r[ni]):
            rec["ccn"] = True
    return out


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    # ---- CIP catalog: definitions / action / examples / CTE flag ----
    # cols: 0 CIPFamily, 1 CIPCode(01.0000), 5 Action, 8 CIPTitle(clean),
    #       9 code-title, 10 CTE Flag, 11 CIPDefinition, 13 Examples
    desc = {}
    for r in list(wb["CIP Descriptions"].iter_rows(values_only=True))[1:]:
        raw = clean(r[1])
        if not raw:
            continue
        code = canon(raw)
        examples = clean(r[13])
        if examples.lower().startswith("examples:"):
            examples = examples[len("examples:"):].strip(" -")
        desc[code] = {
            "t": clean(r[8]).rstrip(".") or split_code_title(r[9])[1],
            "fam": clean(r[0]),
            "act": clean(r[5]),
            "def": clean(r[11]),
            "ex": examples,
            "flag": clean(r[10]),   # this tab's CTE designation (one of two, may disagree)
        }

    # ---- TOP-CIP Data: the OTHER CTE flag per CIP, family titles, TOP→CIP map ----
    # cols: 2 TOP code-title, 11 CIP code-title, 14 CIP family code-title,
    #       15 CIP CTE Flag, 18 relationship source (provenance)
    cross_flags = {}     # cip -> set of the crosswalk-tab CTE flags seen
    fam_titles = {}      # fam2 -> title
    cip_tops = {}        # cip -> set of clean TOP codes that map to it
    top_title = {}       # TOP code -> clean TOP title
    top_cips = {}        # TOP code -> {cip -> strongest provenance tier}
    # provenance -> tier letter (strongest wins per TOP+CIP pair)
    PROV_TIER = {
        "CCCCO TOP-CIP": "o", "COCI": "o", "COE TOP-CIP": "o",
        "Submitted by Field": "f", "Noncredit TOP-CIP": "n",
        "CIP Code File (No TOP Relationship)": "g",
    }
    TIER_RANK = {"o": 0, "f": 1, "n": 2, "g": 3}
    for r in list(wb["TOP-CIP Data"].iter_rows(values_only=True))[1:]:
        top_code, top_ttl = split_code_title(r[2])
        cip_raw, _ = split_code_title(r[11])
        if not cip_raw:
            continue
        cip = canon(cip_raw)
        cross_flags.setdefault(cip, set()).add(clean(r[15]))
        fam_code, fam_title = split_code_title(r[14])
        if len(fam_code) == 2 and fam_code.isdigit() and fam_title:
            fam_titles[fam_code] = fam_title
        if top_code and not top_code.upper().startswith("XXXX"):
            cip_tops.setdefault(cip, set()).add(top_code)
            if top_ttl:
                top_title.setdefault(top_code, top_ttl)
            tier = PROV_TIER.get(clean(r[18]), "g")
            slot = top_cips.setdefault(top_code, {})
            if cip not in slot or TIER_RANK[tier] < TIER_RANK[slot[cip]]:
                slot[cip] = tier
    # a couple of reserved families carry no crosswalk title
    fam_titles.setdefault("21", "Reserved (Canadian CIP)")
    fam_titles.setdefault("55", "Reserved (Canadian CIP)")

    # generic-noncredit boilerplate: CIPs mapped from a large share of ALL TOPs
    # (32.0107 Career Exploration, 32.0111 Workforce Dev — ~280 TOPs each). The
    # tab collapses these out of the ranked list rather than surfacing them.
    from collections import Counter as _C
    _cip_topfreq = _C()
    for _tc, _slot in top_cips.items():
        for _cip in _slot:
            _cip_topfreq[_cip] += 1
    BOILER_MIN_TOPS = 40   # natural break: 32.0107/32.0111 (275+) vs next (32 TOPs)
    boiler = sorted(c for c, n in _cip_topfreq.items() if n >= BOILER_MIN_TOPS)

    # ---- certified CTE designations (authority for the 244 disagreements) ----
    certified = json.load(open(CERT, encoding="utf-8")).get("designations", {})
    certified = {canon(k): v for k, v in certified.items()}

    # ---- C-ID/CCN course-level floor, rolled TOP -> CIP ----
    top_flag = coci_top_flags()

    # ---- resolve each reference row ----
    rows = []
    fams = {}
    cert_hits = uncertified = 0
    for code, d in sorted(desc.items()):
        # category: certified > both-tabs-agree > single-tab > provisional
        xs = cross_flags.get(code)
        xflag = list(xs)[0] if (xs and len(xs) == 1) else ""
        dflag = d["flag"]
        if code in certified:
            src = certified[code]; cert_hits += 1
        elif xflag and dflag and xflag == dflag:
            src = dflag
        elif xflag and not dflag:
            src = xflag
        elif dflag and not xflag:
            src = dflag
        elif xflag and dflag and xflag != dflag:
            uncertified += 1; src = xflag   # provisional; certified list should cover these
        else:
            src = dflag or xflag
        cat = CATMAP.get(src, "")

        # C-ID/CCN floor: any mapped TOP with a C-ID or CCN course
        x = 0
        for tc in cip_tops.get(code, ()):
            f = top_flag.get(tc)
            if f and (f["cid"] or f["ccn"]):
                x = 1
                break

        famt = fam_titles.get(d["fam"], "")
        if d["fam"] and famt:
            fams[d["fam"]] = famt

        row = {"code": code, "t": d["t"], "cat": cat, "fam": d["fam"],
               "def": d["def"], "ex": d["ex"], "act": d["act"]}
        if x:
            row["x"] = 1
        rows.append(row)

    # ---- topcip map: TOP -> {t: title, c: [[cip, tier], ...]} ----
    # Only real CIP rows (present in `desc`) are kept so the tab's per-code lookup
    # always resolves. Candidates are ordered official > field > noncredit >
    # generic (a stable default; the tab re-ranks by description-fit).
    import re as _re
    known = set(desc)
    topcip = {}
    for tc, slot in top_cips.items():
        # skip non-TOP sentinel keys (the workbook carries a "No TOP Match in Records"
        # bucket with ~1,331 pairs — 27% of the map — that no real course ever hits).
        if not _re.match(r"^\d{4}\.\d{2}$", tc):
            continue
        cand = sorted(((cip, tier) for cip, tier in slot.items() if cip in known),
                      key=lambda ct: (TIER_RANK[ct[1]], ct[0]))
        if cand:
            topcip[tc] = {"t": top_title.get(tc, ""), "c": [[c, t] for c, t in cand]}

    payload = {
        "_built_at": BUILT_AT,
        "_built_by": "kb/_build_cip_crosswalk.py",
        "_source": SRC_LABEL,
        "_note": "CIP Code Taxonomy — the faculty-facing reference manual for the "
                 "TOP-to-CIP transition (ESS 26-06, fall 2026). The successor to the "
                 "CCC TOP Code Manual. Full federal CIP-2020 list; COE hosts the "
                 "TOP↔CIP crosswalk.",
        "_category": "rows[].cat = the Chancellor's Office CERTIFIED CIP CTE "
                     "designation (CTE / Both / Non-CTE / Noncredit / Retired / "
                     "Reserved). Authority = kb/reference/cip_cte_certified_260715.json "
                     "for the 244 codes where the workbook tabs disagree.",
        "_transfer": "rows[].x = 1 if any TOP that maps to this CIP has a course "
                     "carrying a C-ID (transfer-model) or CCN (AB 1111). A FLOOR, not "
                     "full CSU/UC transferability (TOP→CIP is one-to-many).",
        "_topcip": "topcip[<TOP>] = {t:title, c:[[cip,tier]]} — the official TOP→CIP "
                   "crosswalk candidates per TOP, ordered official>field>noncredit>"
                   "generic. Powers the 'Find my course's code' easy button (the tab "
                   "re-ranks candidates by description-fit — TOP corroborates, never "
                   "decides). tier: o official, f field-submitted, n noncredit, g "
                   "generic. boiler = universal noncredit CIPs the tab collapses.",
        "fams": fams,
        "rows": rows,
        "topcip": topcip,
        "boiler": boiler,
    }

    # Optional authoritative 4-digit series titles (see load_sub4). Only emitted
    # when a source file is present, so a no-source rebuild stays byte-compatible.
    sub4 = load_sub4([r["code"] for r in rows])
    if sub4:
        payload["_sub4"] = ("sub4[<NN.NN>] = authoritative NCES 4-digit SERIES title "
                            "(the workbook we build from is 6-digit-only; series titles "
                            "come from an operator-supplied NCES all-levels export — never "
                            "inferred).")
        payload["sub4"] = sub4

    body = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("// cip_crosswalk_data.js — GENERATED by kb/_build_cip_crosswalk.py. Do not edit by hand.\n")
        f.write("// Source: CCCCO CIP Searchable Workbook (2026-07-15) + certified CTE designations.\n")
        f.write("// Rebuild: python kb/_build_cip_crosswalk.py\n")
        f.write("window.CIP_CROSSWALK = ")
        f.write(body)
        f.write(";\n")

    from collections import Counter
    dist = Counter(r["cat"] for r in rows)
    flagged = sum(1 for r in rows if r.get("x"))
    goforward = sum(1 for r in rows if r["cat"] in ("CTE", "Both", "Non-CTE", "Noncredit"))
    print(f"CIP codes:        {len(rows)}")
    print(f"Categories:       {dict(dist)}")
    print(f"Go-forward:       {goforward}  (Retired/Reserved hidden by default)")
    print(f"Certified hits:   {cert_hits} of {len(certified)}  |  uncertified conflicts: {uncertified}")
    print(f"C-ID/CCN flagged: {flagged}")
    print(f"Families:         {len(fams)}")
    print(f"4-digit titles:   {len(sub4)}  ({'from ' + ('cip_series4_titles.json / NCES export' ) if sub4 else 'none — drop an NCES all-levels export in kb/reference/ to populate'})")
    npairs = sum(len(v["c"]) for v in topcip.values())
    print(f"TOP→CIP map:      {len(topcip)} TOPs, {npairs} candidate pairs; boiler {boiler}")
    print(f"Wrote {OUT}  ({os.path.getsize(OUT) / 1024:.0f} KB)")


if __name__ == "__main__":
    main()
