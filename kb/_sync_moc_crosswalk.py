"""
Military COOL/MOC crosswalk → kb/reference/moc_crosswalk.json

The MOC authority-anchoring lane (Sam, 2026-07-07: "Let's wire in the COOL/MOC
crosswalk too"): pull the military-occupation → civilian-occupation crosswalk
(MOC = Military Occupational Classification: Army/Marine MOS, Navy/Coast Guard
ratings, Air Force AFSC) so the credential KB can answer the veteran pathway
question — MOC → O*NET-SOC → certifications (the COS registry's cert-finder
occupation filter) → CER credentials via their cos_cert_id anchors → colleges
with articulations. Measured 2026-07-07: the CER's own military exhibit layer
is tiny (25 raws / 16 JST-course families, ZERO occupation codes), so this
lane's consumer is the PATHWAY bridge (Sierra / the recommender scope), not an
exhibit-title matcher.

RUNS ON A GITHUB ACTIONS RUNNER ONLY (.github/workflows/moc-crosswalk-sync.yml)
— onetcenter.org and careeronestop.org are egress-blocked from the agent
sandbox (runner-as-proxy, docs/kb-notes/playbook-runner-as-external-api-proxy.md).

Two lanes, tried in order:
  1. BULK — the O*NET Center's published MOC crosswalk file (public, CC-BY):
     probe scrapes https://www.onetcenter.org/crosswalks.html for the military
     crosswalk link and dumps every candidate href (the COS find_bulk_link
     diagnostic pattern), or set ONET_MOC_URL to pin the file directly.
  2. API — candidate CareerOneStop MOC endpoints with the SAME registered
     credentials as the COS lane (repo secrets COS_USER_ID + COS_API_TOKEN).
     The exact MOC path shape is UNVERIFIED — probe mode tries several and
     prints status + body head for each, so the first probe log pins the real
     contract before any apply parses it.

Probe extras: P3 tests WHICH positional zero in the certification-finder path
is the occupation filter (SOC 49-3023.00, Automotive Service Technicians) —
the SOC→certifications leg of the bridge rides the API we already sync.

Modes:
  python3 kb/_sync_moc_crosswalk.py --probe   # reachability, candidate links,
      endpoint statuses, header/row samples. WRITES NOTHING. Dispatch first.
  python3 kb/_sync_moc_crosswalk.py --apply   # write the reference file
      (idempotent: unchanged crosswalk → file left untouched).

Terms: O*NET data is CC-BY (attribute the U.S. Department of Labor / O*NET);
CareerOneStop data carries the USDOL ETA + MN DEED attribution the CER already
renders. The synced file stays tracker-internal under kb/reference/ pending a
served-consumer decision (the COS-registry precedent, pages.yml prune).
"""
import csv
import io
import json
import os
import re
import ssl
import sys
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime, timezone

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "reference", "moc_crosswalk.json")

CROSSWALKS_PAGE = "https://www.onetcenter.org/crosswalks.html"
ONET_MOC_URL = os.environ.get("ONET_MOC_URL")  # optional pin from the probe log
API_BASE = "https://api.careeronestop.org"
UA = ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
      "Chrome/126.0 Safari/537.36 cpl-project-tracker-sync")

ATTRIBUTION = ("MOC↔O*NET-SOC crosswalk: U.S. Department of Labor, O*NET "
               "(onetcenter.org, CC-BY 4.0). Related certification data: "
               "CareerOneStop (careeronestop.org), sponsored by USDOL ETA; "
               "maintained by Minnesota DEED.")

COS_USER_ID = os.environ.get("COS_USER_ID")
COS_API_TOKEN = os.environ.get("COS_API_TOKEN")


def _get(url, headers=None, timeout=120, attempts=3):
    import time
    req = urllib.request.Request(url, headers={"User-Agent": UA, **(headers or {})})
    ctx = ssl.create_default_context()
    last = None
    for i in range(attempts):
        try:
            with urllib.request.urlopen(req, timeout=timeout, context=ctx) as r:
                return r.read(), dict(r.headers)
        except urllib.error.HTTPError as e:
            last = e
            if e.code not in (403, 429, 500, 502, 503):
                raise
            if i < attempts - 1:
                wait = 5 * (i + 1)
                print(f"  HTTP {e.code} on {url.split('?')[0]} — retry in {wait}s")
                time.sleep(wait)
    raise last


def api_get(path):
    blob, _ = _get(API_BASE + path, headers={
        "Authorization": "Bearer " + COS_API_TOKEN, "Accept": "application/json"})
    return json.loads(blob)


# ── Lane 1: O*NET bulk crosswalk ────────────────────────────────────────────

def find_moc_link(html, probe=False):
    """Scrape the crosswalks page for the military/MOC download link. In probe
    mode dump every candidate href so a pattern drift is diagnosable from the
    run log (the COS find_bulk_link pattern)."""
    if probe:
        print(f"  page bytes: {len(html):,}")
        cands = re.findall(r"""["']([^"']*(?:\.zip|\.xlsx|\.csv|crosswalk[^"']*))["']""",
                           html, flags=re.I)
        seen = []
        for c in cands:
            if c not in seen and re.search(r"military|moc|\.(zip|xlsx|csv)$", c, re.I):
                seen.append(c)
        print(f"  candidate hrefs ({len(seen)}):")
        for c in seen[:40]:
            print(f"    {c}")
    links = re.findall(r"""["']([^"']*(?:military|moc)[^"']*\.(?:zip|xlsx|csv))["']""",
                       html, flags=re.I)
    if not links:
        return None
    best = links[0]
    if best.startswith("http"):
        return best
    return urllib.parse.urljoin(CROSSWALKS_PAGE, best)


HEADER_MAP = {
    # Tolerant header → canonical field mapping; the O*NET crosswalk file's
    # exact headers are unverified until the first probe — extend from the log.
    "branch": ["branch", "service", "svc", "military service"],
    "code": ["moc", "moc code", "military occupation code", "code",
             "military code", "mos code", "moc_code"],
    "title": ["moc title", "military occupation title", "title",
              "military title", "moc_title"],
    "soc": ["o*net-soc code", "onet soc code", "onetsoc code", "soc code",
            "o*net-soc", "onet code", "onetsoc_code", "onetsoc", "onet_soc",
            "onet-soc", "soc", "onet", "onet_soc_code", "onetcode"],
    "soc_title": ["o*net-soc title", "onet soc title", "soc title",
                  "onetsoc_title", "occupation"],
    "status": ["status", "active", "moc status"],
}


def map_headers(headers):
    low = [str(h).strip().lower() for h in headers]
    out = {}
    for field, names in HEADER_MAP.items():
        for n in names:
            if n in low:
                out[field] = low.index(n)
                break
    return out


def read_archive(blob, name, probe):
    """Yield (member_name, headers, rows) for each parseable member (a
    crosswalk zip can carry several service-branch files)."""
    def parse_csvish(data, low):
        text = data.decode("utf-8-sig", errors="replace")
        delim = "\t" if text.count("\t") > text.count(",") else ","
        rows = list(csv.reader(io.StringIO(text), delimiter=delim))
        return (rows[0], rows[1:]) if len(rows) > 1 else (None, None)

    low = name.lower()
    if low.endswith(".zip"):
        zf = zipfile.ZipFile(io.BytesIO(blob))
        names = zf.namelist()
        if probe:
            print(f"  archive members: {names}")
        for n in names:
            data = zf.read(n)
            nl = n.lower()
            if nl.endswith((".csv", ".txt", ".tsv")):
                headers, rows = parse_csvish(data, nl)
                if headers:
                    yield n, headers, rows
            elif nl.endswith((".xlsx", ".xlsm")):
                yield from read_xlsx(data, n)
    elif low.endswith((".xlsx", ".xlsm")):
        yield from read_xlsx(blob, name)
    else:
        headers, rows = parse_csvish(blob, low)
        if headers:
            yield name, headers, rows


def read_xlsx(data, name):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        try:
            headers = [str(c) if c is not None else "" for c in next(it)]
        except StopIteration:
            continue
        yield f"{name}#{sheet}", headers, [list(r) for r in it]


def lane_bulk(probe):
    link = ONET_MOC_URL
    if not link:
        print(f"BULK lane: GET {CROSSWALKS_PAGE}")
        html, _ = _get(CROSSWALKS_PAGE)
        link = find_moc_link(html.decode("utf-8", errors="replace"), probe)
    print(f"  MOC crosswalk link: {link}")
    if not link:
        return None
    blob, hdrs = _get(link, timeout=300)
    print(f"  downloaded {len(blob):,} bytes ({hdrs.get('Content-Type')})")
    mocs = {}  # (branch, code) -> rec
    for member, headers, rows in read_archive(blob, link.rsplit("/", 1)[-1], probe):
        idx = map_headers(headers)
        # Probe prints EVERY header — the first probe truncated at 8 and hid the
        # SOC column, which is exactly the drift this diagnostic exists to catch.
        shown = headers if probe else headers[:8] + (["…"] if len(headers) > 8 else [])
        print(f"  member: {member} | headers: {shown} | rows: {len(rows)} "
              f"| mapped: {sorted(idx)}")
        if probe:
            for r in rows[:2]:
                print(f"    sample: {r[:8]}")
            continue
        if "code" not in idx or "soc" not in idx:
            print(f"    (skipped — no code/soc columns; extend HEADER_MAP if this "
                  f"member should parse)")
            continue
        for row in rows:
            def cell(f):
                i = idx.get(f)
                return (str(row[i]).strip() if i is not None and i < len(row)
                        and row[i] is not None else "")
            code = cell("code")
            if not code:
                continue
            key = (cell("branch"), code)
            rec = mocs.setdefault(key, {
                "branch": cell("branch"), "code": code, "title": cell("title"),
                "socs": []})
            soc = cell("soc")
            if soc and soc not in rec["socs"]:
                rec["socs"].append(soc)
    if probe:
        return "PROBED"
    return sorted(mocs.values(), key=lambda m: (m["branch"], m["code"])) or None


# ── Lane 2 (PROBE-ONLY for now): CareerOneStop MOC endpoints ────────────────

def probe_cos_moc():
    """The exact CareerOneStop MOC path shape is unverified — try candidates
    and print status + body head so the first probe log pins the contract.
    Also P3: find WHICH positional zero in the certification-finder path is
    the occupation filter (the SOC→certs leg of the bridge)."""
    if not (COS_USER_ID and COS_API_TOKEN):
        print("API probes: COS_USER_ID / COS_API_TOKEN not set — skipped.")
        return
    candidates = [
        ("moc keyword", f"/v1/moc/{COS_USER_ID}/91B/1/5"),
        ("moc finder-shaped", f"/v1/moc/{COS_USER_ID}/91B/0/0/Title/ASC/1/5"),
        ("mocs list", f"/v1/mocs/{COS_USER_ID}/91B/1/5"),
        ("military transition", f"/v1/militarytransition/{COS_USER_ID}/91B/1/5"),
    ]
    for label, path in candidates:
        try:
            d = api_get(path)
            print(f"  P2 {label}: 200 — {json.dumps(d)[:280]}")
        except Exception as e:
            print(f"  P2 {label}: {type(e).__name__}: {e}")
    # P3 — occupation-filter position in the certification finder. The COS sync
    # uses {kw}/false/0/0/0/0/0/Name/ASC — five zeros between directFlag and
    # sort; test a SOC in each of positions 3..5 (0=industry..4=agency guess).
    soc = "49-3023.00"
    base = f"/v1/certificationfinder/{COS_USER_ID}"
    trials = [
        # Probe 1 finding: a dotted SOC in any positional-zero slot 404s (the
        # API validates segment formats). Try the SOC as the KEYWORD instead —
        # the finder's search understands occupation terms — plus the undotted
        # and truncated code shapes in the likeliest filter slot.
        ("occ-as-keyword", f"{base}/{urllib.parse.quote(soc)}/false/0/0/0/0/0/Name/ASC/1/3"),
        ("occ-kw-undotted", f"{base}/49-3023/false/0/0/0/0/0/Name/ASC/1/3"),
        ("occ@pos4-undotted", f"{base}/%20/false/0/0/0/493023/0/Name/ASC/1/3"),
    ]
    for label, path in trials:
        try:
            d = api_get(path)
            n = d.get("RecordCount")
            first = (d.get("CertList") or [{}])[0].get("Name", "")
            print(f"  P3 {label}: 200 — RecordCount={n}, first={first!r}")
        except Exception as e:
            print(f"  P3 {label}: {type(e).__name__}: {e}")


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    probe = "--probe" in sys.argv or "--apply" not in sys.argv
    print(f"mode: {'probe' if probe else 'apply'}")

    mocs = None
    try:
        mocs = lane_bulk(probe)
    except Exception as e:
        print(f"bulk lane failed: {type(e).__name__}: {e}")

    if probe:
        probe_cos_moc()
        print("PROBE complete — read the findings above, pin ONET_MOC_URL / extend "
              "HEADER_MAP if needed, then dispatch with mode=apply.")
        return
    if not mocs or mocs == "PROBED":
        sys.exit("Bulk lane produced no crosswalk rows — probe first, then fix.")

    try:
        prior = json.load(open(OUT, encoding="utf-8")).get("mocs")
    except (FileNotFoundError, ValueError):
        prior = None
    if prior == mocs:
        print(f"no change: {len(mocs)} MOC records — file left untouched")
        return

    out = {
        "_about": ("Military Occupational Classification (MOC) → O*NET-SOC "
                   "crosswalk, slimmed for the CER/Sierra veteran-pathway "
                   "bridge (MOC → SOC → certifications via the COS registry's "
                   "occupation filter → CER cos_cert_id anchors). Synced by "
                   "kb/_sync_moc_crosswalk.py on a GitHub Actions runner "
                   "(moc-crosswalk-sync.yml) — the agent sandbox cannot reach "
                   "onetcenter.org. Tracker-internal pending a served-consumer "
                   "decision; display requires the attribution below."),
        "_attribution": ATTRIBUTION,
        "_synced_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "count": len(mocs),
        "mocs": mocs,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=1, ensure_ascii=False)
        f.write("\n")
    print(f"wrote {OUT}: {len(mocs)} MOC records")


if __name__ == "__main__":
    main()
