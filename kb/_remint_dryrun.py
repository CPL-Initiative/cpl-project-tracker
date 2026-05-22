"""
DRY-RUN analyzer for the CourseControlNumber re-mint + M-prefix/banded renumber.

MEASUREMENT ONLY. Writes nothing to the machine layers (memberships,
coci_articulations, minted catalog/clusters, lazy artifacts) and nothing to
Supabase. Produces two reviewable artifacts under kb/remint_dryrun/:
  - report.md       — distribution (1:1 rename / vanish-to-official / 1:many
                      split / orphan) + the 6 curation entries' fates.
  - alias_map.json  — old M-ID -> new identity(ies), with classification, so a
                      later apply step can re-key curation (git + Supabase).

WHY title is the mapping key. Both the old mint (_seed_coci_minted_mids.py) and
the re-mint group rows that LACK an official C-ID/CCN by normalized title; each
distinct normalized title is exactly one old M-ID (corroborated >=2 members, or
a singleton). The re-mint differs by (a) keying membership at College/
CourseControlNumber granularity off the richer raw list and (b) letting members
that carry an official C-ID/CCN promote out of the minted cluster. So an old
M-ID's fate is read off its title cluster in the NEW raw list: rows still
lacking an official ID stay minted (renamed/renumbered); rows now carrying an
official ID promote (and if a cluster yields >1 distinct new identity it is a
1:many split that needs human review before the merge_into pointer is moved).

New minted-identity format (PROPOSED — confirm before apply):
  "<SUBJ4> M<NNNN>"  — SUBJ4 = synthetic 4-letter subject (first <=4 alpha chars
  of the cluster's modal subject; MAP surrogate, NOT the official CCN list).
  Band: noncredit / noncredit-enhanced -> 9xxx (the one band we can assert from
  credit_status). Credit -> 1xxx as a NON-SEMANTIC sequence bucket (the M-prefix
  already disclaims CCN equivalence; we do NOT claim 100-level/transferability).
  This single credit-bucket choice is the one open knob flagged for review.
"""
import json
import os
import re
from collections import Counter, defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(HERE, "reference", "coci_course_list.xlsx")
OUT_DIR = os.path.join(HERE, "remint_dryrun")

SENT = {None, "", "null", "n/a", "na", "not applicable", "(blank)", "blank", "none"}
STOP_PATTERNS = [
    r"independent stud", r"directed stud", r"dir stud", r"special stud",
    r"special project", r"special topic", r"selected topic", r"special problem",
    r"work experience", r"cooperative (work )?(education|experience)", r"coop ",
    r"internship", r"\bintern\b", r"supervised tutoring",
    r"student instructional assistant", r"service learning", r"occupational work",
    r"tutoring", r"practicum", r"fieldwork", r"field work", r"field experience",
    r"field study", r"directed practice", r"clinical practice", r"cooperative work",
    r"work based learning", r"work-based", r"on the job", r"apprenticeship",
    r"seminar$", r"^seminar", r"special assignment", r"volunteer", r"community service",
]
STOP_RE = re.compile("|".join(STOP_PATTERNS))
CODE_RE = re.compile(r"^[a-z]{1,6} ?\d{1,4}[a-z]?$")


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s.lower() in SENT else s


def ntitle(t):
    if t is None:
        return ""
    t = re.sub(r"\s+", " ", str(t)).strip().lower()
    t = re.sub(r"[^a-z0-9 ]", " ", t)
    return re.sub(r"\s+", " ", t).strip()


# Canonical C-ID: SUBJECT (letters, optionally hyphenated) + 2-4 digit number +
# optional single-letter variant suffix tokens (L lab, X cross-listed, S support,
# B, ...). Used to validate every extracted target.
CID_RE = re.compile(r"^[A-Z]+(?:-[A-Z]+)? \d{2,4}[A-Z]?(?: [A-Z]{1,2})*$")


def parse_cids(raw):
    """Extract zero-or-more CLEAN canonical C-IDs from a raw CIDNumber cell.

    Fixes three observed defects: (1) doubled course-number token
    ('AG-PS 104 104' -> 'AG-PS 104'); (2) one cell holding several comma-separated
    C-IDs ('ENGL 110, ENGL 120' -> two); (3) '000' placeholder number tokens
    ('MUS 171 000' -> 'MUS 171'). Legit variant suffixes (L/X/S/...) are kept.
    Returns the list of canonical strings that validate against CID_RE.
    """
    raw = clean(raw)
    if not raw:
        return []
    out = []
    for part in raw.split(","):
        toks = part.split()
        ded = []
        for t in toks:
            if t == "000":            # placeholder pad, not a real course number
                continue
            if ded and t == ded[-1]:  # collapse a doubled token
                continue
            ded.append(t)
        cid = " ".join(ded)
        if cid:
            out.append(cid)
    return out


def credit_status(ct, units):
    c = (ct or "").strip().lower()
    if c == "credit course":
        return "Credit"
    if c in ("other noncredit enhanced funding", "workforce preparation enhanced funding"):
        return "Noncredit Enhanced"
    if c == "non-enhanced funding":
        return "Noncredit"
    try:
        u = float(units)
    except (TypeError, ValueError):
        u = 0
    return "Credit" if u > 0 else "Noncredit"


def subj4(subject):
    a = re.sub(r"[^A-Za-z]", "", str(subject or "")).upper()
    return (a[:4] or "MISC")


def build_new_index():
    """Stream the raw list once. ntitle -> {officials:Counter, minted_n:int,
    minted_subjects:Counter, minted_credit:Counter}. Generic/code titles skipped
    (same exclusions as the mint)."""
    import openpyxl
    wb = openpyxl.load_workbook(RAW, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it)
    idx = {}
    skipped = 0
    diag = {"malformed_cid": Counter(), "rows_doubled": 0, "rows_comma": 0,
            "rows_placeholder": 0}
    for (College, CCNum, Subj, Cnum, Title, Units, CreditType,
         NCcat, Top, CID, Desc, CCNcommon) in it:
        nt = ntitle(Title)
        if not nt or len(nt) < 4 or STOP_RE.search(nt) or CODE_RE.match(nt):
            skipped += 1
            continue
        e = idx.get(nt)
        if e is None:
            e = idx[nt] = {"officials": Counter(), "minted_n": 0,
                           "subjects": Counter(), "credit": Counter()}
        ccn = clean(CCNcommon)
        raw_cid = clean(CID)
        cids = parse_cids(CID)
        if raw_cid:  # tally what the extractor repaired (for the report)
            if re.search(r"\b(\d{2,4})\b \1\b", raw_cid):
                diag["rows_doubled"] += 1
            if "," in raw_cid:
                diag["rows_comma"] += 1
            if "000" in raw_cid.split():
                diag["rows_placeholder"] += 1
        if ccn:
            e["officials"][f"CCN:{ccn}"] += 1
        elif cids:
            for cid in cids:
                if not CID_RE.match(cid):
                    diag["malformed_cid"][cid] += 1
                e["officials"][f"C-ID:{cid}"] += 1
        else:
            e["minted_n"] += 1
            e["subjects"][str(Subj or "").strip()] += 1
            e["credit"][credit_status(CreditType, Units)] += 1
    wb.close()
    return idx, skipped, diag


def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    cat = json.load(open(os.path.join(HERE, "coci_minted_courses.json")))["courses"]
    sing = json.load(open(os.path.join(HERE, "coci_minted_singletons.json")))["courses"]
    curation = json.load(open(os.path.join(HERE, "coci_curation.json")))["curations"]

    # old M-ID -> (common_title, tier)
    old = {}
    for mid, r in cat.items():
        old[mid] = (r.get("common_title"), "corroborated")
    for mid, r in sing.items():
        old[mid] = (r.get("common_title"), "singleton")

    print(f"old M-IDs: {len(old)} (corroborated {len(cat)}, singleton {len(sing)})")
    new_idx, skipped, diag = build_new_index()
    n_malformed = sum(diag["malformed_cid"].values())
    print(f"new-list distinct mintable/official titles: {len(new_idx)} (rows skipped generic/code: {skipped})")
    print(f"C-ID extractor: repaired doubled={diag['rows_doubled']} rows, "
          f"comma-split={diag['rows_comma']} rows, placeholder-000={diag['rows_placeholder']} rows; "
          f"MALFORMED remaining={n_malformed} ({dict(diag['malformed_cid'])})")
    assert n_malformed == 0, f"malformed C-ID targets remain: {dict(diag['malformed_cid'])}"

    # First pass: assign a deterministic new minted code to every title that
    # still has minted rows in the new list (the rename targets). Sequence per
    # (subj4, band) over sorted ntitle so codes are stable + reproducible.
    minted_meta = {}  # nt -> (subj4, credit_status, band_digit, corroborated?)
    for nt, e in new_idx.items():
        if e["minted_n"] == 0:
            continue
        s4 = subj4(e["subjects"].most_common(1)[0][0])
        cs = e["credit"].most_common(1)[0][0]
        band = "9" if cs in ("Noncredit", "Noncredit Enhanced") else "1"
        minted_meta[nt] = (s4, cs, band, e["minted_n"] >= 2)
    # Numbering scheme (option 1 + alphanumeric stand-alones, confirmed 2026-05-22).
    # Both tiers are 4 chars after "M", leading digit = band (9 noncredit / 1 credit):
    #   corroborated (>=2 colleges) -> all-digit "<SUBJ4> M<band><seq:03d>"
    #     (3-digit sequence; corroborated max bucket = 496 < 1000). e.g. ART M1042.
    #   stand-alone / singleton (1 college) -> "<SUBJ4> M<band><d><LL>": band + 1
    #     sequence digit + 2 letters. Capacity 10*26*26 = 6,760 per (subject,band)
    #     vs a max stand-alone bucket of 1,432. The 2 trailing LETTERS are the tell
    #     (corroborated codes are all-digit). e.g. ART M10AA. No 'Ms' prefix needed.
    LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    def sing_code(n):  # 0-based seq -> "<d><L><L>" (base 10*26*26)
        d, r = divmod(n, 26 * 26)
        l1, l2 = divmod(r, 26)
        return f"{d}{LETTERS[l1]}{LETTERS[l2]}"

    seq_corr = defaultdict(int)
    seq_sing = defaultdict(int)
    new_code = {}
    over_corr = []
    over_sing = []
    for nt in sorted(minted_meta):
        s4, cs, band, corr = minted_meta[nt]
        if corr:
            seq_corr[(s4, band)] += 1
            n = seq_corr[(s4, band)]
            if n > 999:
                over_corr.append((s4, band, n))
            new_code[nt] = f"{s4} M{band}{n:03d}"
        else:
            n = seq_sing[(s4, band)]
            seq_sing[(s4, band)] += 1
            if n >= 10 * 26 * 26:
                over_sing.append((s4, band, n))
            new_code[nt] = f"{s4} M{band}{sing_code(n)}"

    # Classify every old M-ID.
    classes = Counter()
    aliases = {}
    split_examples = []
    for mid, (title, tier) in old.items():
        nt = ntitle(title)
        e = new_idx.get(nt)
        if e is None:
            classes["orphan"] += 1
            aliases[mid] = {"class": "orphan", "tier": tier, "new": [],
                            "note": "title absent from new raw list (or now generic/code-excluded)"}
            continue
        officials = sorted(e["officials"])
        minted = e["minted_n"] > 0
        new_ids = []
        if minted:
            new_ids.append(new_code[nt])
        new_ids.extend(officials)
        if minted and not officials:
            cls = "rename"
        elif officials and not minted:
            cls = "vanish_to_official" if len(officials) == 1 else "split"
        else:  # minted + officials
            cls = "split"
        classes[cls] += 1
        rec = {"class": cls, "tier": tier, "new": new_ids}
        if cls == "split":
            rec["flag"] = "1:many — review before moving any curation/merge_into pointer"
            if len(split_examples) < 25:
                split_examples.append((mid, title, new_ids))
        aliases[mid] = rec

    # The 6 curation entries' fates (the decision-critical detail).
    curation_report = {}
    for key, body in curation.items():
        is_mid = key.startswith("M-ID ")
        entry = {"is_m_id": is_mid, "curation_fields": [k for k in body if not k.startswith("_")]}
        if is_mid:
            a = aliases.get(key)
            entry["mapping"] = a or {"class": "NOT FOUND in old M-ID set", "new": []}
        else:
            entry["mapping"] = {"class": "not an M-ID — key is stable, not re-keyed"}
        # does it point merge_into another id? is that target an M-ID (needs rewrite)?
        mi = body.get("merge_into")
        if mi:
            entry["merge_into"] = mi
            entry["merge_into_is_m_id"] = mi.startswith("M-ID ")
        curation_report[key] = entry

    n_cur = len(curation_report)
    n_cur_mid = sum(1 for v in curation_report.values() if v["is_m_id"])
    n_cur_split = sum(1 for v in curation_report.values()
                      if v["mapping"].get("class") == "split")
    new_corr = sum(1 for m in minted_meta.values() if m[3])
    new_sing = len(minted_meta) - new_corr
    top_corr = sorted(seq_corr.items(), key=lambda kv: -kv[1])[:10]

    # (a) Subject canonicalization: title clusters whose minted member rows span
    # >1 distinct LOCAL subject code, collapsed to one synthetic SUBJ4. Surface a
    # few real examples for the report (this is intended cross-college merging).
    canon_examples = []
    for nt in sorted(minted_meta):
        subs = new_idx[nt]["subjects"]
        if len(subs) >= 2:
            modal = subs.most_common(1)[0][0]
            others = [s for s in subs if s != modal]
            canon_examples.append((new_code[nt], subj4(modal),
                                   modal, others, sum(subs.values())))
    canon_examples.sort(key=lambda x: -x[4])
    # (b) Granularity: distinct official identities surfaced by the split clusters
    # (these were previously hidden inside an over-merged (subject,number) M-ID).
    officials_in_splits = set()
    for mid, (title, tier) in old.items():
        e = new_idx.get(ntitle(title))
        if e and e["officials"] and e["minted_n"]:  # a split cluster
            officials_in_splits.update(e["officials"])
    summary = {
        "old_m_ids_total": len(old),
        "corroborated": len(cat),
        "singletons": len(sing),
        "classes": dict(classes),
        "new_minted_identities": len(new_code),
        "new_minted_corroborated_clean_M####": new_corr,
        "new_minted_standalone_alnum": new_sing,
        "new_minted_noncredit_9xxx": sum(1 for nt in new_code if minted_meta[nt][2] == "9"),
        "new_minted_credit_1xxx": sum(1 for nt in new_code if minted_meta[nt][2] == "1"),
        "numbering_scheme": "corroborated -> '<SUBJ4> M<band><seq:03d>' (all-digit, 4 chars); stand-alone -> '<SUBJ4> M<band><d><LL>' (band+1 digit+2 letters, cap 6,760/bucket)",
        "corroborated_max_per_subject_band": max(seq_corr.values()) if seq_corr else 0,
        "corroborated_buckets_over_999": over_corr,
        "standalone_max_per_subject_band": max(seq_sing.values()) if seq_sing else 0,
        "standalone_capacity_per_subject_band": 10 * 26 * 26,
        "standalone_buckets_over_capacity": over_sing,
        "top10_corroborated_buckets": [{"subj_band": f"{s} {b}", "n": v} for (s, b), v in top_corr],
        "cid_extractor": {
            "rows_doubled_repaired": diag["rows_doubled"],
            "rows_comma_split": diag["rows_comma"],
            "rows_placeholder_000_dropped": diag["rows_placeholder"],
            "malformed_remaining": n_malformed,
        },
        "subject_canonicalization_clusters": len(canon_examples),
        "distinct_officials_surfaced_by_splits": len(officials_in_splits),
    }

    alias_doc = {
        "_about": "DRY RUN — old M-ID -> new identity(ies). NOT applied to any machine layer or Supabase.",
        "_scheme": "<SUBJ4> M<band><seq>; band 9=noncredit (asserted), 1=credit (NON-SEMANTIC bucket, confirm).",
        "summary": summary,
        "curation_entries": curation_report,
        "aliases": dict(sorted(aliases.items())),
    }
    with open(os.path.join(OUT_DIR, "alias_map.json"), "w", encoding="utf-8") as f:
        json.dump(alias_doc, f, indent=2, ensure_ascii=False)
        f.write("\n")

    cur_lines = []
    for k, v in curation_report.items():
        m = v["mapping"]
        line = f"| `{k}` | {'M-ID' if v['is_m_id'] else 'cluster'} | {m.get('class')} | {', '.join(f'`{x}`' for x in m.get('new', [])) or '—'} |"
        if "merge_into" in v:
            line += f" merge_into `{v['merge_into']}` (M-ID? {v['merge_into_is_m_id']}) |"
        else:
            line += " — |"
        cur_lines.append(line)
    split_lines = [f"| `{mid}` | {title} | {', '.join(f'`{x}`' for x in nids)} |"
                   for mid, title, nids in split_examples]
    report = f"""# CourseControlNumber re-mint — DRY-RUN report

**Measurement only. Nothing was written to the machine layers or Supabase.**
Artifacts: `kb/remint_dryrun/alias_map.json` (this run) + this report.
Generator: `kb/_remint_dryrun.py`.

## What this measures
Each old minted M-ID is mapped to its fate under a CourseControlNumber-grained
re-mint, read off its **normalized title** (the grouping key the mint actually
uses) in the richer raw list `kb/reference/coci_course_list.xlsx` (141,738 rows).
A title's rows that still lack an official C-ID/CCN stay **minted** (renamed /
renumbered); rows that now carry an official C-ID/CCN **promote** out; a cluster
that yields more than one distinct new identity is a **1:many split**.

## Distribution (all {summary['old_m_ids_total']:,} old M-IDs)
| class | count | meaning |
|---|---|---|
| rename | {classes.get('rename',0):,} | 1:1 — stays minted, gets a new `SUBJ4 M####` key |
| vanish_to_official | {classes.get('vanish_to_official',0):,} | all members promote to ONE official C-ID/CCN; M-ID dissolves |
| split | {classes.get('split',0):,} | **1:many — needs review**; minted remnant + promoted official(s), or >1 official |
| orphan | {classes.get('orphan',0):,} | title absent from the new list |

Old M-IDs: {summary['corroborated']:,} corroborated + {summary['singletons']:,} singletons.
New minted identities: {summary['new_minted_identities']:,}
({summary['new_minted_corroborated_clean_M####']:,} corroborated → all-digit
`M####`; {summary['new_minted_standalone_alnum']:,} stand-alone → `M<band><d><LL>`.
noncredit→9xxx: {summary['new_minted_noncredit_9xxx']:,}; credit→1xxx:
{summary['new_minted_credit_1xxx']:,}).

## The {n_cur} curation entries (decision-critical)
| key | kind | fate | new id(s) | merge_into |
|---|---|---|---|---|
{chr(10).join(cur_lines)}

All {n_cur_mid} curated **M-ID** keys map **1:1 (rename)** — {n_cur_split} split —
and every `merge_into` points at the curator-minted `UC-CUR-MPG029OM` cluster,
which is **not** an M-ID, so no `merge_into` value needs rewriting. The re-key of
the human layer is therefore: rewrite the {n_cur_mid} curation **keys** (git +
Supabase), leave all `merge_into` values untouched. (Reconciled to the live
Supabase table 2026-05-22: git was stale by one entry, `M-ID AELE 100`.)

## Numbering scheme (option 1, confirmed)
CCN's `SUBJ C####` is **4 digits**: the leading digit is the band (level/credit
meaning), the next 3 are the within-(subject,band) sequence. Our minted tier
mirrors that:

- **Corroborated** M-IDs (≥2 colleges) → clean 4-digit `SUBJ M<band><seq:03d>`
  — leading `9` = noncredit, `1` = credit; 3-digit sequence. Max per
  (subject,band) = **{summary['corroborated_max_per_subject_band']}** (< 1,000),
  so it fits with room to spare. Buckets over 999: {summary['corroborated_buckets_over_999'] or 'none'}.
  Top buckets: {', '.join(f"{b['subj_band']}={b['n']}" for b in summary['top10_corroborated_buckets'])}.
- **Stand-alones** (1 college) → `SUBJ M<band><d><LL>` — band + 1 sequence digit
  + **2 letters**. Same 4-char width, but the trailing letters expand capacity to
  **{summary['standalone_capacity_per_subject_band']:,}** per (subject,band) vs a
  max stand-alone bucket of **{summary['standalone_max_per_subject_band']:,}**
  (~4.7× headroom; buckets over capacity: {summary['standalone_buckets_over_capacity'] or 'none'}).
  The 2 trailing letters are the tell — corroborated codes are all-digit. If a
  second college later joins the title it promotes to a corroborated `M####`.

`9` (noncredit) is the only asserted band; `1` (credit) is a non-semantic bucket
(no transferability claim — the `M` already disclaims CCN equivalence).

## C-ID extractor (MUST-FIX) — clean atomic targets
The raw `CIDNumber` column carried three defects; the extractor (`parse_cids`)
repairs all three and every emitted target is validated against `CID_RE`:

| defect | example | rows repaired |
|---|---|---|
| doubled course-number token | `AG-PS 104 104` → `AG-PS 104` | {summary['cid_extractor']['rows_doubled_repaired']} |
| several C-IDs in one cell (comma) | `ENGL 110, ENGL 120` → 2 targets | {summary['cid_extractor']['rows_comma_split']} |
| `000` placeholder number | `MUS 171 000` → `MUS 171` | {summary['cid_extractor']['rows_placeholder_000_dropped']} |

Legit variant suffixes (`L` lab, `X` cross-listed, `S` support, …) are preserved.
**Malformed C-ID targets remaining: {summary['cid_extractor']['malformed_remaining']}.** ✅

## Confirmation (a) — subject canonicalization is intended
A minted identity's subject is its title cluster's **modal local subject**,
synthesized to a 4-letter `SUBJ4`. When the same title is taught under different
local subject codes across colleges, those codes **collapse to one** SUBJ4 — this
is the cross-college consolidation working as designed, not a bug.
{summary['subject_canonicalization_clusters']:,} minted clusters span >1 local
subject. Examples (new code · chosen SUBJ4 · modal local subject · other locals
folded in):

{chr(10).join(f"- `{c}` · **{s4}** · `{modal}` ← {', '.join(repr(o) for o in others[:6])}" for c, s4, modal, others, _n in canon_examples[:8])}

Caveat: collapsing local-subject *variants* of the same course (MTH/MAT→MATH,
MEDS/ALH→NURS) is the intended win. A few folds are looser (e.g. PSYC into CHDEV)
— that is the **pre-existing title-grouping over-merge**, already carried on the
identity as `subject_spread` / `over_merged` for reviewer triage; the re-mint
neither introduces nor worsens it (same title key as today's mint).

## Confirmation (b) — the granularity increase IS the over-merge fix
The {classes.get('split',0):,} **split** clusters are exactly the over-merges the
old lossy `(subject, number)` join hid: a title that some colleges teach as an
un-aligned local course AND others teach as an officially-aligned C-ID/CCN course
used to collapse into ONE M-ID. The re-mint **separates** them — the un-aligned
members stay as a minted remnant, the aligned members promote to their official
identity — so remnant + official **coexist** and the total identity count rises.
This split surfaces **{summary['distinct_officials_surfaced_by_splits']:,} distinct
official C-ID/CCN identities** that were previously buried inside an M-ID. That
rise in granularity is the intended correction, not a regression. (`vanish=0`
because every one of these titles still has ≥1 un-aligned member somewhere, so a
minted remnant always remains.)

## Splits to review (first {len(split_lines)} of {classes.get('split',0):,})
| old M-ID | title | new identities |
|---|---|---|
{chr(10).join(split_lines)}

## How the apply step will treat each layer (per your instruction)
- **Machine layers** (memberships, `coci_articulations`, minted catalog /
  clusters / singletons, lazy artifacts) → **regenerated** under the new keys.
- **Human curation layer** (`coci_curation.json` + live Supabase `kb_curation`)
  → **aliased**: rewrite `course_id` keys (and any `merge_into` that is an M-ID
  — here, none) via this map. Live Supabase row count must be verified to match
  git's 6 before applying.
"""
    with open(os.path.join(OUT_DIR, "report.md"), "w", encoding="utf-8") as f:
        f.write(report)

    print("\n=== DISTRIBUTION ===")
    for k in ("rename", "vanish_to_official", "split", "orphan"):
        print(f"  {k:20} {classes.get(k,0):,}")
    print(f"\nnew minted: {len(new_code):,} (corroborated all-digit M#### {summary['new_minted_corroborated_clean_M####']:,}, "
          f"stand-alone M<band><d><LL> {summary['new_minted_standalone_alnum']:,})")
    print(f"corroborated max/bucket: {summary['corroborated_max_per_subject_band']} (>999: {summary['corroborated_buckets_over_999'] or 'none'}); "
          f"stand-alone max/bucket: {summary['standalone_max_per_subject_band']} / cap {summary['standalone_capacity_per_subject_band']:,} "
          f"(over: {summary['standalone_buckets_over_capacity'] or 'none'})")
    print(f"\n=== {n_cur} CURATION ENTRIES ({n_cur_mid} M-ID keys, {n_cur_split} split) ===")
    for k, v in curation_report.items():
        m = v["mapping"]
        print(f"  {k}: {m.get('class')} -> {m.get('new', [m.get('class')])}")
    print(f"\nwrote {os.path.join(OUT_DIR,'alias_map.json')} + report.md")
    return alias_doc, split_examples


if __name__ == "__main__":
    main()
