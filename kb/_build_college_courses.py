#!/usr/bin/env python3
"""Build the per-college COURSE CATALOGUE for Sierra (`chatbox_college_courses`).

WHY THIS EXISTS
---------------
To tell a college which of ITS OWN courses to articulate against a credit
recommendation, you need that college's course TITLES. Nothing in Supabase had
them:

  coci_college_offerings   is a TOP-aggregated rollup, and it carries two silent
                           caps set in chatbox/build_coci_offerings.py —
                           TITLES_TEXT_CAP = 900 and SAMPLE_PER_TOP = 8.
                           Measured 2026-08-13: 801 of 16,097 rows sit at
                           EXACTLY 900 chars, and 5,678 rows are capped at 8
                           samples while 5,077 rows have MORE courses than that.
                           Cerritos Welding Technology shows 8 of 121 courses,
                           all alphabetically-first AED ironworker rows, so
                           `WELD 214L — Flux Cored Arc Welding (FCAW)
                           Certification Laboratory` — the correct answer for
                           Sam's acceptance case — never appears.

  chatbox_exhibits         is what colleges ARTICULATED, not what they TEACH.

So this publishes the whole COCI course list at per-course grain. It is a
different instrument from the offerings rollup, not a replacement: the rollup
answers "does this college teach construction?" cheaply, this answers "which
exact course of theirs matches this recommendation?".

⚠️ DO NOT scope a candidate set by TOP code. TOP is faculty-entered with no
gatekeeper (~52% of consolidated M-IDs are TOP-mixed), and CLAUDE.md Rule 7
forbids using it for a gatekeeping determination. It is carried here so it can
CORROBORATE a match, never restrict which courses are considered.

Run from repo root:  python3 kb/_build_college_courses.py
Writes: kb/college_courses_payload.json  (gitignored; rebuilt on the runner)
"""
from __future__ import annotations

import json
import os
import re
from collections import Counter

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "kb", "reference", "coci_course_list.xlsx")
OUT = os.path.join(ROOT, "kb", "college_courses_payload.json")

_WS = re.compile(r"\s+")


def fix_moji(s: str) -> str:
    """Repair the double-encoded 'CaÃ±ada College' the COCI export carries, so
    names match the correct-unicode form the rest of the corpus uses."""
    if s and "Ã" in s:
        try:
            return s.encode("latin-1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            return s
    return s


def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("none", "nan", "n/a", "na") else _WS.sub(" ", s)


def build() -> tuple[list, dict]:
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}

    need = ["College", "Subject", "Course_Number", "CourseTitle", "UnitValue",
            "CreditType", "TopCode", "CIDNumber", "CourseControlNumber"]
    missing = [c for c in need if c not in ix]
    if missing:
        raise SystemExit(f"FATAL: coci_course_list.xlsx is missing columns: {missing}")

    seen: set = set()
    rows: list = []
    stats: Counter = Counter()
    for r in ws.iter_rows(min_row=2, values_only=True):
        college = fix_moji(clean(r[ix["College"]]))
        subject = clean(r[ix["Subject"]])
        number = clean(r[ix["Course_Number"]])
        title = clean(r[ix["CourseTitle"]])
        stats["read"] += 1
        # A course with no title cannot participate in title alignment, which is
        # this table's entire purpose. Count them rather than dropping silently.
        if not college or not title:
            stats["skipped_no_college_or_title"] += 1
            continue
        key = (college, subject, number, title)
        if key in seen:
            stats["skipped_duplicate"] += 1
            continue
        seen.add(key)
        units = clean(r[ix["UnitValue"]])
        try:
            units_f = float(units) if units else None
        except ValueError:
            units_f = None
        top_raw = clean(r[ix["TopCode"]])
        rows.append({
            "college": college,
            "subject": subject,
            "course_number": number,
            "course_title": title,
            "units": units_f,
            "credit_type": clean(r[ix["CreditType"]]),
            # "0956.50: Welding Technology" -> code + title, split so a
            # corroborating check can compare codes without string surgery.
            "top_code": top_raw.split(":", 1)[0].strip() if top_raw else "",
            "top_title": top_raw.split(":", 1)[1].strip() if ":" in top_raw else "",
            "cid": clean(r[ix["CIDNumber"]]),
            "control_number": clean(r[ix["CourseControlNumber"]]),
        })
    stats["rows"] = len(rows)
    stats["colleges"] = len({r["college"] for r in rows})
    return rows, dict(stats)


def main() -> int:
    rows, stats = build()
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump({"_generated_by": "kb/_build_college_courses.py",
                   "_stats": stats, "college_courses": rows}, fh, ensure_ascii=False)
    print(f"→ {OUT}")
    for k in sorted(stats):
        print(f"  {k:30} {stats[k]:,}")

    # Acceptance probe — the course the offerings rollup could not show.
    hits = [r for r in rows
            if r["college"] == "Cerritos College" and "FCAW" in r["course_title"].upper()]
    print(f"\nACCEPTANCE — Cerritos courses with 'FCAW' in the title: {len(hits)}")
    for r in hits:
        print(f"  {r['subject']} {r['course_number']:<8} {r['course_title']}  ({r['units']}u)")
    if not hits:
        raise SystemExit("FATAL: Cerritos FCAW course absent — the extract is wrong.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
