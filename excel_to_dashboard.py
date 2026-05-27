"""
Excel -> Dashboard Sync Script
==============================
Reads the CPL_Initiative_Project_List.xlsx and regenerates CPL_Data.js
so the HTML dashboard picks up any changes you made in Excel.

Usage:
    python excel_to_dashboard.py

Requirements:
    pip install openpyxl

The script reads from the Excel file in the same folder and overwrites CPL_Data.js.
Open (or refresh) CPL_Dashboard.html in your browser to see updated data.
"""

import json, os, re, sys, urllib.request
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Timezone ─────────────────────────────────────────────────────────
# CPL Initiative is a California project; all displayed dates, daily
# snapshots, and "last updated" stamps should reflect Pacific Time.
# The daily GitHub Actions workflow runs in UTC, which was rolling the
# date forward by a day in the evening PT. Use zoneinfo when available
# (Python 3.9+) and fall back to a fixed UTC-7 offset otherwise.
try:
    from zoneinfo import ZoneInfo
    _PT = ZoneInfo("America/Los_Angeles")
except ImportError:
    _PT = timezone(timedelta(hours=-7))  # PDT fallback; DST lost

def _now_pt():
    """Return current datetime in Pacific Time, naive (tzinfo stripped).
    Stripping tzinfo keeps arithmetic compatible with existing naive
    datetimes elsewhere in the pipeline."""
    return datetime.now(_PT).replace(tzinfo=None)

# ── Excel source: prefer SharePoint-synced copy, fall back to local ──
# The SharePoint-synced folder is the single source of truth for the Excel.
# Set SHAREPOINT_EXCEL to the path of your synced file. If it exists, the
# pipeline reads from there; otherwise it falls back to the local copy.
SHAREPOINT_EXCEL = os.path.join(
    os.path.expanduser("~"),
    "Riverside Community College District",
    "California MAP Initiative - Documents",
    "CPL Workplan Dashboard",
    "CPL_Initiative_Project_List_v3.xlsx",
)
_LOCAL_EXCEL = os.path.join(SCRIPT_DIR, "CPL_Initiative_Project_List_v3.xlsx")
EXCEL_FILE = SHAREPOINT_EXCEL if os.path.exists(SHAREPOINT_EXCEL) else _LOCAL_EXCEL
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "CPL_Data.js")

# SharePoint Excel-for-the-Web URL for the workbook. Used by the "Update"
# buttons so they deep-link to the exact row/cell instead of just downloading.
SHAREPOINT_EXCEL_URL = (
    "https://studentrcc.sharepoint.com/:x:/r/sites/MilitaryArticulationPlatform/"
    "Shared%20Documents/CCCCO/Claude%20Prompts/Projects/CPL%20Projects/"
    "CPL%20Project%20Tracker/CPL_Initiative_Project_List_v3.xlsx"
    "?d=wb836315f451d4e2cb03e6cf2b6af4568&csf=1&web=1&e=WJpt4H"
)
EXCEL_SHEET_NAME = "Project List"

def excel_cell_url(row, col="P"):
    """Build a SharePoint Excel-for-the-Web URL that opens at a specific cell."""
    if not row:
        return SHAREPOINT_EXCEL_URL
    # activeCell uses 'Sheet Name'!Cell form; URL-encode the sheet name.
    from urllib.parse import quote
    sheet = quote(f"'{EXCEL_SHEET_NAME}'", safe="")
    return f"{SHAREPOINT_EXCEL_URL}&activeCell={sheet}!{col}{row}"

# ── CPL Knowledge Base (public GitHub repo) ──
# Fetched at pipeline build time and embedded into the dashboard as window.CPL_KB
# so the College Custom Report generator can ground Claude's commentary in
# authoritative CPL framing without bundling stale copies in this repo.
CPL_KB_REPO   = "CPL-Initiative/cpl-knowledge-base"
CPL_KB_BRANCH = "main"
# Explicit allowlist of files to embed in the prompt. We avoid the GitHub
# directory-listing API because unauthenticated requests are aggressively
# rate-limited (60/hr/IP); raw.githubusercontent.com has a much higher budget.
# If new KB content is added, append it here.
CPL_KB_FILES = [
    "README.md",
    "glossary.md",
    "overview/overview.md",
    "overview/map-platform.md",
    "overview/ai-ready-california.md",
    "methodology/three-pillar-initiative-design.md",
    "methodology/sprint-based-execution.md",
    "methodology/evidence-first-advocacy.md",
    "methodology/infrastructure-first-scaling.md",
    "methodology/knowledge-consolidation.md",
    "methodology/pattern-scattered-to-systematic.md",
    "methodology/pattern-ai-as-accelerant-with-faculty-validation.md",
    "current-status/README.md",
]
CPL_KB_MAX_CHARS = 24000   # prompt budget for the embedded KB excerpt

def fetch_cpl_kb(max_chars=CPL_KB_MAX_CHARS):
    """Return a concatenated string of CPL KB markdown for prompt injection.
    Best-effort: any fetch failures are logged and skipped so the pipeline never
    breaks if GitHub is briefly unreachable."""
    sections = []
    for path in CPL_KB_FILES:
        raw_url = f"https://raw.githubusercontent.com/{CPL_KB_REPO}/{CPL_KB_BRANCH}/{path}"
        try:
            req = urllib.request.Request(raw_url, headers={"User-Agent": "cpl-dashboard-pipeline"})
            with urllib.request.urlopen(req, timeout=10) as r:
                body = r.read().decode("utf-8", errors="replace")
            sections.append(f"## {path}\n\n{body.strip()}")
        except Exception as e:
            print(f"  KB fetch warning ({path}): {e}")
    if not sections:
        return ""
    full = "\n\n---\n\n".join(sections)
    if len(full) > max_chars:
        full = full[:max_chars].rsplit("\n", 1)[0] + "\n\n[KB excerpt truncated for prompt size]"
    return full

HTML_FILE   = os.path.join(SCRIPT_DIR, "CPL_Dashboard.html")
LIVE_FILE    = os.path.join(SCRIPT_DIR, "live_metrics.json")
HISTORY_FILE = os.path.join(SCRIPT_DIR, "kpi_history.json")

# ── MAP Exhibit / Custom Report data source ─────────────────────────
# The Custom Reporting Module (customreportingmodule.azurewebsites.net) exports
# a combined JSON file containing multiple datasets.  The pipeline reads all
# available datasets and extracts KPIs from each.
#
# Search order for the JSON file:
#   1. CPL Project Tracker folder (primary — consolidated location)
#   2. Legacy MAP Exhibit Project folder paths (backward compatibility)
_EXHIBIT_LOCATIONS = [
    SCRIPT_DIR,  # primary: same folder as pipeline (consolidated)
    os.path.join(os.path.expanduser("~"), "Documents", "Claude", "Projects", "MAP Exhibit Project"),
    os.path.join(os.path.dirname(SCRIPT_DIR), "..", "MAP Exhibit Project"),  # sibling mount in Cowork
]

def _find_exhibit_json(folder):
    """Find the most recent CustomReport_*.json file in a folder.
    Prefers combined multi-dataset exports (largest file) over single-view exports.
    """
    if not os.path.isdir(folder):
        return None
    candidates = [f for f in os.listdir(folder)
                  if f.startswith("CustomReport_") and f.endswith(".json")]
    if not candidates:
        # Also check for Articulations-prefixed exports (older naming convention)
        candidates = [f for f in os.listdir(folder)
                      if f.startswith("Articulations_CustomReport_") and f.endswith(".json")]
    if not candidates:
        return None
    # Sort by date-stamp in filename (newest first)
    candidates.sort(reverse=True)
    return os.path.join(folder, candidates[0])

EXHIBIT_FILE = None
for _loc in _EXHIBIT_LOCATIONS:
    EXHIBIT_FILE = _find_exhibit_json(_loc)
    if EXHIBIT_FILE:
        break

# ── Column map (1-indexed) ──────────────────────────────────────────
COL_ID          = 1
COL_NAME        = 2
COL_DESC        = 3
COL_ACTIVITY    = 4
COL_V2030       = 5
COL_GOAL        = 6
COL_BUDGET_SRC  = 7
COL_BUDGET      = 8
COL_LEAD        = 9
COL_TEAM        = 10
COL_STATUS      = 11
COL_PCT         = 12
COL_START       = 13
COL_END         = 14
COL_MILESTONES  = 15
COL_UPDATE      = 16
COL_UPDATE_DATE = 17
COL_KPI_METRIC  = 18   # R
COL_KPI_UNIT    = 19   # S  — KPI Unit (single column for all years)
COL_KPI_G2526   = 20   # T  — KPI Goal 25-26
COL_KPI_S2526   = 21   # U  — KPI Stretch 25-26
COL_KPI_G2627   = 22   # V  — KPI Goal 26-27
COL_KPI_S2627   = 23   # W  — KPI Stretch 26-27
COL_KPI_G2728   = 24   # X  — KPI Goal 27-28
COL_KPI_S2728   = 25   # Y  — KPI Stretch 27-28
COL_KPI_G2829   = 26   # Z  — KPI Goal 28-29
COL_KPI_S2829   = 27   # AA — KPI Stretch 28-29
COL_KPI_G2930   = 28   # AB — KPI Goal 29-30 (Vision 2030)
COL_KPI_S2930   = 29   # AC — KPI Stretch 29-30
COL_WP_NOTES    = 30   # AD
COL_KPI_ORDER   = 31   # AE — display order for headline KPI cards (lower = first)
COL_SOURCE_LOGIC = 32  # AF — Source & Logic documentation (read-only reference)
COL_OVERRIDE     = 33  # AG — Algorithm Override (user-editable values)


# ── KPI Tunable Parameters & Algorithm Descriptions ─────────────────
# Default values for parameters that control KPI computation. Overridden
# by values in the "KPI_Config" sheet of the Excel workbook when present.
#
# Where each parameter is ACTUALLY applied varies:
#   - SAVINGS_DEFAULT       → this file (fallback when live scrape unavailable)
#   - TIER_*                → cloudflare-worker-proxy.js (tier classification)
#   - BEACON_PER_UNIT       → CCCCO dashboard (reference only; not recomputed here)
#   - TWENTY_YEAR_MULTIPLIER → CCCCO dashboard (reference only)
KPI_PARAMETERS_DEFAULTS = {
    "SAVINGS_DEFAULT":                       "$269M",
    "TIER_MIN_STUDENTS":                     500,
    "TIER_MIN_UNITS":                        3000,
    "TIER_MIN_AVG_UNITS_PER_STUDENT":        5,
    "TIER_MIN_TRANSCRIPTION_RATE":           0.25,
    "TIER_MIN_AVG_TRANSCRIBED_PER_STUDENT":  3,
    "BEACON_PER_UNIT_SAVINGS":               1420,
    "TWENTY_YEAR_MULTIPLIER":                4.0,
}

KPI_PARAMETER_META = {
    "SAVINGS_DEFAULT":                       ("this file (fallback only)",
        "Fallback value for Estimated Savings when live scrape unavailable."),
    "TIER_MIN_STUDENTS":                     ("cloudflare-worker-proxy.js L184",
        "Min CPL students for one tier criterion (college-level)."),
    "TIER_MIN_UNITS":                        ("cloudflare-worker-proxy.js L185",
        "Min CPL eligible units for one tier criterion."),
    "TIER_MIN_AVG_UNITS_PER_STUDENT":        ("cloudflare-worker-proxy.js L186",
        "Min average eligible units per student for one tier criterion."),
    "TIER_MIN_TRANSCRIPTION_RATE":           ("cloudflare-worker-proxy.js L187",
        "Min TranscribedUnits/Units ratio for one tier criterion (0.25 = 25%)."),
    "TIER_MIN_AVG_TRANSCRIBED_PER_STUDENT":  ("cloudflare-worker-proxy.js L188",
        "Min average transcribed units per student for one tier criterion."),
    "BEACON_PER_UNIT_SAVINGS":               ("CCCCO dashboard (reference)",
        "Per-unit savings factor (Beacon Economics). CCCCO computes the savings; this value is documentation only."),
    "TWENTY_YEAR_MULTIPLIER":                ("CCCCO dashboard (reference)",
        "Multiplier applied by CCCCO to one-year impact for 20-year projection."),
}


def read_kpi_parameters(wb):
    """Read tunable KPI parameters from the 'KPI_Config' sheet.
    Falls back to KPI_PARAMETERS_DEFAULTS for any missing entries.
    Returns a dict of {param: value} with all parameters populated,
    plus a '_last_modified' dict of {param: last_modified_str}.
    """
    params = dict(KPI_PARAMETERS_DEFAULTS)
    last_modified = {}
    if "KPI_Config" not in wb.sheetnames:
        return {**params, "_last_modified": last_modified}

    ws = wb["KPI_Config"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        if key not in params:
            continue
        val = row[1] if len(row) > 1 else None
        lm  = row[3] if len(row) > 3 else None
        if val is not None and val != "":
            default = KPI_PARAMETERS_DEFAULTS[key]
            if isinstance(default, int) and not isinstance(default, bool):
                try: params[key] = int(float(val))
                except (ValueError, TypeError): pass
            elif isinstance(default, float):
                try: params[key] = float(val)
                except (ValueError, TypeError): pass
            else:
                params[key] = str(val)
        if lm:
            last_modified[key] = str(lm)[:10] if hasattr(lm, "strftime") else str(lm)

    return {**params, "_last_modified": last_modified}


def ensure_kpi_config_sheet(wb):
    """Create the 'KPI_Config' sheet populated with defaults if it doesn't exist.
    Returns True if a new sheet was created (caller should save the workbook).
    """
    if "KPI_Config" in wb.sheetnames:
        return False
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False

    ws = wb.create_sheet("KPI_Config")
    ws.append(["Parameter", "Value", "Description", "Last Modified", "Where Applied"])

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0A2240", end_color="0A2240", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    today = _now_pt().strftime("%Y-%m-%d")
    for key, default_val in KPI_PARAMETERS_DEFAULTS.items():
        where, desc = KPI_PARAMETER_META.get(key, ("", ""))
        ws.append([key, default_val, desc, today, where])

    # Column widths
    widths = {"A": 42, "B": 12, "C": 60, "D": 16, "E": 38}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return True


# Structured algorithm descriptions for each card shown on the dashboard.
# Fields support Python format-string interpolation against KPI parameters
# (e.g. "{TIER_MIN_STUDENTS}" expands to the current configured value).
# Keep descriptions concise — one or two sentences per field.
ALGO_DESCRIPTIONS = {
    # ── Headline KPI cards ──────────────────────────────────────────
    "cumulative_students": {
        "source":      "CCCCO MAP CPL Dashboard (live scrape via Cloudflare worker).",
        "formula":     "Direct value pull of the STUDENTS SERVED metric.",
        "assumptions": "Breakdowns (Military / Workforce-Other / Apprentice) are reported by CCCCO directly; not recomputed.",
        "caveats":     "Falls back to the Excel KPI metric on project 3.1 if the live scrape is unavailable.",
        "last_modified": "2026-04-19",
    },
    "eligible_units": {
        "source":      "CCCCO MAP CPL Dashboard (live scrape).",
        "formula":     "Direct value pull of the ELIGIBLE UNITS metric.",
        "assumptions": "Population breakdowns mirror the CCCCO dashboard.",
        "caveats":     "Falls back to the Excel KPI metric on project 3.2 if the live scrape is unavailable.",
        "last_modified": "2026-04-19",
    },
    "transcripted_units": {
        "source":      "CCCCO MAP CPL Dashboard (live scrape).",
        "formula":     "Direct value pull of the TRANSCRIBED UNITS metric. 'Applied Credits' is a secondary number posted to student records.",
        "assumptions": "Transcribed = credits actually posted to a student record (distinct from 'eligible').",
        "caveats":     "Population breakdowns are reported directly by CCCCO.",
        "last_modified": "2026-04-19",
    },
    "credit_recommendations": {
        "source":      "MAP Custom Reporting Module (CustomReport_latest.json → View_ArticulatedMAPExhibits).",
        "formula":     "Count of articulation rows across all colleges (each row = one college-to-course articulation).",
        "assumptions": "Each row represents a distinct (Exhibit, College, Course) mapping.",
        "caveats":     "Known issue: the raw export contains some duplicates, so the CR count can exceed the implied product of exhibits × colleges. The underlying MAP Articulation Analysis section also inherits these duplicates.",
        "last_modified": "2026-04-19",
    },
    "map_exhibits": {
        "source":      "MAP Custom Reporting Module (View_ArticulatedMAPExhibits).",
        "formula":     "Distinct count of ExhibitID values across the dataset.",
        "assumptions": "ExhibitID is the primary key for an exhibit (same exhibit articulated at multiple colleges counts once).",
        "caveats":     "'Originating colleges' (62) counts colleges that created exhibits, not colleges that adopted them.",
        "last_modified": "2026-04-19",
    },
    "ccc_collaborative": {
        "source":      "MAP Custom Reporting Module (View_ArticulatedMAPExhibits).",
        "formula":     "Distinct count of colleges that have adopted at least one exhibit whose Collaborative Type contains 'CCC'.",
        "assumptions": "Adoption = a row exists where the college articulates a CCC Collaborative exhibit to a local course.",
        "caveats":     "Counts adopting colleges only, not colleges that created collaborative exhibits.",
        "last_modified": "2026-04-19",
    },
    "active_colleges": {
        "source":      "CCCCO MAP CPL Dashboard (tier classification computed by the Cloudflare worker).",
        "formula":     "Leading + Advancing tiers. A college is Leading if it meets ≥3 of 5 criteria, Advancing if 1-2, Inactive otherwise.",
        "assumptions": "Criteria: Students ≥ {TIER_MIN_STUDENTS}; Units ≥ {TIER_MIN_UNITS}; Avg Units/Student ≥ {TIER_MIN_AVG_UNITS_PER_STUDENT}; Transcription Rate ≥ {TIER_MIN_TRANSCRIPTION_RATE_PCT}; Avg Transcribed/Student ≥ {TIER_MIN_AVG_TRANSCRIBED_PER_STUDENT}.",
        "caveats":     "Thresholds live in cloudflare-worker-proxy.js (lines 184-188). Editing the KPI_Config sheet updates this description but does NOT automatically redeploy the worker.",
        "last_modified": "2026-04-19",
    },
    "articulation_colleges": {
        "source":      "MAP Custom Reporting Module (View_ArticulatedMAPExhibits).",
        "formula":     "Distinct count of colleges appearing in the 'Articulation College' field.",
        "assumptions": "A college is 'articulating' if it appears as the receiving college on at least one exhibit row.",
        "caveats":     "'Originating Colleges' (62) = colleges that created exhibits; 'Adopting CCC Collaborative' (55) = subset that adopted statewide exhibits.",
        "last_modified": "2026-04-19",
    },
    "estimated_savings": {
        "source":      "CCCCO MAP CPL Dashboard (live scrape). Fallback: {SAVINGS_DEFAULT} from KPI_Config sheet.",
        "formula":     "Live value pull. CCCCO computes savings using the Beacon Economics per-unit factor (≈ ${BEACON_PER_UNIT_SAVINGS}/unit).",
        "assumptions": "Per-population breakdowns (Military/Workforce/Apprentice) are proportional splits by eligible-unit share, not independent calculations.",
        "caveats":     "BEACON_PER_UNIT_SAVINGS is a reference value; the actual savings math happens at CCCCO.",
        "last_modified": "2026-04-19",
    },
    "veteran_sprint": {
        "source":      "Live CCCCO MAP CPL Dashboard scrape (star_college_count). Augmented with live military data from the same scrape.",
        "formula":     "Headline = live star_college_count from the CCCCO scrape. JST Credits = military students (from cumulative_students breakdowns). Eligible CPL = military eligible units (from eligible_units breakdowns).",
        "assumptions": "Goal for JST Credits = 30,000 at Veteran Star Colleges.",
        "caveats":     "Headline value is the number of colleges participating, not students served. Pulled live from the CCCCO MAP CPL Dashboard each pipeline run.",
        "last_modified": "2026-04-19",
    },
    "twenty_year_impact": {
        "source":      "CCCCO MAP CPL Dashboard (live scrape).",
        "formula":     "Direct pull of the 20-YEAR IMPACT metric. CCCCO applies a {TWENTY_YEAR_MULTIPLIER}x multiplier to one-year impact.",
        "assumptions": "Beacon Economics long-term earnings model; breakdowns by population are proportional splits.",
        "caveats":     "The multiplier lives on the CCCCO side — changing TWENTY_YEAR_MULTIPLIER here is for documentation only.",
        "last_modified": "2026-04-19",
    },

    # ── KPI Trends card ────────────────────────────────────────────
    "kpi_trends": {
        "source":      "kpi_history.json — one snapshot appended per day (idempotent; same-day runs overwrite).",
        "formula":     "For each metric, render: today's value, deltas at 1d/7d/30d/QTD/1yr, and a 30-day sparkline.",
        "assumptions": "QTD resets at the start of each academic quarter (Jul/Oct/Jan/Apr). Deltas are % for volumetric metrics and absolute counts for Active Colleges, MAP Exhibits, CCC Collaborative, Articulating Colleges, and Star Colleges.",
        "caveats":     "A dash (—) means no historical value exists yet for that period; tracking begins the first time the pipeline ran. Sparklines need 2+ snapshots to render.",
        "last_modified": "2026-04-19",
    },

    # ── College Activity card ─────────────────────────────────────
    "college_activity": {
        "source":      "CCCCO MAP CPL Dashboard via Cloudflare worker — per-college tier classifications.",
        "formula":     "Same 3-of-5 criteria as Active Colleges. Leading (≥3 met), Advancing (1-2 met), Inactive (<10 students and 0 units). Tiers sorted by criteria met desc, then by students desc.",
        "assumptions": "Criteria: Students ≥ {TIER_MIN_STUDENTS}; Units ≥ {TIER_MIN_UNITS}; Avg Units/Student ≥ {TIER_MIN_AVG_UNITS_PER_STUDENT}; Transcription Rate ≥ {TIER_MIN_TRANSCRIPTION_RATE_PCT}; Avg Transcribed/Student ≥ {TIER_MIN_AVG_TRANSCRIBED_PER_STUDENT}.",
        "caveats":     "⭐ indicates MAP Star College status (at least one criterion met). JST column derives from View_StudentAggregatedValues rows where Military Credits > 0.",
        "last_modified": "2026-04-19",
    },

    # ── MAP Articulation Analysis detail cards ────────────────────
    "exhibit-collaborative": {
        "source":      "View_ArticulatedMAPExhibits (MAP Custom Reporting Module).",
        "formula":     "Group rows by Collaborative Type. Count distinct exhibits, rows (credit recs), colleges, and disciplines per group.",
        "assumptions": "Categories: Local (college-specific), CCC Collaborative (statewide faculty workgroup), Industry/Other (anything else).",
        "caveats":     "An exhibit has exactly one Collaborative Type. '%' is each group's share of total credit recs.",
        "last_modified": "2026-04-19",
    },
    "exhibit-by-cpl-type": {
        "source":      "View_ArticulatedMAPExhibits.",
        "formula":     "Group by CPL Type. Count distinct exhibits, rows (credit recs), and colleges per type.",
        "assumptions": "Six CPL types: Standardized Assessment, Industry Certification, Credit By Exam, Portfolio Review, Military, Other.",
        "caveats":     "Each exhibit belongs to exactly one CPL type.",
        "last_modified": "2026-04-19",
    },
    "exhibit-by-mol": {
        "source":      "View_ArticulatedMAPExhibits.",
        "formula":     "Group by Mode of Learning code (S/I/A/H/N/M/O/G/P/J). Count exhibits, rows, colleges.",
        "assumptions": "Mode taxonomy defined by CCCCO; each exhibit carries one MoL code.",
        "caveats":     "Codes with very low counts (e.g. Justice-involved learning) reflect limited adoption, not missing data.",
        "last_modified": "2026-04-19",
    },
    "exhibit-by-discipline": {
        "source":      "View_ArticulatedMAPExhibits.",
        "formula":     "Group by CCC Discipline. Count exhibits, rows (credit recs), courses, colleges, and CCC Collaborative adoptions.",
        "assumptions": "'Not Mapped' = exhibits without a discipline tag.",
        "caveats":     "~31% of credit recs are Not Mapped — this flags a tagging backlog, not missing articulations.",
        "last_modified": "2026-04-19",
    },
    "exhibit-by-college": {
        "source":      "View_ArticulatedMAPExhibits.",
        "formula":     "Group rows by Articulation College. Count credit recs (rows), distinct exhibits, disciplines, CCC Collaborative rows, and Industry Certs.",
        "assumptions": "Each row represents one college-to-course articulation of one exhibit.",
        "caveats":     "Credit Recs > Exhibits is expected: one exhibit can articulate to multiple courses at the same college. Known inflation from duplicates in the raw export.",
        "last_modified": "2026-04-19",
    },
    "exhibit-top-50": {
        "source":      "View_ArticulatedMAPExhibits.",
        "formula":     "Group rows by ExhibitID, sum credit recommendations across all colleges, sort desc, take top 50.",
        "assumptions": "Ranking = total articulation count statewide.",
        "caveats":     "Ties broken by exhibit ID order. Counts reflect raw row counts and can include the known duplicate inflation.",
        "last_modified": "2026-04-19",
    },

    "articulations-by-course": {
        "source":      "kb/coci_articulations.json — earned MAP articulations (View_ArticulatedMAPExhibits) resolved to the unified course-identity layer.",
        "formula":     "Group earned articulations by unified course identity (C-ID/CCN when the row carries a CID Number, else a minted M-ID; multi-M-ID rows disambiguated by local Course Title). Per identity: count colleges that earned it, the modal credit recommendation, linked credential(s), and adoption leverage = peer colleges teaching the same identity that have not yet earned the articulation.",
        "assumptions": "One identity = one common course across colleges. Adoption leverage is a candidate list, not a guarantee — title-consistent membership only.",
        "caveats":     "GE-area articulations are excluded (routed separately). Adoption leverage is suppressed (shown as flagged) for identities marked over_merged, since the cluster may conflate distinct courses and would yield bogus adoption targets.",
        "last_modified": "2026-05-22",
    },

    # ── Exhibit Adoption & Credit Recommendations (statewide_interactive.js) ──
    "statewide_adoption": {
        "source":      "View_ArticulatedMAPExhibits joined with college/district/region lookups from college_lookup.js.",
        "formula":     "For each statewide (CCC Collaborative) exhibit: count adopting colleges and list potential adopters (colleges that could adopt but haven't yet).",
        "assumptions": "Potential adopters = colleges in the CCC system not currently articulating this exhibit. Credit recs count each college-course pair separately.",
        "caveats":     "Interactive filters (CPL Type, Discipline, District, SW Region) narrow results client-side. Exports reflect current filter state.",
        "last_modified": "2026-04-19",
    },
}


def render_algo_details(card_id, params=None, container_style=None):
    """Return a collapsible <details> block with the algorithm description for a card.
    Returns empty string if no description is defined. Always collapsed by default.

    params: dict of KPI parameter values used for {PLACEHOLDER} interpolation.
            Defaults to KPI_PARAMETERS_DEFAULTS.
    container_style: optional inline CSS for the outer <details> element (use to
                     match the parent card's background if it sits outside a card).
    """
    desc = ALGO_DESCRIPTIONS.get(card_id)
    if not desc:
        return ""

    p = dict(KPI_PARAMETERS_DEFAULTS)
    if params:
        p.update({k: v for k, v in params.items() if not k.startswith("_")})
    # Derived format helpers
    try:
        p["TIER_MIN_TRANSCRIPTION_RATE_PCT"] = f"{float(p['TIER_MIN_TRANSCRIPTION_RATE']) * 100:.0f}%"
    except (ValueError, TypeError, KeyError):
        p["TIER_MIN_TRANSCRIPTION_RATE_PCT"] = "25%"

    def _fmt(s):
        if not isinstance(s, str):
            return s
        try:
            return s.format(**p)
        except (KeyError, IndexError, ValueError):
            return s

    rows = []
    for label, key in (("Source", "source"), ("Formula", "formula"),
                       ("Assumptions", "assumptions"), ("Caveats", "caveats")):
        v = desc.get(key)
        if v:
            rows.append(
                f'<div class="algo-row">'
                f'<span class="algo-label">{label}:</span> '
                f'<span class="algo-value">{_fmt(v)}</span>'
                f'</div>'
            )

    # Use sheet-provided last_modified if present, else fall back to desc
    lm_sheet = (params or {}).get("_last_modified", {}).get(card_id) if params else None
    lm = lm_sheet or desc.get("last_modified", "")
    meta = f'<div class="algo-meta">Description last updated: {lm}</div>' if lm else ""

    style_attr = f' style="{container_style}"' if container_style else ""
    return (
        f'<details class="algo-details"{style_attr}>'
        f'<summary>How this is calculated</summary>'
        f'<div class="algo-body">'
        + "".join(rows)
        + meta
        + '</div></details>'
    )


ALGO_DETAILS_CSS = """
/* ═══ Collapsible Algorithm Descriptions ═══ */
.algo-details {
    margin-top: 0.6rem;
    padding-top: 0.4rem;
    border-top: 1px dashed rgba(255,255,255,0.12);
    font-size: 0.72rem;
}
.algo-details summary {
    cursor: pointer;
    color: rgba(255,255,255,0.45);
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    font-size: 0.6rem;
    padding: 0.3rem 0;
    user-select: none;
    list-style: none;
    display: flex;
    align-items: center;
    gap: 0.35rem;
}
.algo-details summary::-webkit-details-marker { display: none; }
.algo-details summary::before {
    content: "▸";
    display: inline-block;
    font-size: 0.65rem;
    transition: transform 0.15s ease;
    color: rgba(201,168,76,0.7);
}
.algo-details[open] summary::before { transform: rotate(90deg); }
.algo-details summary:hover { color: rgba(201,168,76,0.9); }
.algo-details[open] summary { color: #C9A84C; }
.algo-details .algo-body {
    padding: 0.55rem 0 0.3rem;
    color: rgba(255,255,255,0.72);
    line-height: 1.55;
}
.algo-details .algo-row { margin-bottom: 0.35rem; }
.algo-details .algo-label {
    font-weight: 700;
    color: rgba(201,168,76,0.9);
    margin-right: 0.3rem;
    text-transform: uppercase;
    font-size: 0.62rem;
    letter-spacing: 0.4px;
}
.algo-details .algo-value { color: rgba(255,255,255,0.78); }
.algo-details .algo-meta {
    margin-top: 0.6rem;
    font-size: 0.62rem;
    color: rgba(255,255,255,0.35);
    font-style: italic;
}
"""


def scan_attachments(attachments_dir):
    """Scan a local attachments directory with subfolders and return counts.
    Expected structure:
      Attachments/
        Activity 1/         ← activity-level attachments
        Activity 2/
        1.1 MAP Platform Development/  ← project-level attachments
        2.1 Credit Recommendations/
        (root-level files count toward total only)
    Returns dict:
      {"total": N, "root_files": N,
       "by_activity": {"1": N, "2": N, ...},
       "by_project": {"1.1": N, "2.1": N, ...}}
    """
    import re as _re
    result = {"total": 0, "root_files": 0, "by_activity": {}, "by_project": {}}
    if not attachments_dir or not os.path.isdir(attachments_dir):
        return result

    def _count_files(folder):
        """Count non-hidden files in a directory (non-recursive)."""
        if not os.path.isdir(folder):
            return 0
        return sum(1 for f in os.listdir(folder)
                   if os.path.isfile(os.path.join(folder, f)) and not f.startswith('.'))

    for entry in os.listdir(attachments_dir):
        fpath = os.path.join(attachments_dir, entry)
        if entry.startswith('.'):
            continue
        if os.path.isfile(fpath):
            result["root_files"] += 1
            result["total"] += 1
        elif os.path.isdir(fpath):
            count = _count_files(fpath)
            result["total"] += count
            # Match "Activity N" folders
            m_act = _re.match(r'^Activity\s+(\d+)', entry, _re.IGNORECASE)
            if m_act:
                result["by_activity"][m_act.group(1)] = count
                continue
            # Match "N.N Project Name" folders
            m_proj = _re.match(r'^(\d+\.\d+)', entry)
            if m_proj:
                pid = m_proj.group(1)
                result["by_project"][pid] = count
                # Also roll up to activity
                act_num = pid.split(".")[0]
                result["by_activity"][act_num] = result["by_activity"].get(act_num, 0) + count
    return result


def ensure_attachment_subfolders(attachments_dir, projects):
    """Create subfolders for each Activity and Project if they don't exist.
    Called during pipeline run to auto-create folders for new projects.
    """
    if not attachments_dir or not os.path.isdir(attachments_dir):
        return 0
    created = 0
    activities_seen = set()
    for p in projects:
        pid = p.get("id", "")
        name = p.get("name", "")
        if not pid:
            continue
        act_num = pid.split(".")[0]
        # Create Activity folder
        if act_num not in activities_seen:
            activities_seen.add(act_num)
            act_folder = os.path.join(attachments_dir, f"Activity {act_num}")
            if not os.path.exists(act_folder):
                os.makedirs(act_folder, exist_ok=True)
                created += 1
        # Create project folder (skip sub-population rows like 3.1.1)
        if pid.count(".") == 1 and name:
            # Sanitize folder name
            safe_name = "".join(c for c in name if c.isalnum() or c in " -_&().").strip()
            proj_folder = os.path.join(attachments_dir, f"{pid} {safe_name}")
            if not os.path.exists(proj_folder):
                os.makedirs(proj_folder, exist_ok=True)
                created += 1
    return created


def read_project_config(wb):
    """Read project configuration from rows 1-2 of the Project List tab.
    Layout:
      A1='Project Title:', B1=title, E1='Project ID:', F1=id,
      H1='Attachments URL:', I1=url
      A2='Description:', B2=description text
    Returns dict with keys: title, project_id, description, attachments_url
    """
    ws = wb["Project List"]
    config = {
        "title": str(ws.cell(1, 2).value or "CPL Initiative").strip(),
        "project_id": str(ws.cell(1, 6).value or "").strip(),
        "attachments_url": str(ws.cell(1, 9).value or "").strip(),
        "description": str(ws.cell(2, 2).value or "").strip(),
    }
    return config


# ── Official CPL Workplan Goal Definitions ─────────────────────────
# Source: CCCCO Vision 2030 Credit for Prior Learning Workplan
CPL_GOALS = {
    "Goal 1": {
        "title": "Expand Equitable CPL Access & Boost Student Success",
        "target": "By 2030, expand CPL opportunities for at least 250,000 Californians, including 220,000 working adults and apprentices, and 30,000 veterans.",
        "bullets": [
            "Offer or award CPL using 1,000 statewide credit recommendations from 25 faculty/industry workgroups",
            "80% of CPL recipients report improvements in savings, belonging, well-being, completion, and career prospects",
            "500 personal success stories documented in MAP",
        ],
        "stretch": "Stretch: Serve 500,000 learners (450K adults + 50K veterans); 3,000 credit recommendations from 40 workgroups; 90% reporting improvements; 1,000 success stories",
    },
    "Goal 2": {
        "title": "Build a Unified, Interoperable, Student-Centered CPL System",
        "target": "By 2030, build a unified statewide CPL system with CPL-embedded outreach pathways, interoperable data, and seamless student experiences.",
        "bullets": [
            "MAP platform serving all 116 colleges with AI-enhanced tools and credential matching",
            "Full interoperability with CCCApply, eTranscript, MIS, and Credential Engine",
            "Statewide CPL data infrastructure tracking offers, awards, transcription, and student outcomes",
        ],
        "stretch": "",
    },
    "Goal 3": {
        "title": "Sustainable Policies, Resources & Professional Learning",
        "target": "By 2030, establish clear CPL policies, sustainable resources, professional development, and best practices across all colleges.",
        "bullets": [
            "Permanent CPL operational funding secured ($5\u20137.5M ongoing)",
            "Title 5 \u00a755050 regulatory updates reducing residency and access barriers",
            "1,000+ staff, faculty, and administrators trained in CPL implementation",
            "Scalable technical assistance model for all 116 colleges",
        ],
        "stretch": "",
    },
}


def fmt_date(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val) if val else ""


def fmt_number(val):
    """Format a number with commas for display (e.g. 42620 -> '42,620').
    Handles both numeric types and string representations of numbers.
    """
    if val is None or val == "":
        return ""
    # If it's already numeric, format directly
    if isinstance(val, (int, float)):
        if float(val) == int(val):
            return f"{int(val):,}"
        return f"{val:,.1f}"
    # Try to parse string as number (e.g. '42620' -> 42,620)
    s = str(val).strip().replace(",", "")
    try:
        n = float(s)
        if n == int(n):
            return f"{int(n):,}"
        return f"{n:,.1f}"
    except ValueError:
        return str(val)


def fmt_dollars(val):
    """Format a numeric value as dollars with smart scaling.
    - Millions: $X.XM (e.g., 6000000 -> $6.0M)
    - Thousands: $XXK (e.g., 50000 -> $50K)
    - Exact: $XXX (e.g., 999 -> $999)
    Handles None and non-numeric gracefully (returns $0).
    """
    if val is None or val == "":
        return "$0"
    # Convert to float
    try:
        if isinstance(val, str):
            num = float(val.replace(",", "").replace("$", "").strip())
        else:
            num = float(val)
    except (ValueError, AttributeError):
        return "$0"

    if num == 0:
        return "$0"

    # Determine scaling
    abs_num = abs(num)
    if abs_num >= 1_000_000:
        # Millions
        return f"${num / 1_000_000:.1f}M".rstrip('0').rstrip('.')
    elif abs_num >= 1_000:
        # Thousands
        return f"${num / 1_000:.0f}K"
    else:
        # Exact dollars
        return f"${num:,.0f}"


def substitute_kpi(text, kpi_value):
    """Replace {KPI} tokens in text with the formatted KPI metric value."""
    if not text or not kpi_value:
        return text or ""
    formatted = fmt_number(kpi_value)  # handles both numeric and string values
    return str(text).replace("{KPI}", formatted)


def cell_val(ws, row, col, default=""):
    v = ws.cell(row=row, column=col).value
    return v if v is not None else default


def archive_updates_to_log(excel_path):
    """Auto-archive: copy current Project List notes (col P/Q) into the Update Log tab.
    Only adds a row if the project ID + date + note combo doesn't already exist.
    This runs before the read-only passes so the log is always up to date.
    """
    from openpyxl import load_workbook as _lwb
    try:
        wb = _lwb(excel_path)
    except Exception as e:
        print(f"  (Skipping update archive — cannot open Excel for writing: {e})")
        return

    # Ensure Update Log tab exists with headers
    if "Update Log" not in wb.sheetnames:
        ws_log = wb.create_sheet("Update Log")
    else:
        ws_log = wb["Update Log"]

    # Set up headers if missing or wrong
    # Col 4 = Type: "update" or "workplan" to distinguish note sources
    if ws_log.cell(1, 1).value != "Project ID" or ws_log.cell(1, 4).value != "Type":
        ws_log.cell(1, 1, "Project ID")
        ws_log.cell(1, 2, "Date")
        ws_log.cell(1, 3, "Note")
        ws_log.cell(1, 4, "Type")
        # Bold headers
        from openpyxl.styles import Font
        bold = Font(bold=True)
        for c in range(1, 5):
            ws_log.cell(1, c).font = bold

    # Read existing log entries to avoid duplicates
    # Also track the most recent note text per project per type for auto-stamp detection
    existing = set()
    latest_by_pid = {}  # pid -> {"update": (date, note), "workplan": (date, note)}
    for r in range(2, ws_log.max_row + 1):
        eid = str(ws_log.cell(r, 1).value or "").strip()
        edate = ws_log.cell(r, 2).value
        enote = str(ws_log.cell(r, 3).value or "").strip()
        etype = str(ws_log.cell(r, 4).value or "update").strip()
        if isinstance(edate, datetime):
            edate = edate.strftime("%Y-%m-%d")
        else:
            edate = str(edate or "").strip()
        existing.add((eid, edate, enote, etype))
        # Track latest note per project per type
        if eid not in latest_by_pid:
            latest_by_pid[eid] = {}
        prev = latest_by_pid[eid].get(etype)
        if prev is None or edate >= prev[0]:
            latest_by_pid[eid][etype] = (edate, enote)

    # Read Project List and archive notes from both col P (update) and col V (workplan)
    # Auto-stamp col Q if either P or V changed since last log entry
    ws_pl = wb["Project List"]
    today_str = _now_pt().strftime("%Y-%m-%d")
    archived = 0
    auto_stamped = 0
    for r in range(4, ws_pl.max_row + 1):
        pid = cell_val(ws_pl, r, COL_ID, "")
        if not pid:
            continue
        pid = str(pid).strip()
        update_note = str(cell_val(ws_pl, r, COL_UPDATE, "")).strip()
        wp_note = str(cell_val(ws_pl, r, COL_WP_NOTES, "")).strip()

        if not update_note and not wp_note:
            continue

        # Check if either note changed → auto-stamp column Q
        pid_latest = latest_by_pid.get(pid, {})
        prev_update = pid_latest.get("update")
        prev_wp = pid_latest.get("workplan")
        changed = False
        if update_note and (prev_update is None or update_note != prev_update[1]):
            changed = True
        if wp_note and (prev_wp is None or wp_note != prev_wp[1]):
            changed = True

        if changed:
            ws_pl.cell(row=r, column=COL_UPDATE_DATE, value=today_str)
            auto_stamped += 1
            date_str = today_str
        else:
            date_val = cell_val(ws_pl, r, COL_UPDATE_DATE, "")
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val or "").strip()
            if not date_str:
                ws_pl.cell(row=r, column=COL_UPDATE_DATE, value=today_str)
                auto_stamped += 1
                date_str = today_str

        # Archive update note (col P)
        if update_note:
            key = (pid, date_str, update_note, "update")
            if key not in existing:
                next_row = ws_log.max_row + 1
                ws_log.cell(next_row, 1, pid)
                ws_log.cell(next_row, 2, date_str)
                ws_log.cell(next_row, 3, update_note)
                ws_log.cell(next_row, 4, "update")
                existing.add(key)
                archived += 1

        # Archive workplan note (col V)
        if wp_note:
            key = (pid, date_str, wp_note, "workplan")
            if key not in existing:
                next_row = ws_log.max_row + 1
                ws_log.cell(next_row, 1, pid)
                ws_log.cell(next_row, 2, date_str)
                ws_log.cell(next_row, 3, wp_note)
                ws_log.cell(next_row, 4, "workplan")
                existing.add(key)
                archived += 1

    needs_save = archived > 0 or auto_stamped > 0
    if needs_save:
        try:
            wb.save(excel_path)
            if archived > 0:
                print(f"  Archived {archived} new note(s) to Update Log")
            if auto_stamped > 0:
                print(f"  Auto-stamped {auto_stamped} date(s) in column Q")
        except PermissionError:
            print(f"  (Skipping update archive — Excel file is locked)")
    else:
        print(f"  Update Log up to date (no new notes)")
    wb.close()


def read_update_log(wb):
    """Read the Update Log tab and return a dict of project_id -> list of
    {date, note, type}, sorted from newest to oldest.
    Type is 'update' (col P) or 'workplan' (col V)."""
    log = {}
    if "Update Log" not in wb.sheetnames:
        return log

    ws = wb["Update Log"]
    for r in range(2, ws.max_row + 1):
        pid = str(ws.cell(r, 1).value or "").strip()
        if not pid:
            continue
        date_val = ws.cell(r, 2).value
        note = str(ws.cell(r, 3).value or "").strip()
        ntype = str(ws.cell(r, 4).value or "update").strip()
        if not note:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val or "").strip()
        log.setdefault(pid, []).append({"date": date_str, "note": note, "type": ntype})

    # Sort each project's notes newest-first
    for pid in log:
        log[pid].sort(key=lambda x: x["date"], reverse=True)

    return log


def read_projects(wb):
    ws = wb["Project List"]
    projects = []

    for r in range(4, ws.max_row + 1):
        pid = cell_val(ws, r, COL_ID, None)
        if not pid:
            break

        # ── percentage ──
        pct_raw = cell_val(ws, r, COL_PCT, 0)
        if isinstance(pct_raw, (int, float)):
            pct = int(round(pct_raw * 100)) if pct_raw <= 1 else int(pct_raw)
        else:
            pct = 0

        # ── KPI columns ──
        kpi_metric = cell_val(ws, r, COL_KPI_METRIC, None)
        kpi_unit   = cell_val(ws, r, COL_KPI_UNIT, "")
        kpi_g2526  = cell_val(ws, r, COL_KPI_G2526, "")
        kpi_s2526  = cell_val(ws, r, COL_KPI_S2526, "")
        kpi_g2627  = cell_val(ws, r, COL_KPI_G2627, "")
        kpi_s2627  = cell_val(ws, r, COL_KPI_S2627, "")
        kpi_g2728  = cell_val(ws, r, COL_KPI_G2728, "")
        kpi_s2728  = cell_val(ws, r, COL_KPI_S2728, "")
        kpi_g2829  = cell_val(ws, r, COL_KPI_G2829, "")
        kpi_s2829  = cell_val(ws, r, COL_KPI_S2829, "")
        kpi_g2930  = cell_val(ws, r, COL_KPI_G2930, "")
        kpi_s2930  = cell_val(ws, r, COL_KPI_S2930, "")
        wp_notes   = cell_val(ws, r, COL_WP_NOTES, "")
        kpi_order  = cell_val(ws, r, COL_KPI_ORDER, None)
        override   = cell_val(ws, r, COL_OVERRIDE, None)

        # ── raw text fields (before substitution) ──
        desc_raw       = str(cell_val(ws, r, COL_DESC, ""))
        milestones_raw = str(cell_val(ws, r, COL_MILESTONES, ""))
        update_raw     = str(cell_val(ws, r, COL_UPDATE, ""))

        # ── {KPI} substitution ──
        desc       = substitute_kpi(desc_raw, kpi_metric)
        milestones = substitute_kpi(milestones_raw, kpi_metric)
        update     = substitute_kpi(update_raw, kpi_metric)

        projects.append({
            "id":            str(pid),
            "name":          cell_val(ws, r, COL_NAME, ""),
            "desc":          desc,
            "activity":      cell_val(ws, r, COL_ACTIVITY, ""),
            "v2030":         cell_val(ws, r, COL_V2030, ""),
            "goal":          cell_val(ws, r, COL_GOAL, ""),
            "budget_source": cell_val(ws, r, COL_BUDGET_SRC, ""),
            "budget":        cell_val(ws, r, COL_BUDGET, ""),
            "lead":          cell_val(ws, r, COL_LEAD, ""),
            "team":          cell_val(ws, r, COL_TEAM, ""),
            "status":        cell_val(ws, r, COL_STATUS, ""),
            "pct":           pct,
            "start":         fmt_date(cell_val(ws, r, COL_START, "")),
            "end":           fmt_date(cell_val(ws, r, COL_END, "")),
            "milestones":    milestones,
            "update":        update,
            "update_date":   fmt_date(cell_val(ws, r, COL_UPDATE_DATE, "")),
            "kpi_metric":    fmt_number(kpi_metric) if kpi_metric is not None else "",
            "kpi_unit":      str(kpi_unit),
            "kpi_goal_2526": fmt_number(kpi_g2526) if isinstance(kpi_g2526, (int, float)) else str(kpi_g2526) if kpi_g2526 else "",
            "kpi_stretch_2526": fmt_number(kpi_s2526) if isinstance(kpi_s2526, (int, float)) else str(kpi_s2526) if kpi_s2526 else "",
            "kpi_goal_2627": fmt_number(kpi_g2627) if isinstance(kpi_g2627, (int, float)) else str(kpi_g2627) if kpi_g2627 else "",
            "kpi_stretch_2627": fmt_number(kpi_s2627) if isinstance(kpi_s2627, (int, float)) else str(kpi_s2627) if kpi_s2627 else "",
            "kpi_goal_2728": fmt_number(kpi_g2728) if isinstance(kpi_g2728, (int, float)) else str(kpi_g2728) if kpi_g2728 else "",
            "kpi_stretch_2728": fmt_number(kpi_s2728) if isinstance(kpi_s2728, (int, float)) else str(kpi_s2728) if kpi_s2728 else "",
            "kpi_goal_2829": fmt_number(kpi_g2829) if isinstance(kpi_g2829, (int, float)) else str(kpi_g2829) if kpi_g2829 else "",
            "kpi_stretch_2829": fmt_number(kpi_s2829) if isinstance(kpi_s2829, (int, float)) else str(kpi_s2829) if kpi_s2829 else "",
            "kpi_goal_2930": fmt_number(kpi_g2930) if isinstance(kpi_g2930, (int, float)) else str(kpi_g2930) if kpi_g2930 else "",
            "kpi_stretch_2930": fmt_number(kpi_s2930) if isinstance(kpi_s2930, (int, float)) else str(kpi_s2930) if kpi_s2930 else "",
            "workplan_notes": substitute_kpi(str(wp_notes), kpi_metric),
            "kpi_order": int(kpi_order) if isinstance(kpi_order, (int, float)) and kpi_order else None,
            "override": override,
            "excel_row": r,
        })

    return projects


def read_config_overrides(wb):
    """
    Read dashboard configuration overrides from the Project List tab.
    Scans rows BELOW the last project (after the first empty ID cell)
    looking for config key-value pairs in Col A (key) and Col AG (value).

    Returns a dict of config_key → value.

    Supported config keys (case-insensitive):
      BASE_MIL, BASE_WF, BASE_APP         — workplan chart 2024 baselines
      ACTUAL_2025_MIL, ACTUAL_2025_WF,     — workplan chart 2025 actuals
      ACTUAL_2025_APP, ACTUAL_2025_TOTAL
      V2030_G1_PROGRESS, V2030_G2_PROGRESS, V2030_G3_PROGRESS  — Vision 2030 %
      V2030_G2_CURRENT, V2030_G3_CURRENT   — Vision 2030 status text
      SAVINGS_DEFAULT                       — estimated savings fallback
    """
    ws = wb["Project List"]
    overrides = {}

    # Skip past project rows (they end at first empty Col A after row 3)
    config_start = None
    for r in range(4, ws.max_row + 1):
        pid = cell_val(ws, r, COL_ID, None)
        if not pid:
            config_start = r
            break

    if config_start is None:
        return overrides

    # Scan remaining rows for config entries (skip section headers)
    _skip = {"DASHBOARD CONFIGURATION OVERRIDES", "VALUE"}
    for r in range(config_start, min(config_start + 50, ws.max_row + 1)):
        key = cell_val(ws, r, COL_ID, None)
        val = cell_val(ws, r, COL_OVERRIDE, None)
        if key and val is not None:
            k = str(key).strip().upper()
            if k not in _skip and not k.startswith("──"):
                overrides[k] = val

    if overrides:
        print(f"  Config overrides from Col AG: {list(overrides.keys())}")

    return overrides


def _override_int(proj_map, pid):
    """Return the Col AG override for a project as an int, or None if not set."""
    p = proj_map.get(pid)
    if p and p.get("override") is not None:
        try:
            return int(float(p["override"]))
        except (ValueError, TypeError):
            pass
    return None


def read_annual_goals(wb):
    """
    Read the 'Annual Workplan Goals' tab.
    Returns a list of dicts, one per sub-activity, each with:
      id, name, activity_label, goal/stretch/current for each year, and totals.
    Activity header rows (no ID) are kept as group separators.
    """
    if "Annual Workplan Goals" not in wb.sheetnames:
        print("  WARNING: 'Annual Workplan Goals' tab not found — skipping")
        return []

    ws = wb["Annual Workplan Goals"]
    year_cols = ["2025-26", "2026-27", "2027-28", "2028-29", "2029-30"]

    rows = []
    current_activity = ""
    r = 2  # skip header row
    while r <= ws.max_row:
        pid = cell_val(ws, r, 1, "")
        name = cell_val(ws, r, 2, "")
        row_type = str(cell_val(ws, r, 3, "")).strip().upper()

        # Activity header row (no ID, no row type)
        if not pid and not row_type and name:
            current_activity = str(name)
            r += 1
            continue

        if not pid or row_type != "GOAL":
            r += 1
            continue

        # Read GOAL row
        goal_vals = {}
        for ci, yr in enumerate(year_cols, start=4):
            v = cell_val(ws, r, ci, 0)
            goal_vals[yr] = float(v) if isinstance(v, (int, float)) else 0
        goal_vals["total"] = float(cell_val(ws, r, 9, 0) or 0)

        # Next row should be CURRENT
        current_vals = {}
        if r + 1 <= ws.max_row and str(cell_val(ws, r + 1, 3, "")).strip().upper() == "CURRENT":
            for ci, yr in enumerate(year_cols, start=4):
                v = cell_val(ws, r + 1, ci, 0)
                current_vals[yr] = float(v) if isinstance(v, (int, float)) else 0
            current_vals["total"] = float(cell_val(ws, r + 1, 9, 0) or 0)

        # Next row should be STRETCH
        stretch_vals = {}
        if r + 2 <= ws.max_row and str(cell_val(ws, r + 2, 3, "")).strip().upper() == "STRETCH":
            for ci, yr in enumerate(year_cols, start=4):
                v = cell_val(ws, r + 2, ci, 0)
                stretch_vals[yr] = float(v) if isinstance(v, (int, float)) else 0
            stretch_vals["total"] = float(cell_val(ws, r + 2, 9, 0) or 0)

        rows.append({
            "id": str(pid),
            "name": str(name),
            "activity": current_activity,
            "goal": goal_vals,
            "current": current_vals,
            "stretch": stretch_vals,
        })
        r += 3  # skip GOAL/CURRENT/STRETCH triplet

    return rows


def populate_current_metrics(annual_goals, projects):
    """
    Fill in the 'current' 2025-26 column for each annual goal row
    using actual metrics from the Project List and live data.
    """
    proj_map = {p["id"]: p for p in projects}
    # Sub-population data rows
    dpop = {p["id"]: p for p in projects if p["id"].startswith("D.")}

    # Mapping: annual goal ID → how to compute current metric
    # Most map to the Project List KPI metric for the matching or related project ID.
    # Col AG overrides take priority when present (checked via _override_int).
    metric_map = {
        "1.1": lambda: _override_int(proj_map, "1.1") if _override_int(proj_map, "1.1") is not None else 1,  # MAP platform operational (default 1)
        "1.2": lambda: _override_int(proj_map, "1.2") if _override_int(proj_map, "1.2") is not None else _pcount(proj_map, "1.2"),
        "1.3": lambda: _override_int(proj_map, "1.3") if _override_int(proj_map, "1.3") is not None else (1 if _ppct(proj_map, "1.3") >= 50 else 0),
        "1.4": lambda: _override_int(proj_map, "1.4") if _override_int(proj_map, "1.4") is not None else _pmetric_int(proj_map, "1.4"),
        "2.1": lambda: _override_int(proj_map, "2.1") if _override_int(proj_map, "2.1") is not None else _pmetric_int(proj_map, "2.1"),
        "2.2": lambda: _override_int(proj_map, "2.2") if _override_int(proj_map, "2.2") is not None else _pmetric_int(proj_map, "2.2"),
        "2.3": lambda: _override_int(proj_map, "2.3") if _override_int(proj_map, "2.3") is not None else _pmetric_int(proj_map, "2.3"),
        "2.4": lambda: _override_int(proj_map, "2.4") if _override_int(proj_map, "2.4") is not None else (1 if _ppct(proj_map, "5.1") > 0 else 0),  # AI-Ready = project 5.1
        "3.1": lambda: _override_int(proj_map, "3.1") if _override_int(proj_map, "3.1") is not None else _pmetric_int(proj_map, "3.1"),
        "3.1.1": lambda: _override_int(dpop, "D.1") if _override_int(dpop, "D.1") is not None else _pmetric_int(dpop, "D.1"),
        "3.1.2": lambda: _override_int(dpop, "D.2") if _override_int(dpop, "D.2") is not None else _pmetric_int(dpop, "D.2"),
        "3.1.2b": lambda: _override_int(dpop, "D.3") if _override_int(dpop, "D.3") is not None else _pmetric_int(dpop, "D.3"),
        "3.2": lambda: _override_int(proj_map, "3.2") if _override_int(proj_map, "3.2") is not None else _pmetric_int(proj_map, "3.2"),
        "3.3": lambda: _override_int(proj_map, "3.3") if _override_int(proj_map, "3.3") is not None else _pmetric_int(proj_map, "3.3"),
        "3.4": lambda: _override_int(proj_map, "3.5") if _override_int(proj_map, "3.5") is not None else _pmetric_int(proj_map, "3.5"),  # stories
        "4.1": lambda: _override_int(proj_map, "4.1") if _override_int(proj_map, "4.1") is not None else 2,  # Title 5 updates (default 2)
        "4.2": lambda: _override_int(proj_map, "4.2") if _override_int(proj_map, "4.2") is not None else 5,  # $5M ongoing (default 5)
        "4.3": lambda: _override_int(proj_map, "4.3") if _override_int(proj_map, "4.3") is not None else _pmetric_int(proj_map, "4.3"),
        "4.4": lambda: _override_int(proj_map, "4.4") if _override_int(proj_map, "4.4") is not None else _pmetric_int(proj_map, "4.3"),  # TA ~ trainings
        "4.5": lambda: _override_int(proj_map, "4.5") if _override_int(proj_map, "4.5") is not None else _pmetric_int(proj_map, "5.4"),  # research ~ RP Group
    }

    for row in annual_goals:
        getter = metric_map.get(row["id"])
        if getter:
            try:
                val = getter()
            except Exception:
                val = 0
            row["current"]["2025-26"] = val
            row["current"]["total"] = val  # only first year has actuals so far


def _pmetric_int(pmap, pid):
    """Extract integer metric from a project map entry."""
    p = pmap.get(pid)
    if not p:
        return 0
    m = p.get("kpi_metric", "")
    try:
        return int(str(m).replace(",", "").replace("k", "000").replace("K", "000").rstrip("+"))
    except (ValueError, TypeError):
        return 0


def _ppct(pmap, pid):
    p = pmap.get(pid)
    return p.get("pct", 0) if p else 0


def _pcount(pmap, pid):
    """Count non-empty integrations for a project."""
    p = pmap.get(pid)
    if not p:
        return 0
    m = p.get("kpi_metric", "")
    try:
        return int(str(m).replace(",", ""))
    except (ValueError, TypeError):
        return 0


def render_annual_goals_table_html(annual_goals):
    """
    Render the Annual Workplan Goals as a static HTML table
    at the bottom of the dashboard. Shows GOAL, CURRENT, STRETCH for each row.
    """
    if not annual_goals:
        return ""

    year_cols = ["2025-26", "2026-27", "2027-28", "2028-29", "2029-30"]

    html = '''        <div style="margin:2rem 0;padding:1.5rem;background:#fff;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,0.06);">
            <h2 style="color:#0A2240;margin:0 0 1rem 0;">Annual Workplan Goals</h2>
            <div style="overflow-x:auto;">
            <table style="width:100%;border-collapse:collapse;font-size:0.82rem;">
                <thead>
                    <tr style="background:#0A2240;color:#fff;">
                        <th style="padding:0.5rem 0.6rem;text-align:left;white-space:nowrap;border:1px solid #163A5F;">Sub-Activity</th>
                        <th style="padding:0.5rem 0.4rem;text-align:center;width:55px;border:1px solid #163A5F;">Type</th>
'''
    for yr in year_cols:
        html += f'                        <th style="padding:0.5rem 0.4rem;text-align:right;white-space:nowrap;border:1px solid #163A5F;">{yr}</th>\n'
    html += '                        <th style="padding:0.5rem 0.4rem;text-align:right;border:1px solid #163A5F;">TOTAL</th>\n'
    html += '                    </tr>\n                </thead>\n                <tbody>\n'

    current_activity = ""
    for row in annual_goals:
        # Activity group header
        if row["activity"] != current_activity:
            current_activity = row["activity"]
            html += (f'                    <tr style="background:#163A5F;color:#fff;">\n'
                     f'                        <td colspan="8" style="padding:0.5rem 0.6rem;font-weight:700;border:1px solid #0A2240;">{current_activity}</td>\n'
                     f'                    </tr>\n')

        # Three rows per sub-activity: GOAL, CURRENT, STRETCH
        for rtype, vals, style in [
            ("Goal", row["goal"], "background:#FAF8F4;font-weight:600;color:#0A2240;"),
            ("Current", row["current"], "background:#fff;color:#2A7D4F;font-weight:700;"),
            ("Stretch", row["stretch"], "background:#FAF8F4;color:#C9A84C;font-style:italic;"),
        ]:
            is_first = rtype == "Goal"
            name_cell = ""
            if is_first:
                name_cell = (f'<td rowspan="3" style="padding:0.4rem 0.6rem;border:1px solid #ddd;'
                             f'vertical-align:top;font-weight:600;background:#fff;">'
                             f'<span style="color:#888;font-size:0.75rem;">{row["id"]}</span> {row["name"]}</td>')

            html += f'                    <tr style="{style}">\n'
            html += f'                        {name_cell}\n'
            html += (f'                        <td style="padding:0.3rem 0.4rem;text-align:center;border:1px solid #ddd;'
                     f'font-size:0.75rem;">{rtype}</td>\n')
            for yr in year_cols:
                v = vals.get(yr, 0)
                # Format: integers as commas, decimals with 1 place
                if isinstance(v, float) and v == int(v):
                    v = int(v)
                disp = fmt_number(v) if v else ""
                # Highlight current year (2025-26) with special styling
                yr_style = "font-weight:700;" if yr == "2025-26" and rtype == "Current" and v else ""
                html += (f'                        <td style="padding:0.3rem 0.4rem;text-align:right;border:1px solid #ddd;{yr_style}">'
                         f'{disp}</td>\n')
            total = vals.get("total", 0)
            if isinstance(total, float) and total == int(total):
                total = int(total)
            html += (f'                        <td style="padding:0.3rem 0.4rem;text-align:right;border:1px solid #ddd;font-weight:700;">'
                     f'{fmt_number(total) if total else ""}</td>\n')
            html += '                    </tr>\n'

    html += '                </tbody>\n            </table>\n            </div>\n        </div>\n'
    return html


def build_activity_kpis(projects):
    """
    Build 19 activity-level KPI cards from the core sub-activities (1.1-4.5).
    Groups them by Activity 1-4 for the dashboard.
    """
    # Core sub-activity IDs in order (19 Workplan sub-activities)
    # Note: 4.1 is represented by 4.1a-4.1d in the project list (sprint components)
    core_ids = [
        "1.1", "1.2", "1.3", "1.4",
        "2.1", "2.2", "2.3", "2.4",
        "3.1", "3.2", "3.3", "3.4", "3.5", "3.6",
        "4.1",  # aggregated from 4.1a-4.1d
        "4.2", "4.3", "4.4", "4.5",
    ]

    # Sub-IDs that compose 4.1
    sprint_ids = ["4.1a", "4.1b", "4.1c", "4.1d"]

    activity_labels = {
        "1": "Activity 1: Build AI-Enhanced CPL Infrastructure",
        "2": "Activity 2: Faculty Workgroups & Credit Recommendations",
        "3": "Activity 3: Build CPL Data Infrastructure",
        "4": "Activity 4: Sprints, Projects, Partnerships & Scale",
    }

    # Build a lookup by project ID
    proj_map = {p["id"]: p for p in projects}

    # Build composite entry for 4.1 (Sprints) from 4.1a-4.1d
    sprint_projects = [proj_map[sid] for sid in sprint_ids if sid in proj_map]
    if sprint_projects:
        sprint_details = "; ".join(
            f"{sp['id']} {sp['name']}: {sp['kpi_metric'] or 'N/A'} {sp['kpi_unit']}"
            for sp in sprint_projects
        )
        # Use the first sprint with a workplan note, or compose one
        wp_note = next((sp["workplan_notes"] for sp in sprint_projects if sp["workplan_notes"]), "")
        composite_41 = {
            "id": "4.1",
            "name": "Sprints (Veteran, Apprenticeship, Adoption, 29 Palms)",
            "kpi_metric": str(len(sprint_projects)),
            "kpi_unit": "active sprints",
            "kpi_goal_2526": "",
            "kpi_stretch_2526": "",
            "kpi_goal_2627": "",
            "kpi_stretch_2627": "",
            "kpi_goal_2728": "",
            "kpi_stretch_2728": "",
            "kpi_goal_2829": "",
            "kpi_stretch_2829": "",
            "kpi_goal_2930": "",
            "kpi_stretch_2930": "",
            "workplan_notes": wp_note or f"Components: {sprint_details}",
            "status": "In Progress",
            "pct": sum(sp["pct"] for sp in sprint_projects) // len(sprint_projects),
            "excel_row": sprint_projects[0].get("excel_row", 0),
            "sprint_components": [
                {
                    "id": sp["id"],
                    "name": sp["name"],
                    "metric": sp["kpi_metric"],
                    "unit": sp["kpi_unit"],
                    "status": sp["status"],
                    "pct": sp["pct"],
                }
                for sp in sprint_projects
            ],
        }
        proj_map["4.1"] = composite_41

    groups = {}  # key = "1","2","3","4"
    for pid in core_ids:
        p = proj_map.get(pid)
        if not p:
            continue
        act_num = pid.split(".")[0]
        if act_num not in groups:
            groups[act_num] = {
                "activity_id": f"Activity {act_num}",
                "activity_name": activity_labels.get(act_num, f"Activity {act_num}"),
                "kpis": [],
            }
        entry = {
            "id":            p["id"],
            "name":          p.get("name", ""),
            "goal":          p.get("goal", ""),
            "metric":        p.get("kpi_metric", ""),
            "unit":          p.get("kpi_unit", ""),
            "goal_2526":   p.get("kpi_goal_2526", ""),
            "stretch_2526":  p.get("kpi_stretch_2526", ""),
            "goal_2627":   p.get("kpi_goal_2627", ""),
            "stretch_2627":  p.get("kpi_stretch_2627", ""),
            "goal_2728":   p.get("kpi_goal_2728", ""),
            "stretch_2728":  p.get("kpi_stretch_2728", ""),
            "goal_2829":   p.get("kpi_goal_2829", ""),
            "stretch_2829":  p.get("kpi_stretch_2829", ""),
            "goal_2930":   p.get("kpi_goal_2930", ""),
            "stretch_2930":  p.get("kpi_stretch_2930", ""),
            "workplan_notes": p.get("workplan_notes", ""),
            "update":        p.get("update", ""),
            "update_date":   p.get("update_date", ""),
            "status":        p.get("status", ""),
            "pct":           p.get("pct", 0),
            "excel_row":     p.get("excel_row", 0),
        }
        # Include sprint components for 4.1
        if pid == "4.1" and "sprint_components" in p:
            entry["sprint_components"] = p["sprint_components"]
        groups[act_num]["kpis"].append(entry)

    # Return as ordered list
    return [groups[k] for k in sorted(groups.keys()) if k in groups]


def render_kpi_section_html(kpis, kpi_display_order=None, kpi_params=None):
    """
    Generate static HTML for the headline KPI cards section,
    including LIVE badges, subtitles, and population breakdowns.
    kpi_display_order: list of KPI keys in desired display order.
    kpi_params: tunable parameters from the KPI_Config sheet (used to
                interpolate placeholders in algo descriptions).
    """
    default_order = ['cumulative_students', 'eligible_units', 'transcripted_units',
                     'credit_recommendations', 'map_exhibits', 'ccc_collaborative',
                     'active_colleges', 'articulation_colleges',
                     'estimated_savings', 'veteran_sprint', 'twenty_year_impact']
    display_keys = kpi_display_order if kpi_display_order else default_order

    cards_html = ""
    for key in display_keys:
        kpi = kpis.get(key)
        if not kpi:
            continue

        # ── Standard KPI card rendering ──
        sub_html = ""
        if kpi.get("sub"):
            sub_html = f'<div class="kpi-sub" style="font-size:0.75rem;opacity:0.85;margin-top:2px;color:#fff;">{kpi["sub"]}</div>'
        bd_html = ""
        if kpi.get("breakdowns"):
            rows = ""
            for bd in kpi["breakdowns"]:
                note_html = f' <span class="kpi-bd-note">({bd["note"]})</span>' if bd.get("note") else ""
                rows += (f'<div class="kpi-bd-row">'
                         f'<span class="kpi-bd-label">{bd["label"]}{note_html}</span>'
                         f'<span class="kpi-bd-value">{bd["value"]}</span>'
                         f'</div>\n')
            bd_html = f'<div class="kpi-breakdowns">{rows}</div>'

        # Footnote (e.g. tier criteria for Active Colleges)
        fn_html = ""
        if kpi.get("footnote"):
            fn_items = "".join(f'<div class="kpi-fn-item">{item}</div>' for item in kpi["footnote"])
            fn_html = f'<div class="kpi-footnote">{fn_items}</div>'

        algo_html = render_algo_details(key, params=kpi_params)

        cards_html += (
            f'        <div class="kpi-card">\n'
            f'            <div class="kpi-number">{kpi["value"]}</div>\n'
            f'            <div class="kpi-label">{kpi["label"]}</div>\n'
            f'            {sub_html}\n'
            f'            {bd_html}\n'
            f'            {fn_html}\n'
            f'            {algo_html}\n'
            f'        </div>\n'
        )
    return cards_html


def _att_badge(attachments, act_num=None, project_id=None):
    """Return a small count badge HTML if there are attachments for an activity or project."""
    if not attachments:
        return ""
    count = 0
    if project_id:
        count = attachments.get("by_project", {}).get(str(project_id), 0)
    elif act_num:
        count = attachments.get("by_activity", {}).get(str(act_num), 0)
    if count <= 0:
        return ""
    return (f' <span style="background:#C9A84C;color:#0A2240;font-size:0.6rem;'
            f'font-weight:700;padding:1px 5px;border-radius:8px;margin-left:2px;">'
            f'{count}</span>')


def render_activity_kpis_html(activity_kpis, annual_goals=None, update_log=None, attachments=None):
    """
    Generate static HTML for the 19 activity-level KPI cards section.
    KPI cards are grouped under Goal sub-headers within each Activity.
    annual_goals: list from read_annual_goals() — used to show 2025-26 targets.
    """
    if not activity_kpis:
        return ""

    # Build a lookup: activity number → list of annual goal rows for that activity
    ag_by_activity = {}
    if annual_goals:
        for ag in annual_goals:
            act_num = ag["id"].split(".")[0]
            ag_by_activity.setdefault(act_num, []).append(ag)

    # Title now lives in the collapsible wrapper header — do not emit it here.
    html = ''

    # Activity-level goals for progress bars
    activity_goals = {
        "Activity 1": "AI-enhanced MAP platform with student portal, integrations, and credential registry",
        "Activity 2": "1,000 statewide credit recommendations from 25 faculty/industry workgroups",
        "Activity 3": "Statewide CPL data infrastructure tracking 250K students across 116 colleges",
        "Activity 4": "Scale CPL through sprints, partnerships, training, policy, and sustainable funding",
    }

    for group in activity_kpis:
        act_id = group["activity_id"]
        act_name = group["activity_name"].replace(act_id + ": ", "")
        act_num = act_id.replace("Activity ", "")

        # Compute average progress across sub-activities in this group
        pcts = [kpi.get("pct", 0) for kpi in group["kpis"]]
        avg_pct = round(sum(pcts) / len(pcts)) if pcts else 0
        completed = sum(1 for p in pcts if p >= 100)
        total_kpis = len(pcts)
        act_goal_text = activity_goals.get(act_id, "")

        # Determine bar color based on avg progress
        if avg_pct >= 75:
            bar_color = "#2A7D4F"  # green
        elif avg_pct >= 50:
            bar_color = "#C9A84C"  # gold
        elif avg_pct >= 25:
            bar_color = "#4A90D9"  # blue
        else:
            bar_color = "#888"     # gray

        # Build 2025-26 annual goals summary for this activity
        ag_items = ag_by_activity.get(act_num, [])
        annual_summary = ""
        if ag_items:
            parts = []
            for ag in ag_items:
                g_val = ag["goal"].get("2025-26", 0)
                c_val = ag["current"].get("2025-26", 0)
                if g_val:
                    g_int = int(g_val) if float(g_val) == int(g_val) else g_val
                    c_int = int(c_val) if float(c_val) == int(c_val) else c_val
                    # Shorten name for display
                    short_name = ag["name"]
                    if len(short_name) > 30:
                        short_name = short_name[:28] + "…"
                    parts.append(f'{short_name}: <strong>{fmt_number(c_int)}</strong>/{fmt_number(g_int)}')
            if parts:
                annual_summary = (
                    '<div style="margin-top:0.3rem;display:flex;flex-wrap:wrap;gap:0.3rem 1rem;">'
                    + ''.join(f'<span style="font-size:0.7rem;color:#555;">{p}</span>' for p in parts)
                    + '</div>'
                )

        html += f'            <div class="activity-group">\n'
        html += (f'            <div class="activity-group-header">\n'
                 f'                <h3><span style="color:#888;font-weight:600;">{act_id}:</span> {act_name}</h3>\n'
                 f'            </div>\n')
        # Activity progress bar with goal + annual targets
        html += (f'            <div style="padding:0 0.5rem 0.6rem 0.5rem;">\n'
                 f'                <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:0.25rem;">\n'
                 f'                    <span style="font-size:0.75rem;color:#555;">{act_goal_text}</span>\n'
                 f'                    <span style="font-size:0.75rem;font-weight:700;color:{bar_color};white-space:nowrap;margin-left:0.5rem;">{avg_pct}% avg &middot; {completed}/{total_kpis} complete (toward 2030 goal)</span>\n'
                 f'                </div>\n'
                 f'                <div style="height:6px;background:#e8e8e8;border-radius:3px;overflow:hidden;">\n'
                 f'                    <div style="height:100%;width:{avg_pct}%;background:{bar_color};border-radius:3px;transition:width 0.3s;"></div>\n'
                 f'                </div>\n'
                 f'            </div>\n')
        # Insert 2025-26 annual goal summary below progress bar
        if annual_summary:
            html += f'            <div style="padding:0 0.5rem 0.3rem 0.5rem;">{annual_summary}</div>\n'

        # Group KPIs by their primary Goal
        goal_groups = {}
        for kpi in group["kpis"]:
            goal_raw = kpi.get("goal", "")
            # Extract first goal (e.g., "Goal 1" from "Goal 1; Goal 2")
            primary_goal = ""
            if goal_raw:
                gm = re.match(r'(Goal\s*\d+)', goal_raw)
                if gm:
                    primary_goal = gm.group(1)
            if not primary_goal:
                primary_goal = "Other"
            goal_groups.setdefault(primary_goal, []).append(kpi)

        # Sort goal groups: Goal 1, Goal 2, Goal 3, then Other
        sorted_goals = sorted(goal_groups.keys(),
                              key=lambda g: (0, int(re.search(r'\d+', g).group())) if re.search(r'\d+', g) else (1, 0))

        for goal_key in sorted_goals:
            goal_kpis = goal_groups[goal_key]
            goal_info = CPL_GOALS.get(goal_key, {})
            goal_title = goal_info.get("title", "")
            if goal_key != "Other" and goal_title:
                html += (f'            <div class="goal-subheader" style="margin:1rem 0 0.5rem 0;padding:0.5rem 0.8rem;'
                         f'background:linear-gradient(90deg,#163A5F 0%,#0A2240 100%);border-radius:6px;">\n'
                         f'                <div style="display:flex;align-items:center;gap:0.6rem;">\n'
                         f'                    <span style="color:#C9A84C;font-weight:700;font-size:0.85rem;white-space:nowrap;">{goal_key}</span>\n'
                         f'                    <span style="color:#fff;font-size:0.8rem;font-weight:600;">{goal_title}</span>\n'
                         f'                </div>\n')
                # Add target text
                target = goal_info.get("target", "")
                if target:
                    html += f'                <div style="color:rgba(255,255,255,0.75);font-size:0.72rem;margin-top:0.2rem;line-height:1.3;">{target}</div>\n'
                html += f'            </div>\n'

            html += '            <div class="activity-kpi-grid">\n'

            for kpi in goal_kpis:
                has_metric = kpi.get("metric") and kpi["metric"] not in ("0", "")
                card_class = "has-metric" if has_metric else "no-metric"
                status_raw = kpi.get("status", "")
                status_class = status_raw.lower().replace(" ", "-")

                html += f'            <div class="activity-kpi-card {card_class}">\n'
                html += (f'                <div class="akpi-header">\n'
                         f'                    <span class="akpi-id">{kpi["id"]}</span>\n'
                         f'                    <span class="akpi-status status-badge status-{status_class}" '
                         f'style="margin:0;font-size:0.7rem;padding:0.15rem 0.5rem;">{status_raw}</span>\n'
                         f'                </div>\n')
                html += f'                <div class="akpi-name">{kpi["name"]}</div>\n'

                # 2030 Goal & Stretch subtitle
                g2930 = kpi.get("goal_2930", "")
                s2930_val = kpi.get("stretch_2930", "")
                if g2930:
                    sub_parts = [f'2030 Goal: {g2930}']
                    if s2930_val:
                        sub_parts.append(f'Stretch: {s2930_val}')
                    html += (f'                <div style="font-size:0.7rem;color:#666;margin:-0.2rem 0 0.3rem 0;">'
                             f'{" &nbsp; ".join(sub_parts)}</div>\n')

                if has_metric:
                    html += (f'                <div class="akpi-metric-row">\n'
                             f'                    <span class="akpi-metric-value">{kpi["metric"]}</span>\n'
                             f'                    <span class="akpi-metric-unit">{kpi["unit"]}</span>\n'
                             f'                </div>\n')
                else:
                    html += (f'                <div class="akpi-metric-row">\n'
                             f'                    <span class="akpi-metric-unit" style="opacity:0.5;">{kpi["unit"]}</span>\n'
                             f'                </div>\n')

                # Annual cumulative goals — mini table: Year / Goal / Stretch
                g2526 = kpi.get("goal_2526", "")
                g2627 = kpi.get("goal_2627", "")
                g2728 = kpi.get("goal_2728", "")
                g2829 = kpi.get("goal_2829", "")
                g2930_v = kpi.get("goal_2930", "")
                s2526 = kpi.get("stretch_2526", "")
                s2627 = kpi.get("stretch_2627", "")
                s2728 = kpi.get("stretch_2728", "")
                s2829 = kpi.get("stretch_2829", "")
                s2930 = kpi.get("stretch_2930", "")
                any_goal = g2526 or g2627 or g2728 or g2829 or g2930_v
                if any_goal:
                    tbl_style = ('font-size:0.65rem;border-collapse:collapse;width:100%;'
                                 'margin-top:0.3rem;text-align:center;')
                    th_style = 'padding:2px 4px;color:#666;font-weight:600;border-bottom:1px solid #ddd;'
                    td_style = 'padding:2px 4px;'
                    html += f'                <table style="{tbl_style}">\n'
                    # Header row: YEAR
                    html += '                  <tr>\n'
                    html += f'                    <td style="{th_style}font-weight:700;text-align:left;">YEAR</td>\n'
                    for yr in ['2526', '2627', '2728', '2829', '2930']:
                        html += f'                    <td style="{th_style}">{yr}</td>\n'
                    html += '                  </tr>\n'
                    # Goal row
                    html += '                  <tr>\n'
                    html += f'                    <td style="{td_style}font-weight:600;text-align:left;color:#0A2240;">GOAL</td>\n'
                    for v in [g2526, g2627, g2728, g2829, g2930_v]:
                        html += f'                    <td style="{td_style}">{v}</td>\n'
                    html += '                  </tr>\n'
                    # Stretch row
                    html += '                  <tr>\n'
                    html += f'                    <td style="{td_style}font-weight:600;text-align:left;color:#C9A84C;">STRETCH</td>\n'
                    for v in [s2526, s2627, s2728, s2829, s2930]:
                        html += f'                    <td style="{td_style}color:#C9A84C;">{v}</td>\n'
                    html += '                  </tr>\n'
                    html += '                </table>\n'
                    html += ('                <div style="font-size:0.55rem;color:#999;'
                             'font-style:italic;margin-top:0.15rem;">Note: All values are cumulative</div>\n')

                # Progress bar with percentage label
                pct = kpi.get("pct", 0)
                if pct >= 75:
                    pbar_color = "#2A7D4F"
                elif pct >= 50:
                    pbar_color = "#C9A84C"
                elif pct >= 25:
                    pbar_color = "#4A90D9"
                else:
                    pbar_color = "#888"
                html += (f'                <div class="akpi-progress" style="margin-top:0.5rem;">\n'
                         f'                    <div style="display:flex;justify-content:space-between;font-size:0.7rem;margin-bottom:0.15rem;">\n'
                         f'                        <span style="color:#666;">Progress</span>\n'
                         f'                        <span style="font-weight:700;color:{pbar_color};">{pct}%</span>\n'
                         f'                    </div>\n'
                         f'                    <div style="height:5px;background:#e8e8e8;border-radius:3px;overflow:hidden;">\n'
                         f'                        <div style="height:100%;width:{pct}%;background:{pbar_color};border-radius:3px;"></div>\n'
                         f'                    </div>\n'
                         f'                </div>\n')

                # Sprint components for 4.1
                if kpi.get("sprint_components"):
                    html += '                <div class="akpi-sprint-components" style="margin-top:0.5rem;font-size:0.75rem;">\n'
                    for sc in kpi["sprint_components"]:
                        sc_status_class = sc.get("status", "").lower().replace(" ", "-")
                        html += (f'                    <div style="display:flex;justify-content:space-between;padding:0.2rem 0;border-bottom:1px solid rgba(0,0,0,0.05);">\n'
                                 f'                        <span><strong>{sc["id"]}</strong> {sc["name"]}</span>\n'
                                 f'                        <span class="status-badge status-{sc_status_class}" style="font-size:0.65rem;padding:0.1rem 0.4rem;">{sc.get("status","")}</span>\n'
                                 f'                    </div>\n')
                    html += '                </div>\n'

                # ── Notes section: Last updated + Latest Update + Workplan Note + history toggle ──
                kpi_update = kpi.get("update", "")
                kpi_wp = kpi.get("workplan_notes", "")
                kpi_date = kpi.get("update_date", "")
                kpi_pid = kpi["id"]
                kpi_notes_list = (update_log or {}).get(kpi_pid, [])

                if kpi_update or kpi_wp:
                    # Toggle for full history
                    toggle_part = ""
                    if len(kpi_notes_list) > 1:
                        toggle_part = (
                            f'                        <label style="display:inline-flex;align-items:center;gap:0.3rem;'
                            f'font-size:0.68rem;color:#163A5F;cursor:pointer;margin-left:auto;">'
                            f'<input type="checkbox" class="notes-history-toggle" data-pid="{kpi_pid}" '
                            f'style="accent-color:#C9A84C;cursor:pointer;"> '
                            f'Show all ({len(kpi_notes_list)})</label>\n'
                        )

                    html += (f'                <div style="margin-top:0.5rem;border-top:1px solid #e8e8e8;padding-top:0.4rem;">\n'
                             f'                    <div style="display:flex;align-items:center;margin-bottom:0.3rem;">\n'
                             f'                        <span style="font-size:0.7rem;color:#888;">Last updated: '
                             f'<strong style="color:#0A2240;">{kpi_date}</strong></span>\n'
                             f'{toggle_part}'
                             f'                    </div>\n')

                    # Latest Update (col P)
                    if kpi_update:
                        html += (f'                    <div style="font-size:0.75rem;color:#444;line-height:1.4;margin-bottom:0.3rem;">'
                                 f'<span style="font-size:0.62rem;font-weight:600;background:#163A5F;color:#fff;'
                                 f'padding:0.1rem 0.3rem;border-radius:3px;margin-right:0.25rem;">Latest Update</span>'
                                 f'{kpi_update}</div>\n')

                    # Workplan Note (col V)
                    if kpi_wp:
                        html += (f'                    <div style="font-size:0.75rem;color:#444;line-height:1.4;margin-bottom:0.3rem;">'
                                 f'<span style="font-size:0.62rem;font-weight:600;background:#C9A84C;color:#0A2240;'
                                 f'padding:0.1rem 0.3rem;border-radius:3px;margin-right:0.25rem;">Workplan Note</span>'
                                 f'{kpi_wp}</div>\n')

                    # Full history (hidden by default)
                    if len(kpi_notes_list) > 1:
                        type_badge_css = {
                            "update":   "background:#163A5F;color:#fff;",
                            "workplan": "background:#C9A84C;color:#0A2240;",
                        }
                        type_labels = {"update": "Progress Update", "workplan": "Workplan Note"}
                        html += (f'                    <div class="notes-history" data-pid="{kpi_pid}" '
                                 f'style="display:none;margin-top:0.3rem;max-height:200px;overflow-y:auto;'
                                 f'border-left:3px solid #C9A84C;padding-left:0.5rem;">\n')
                        for n in kpi_notes_list:
                            ntype = n.get("type", "update")
                            badge_style = type_badge_css.get(ntype, type_badge_css["update"])
                            badge_label = type_labels.get(ntype, "Update")
                            html += (f'                        <div style="margin-bottom:0.4rem;">'
                                     f'<span style="font-size:0.68rem;font-weight:700;color:#0A2240;'
                                     f'background:#f0f0f0;padding:0.1rem 0.35rem;border-radius:3px;">{n["date"]}</span>'
                                     f' <span style="font-size:0.6rem;font-weight:600;{badge_style}'
                                     f'padding:0.1rem 0.3rem;border-radius:3px;">{badge_label}</span>'
                                     f'<div style="font-size:0.73rem;color:#444;margin-top:0.1rem;line-height:1.3;">{n["note"]}</div>'
                                     f'</div>\n')
                        html += '                    </div>\n'

                    html += '                </div>\n'

                # Report + Update buttons
                safe_id = kpi_pid.replace(".", "_")
                excel_row = kpi.get("excel_row", 0)
                btn_style = ('display:inline-flex;align-items:center;gap:0.3rem;margin-top:0.5rem;'
                             'font-size:0.7rem;text-decoration:none;font-weight:600;'
                             'padding:0.25rem 0.5rem;border:1px solid #ddd;border-radius:4px;'
                             'cursor:pointer;transition:background 0.2s;margin-right:0.4rem;')
                html += (f'                <a href="reports/projects/{kpi_pid}_Report.docx" '
                         f'download class="report-btn" '
                         f'style="{btn_style}color:#163A5F;background:#fafafa;"'
                         f' onmouseover="this.style.background=\'#e8e8e8\'" onmouseout="this.style.background=\'#fafafa\'">'
                         f'<span style="font-size:0.8rem;">&#128196;</span> Report</a>'
                         f'<a href="{excel_cell_url(excel_row)}" target="_blank" rel="noopener" '
                         f'class="update-btn" data-row="{excel_row}" data-col="P" '
                         f'style="{btn_style}color:#FFFFFF;background:#C9A84C;"'
                         f' onmouseover="this.style.background=\'#b89540\'" onmouseout="this.style.background=\'#C9A84C\'"'
                         f' title="Open Excel for the Web at cell P{excel_row} ({EXCEL_SHEET_NAME})">'
                         f'<span style="font-size:0.8rem;">&#9998;</span> Update</a>'
                         f'<a href="#" '
                         f'class="attach-btn" '
                         f'data-folder="{kpi_pid} {kpi["name"]}" '
                         f'style="{btn_style}color:#163A5F;background:#fafafa;"'
                         f' onmouseover="this.style.background=\'#e8e8e8\'" onmouseout="this.style.background=\'#fafafa\'"'
                         f' title="Open SharePoint folder — use Upload or drag &amp; drop to add files">'
                         f'<span style="font-size:0.8rem;">&#128206;</span> Attach'
                         f'{_att_badge(attachments, act_num)}</a>\n')

                html += '            </div>\n'  # close activity-kpi-card

            html += '            </div>\n'  # close activity-kpi-grid (per goal)

        html += '            </div>\n'  # close activity-group

    return html


def render_workplan_goals_html(workplan_goals):
    """
    Render the Annual Workplan Goals as a dashboard section with
    grouped tables showing year-by-year goals and stretch goals.
    Activities are grouped by their primary number (1.x, 2.x, 3.x, 4.x).
    """
    if not workplan_goals:
        return ""

    activity_group_labels = {
        "1": "Activity 1: Build AI-Enhanced CPL Infrastructure",
        "2": "Activity 2: Faculty Workgroups & Credit Recommendations",
        "3": "Activity 3: Build CPL Data Infrastructure",
        "4": "Activity 4: Sprints, Projects, Partnerships & Scale",
    }

    # Group by activity number
    groups = {}
    for act in workplan_goals:
        grp = act["id"].split(".")[0]
        groups.setdefault(grp, []).append(act)

    html = '''        <div class="workplan-goals-section" style="margin:2.5rem 0;">
            <h2 style="color:#0A2240;margin-bottom:0.5rem;">Annual Workplan Goals & Stretch Targets</h2>
            <p style="color:#666;font-size:0.85rem;margin-bottom:1.5rem;">Five-year trajectory from the CCCCO CPL Workplan — Goal and Stretch targets per activity per year.</p>
'''

    for grp_num in sorted(groups.keys()):
        acts = groups[grp_num]
        grp_label = activity_group_labels.get(grp_num, f"Activity {grp_num}")

        html += f'''            <div style="margin-bottom:2rem;">
            <div style="background:linear-gradient(135deg,#163A5F 0%,#0A2240 100%);border-radius:8px 8px 0 0;padding:0.6rem 1rem;">
                <span style="color:#C9A84C;font-weight:700;font-size:0.9rem;">{grp_label}</span>
            </div>
            <div style="overflow-x:auto;">
            <table style="width:100%;border-collapse:collapse;font-size:0.8rem;background:#fff;box-shadow:0 2px 8px rgba(0,0,0,0.06);border-radius:0 0 8px 8px;table-layout:fixed;">
                <colgroup>
                    <col style="width:30%;">
                    <col style="width:8%;">
                    <col style="width:10.4%;">
                    <col style="width:10.4%;">
                    <col style="width:10.4%;">
                    <col style="width:10.4%;">
                    <col style="width:10.4%;">
                    <col style="width:10%;">
                </colgroup>
                <thead>
                    <tr style="background:#f0f4f8;">
                        <th style="text-align:left;padding:0.5rem 0.7rem;border-bottom:2px solid #ddd;">Activity</th>
                        <th style="text-align:center;padding:0.5rem 0.4rem;border-bottom:2px solid #ddd;color:#666;">Type</th>
                        <th style="text-align:right;padding:0.5rem 0.4rem;border-bottom:2px solid #ddd;">2025-26</th>
                        <th style="text-align:right;padding:0.5rem 0.4rem;border-bottom:2px solid #ddd;">2026-27</th>
                        <th style="text-align:right;padding:0.5rem 0.4rem;border-bottom:2px solid #ddd;">2027-28</th>
                        <th style="text-align:right;padding:0.5rem 0.4rem;border-bottom:2px solid #ddd;">2028-29</th>
                        <th style="text-align:right;padding:0.5rem 0.4rem;border-bottom:2px solid #ddd;">2029-30</th>
                        <th style="text-align:right;padding:0.5rem 0.7rem;border-bottom:2px solid #ddd;font-weight:700;">Total</th>
                    </tr>
                </thead>
                <tbody>
'''
        for i, act in enumerate(acts):
            is_pct = act.get("is_percentage", False)
            bg = "#fff" if i % 2 == 0 else "#fafbfc"

            def fmt_val(v, pct=False):
                if pct and isinstance(v, (int, float)):
                    return f"{int(v * 100)}%"
                if isinstance(v, (int, float)):
                    if float(v) == int(v):
                        return f"{int(v):,}"
                    return f"{v:,.1f}"
                return str(v) if v else "—"

            # GOAL row
            html += f'''                    <tr style="background:{bg};">
                        <td rowspan="2" style="padding:0.5rem 0.7rem;border-bottom:1px solid #eee;vertical-align:top;font-weight:600;color:#0A2240;">{act["id"]} {act["name"]}</td>
                        <td style="text-align:center;padding:0.3rem 0.4rem;color:#163A5F;font-weight:600;font-size:0.72rem;">GOAL</td>
'''
            for v in act["goal"]:
                html += f'                        <td style="text-align:right;padding:0.3rem 0.4rem;">{fmt_val(v, is_pct)}</td>\n'
            html += f'                        <td style="text-align:right;padding:0.3rem 0.7rem;font-weight:700;color:#0A2240;">{fmt_val(act["goal_total"], is_pct)}</td>\n'
            html += '                    </tr>\n'

            # STRETCH row
            html += f'''                    <tr style="background:{bg};">
                        <td style="text-align:center;padding:0.3rem 0.4rem;border-bottom:1px solid #eee;color:#C9A84C;font-weight:600;font-size:0.72rem;">STRETCH</td>
'''
            for v in act["stretch"]:
                html += f'                        <td style="text-align:right;padding:0.3rem 0.4rem;border-bottom:1px solid #eee;color:#C9A84C;">{fmt_val(v, is_pct)}</td>\n'
            html += f'                        <td style="text-align:right;padding:0.3rem 0.7rem;border-bottom:1px solid #eee;font-weight:700;color:#C9A84C;">{fmt_val(act["stretch_total"], is_pct)}</td>\n'
            html += '                    </tr>\n'

        html += '''                </tbody>
            </table>
            </div>
            </div>
'''

    html += '        </div>\n'
    return html


def _render_single_project_card(p, update_log=None, attachments=None):
    """Render a single project card HTML string with full notes history."""
    pid = p["id"]
    status = p.get("status", "")
    status_class = status.lower().replace(" ", "-")
    status_css_map = {
        "goal-met": "completed",
        "stretch-met": "completed",
        "on-track": "on-track",
        "in-progress": "in-progress",
        "foundational-year": "foundational",
        "not-started": "not-started",
        # Legacy fallbacks
        "completed": "completed",
        "exceeded-target": "completed",
        "ahead-of-schedule": "on-track",
        "strong-progress": "on-track",
        "active": "in-progress",
        "at-risk": "at-risk",
        "not-yet-started": "not-started",
        "proposed": "not-started",
    }
    css_class = status_css_map.get(status_class, "in-progress")
    pct = p.get("pct", 0)
    activity = p.get("activity", "")
    v2030 = p.get("v2030", "")
    goal = p.get("goal", "")
    lead = p.get("lead", "")

    # ── Current notes (always visible) ──
    update_text = p.get("update", "")
    wp_text = p.get("workplan_notes", "")
    update_date = p.get("update_date", "")

    # Current update + workplan note section
    current_notes_html = ""
    if update_text:
        current_notes_html += (
            f'            <div style="font-size:0.8rem;color:#444;line-height:1.4;margin-bottom:0.35rem;">'
            f'<span style="font-size:0.65rem;font-weight:600;background:#163A5F;color:#fff;'
            f'padding:0.1rem 0.35rem;border-radius:3px;margin-right:0.3rem;">Latest Update</span>'
            f'{update_text}</div>\n'
        )
    if wp_text:
        current_notes_html += (
            f'            <div style="font-size:0.8rem;color:#444;line-height:1.4;margin-bottom:0.35rem;">'
            f'<span style="font-size:0.65rem;font-weight:600;background:#C9A84C;color:#0A2240;'
            f'padding:0.1rem 0.35rem;border-radius:3px;margin-right:0.3rem;">Workplan Note</span>'
            f'{wp_text}</div>\n'
        )

    # ── Full history toggle (hidden by default, shown via JS checkbox) ──
    all_notes = (update_log or {}).get(pid, [])
    history_html = ""
    if len(all_notes) > 1:
        type_badge_css = {
            "update":   "background:#163A5F;color:#fff;",
            "workplan": "background:#C9A84C;color:#0A2240;",
        }
        type_labels = {"update": "Progress Update", "workplan": "Workplan Note"}
        history_html += (
            f'            <div class="notes-history" data-pid="{pid}" '
            f'style="display:none;margin-top:0.4rem;max-height:240px;overflow-y:auto;'
            f'border-left:3px solid #C9A84C;padding-left:0.6rem;">\n'
        )
        for n in all_notes:
            ntype = n.get("type", "update")
            badge_style = type_badge_css.get(ntype, type_badge_css["update"])
            badge_label = type_labels.get(ntype, "Update")
            history_html += (
                f'                <div style="margin-bottom:0.5rem;">'
                f'<span style="font-size:0.72rem;font-weight:700;color:#0A2240;'
                f'background:#f0f0f0;padding:0.1rem 0.4rem;border-radius:3px;">{n["date"]}</span>'
                f' <span style="font-size:0.65rem;font-weight:600;{badge_style}'
                f'padding:0.1rem 0.35rem;border-radius:3px;">{badge_label}</span>'
                f'<div style="font-size:0.8rem;color:#444;margin-top:0.15rem;line-height:1.4;">{n["note"]}</div>'
                f'</div>\n'
            )
        history_html += '            </div>\n'

    # Combine into notes section
    notes_html = ""
    if current_notes_html or history_html:
        toggle_html = ""
        if len(all_notes) > 1:
            toggle_html = (
                f'                <label style="display:inline-flex;align-items:center;gap:0.3rem;'
                f'font-size:0.72rem;color:#163A5F;cursor:pointer;margin-left:auto;">'
                f'<input type="checkbox" class="notes-history-toggle" data-pid="{pid}" '
                f'style="accent-color:#C9A84C;cursor:pointer;"> '
                f'Show all ({len(all_notes)})</label>\n'
            )
        notes_html = (
            f'            <div style="margin-top:0.5rem;border-top:1px solid #e8e8e8;padding-top:0.5rem;">\n'
            f'                <div style="display:flex;align-items:center;margin-bottom:0.35rem;">\n'
            f'                    <span style="font-size:0.75rem;color:#888;">Last updated: '
            f'<strong style="color:#0A2240;">{update_date}</strong></span>\n'
            f'{toggle_html}'
            f'                </div>\n'
            f'{current_notes_html}'
            f'{history_html}'
            f'            </div>\n'
        )

    return f'''        <div class="project-card" data-activity="{activity}" data-v2030="{v2030}" data-goal="{goal}" data-status="{status}" data-lead="{lead}">
            <div class="project-name">{p.get("name", "")}</div>
            <div class="project-desc">{p.get("desc", "")}</div>
            <span class="status-badge status-{css_class}">{status}</span>
            <div class="progress-container">
                <div class="progress-label">
                    <span>Progress</span>
                    <span>{pct}%</span>
                </div>
                <div class="progress-bar-bg">
                    <div class="progress-bar-fill" style="width:{pct}%;--progress-width:{pct}%;"></div>
                </div>
            </div>
            <div style="font-size:0.85rem;color:#555;margin-bottom:0.5rem;">
                <strong>Lead:</strong> {lead}
            </div>
            <div style="font-size:0.85rem;color:#555;margin-bottom:0.5rem;">
                <strong>Activity:</strong> {activity}
            </div>
            <div style="font-size:0.85rem;color:#555;margin-bottom:0.5rem;">
                <strong>Budget:</strong> {p.get("budget", "")} ({p.get("budget_source", "")})
            </div>
{notes_html}            <div style="display:flex;gap:0.4rem;flex-wrap:wrap;margin-top:0.5rem;">
                <a href="reports/projects/{pid}_Report.docx" download class="report-btn"
                    style="display:inline-flex;align-items:center;gap:0.3rem;
                    font-size:0.75rem;color:#163A5F;text-decoration:none;font-weight:600;
                    padding:0.3rem 0.6rem;border:1px solid #ddd;border-radius:4px;
                    background:#fafafa;cursor:pointer;transition:background 0.2s;"
                    onmouseover="this.style.background='#e8e8e8'" onmouseout="this.style.background='#fafafa'">
                    <span style="font-size:0.85rem;">&#128196;</span> Report</a>
                <a href="{excel_cell_url(p.get('excel_row', 0))}" target="_blank" rel="noopener"
                    class="update-btn" data-row="{p.get('excel_row', 0)}" data-col="P"
                    style="display:inline-flex;align-items:center;gap:0.3rem;
                    font-size:0.75rem;color:#FFFFFF;text-decoration:none;font-weight:600;
                    padding:0.3rem 0.6rem;border:1px solid #b89540;border-radius:4px;
                    background:#C9A84C;cursor:pointer;transition:background 0.2s;"
                    onmouseover="this.style.background='#b89540'" onmouseout="this.style.background='#C9A84C'"
                    title="Open Excel for the Web at cell P{p.get('excel_row', '')} ({EXCEL_SHEET_NAME})">
                    <span style="font-size:0.85rem;">&#9998;</span> Update</a>
                <a href="#" class="attach-btn"
                    data-folder="{pid} {p['name']}"
                    style="display:inline-flex;align-items:center;gap:0.3rem;
                    font-size:0.75rem;color:#163A5F;text-decoration:none;font-weight:600;
                    padding:0.3rem 0.6rem;border:1px solid #ddd;border-radius:4px;
                    background:#fafafa;cursor:pointer;transition:background 0.2s;"
                    onmouseover="this.style.background='#e8e8e8'" onmouseout="this.style.background='#fafafa'"
                    title="Open SharePoint folder — use Upload or drag &amp; drop to add files">
                    <span style="font-size:0.85rem;">&#128206;</span> Attach{_att_badge(attachments, project_id=pid)}</a>
            </div>
        </div>
'''


def render_projects_grid_html(projects, update_log=None, attachments=None):
    """
    Generate static HTML for the projects grid with cards,
    grouped under Goal headers (Goal 1, Goal 2, Goal 3).
    Excludes subpopulation rows (D.x IDs).
    Each card has data attributes for JS filtering.
    Multi-goal projects appear under their primary (first) goal.
    """
    # Group projects by primary goal
    goal_groups = {}
    for p in projects:
        if p["id"].startswith("D."):
            continue
        goal_raw = p.get("goal", "")
        primary_goal = ""
        if goal_raw:
            gm = re.match(r'(Goal\s*\d+)', goal_raw)
            if gm:
                primary_goal = gm.group(1)
        if not primary_goal:
            primary_goal = "Other"
        goal_groups.setdefault(primary_goal, []).append(p)

    # Sort: Goal 1, Goal 2, Goal 3, then Other
    sorted_goals = sorted(goal_groups.keys(),
                          key=lambda g: (0, int(re.search(r'\d+', g).group())) if re.search(r'\d+', g) else (1, 0))

    html = ""
    for goal_key in sorted_goals:
        goal_projects = goal_groups[goal_key]
        goal_info = CPL_GOALS.get(goal_key, {})
        goal_title = goal_info.get("title", "")
        goal_target = goal_info.get("target", "")
        goal_bullets = goal_info.get("bullets", [])
        goal_stretch = goal_info.get("stretch", "")

        # Goal header bar with full workplan content
        if goal_key != "Other" and goal_title:
            bullets_html = ""
            if goal_bullets:
                bullets_html = '<ul style="margin:0.3rem 0 0 1.2rem;padding:0;list-style:disc;">'
                for b in goal_bullets:
                    bullets_html += f'<li style="color:rgba(255,255,255,0.8);font-size:0.75rem;margin-bottom:0.15rem;line-height:1.3;">{b}</li>'
                bullets_html += '</ul>'
            stretch_html = ""
            if goal_stretch:
                stretch_html = f'<div style="color:#C9A84C;font-size:0.7rem;margin-top:0.25rem;font-style:italic;">{goal_stretch}</div>'

            html += (f'        <div class="goal-header" data-goal-key="{goal_key}" style="margin:1.5rem 0 0.75rem 0;padding:0.75rem 1rem;'
                     f'background:linear-gradient(135deg,#163A5F 0%,#0A2240 100%);border-radius:8px;">\n'
                     f'            <div style="display:flex;align-items:center;gap:0.8rem;margin-bottom:0.3rem;">\n'
                     f'                <span style="color:#C9A84C;font-weight:700;font-size:1rem;white-space:nowrap;">{goal_key}</span>\n'
                     f'                <span style="color:#fff;font-size:0.95rem;font-weight:600;">{goal_title}</span>\n'
                     f'                <span style="color:rgba(255,255,255,0.5);font-size:0.8rem;margin-left:auto;white-space:nowrap;">'
                     f'{len(goal_projects)} projects</span>\n'
                     f'            </div>\n'
                     f'            <div style="color:rgba(255,255,255,0.75);font-size:0.78rem;line-height:1.3;">{goal_target}</div>\n'
                     f'            {bullets_html}\n'
                     f'            {stretch_html}\n'
                     f'        </div>\n')
        else:
            html += (f'        <div class="goal-header" data-goal-key="Other" style="margin:1.5rem 0 0.75rem 0;padding:0.6rem 1rem;'
                     f'background:#888;border-radius:8px;color:#fff;font-weight:700;">\n'
                     f'            Other Projects ({len(goal_projects)})\n'
                     f'        </div>\n')

        # Project cards grid for this goal
        html += '        <div class="projects-grid goal-project-group">\n'
        for p in goal_projects:
            html += _render_single_project_card(p, update_log, attachments=attachments)
        html += '        </div>\n'

    return html


def render_workplan_charts_html(current_students, sub_pops=None, workplan_goals=None, config_overrides=None):
    """
    Render two side-by-side canvas-based trend charts:
      Left:  Goal trajectory (250K target)
      Right: Stretch trajectory (500K target)
    Lines are solid for actual data (up to current year) and dashed for projections.
    Charts are rendered to <canvas> so users can right-click → Copy Image.
    """
    import json as _json

    def parse_int(v, fallback=0):
        try:
            return int(str(v).replace(",", ""))
        except (ValueError, TypeError):
            return fallback

    sub_pops = sub_pops or {}
    mil_now = parse_int(sub_pops.get("military", 21866))
    wf_now = parse_int(sub_pops.get("workforce", 21526))
    app_now = parse_int(sub_pops.get("apprentice", 681))
    total_now = parse_int(current_students, mil_now + wf_now + app_now)

    # 2024 baselines — read from Col AG config overrides, fallback to original values
    cfg = config_overrides or {}
    BASE_MIL = int(cfg.get("BASE_MIL", 8248))
    BASE_WF  = int(cfg.get("BASE_WF", 9181))
    BASE_APP = int(cfg.get("BASE_APP", 196))

    # ── Look up workplan goal entries by name ──
    mil_wg, wf_wg, app_wg = {}, {}, {}
    for wg in (workplan_goals or []):
        name_l = wg.get("name", "").lower()
        if "military" in name_l:
            mil_wg = wg
        elif "apprentice" in name_l:
            app_wg = wg
        elif "working" in name_l or "workforce" in name_l:
            wf_wg = wg

    def get_wg(wg_dict, def_goal, def_stretch, def_gt, def_st):
        return (
            wg_dict.get("goal", def_goal),
            wg_dict.get("stretch", def_stretch),
            int(wg_dict.get("goal_total", def_gt)),
            int(wg_dict.get("stretch_total", def_st)),
        )

    mil_goal, mil_str, mil_gt, mil_st = get_wg(mil_wg, [30000]*5, [30000]*5, 70000, 100000)
    wf_goal, wf_str, wf_gt, wf_st = get_wg(wf_wg, [5000]*5, [10000]*5, 150000, 320000)
    app_goal, app_str, app_gt, app_st = get_wg(app_wg, [500]*5, [1000]*5, 20000, 80000)

    def build_trajectory(baseline, annual_vals, endpoint):
        pts = [(2024, baseline)]
        if not annual_vals:
            return pts
        is_cumulative = all(annual_vals[i] >= annual_vals[i-1] for i in range(1, len(annual_vals)) if annual_vals[i] > 0)
        if is_cumulative and annual_vals[-1] > annual_vals[0]:
            pts.append((2025, int((baseline + annual_vals[0]) / 2)))
            for i, val in enumerate(annual_vals):
                pts.append((2026 + i, int(val)))
        else:
            annual_sum = sum(annual_vals) if annual_vals else 1
            delta = endpoint - baseline
            pts.append((2025, int(baseline + delta * (annual_vals[0] * 0.5) / annual_sum)))
            cumul_prop = 0
            for i, val in enumerate(annual_vals):
                cumul_prop += val / annual_sum
                pts.append((2026 + i, int(baseline + delta * cumul_prop)))
        return pts

    # Build GOAL trajectories
    mil_goal_traj = build_trajectory(BASE_MIL, mil_goal, mil_gt)
    wf_goal_traj = build_trajectory(BASE_WF, wf_goal, wf_gt)
    app_goal_traj = build_trajectory(BASE_APP, app_goal, app_gt)

    # Build STRETCH trajectories
    mil_str_traj = build_trajectory(BASE_MIL, mil_str, mil_st)
    wf_str_traj = build_trajectory(BASE_WF, wf_str, wf_st)
    app_str_traj = build_trajectory(BASE_APP, app_str, app_st)

    # Actuals up to current year — read 2025 values from config overrides if available
    act_2025_mil   = int(cfg.get("ACTUAL_2025_MIL", 18500))
    act_2025_wf    = int(cfg.get("ACTUAL_2025_WF", 19200))
    act_2025_app   = int(cfg.get("ACTUAL_2025_APP", 300))
    act_2025_total = int(cfg.get("ACTUAL_2025_TOTAL", act_2025_mil + act_2025_wf + act_2025_app))
    mil_actual = [[2024, BASE_MIL], [2025, act_2025_mil], [2026, mil_now]]
    wf_actual = [[2024, BASE_WF], [2025, act_2025_wf], [2026, wf_now]]
    app_actual = [[2024, BASE_APP], [2025, act_2025_app], [2026, app_now]]
    total_actual = [[2024, BASE_MIL + BASE_WF + BASE_APP], [2025, act_2025_total], [2026, total_now]]

    def val_at(series, yr):
        for y, v in series:
            if y == yr:
                return v
        return 0

    def merge_series(actual, proj):
        merged = [list(a) for a in actual]
        last_yr = actual[-1][0]
        for yr, val in proj:
            if yr > last_yr:
                merged.append([yr, val])
        return merged

    all_years = list(range(2024, 2031))

    # Total goal / stretch trajectories
    total_goal_traj = [[yr, val_at(mil_goal_traj, yr) + val_at(wf_goal_traj, yr) + val_at(app_goal_traj, yr)] for yr in all_years]
    total_str_traj = [[yr, val_at(mil_str_traj, yr) + val_at(wf_str_traj, yr) + val_at(app_str_traj, yr)] for yr in all_years]

    # Full series for each (actual + projected)
    mil_goal_full = merge_series(mil_actual, mil_goal_traj)
    wf_goal_full = merge_series(wf_actual, wf_goal_traj)
    app_goal_full = merge_series(app_actual, app_goal_traj)
    total_goal_full = merge_series(total_actual, total_goal_traj)

    mil_str_full = merge_series(mil_actual, mil_str_traj)
    wf_str_full = merge_series(wf_actual, wf_str_traj)
    app_str_full = merge_series(app_actual, app_str_traj)
    total_str_full = merge_series(total_actual, total_str_traj)

    # Compute totals for targets
    goal_total = mil_gt + wf_gt + app_gt
    stretch_total = mil_st + wf_st + app_st

    # Serialize data for JS
    chart_data = {
        "goal": {
            "total": total_goal_full, "military": mil_goal_full,
            "workforce": wf_goal_full, "apprentice": app_goal_full,
            "target": goal_total,
            "mil_target": mil_gt, "wf_target": wf_gt, "app_target": app_gt,
        },
        "stretch": {
            "total": total_str_full, "military": mil_str_full,
            "workforce": wf_str_full, "apprentice": app_str_full,
            "target": stretch_total,
            "mil_target": mil_st, "wf_target": wf_st, "app_target": app_st,
        },
        "currentYear": 2026,
    }

    data_json = _json.dumps(chart_data)

    html = f'''        <div style="margin:2rem 0;padding:1.5rem;background:#fff;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,0.06);">
            <h3 style="color:#0A2240;margin:0 0 0.25rem 0;font-size:1.1rem;">CPL Workplan Progress — Path to 2030</h3>
            <p style="color:#888;font-size:0.8rem;margin:0 0 1rem 0;">Solid lines = actuals &middot; Dashed lines = projected &middot; Right-click any chart to copy image</p>
            <div style="display:flex;gap:1.5rem;flex-wrap:wrap;">
                <div style="flex:1;min-width:340px;">
                    <h4 style="color:#0A2240;font-size:0.9rem;margin:0 0 0.5rem 0;text-align:center;">Goal Trajectory (250K Target)</h4>
                    <canvas id="goalChart" width="640" height="400" style="width:100%;height:auto;border-radius:6px;background:#FAF8F4;"></canvas>
                </div>
                <div style="flex:1;min-width:340px;">
                    <h4 style="color:#0A2240;font-size:0.9rem;margin:0 0 0.5rem 0;text-align:center;">Stretch Trajectory (500K Target)</h4>
                    <canvas id="stretchChart" width="640" height="400" style="width:100%;height:auto;border-radius:6px;background:#FAF8F4;"></canvas>
                </div>
            </div>
            <div style="display:flex;flex-wrap:wrap;gap:0.8rem 1.5rem;margin-top:0.8rem;font-size:0.8rem;color:#555;align-items:center;">
                <span><span style="display:inline-block;width:20px;height:3px;background:#0A2240;vertical-align:middle;margin-right:4px;"></span> Total</span>
                <span><span style="display:inline-block;width:20px;height:3px;background:#C9A84C;vertical-align:middle;margin-right:4px;"></span> Military</span>
                <span><span style="display:inline-block;width:20px;height:3px;background:#4A90D9;vertical-align:middle;margin-right:4px;"></span> Workforce/Other</span>
                <span><span style="display:inline-block;width:20px;height:3px;background:#2A7D4F;vertical-align:middle;margin-right:4px;"></span> Apprentice</span>
                <span style="color:#aaa;">&mdash; Solid = actual &nbsp; - - - Dashed = projected</span>
            </div>
        </div>
        <script>
        (function() {{
            var D = {data_json};
            var COLORS = {{total:'#0A2240',military:'#C9A84C',workforce:'#4A90D9',apprentice:'#2A7D4F'}};
            var SERIES = ['total','military','workforce','apprentice'];
            var LABELS = {{total:'Total',military:'Military',workforce:'Workforce',apprentice:'Apprentice'}};

            function drawChart(canvasId, data, targetVal) {{
                var canvas = document.getElementById(canvasId);
                if (!canvas) return;
                var dpr = window.devicePixelRatio || 1;
                var W = 640, H = 400;
                canvas.width = W * dpr;
                canvas.height = H * dpr;
                canvas.style.width = '100%';
                canvas.style.height = 'auto';
                var ctx = canvas.getContext('2d');
                ctx.scale(dpr, dpr);

                var PAD_L = 55, PAD_R = 20, PAD_T = 20, PAD_B = 35;
                var CW = W - PAD_L - PAD_R;
                var CH = H - PAD_T - PAD_B;
                var START_YR = 2024, END_YR = 2030, SPAN = 6;
                var maxVal = Math.max(targetVal * 1.15, 300000);

                // Round maxVal up to a nice number
                var step = maxVal > 300000 ? 100000 : 50000;
                maxVal = Math.ceil(maxVal / step) * step;

                function xPos(yr) {{ return PAD_L + (yr - START_YR) / SPAN * CW; }}
                function yPos(val) {{ return PAD_T + CH - (val / maxVal * CH); }}
                function fmtK(v) {{
                    if (v >= 1000) {{
                        var k = v / 1000;
                        return (k === Math.floor(k)) ? k + 'K' : k.toFixed(1) + 'K';
                    }}
                    return '' + v;
                }}

                // Background
                ctx.fillStyle = '#FAF8F4';
                ctx.fillRect(0, 0, W, H);

                // Grid lines
                ctx.strokeStyle = '#e8e8e8';
                ctx.lineWidth = 0.7;
                ctx.font = '10px Calibri, sans-serif';
                ctx.fillStyle = '#999';
                ctx.textAlign = 'right';
                for (var tick = 0; tick <= maxVal; tick += step) {{
                    var yy = yPos(tick);
                    ctx.beginPath();
                    ctx.moveTo(PAD_L, yy);
                    ctx.lineTo(PAD_L + CW, yy);
                    ctx.stroke();
                    ctx.fillText(fmtK(tick), PAD_L - 6, yy + 3);
                }}

                // Year labels
                ctx.textAlign = 'center';
                ctx.fillStyle = '#888';
                for (var yr = 2024; yr <= 2030; yr++) {{
                    ctx.fillText('' + yr, xPos(yr), H - 10);
                }}

                // Target ceiling line
                ctx.strokeStyle = '#0A2240';
                ctx.lineWidth = 1.5;
                ctx.setLineDash([8, 4]);
                ctx.globalAlpha = 0.3;
                ctx.beginPath();
                ctx.moveTo(PAD_L, yPos(targetVal));
                ctx.lineTo(PAD_L + CW, yPos(targetVal));
                ctx.stroke();
                ctx.setLineDash([]);
                ctx.globalAlpha = 1;
                ctx.fillStyle = '#0A2240';
                ctx.textAlign = 'right';
                ctx.font = 'bold 10px Calibri, sans-serif';
                ctx.fillText(fmtK(targetVal) + ' Target', PAD_L + CW - 4, yPos(targetVal) - 5);

                // "Now" vertical marker
                ctx.strokeStyle = '#0A2240';
                ctx.lineWidth = 1;
                ctx.setLineDash([3, 3]);
                ctx.globalAlpha = 0.15;
                ctx.beginPath();
                ctx.moveTo(xPos(D.currentYear), PAD_T);
                ctx.lineTo(xPos(D.currentYear), PAD_T + CH);
                ctx.stroke();
                ctx.setLineDash([]);
                ctx.globalAlpha = 1;
                ctx.font = '9px Calibri, sans-serif';
                ctx.fillStyle = 'rgba(10,34,64,0.5)';
                ctx.textAlign = 'center';
                ctx.fillText('Now', xPos(D.currentYear), PAD_T - 4);

                // Draw each series
                for (var si = 0; si < SERIES.length; si++) {{
                    var key = SERIES[si];
                    var pts = data[key];
                    var color = COLORS[key];
                    var lw = key === 'total' ? 3 : 2.2;

                    if (!pts || pts.length < 2) continue;

                    // Split into actual (up to currentYear) and projected (after)
                    var actualPts = [];
                    var projPts = [];
                    for (var i = 0; i < pts.length; i++) {{
                        if (pts[i][0] <= D.currentYear) {{
                            actualPts.push(pts[i]);
                        }}
                        if (pts[i][0] >= D.currentYear) {{
                            projPts.push(pts[i]);
                        }}
                    }}

                    // Draw actual (solid)
                    if (actualPts.length >= 2) {{
                        ctx.strokeStyle = color;
                        ctx.lineWidth = lw;
                        ctx.setLineDash([]);
                        ctx.beginPath();
                        ctx.moveTo(xPos(actualPts[0][0]), yPos(actualPts[0][1]));
                        for (var i = 1; i < actualPts.length; i++) {{
                            ctx.lineTo(xPos(actualPts[i][0]), yPos(actualPts[i][1]));
                        }}
                        ctx.stroke();
                    }}

                    // Draw projected (dashed)
                    if (projPts.length >= 2) {{
                        ctx.strokeStyle = color;
                        ctx.lineWidth = lw;
                        ctx.setLineDash([8, 5]);
                        ctx.beginPath();
                        ctx.moveTo(xPos(projPts[0][0]), yPos(projPts[0][1]));
                        for (var i = 1; i < projPts.length; i++) {{
                            ctx.lineTo(xPos(projPts[i][0]), yPos(projPts[i][1]));
                        }}
                        ctx.stroke();
                        ctx.setLineDash([]);
                    }}

                    // Data dots
                    var dotR = key === 'total' ? 4 : 3;
                    for (var i = 0; i < pts.length; i++) {{
                        ctx.fillStyle = color;
                        ctx.beginPath();
                        ctx.arc(xPos(pts[i][0]), yPos(pts[i][1]), dotR, 0, Math.PI * 2);
                        ctx.fill();
                    }}

                    // Value labels from currentYear onward
                    ctx.font = 'bold 9px Calibri, sans-serif';
                    ctx.fillStyle = color;
                    ctx.textAlign = 'center';
                    for (var i = 0; i < pts.length; i++) {{
                        if (pts[i][0] >= D.currentYear) {{
                            var dy = (key === 'total' || key === 'military') ? -10 : 14;
                            ctx.fillText(fmtK(pts[i][1]), xPos(pts[i][0]), yPos(pts[i][1]) + dy);
                        }}
                    }}

                    // Endpoint label at 2030
                    var lastPt = pts[pts.length - 1];
                    if (lastPt[0] === 2030) {{
                        ctx.font = 'bold 9.5px Calibri, sans-serif';
                        ctx.fillStyle = color;
                        ctx.textAlign = 'left';
                        // Only show label if total — sub-pop labels via value labels already
                    }}
                }}
            }}

            drawChart('goalChart', D.goal, D.goal.target);
            drawChart('stretchChart', D.stretch, D.stretch.target);
        }})();
        </script>
'''
    return html


def compute_headline_kpis(projects, budget, config_overrides=None, live_data=None):
    """
    Compute the 6 headline KPIs from sub-activity data.
    Falls back to hardcoded values if sub-activity data is missing.
    Col AG overrides take priority when present.
    """
    proj_map = {p["id"]: p for p in projects}
    cfg = config_overrides or {}

    def get_metric(pid, fallback=""):
        # Check Col AG override first, then KPI metric, then fallback
        p = proj_map.get(pid)
        if p and p.get("override") is not None:
            return str(p["override"])
        return p["kpi_metric"] if p and p["kpi_metric"] else fallback

    students   = get_metric("3.1", "42,620")
    units      = get_metric("3.2", "96,449")
    recs       = get_metric("2.1", "576")
    colleges   = get_metric("3.3", "84")

    # Estimated savings — read from config override or use static fallback
    savings = str(cfg.get("SAVINGS_DEFAULT", "$269M"))

    return {
        "cumulative_students": {
            "value": students,
            "label": "Cumulative CPL Students",
            "sub": "Target: 250,000 by 2030"
        },
        "transcripted_units": {
            "value": units,
            "label": "Transcripted Units",
            "sub": "Target: 85,000 (exceeded)"
        },
        "credit_recommendations": {
            "value": recs,
            "label": "Credit Recommendations",
            "sub": "281 adopted this year"
        },
        "active_colleges": {
            "value": colleges,
            "label": "Active Colleges",
            "sub": "of 116 system colleges"
        },
        "estimated_savings": {
            "value": savings,
            "label": "Estimated Savings",
            "sub": "Eligible units, Beacon Economics"
        },
        "veteran_sprint": (lambda: {
            "value": (lambda v: str(v) if v else "—")(int((live_data or {}).get("star_college_count", 0))),
            "label": "VETERAN SPRINT",
            "sub": "Star Colleges",
            "breakdowns": [
                {"label": "JST Credits", "value": "{jst_credits} / 30,000"},
                {"label": "Basic Training Credit",
                 "value": (lambda v: f'{v} Colleges' if v else '— Colleges')(int((live_data or {}).get("star_college_count", 0)))},
                {"label": "Eligible CPL", "value": "{eligible_cpl} Units"},
            ],
        })(),
    }


# ── KPI History Logging & Trend Card ────────────────────────────────

def _academic_quarter(dt):
    """Return academic quarter label for a datetime (Q1=Jul-Sep, Q2=Oct-Dec, Q3=Jan-Mar, Q4=Apr-Jun)."""
    m = dt.month
    if   m in (7, 8, 9):   return f"Q1 {dt.year}"
    elif m in (10, 11, 12): return f"Q2 {dt.year}"
    elif m in (1, 2, 3):    return f"Q3 {dt.year}"
    else:                   return f"Q4 {dt.year}"

def _quarter_start(dt):
    """Return the first day of the academic quarter containing dt."""
    m = dt.month
    if   m in (7, 8, 9):   return dt.replace(month=7,  day=1)
    elif m in (10, 11, 12): return dt.replace(month=10, day=1)
    elif m in (1, 2, 3):    return dt.replace(month=1,  day=1)
    else:                   return dt.replace(month=4,  day=1)


def log_daily_snapshot(live_data, exhibit_data):
    """Append today's metric snapshot to kpi_history.json (idempotent — overwrites same-day entry).
    Returns the full sorted history list.
    """
    today = _now_pt().strftime("%Y-%m-%d")
    raw = (live_data or {}).get("raw", {})
    tiers = (live_data or {}).get("tiers", {})

    snapshot = {
        "date":                  today,
        # Scraped live metrics
        "students":              int(raw.get("Students", 0)),
        "students_military":     int(raw.get("MilitaryStudents", 0)),
        "students_workforce":    int(raw.get("NonMilitaryStudents", 0)),
        "students_apprentice":   int(raw.get("AprenticeStudents", 0)),   # API typo preserved
        "eligible_units":        int(raw.get("Units", 0)),
        "transcribed_units":     int(raw.get("TranscribedUnits", 0)),
        "savings_m":             round(raw.get("Savings", 0) / 1_000_000, 1),
        "year_impact_b":         round(raw.get("YearImpact", 0) / 1_000_000_000, 2),
        "active_colleges":       int((live_data or {}).get("active_college_count", 0)),
        "leading_colleges":      int(tiers.get("leading", {}).get("count", 0)),
        "star_colleges":         int((live_data or {}).get("star_college_count", 0)),
        # Exhibit / CustomReport metrics
        "credit_recs":           int((exhibit_data or {}).get("total_credit_recs", 0)),
        "map_exhibits":          int((exhibit_data or {}).get("unique_exhibits", 0)),
        "ccc_collaborative":     int((exhibit_data or {}).get("ccc_collaborative", {}).get("adopting_colleges", 0)),
        "articulating_colleges": int((exhibit_data or {}).get("articulation_colleges", 0)),
    }

    history = []
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                history = json.load(f)
        except (json.JSONDecodeError, IOError):
            history = []

    history = [e for e in history if e.get("date") != today]
    history.append(snapshot)
    history.sort(key=lambda e: e.get("date", ""))

    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, indent=2, ensure_ascii=False)

    print(f"  Logged daily KPI snapshot ({len(history)} total entries in kpi_history.json)")
    return history


def _sparkline_svg(values, width=110, height=30, color="#C9A84C"):
    """Return an inline SVG sparkline for a list of numeric values."""
    vals = [v for v in values if v is not None and v > 0]
    if len(vals) < 2:
        return f'<svg width="{width}" height="{height}"></svg>'
    mn, mx = min(vals), max(vals)
    rng = mx - mn if mx != mn else max(mx, 1)
    pad = 4
    pts = [
        (pad + i * (width - 2*pad) / (len(vals) - 1),
         pad + (1 - (v - mn) / rng) * (height - 2*pad))
        for i, v in enumerate(vals)
    ]
    line = "M " + " L ".join(f"{x:.1f},{y:.1f}" for x, y in pts)
    area_pts = [(pad, height - pad)] + pts + [(width - pad, height - pad)]
    area = "M " + " L ".join(f"{x:.1f},{y:.1f}" for x, y in area_pts) + " Z"
    cx, cy = pts[-1]
    return (
        f'<svg width="{width}" height="{height}" style="display:block;overflow:visible">'
        f'<defs><linearGradient id="sg" x1="0" y1="0" x2="0" y2="1">'
        f'<stop offset="0%" stop-color="{color}" stop-opacity="0.25"/>'
        f'<stop offset="100%" stop-color="{color}" stop-opacity="0.02"/>'
        f'</linearGradient></defs>'
        f'<path d="{area}" fill="url(#sg)" stroke="none"/>'
        f'<path d="{line}" fill="none" stroke="{color}" stroke-width="1.8" '
        f'stroke-linejoin="round" stroke-linecap="round"/>'
        f'<circle cx="{cx:.1f}" cy="{cy:.1f}" r="3" fill="{color}"/>'
        f'</svg>'
    )


def _history_lookup(history, key, days_ago=None, date_str=None):
    """Return the value of `key` from `days_ago` days before today, or from a specific date."""
    if date_str is None:
        target = (_now_pt() - timedelta(days=days_ago)).strftime("%Y-%m-%d")
    else:
        target = date_str
    candidates = [e for e in history if e.get("date", "") <= target and e.get(key, 0) > 0]
    if not candidates:
        return None
    return max(candidates, key=lambda e: e["date"]).get(key)


def _delta_badge(current, previous, abs_format=False):
    """Return (html_badge, sort_key) showing change from previous to current."""
    if previous is None or previous == 0 or current == 0:
        return '<span style="color:rgba(255,255,255,0.35);font-size:0.68rem;">—</span>', 0
    diff = current - previous
    pct  = diff / previous * 100
    if abs(pct) < 0.05:
        return '<span style="color:rgba(255,255,255,0.35);font-size:0.68rem;">—</span>', 0
    color = "#4CAF50" if diff >= 0 else "#f44336"
    arrow = "▲" if diff >= 0 else "▼"
    if abs_format:
        label = f"{arrow}{abs(diff):,.0f}"
    else:
        label = f"{arrow}{abs(pct):.1f}%"
    badge = (f'<span style="color:{color};font-size:0.68rem;font-weight:700;'
             f'white-space:nowrap;">{label}</span>')
    return badge, pct


def render_kpi_history_card(history, kpi_params=None):
    """Render a full-width KPI Trends card from kpi_history.json data."""
    if not history:
        return ""

    today_entry = history[-1]
    today_str   = today_entry.get("date", "")
    today_dt    = datetime.strptime(today_str, "%Y-%m-%d") if today_str else _now_pt()

    # Academic quarter start date for current quarter
    q_start = _quarter_start(today_dt).strftime("%Y-%m-%d")

    # Periods: label → days_ago (None means use date_str for quarter)
    PERIODS = [
        ("1d",  1,   None),
        ("7d",  7,   None),
        ("30d", 30,  None),
        ("QTD", None, q_start),
        ("1yr", 365, None),
    ]

    # Metrics to display: (key, display_label, format_fn, abs_delta?)
    def _fmt_students(v): return f"{v:,}"
    def _fmt_units(v):    return f"{int(v/1000)}k" if v >= 1000 else str(v)
    def _fmt_savings(v):  return f"${v:.0f}M"
    def _fmt_int(v):      return f"{v:,}"
    def _fmt_recs(v):     return f"{v:,}"

    def _fmt_sub(v): return f"  {v:,}"   # indented to signal sub-row

    METRICS = [
        ("students",              "CPL Students",           _fmt_students, False),
        ("students_military",     "↳ Military",             _fmt_sub,      False),
        ("students_workforce",    "↳ Workforce / Other",    _fmt_sub,      False),
        ("students_apprentice",   "↳ Apprentice",           _fmt_sub,      False),
        ("credit_recs",           "Credit Recs",            _fmt_recs,    False),
        ("active_colleges",       "Active Colleges",        _fmt_int,      True),
        ("savings_m",             "Est. Savings",           _fmt_savings,  False),
        ("map_exhibits",          "MAP Exhibits",           _fmt_int,      True),
        ("ccc_collaborative",     "CCC Collaborative",      _fmt_int,      True),
        ("articulating_colleges", "Articulating Colleges",  _fmt_int,      True),
        ("star_colleges",         "Veteran Star Colleges",  _fmt_int,      True),
    ]

    # Build sparkline values (last 30 entries) for each metric
    recent = history[-30:]

    # Header row
    period_headers = "".join(
        f'<th style="font-size:0.65rem;color:rgba(255,255,255,0.5);font-weight:600;'
        f'text-transform:uppercase;padding:0.2rem 0.5rem;text-align:center;'
        f'white-space:nowrap;">{p[0]}</th>'
        for p in PERIODS
    )
    header = (
        f'<tr>'
        f'<th style="font-size:0.65rem;color:rgba(255,255,255,0.5);font-weight:600;'
        f'text-transform:uppercase;padding:0.2rem 0.5rem;text-align:left;">Metric</th>'
        f'<th style="font-size:0.65rem;color:rgba(255,255,255,0.5);font-weight:600;'
        f'text-transform:uppercase;padding:0.2rem 0.5rem;text-align:right;">Today</th>'
        f'{period_headers}'
        f'<th style="font-size:0.65rem;color:rgba(255,255,255,0.5);font-weight:600;'
        f'text-transform:uppercase;padding:0.2rem 0.5rem;text-align:center;">30-Day Trend</th>'
        f'</tr>'
    )

    rows_html = ""
    for key, label, fmt_fn, use_abs in METRICS:
        current = today_entry.get(key, 0)
        if current == 0:
            continue
        current_str = fmt_fn(current)

        is_sub = label.startswith("↳")
        row_style    = "border-top:1px solid rgba(255,255,255,0.04);" if is_sub else "border-top:1px solid rgba(255,255,255,0.06);"
        label_style  = (f"padding:0.15rem 0.5rem 0.15rem 1.2rem;font-size:0.68rem;"
                        f"color:rgba(255,255,255,0.5);white-space:nowrap;{row_style}") if is_sub else \
                       (f"padding:0.3rem 0.5rem;font-size:0.75rem;"
                        f"color:rgba(255,255,255,0.8);white-space:nowrap;{row_style}")
        value_style  = (f"padding:0.15rem 0.5rem;font-size:0.68rem;font-weight:600;"
                        f"color:rgba(201,168,76,0.65);text-align:right;white-space:nowrap;{row_style}") if is_sub else \
                       (f"padding:0.3rem 0.5rem;font-size:0.8rem;font-weight:700;"
                        f"color:#C9A84C;text-align:right;white-space:nowrap;{row_style}")
        cell_style   = f"padding:{'0.15rem' if is_sub else '0.3rem'} 0.5rem;text-align:center;{row_style}"
        spark_style  = f"padding:{'0.15rem' if is_sub else '0.3rem'} 0.8rem;{row_style}"

        # Delta badges for each period
        delta_cells = ""
        for _, days_ago, date_str in PERIODS:
            prev = _history_lookup(history, key, days_ago=days_ago, date_str=date_str)
            badge, _ = _delta_badge(current, prev, abs_format=use_abs)
            if is_sub:
                badge = badge.replace("font-size:0.68rem", "font-size:0.62rem")
            delta_cells += f'<td style="{cell_style}">{badge}</td>'

        # Sparkline (smaller for sub-rows)
        spark_vals = [e.get(key, 0) for e in recent]
        spark_color = "rgba(201,168,76,0.5)" if is_sub else "#C9A84C"
        spark_svg   = _sparkline_svg(spark_vals, height=20 if is_sub else 30, color=spark_color)

        rows_html += (
            f'<tr>'
            f'<td style="{label_style}">{label}</td>'
            f'<td style="{value_style}">{current_str}</td>'
            f'{delta_cells}'
            f'<td style="{spark_style}">'
            f'{spark_svg}</td>'
            f'</tr>'
        )

    n_entries = len(history)
    first_date = history[0].get("date", "") if history else ""
    since_note = f"Tracking since {first_date} &nbsp;·&nbsp; {n_entries} daily snapshot{'s' if n_entries != 1 else ''}" if first_date else ""
    aq_label   = _academic_quarter(today_dt)

    html = f"""
    <div style="grid-column:1/-1;background:var(--navy-primary);border-radius:8px;
                border-top:4px solid var(--gold-accent);padding:1.2rem 1.5rem;
                box-shadow:0 2px 6px rgba(0,0,0,0.15);">
      <div style="display:flex;align-items:baseline;justify-content:space-between;
                  margin-bottom:0.8rem;">
        <div>
          <span style="font-family:Georgia,serif;font-size:1rem;font-weight:bold;
                       color:var(--gold-accent);">&#128200; KPI Trends</span>
          <span style="font-size:0.7rem;color:rgba(255,255,255,0.45);margin-left:0.8rem;">
            Academic {aq_label} &nbsp;·&nbsp; QTD resets each academic quarter (Jul/Oct/Jan/Apr)
          </span>
        </div>
        <span style="font-size:0.65rem;color:rgba(255,255,255,0.35);">{since_note}</span>
      </div>
      <div style="overflow-x:auto;">
        <table style="width:100%;border-collapse:collapse;">
          <thead>{header}</thead>
          <tbody>{rows_html}</tbody>
        </table>
      </div>
      {render_algo_details("kpi_trends", params=kpi_params)}
    </div>
"""
    return html


_COLLEGE_DISTRICT_LOOKUP = None
_COLLEGE_ACTIVITY_TEMPLATE = None


def _load_college_district_lookup():
    """Parse college_lookup.js into a {college_name: district} dict (cached).
    Returns an empty dict if the file is missing or unparseable.
    """
    global _COLLEGE_DISTRICT_LOOKUP
    if _COLLEGE_DISTRICT_LOOKUP is not None:
        return _COLLEGE_DISTRICT_LOOKUP

    _COLLEGE_DISTRICT_LOOKUP = {}
    path = os.path.join(SCRIPT_DIR, "college_lookup.js")
    if not os.path.isfile(path):
        return _COLLEGE_DISTRICT_LOOKUP

    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
    except IOError:
        return _COLLEGE_DISTRICT_LOOKUP

    # Match:  "College Name": { district: "District Name", swRegion: "..." }
    pattern = re.compile(
        r'"([^"]+)"\s*:\s*\{\s*district\s*:\s*"([^"]+)"',
        re.MULTILINE,
    )
    for college, district in pattern.findall(content):
        _COLLEGE_DISTRICT_LOOKUP[college] = district

    return _COLLEGE_DISTRICT_LOOKUP


def _load_college_activity_template():
    """Read college_activity_template.html (cached)."""
    global _COLLEGE_ACTIVITY_TEMPLATE
    if _COLLEGE_ACTIVITY_TEMPLATE is not None:
        return _COLLEGE_ACTIVITY_TEMPLATE
    path = os.path.join(SCRIPT_DIR, "college_activity_template.html")
    try:
        with open(path, "r", encoding="utf-8") as f:
            _COLLEGE_ACTIVITY_TEMPLATE = f.read()
    except IOError:
        _COLLEGE_ACTIVITY_TEMPLATE = ""
    return _COLLEGE_ACTIVITY_TEMPLATE


def render_college_activity_card(live_data, last_activity=None, military_students=None,
                                 exhibit_tables=None, kpi_params=None):
    """Render the full-width College Activity card — rich single-table layout.

    Emits:
      - HTML shell (header, filter bar, table, legend) read from
        college_activity_template.html
      - <script> block with window.COLLEGE_ACTIVITY_DATA and
        window.COLLEGE_DISCIPLINE_DETAIL populated from live_data + exhibit_tables
      - <script src="college_activity.js"></script> to run the interactive logic

    Parameters:
      live_data:         parsed live_metrics.json (tiers + per-college metrics)
      last_activity:     {college_name: datetime} from CustomReport
      military_students: {college_name: int} from CustomReport (used as veterans fallback)
      exhibit_tables:    output of build_exhibit_analysis_tables() — supplies
                         per-college exhibit counts and discipline detail
    """
    if not live_data:
        return ""

    tiers     = live_data.get("tiers", {})
    leading   = tiers.get("leading",   {}).get("colleges", [])
    advancing = tiers.get("advancing", {}).get("colleges", [])
    inactive  = tiers.get("inactive",  {}).get("colleges", [])
    # Worker may return inactive as [str, ...] or [{"college": str, ...}, ...];
    # we also accept full dicts with metrics when present.
    inactive_objs = []
    for c in inactive:
        if isinstance(c, dict):
            inactive_objs.append(c)
        else:
            inactive_objs.append({"college": c})
    inactive = inactive_objs

    if not leading and not advancing and not inactive:
        return ""

    template = _load_college_activity_template()
    if not template:
        return ""  # Missing template file — fail closed rather than emit the wrong layout

    scraped_at        = live_data.get("scraped_at", "")[:10]
    district_lookup   = _load_college_district_lookup()
    la                = last_activity or {}
    ms                = military_students or {}
    today_dt          = _now_pt()

    # Per-college exhibit counts (from build_exhibit_analysis_tables output)
    by_college = {}
    college_discipline_detail = {}
    if exhibit_tables:
        for row in exhibit_tables.get("by_college", []):
            by_college[row["college"]] = {
                "exhibits":    row.get("exhibits", 0),
                "credit_recs": row.get("credit_recs", 0),
                "disciplines": row.get("disciplines", 0),
            }
        college_discipline_detail = exhibit_tables.get("college_discipline_detail", {})

    def _record(c, tier_name):
        name = c.get("college", "") if isinstance(c, dict) else str(c)
        if not isinstance(c, dict):
            c = {}
        students    = c.get("students", 0) or 0
        units       = c.get("units", 0) or 0
        transcribed = c.get("transcribedUnits", 0) or 0
        trans_rate  = c.get("transcriptionRate", 0) or 0
        criteria    = c.get("criteriaMetCount", 0) or 0
        military    = c.get("militaryStudents", ms.get(name, 0)) or 0
        non_mil     = c.get("nonMilitaryStudents", 0) or 0
        apprentice  = c.get("apprenticeStudents", 0) or 0
        savings     = c.get("savings", 0) or 0
        year_impact = c.get("yearImpact", 0) or 0

        ex = by_college.get(name, {"exhibits": 0, "credit_recs": 0, "disciplines": 0})

        last_dt = la.get(name)
        last_days = (today_dt - last_dt).days if last_dt else None

        return {
            "tier":               tier_name,
            "college":            name,
            "district":           district_lookup.get(name, ""),
            "students":           students,
            "veterans":           military,
            "working_adults":     non_mil,
            "apprentices":        apprentice,
            "eligible_units":     units,
            "transcribed_units":  transcribed,
            "exhibits":           ex["exhibits"],
            "credit_recs":        ex["credit_recs"],
            "disciplines":        ex["disciplines"],
            "savings":            savings,
            "year_impact":        year_impact,
            "trans_rate":         trans_rate,
            "criteria_met":       criteria,
            "last_activity_days": last_days,
        }

    all_data = (
        [_record(c, "Leading")   for c in leading] +
        [_record(c, "Advancing") for c in advancing] +
        [_record(c, "Inactive")  for c in inactive]
    )

    html = template.replace("__AS_OF_DATE__", scraped_at)

    # Inject collapsible algorithm description inside the card, before the
    # final closing </div>. The template ends with a legend div then the
    # outer card's </div>, so we insert right before that last boundary.
    algo_html = render_algo_details("college_activity", params=kpi_params)
    if algo_html:
        # Replace the final occurrence of "    </div>" (outer wrapper close)
        # with algo + same closing div.
        marker = "      </div>\n\n    </div>"
        if marker in html:
            html = html.replace(marker, f"      </div>\n      {algo_html}\n\n    </div>", 1)
        else:
            html = html.rstrip() + "\n" + algo_html + "\n"

    # Emit the data blob immediately before the external script so it loads first.
    data_script = (
        "    <script>\n"
        "    window.COLLEGE_ACTIVITY_DATA = " + json.dumps(all_data, ensure_ascii=False) + ";\n"
        "    window.COLLEGE_DISCIPLINE_DETAIL = " + json.dumps(college_discipline_detail, ensure_ascii=False) + ";\n"
        "    </script>\n"
        "    <script src=\"college_activity.js\"></script>\n"
    )

    return html + "\n" + data_script


def read_live_metrics():
    """
    Read live_metrics.json (written by the daily scraper agent).
    Returns the metrics dict or None if the file doesn't exist.
    Format: {
        "scraped_at": "2026-04-01T09:00:00",
        "metrics": [
            {"title": "STUDENTS SERVED", "value": "43,321",
             "breakdowns": [{"label":"Military","value":"21,866"}, ...]},
            ...
        ]
    }
    """
    if not os.path.exists(LIVE_FILE):
        return None
    try:
        with open(LIVE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"WARNING: Could not read {LIVE_FILE}: {e}")
        return None


def merge_live_metrics(kpis, live_data):
    """
    Merge live dashboard metrics into the headline KPIs,
    adding population breakdowns.
    """
    if not live_data or "metrics" not in live_data:
        return kpis

    # Map live metric titles to our KPI keys
    title_map = {
        "STUDENTS SERVED": "cumulative_students",
        "ELIGIBLE UNITS": "eligible_units",
        "TRANSCRIBED UNITS": "transcripted_units",
        "SAVINGS": "estimated_savings",
        "20-YEAR IMPACT": "twenty_year_impact",
        "ACTIVE COLLEGES": "active_colleges",
    }

    for metric in live_data["metrics"]:
        title = metric.get("title", "").upper().strip()
        key = title_map.get(title)
        if not key:
            continue

        breakdowns = []
        for b in metric.get("breakdowns", []):
            bd_entry = {"label": b["label"], "value": b["value"]}
            if b.get("note"):
                bd_entry["note"] = b["note"]
            breakdowns.append(bd_entry)

        if key in kpis:
            # Update existing KPI with live value + breakdowns
            kpis[key]["value"] = metric["value"]
            kpis[key]["breakdowns"] = breakdowns
            kpis[key]["live"] = True
        else:
            # Add new KPI (eligible_units, twenty_year_impact, active_colleges)
            kpis[key] = {
                "value": metric["value"],
                "label": metric.get("title", title.title()),
                "sub": f"Source: MAP CPL Dashboard",
                "breakdowns": breakdowns,
                "live": True,
            }

        # Preserve footnote (e.g. Active Colleges criteria list)
        if metric.get("footnote"):
            kpis[key]["footnote"] = metric["footnote"]

    # Update Veteran Sprint card from live scraped values
    if "veteran_sprint" in kpis:
        vs_bds = kpis["veteran_sprint"].get("breakdowns", [])
        # JST Credits = Military students count
        if "cumulative_students" in kpis:
            mil_students = [b for b in kpis["cumulative_students"].get("breakdowns", [])
                            if b["label"].lower().startswith("military")]
            if mil_students:
                for bd in vs_bds:
                    if "{jst_credits}" in bd["value"]:
                        bd["value"] = bd["value"].replace("{jst_credits}", str(mil_students[0]["value"]))
        # Eligible CPL = Military eligible units
        if "eligible_units" in kpis:
            mil_units = [b for b in kpis["eligible_units"].get("breakdowns", [])
                         if b["label"].lower().startswith("military")]
            if mil_units:
                for bd in vs_bds:
                    if "{eligible_cpl}" in bd["value"]:
                        bd["value"] = bd["value"].replace("{eligible_cpl}", str(mil_units[0]["value"]))
        kpis["veteran_sprint"]["live"] = True

    kpis["_live_updated"] = live_data.get("scraped_at", "")
    return kpis


# ── MAP Exhibit Metrics ─────────────────────────────────────────────

def read_exhibit_metrics():
    """
    Read the combined CustomReport JSON exported from the MAP Custom Reporting Module.
    The file contains multiple datasets (views) in a single JSON array.

    Recognized datasets:
      - View_ArticulatedMAPExhibits_APIDataset      → exhibit-level credit recs + Collaborative Type
      - View_ArticulatedCollegeCourses_APIDataset    → course-level articulation detail
      - View_CreditDistributionByCollege_APIDataset  → per-college credit breakdown
      - View_StudentAggregatedValues_APIDataset      → student-level credit data
      - View_PointInTime_StudentAggregatedValues_APIDataset → student aggregates by year/type
      - View_CollegeCourses_APIDataset               → college course catalog
      - View_ProgramsofStudy_APIDataset              → programs of study

    Returns a dict of computed KPIs or None if file unavailable.
    """
    if not EXHIBIT_FILE or not os.path.exists(EXHIBIT_FILE):
        return None
    try:
        with open(EXHIBIT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"WARNING: Could not read exhibit file {EXHIBIT_FILE}: {e}")
        return None

    if not data or not isinstance(data, list):
        print("WARNING: Exhibit JSON has unexpected structure")
        return None

    # ── Index datasets by view name ──
    datasets = {}
    for report in data:
        view = report.get("viewName", "")
        if view and report.get("columnValue"):
            col_map = {c: i for i, c in enumerate(report.get("columnName", []))}
            datasets[view] = {
                "rows": report["columnValue"],
                "col_map": col_map,
                "generated_at": report.get("generatedAt", ""),
                "count": report.get("dataCount", len(report["columnValue"])),
            }
            print(f"  Dataset: {view} ({len(report['columnValue']):,} rows)")

    if not datasets:
        print("WARNING: No usable datasets in exhibit JSON")
        return None

    generated_at = data[0].get("generatedAt", "")

    # ── Parse exhibits from ArticulatedMAPExhibits (preferred — has Collaborative Type) ──
    exhibits_result = _parse_exhibits(datasets)

    # ── Parse credit distribution by college ──
    credit_dist = _parse_credit_distribution(datasets)

    # ── Compute last MAP activity per college ──
    last_activity = _compute_college_last_activity(datasets)

    # ── Compute per-college military/JST student counts ──
    military_students = _compute_college_military_students(datasets)

    # ── Combine into unified result ──
    result = exhibits_result or {}
    result["source_file"] = os.path.basename(EXHIBIT_FILE)
    result["generated_at"] = generated_at
    result["datasets_found"] = list(datasets.keys())
    result["college_last_activity"] = last_activity
    result["college_military_students"] = military_students

    if credit_dist:
        result["credit_distribution"] = credit_dist

    return result if result.get("total_credit_recs") else None


_TEST_COLLEGES = {"RivTest City College", "MorTest City College", "Nortest City College", "CA MAP INITIATIVE COLLEGE",
                  "RivTest", "MorTest", "Nortest"}

def _compute_college_last_activity(datasets):
    """Return {college_name: datetime} for the most recent student record upload
    per college. Uses View_StudentAggregatedValues_APIDataset's 'Uploaded Date'
    column (the newer CustomReport export replaced the old
    View_ArticulatedCollegeCourses_APIDataset which had 'Last Submitted On').
    Filters out test and potential-student records.
    Returns empty dict if the dataset is unavailable.
    """
    ds = datasets.get("View_StudentAggregatedValues_APIDataset")
    if not ds:
        # Backward compatibility: older CustomReport exports used a different dataset
        ds = datasets.get("View_ArticulatedCollegeCourses_APIDataset")
        if not ds:
            return {}
        rows  = ds["rows"]
        cm    = ds["col_map"]
        i_col  = cm.get("College", 0)
        i_date = cm.get("Last Submitted On", 18)
        i_pot  = -1
        i_test = -1
    else:
        rows   = ds["rows"]
        cm     = ds["col_map"]
        i_col  = cm.get("College", 0)
        i_date = cm.get("Uploaded Date", 22)
        i_pot  = cm.get("Potential Student", 18)
        i_test = cm.get("Test Student", 20)

    college_latest = {}  # college_name -> datetime
    for row in rows:
        college = (row[i_col] or "").strip()
        if not college or college in _TEST_COLLEGES:
            continue
        if i_pot >= 0 and str(row[i_pot]).strip().lower() in ("true", "1", "yes"):
            continue
        if i_test >= 0 and str(row[i_test]).strip().lower() in ("true", "1", "yes"):
            continue
        raw_date = (row[i_date] or "").strip()
        if not raw_date:
            continue
        # Accept either "M/D/YYYY h:mm:ss AM/PM" (new) or with seconds (old)
        dt = None
        for fmt in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y"):
            try:
                dt = datetime.strptime(raw_date, fmt)
                break
            except ValueError:
                continue
        if dt is None:
            continue
        if college not in college_latest or dt > college_latest[college]:
            college_latest[college] = dt

    return college_latest


def _compute_college_military_students(datasets):
    """Return {college_name: int} of military/JST student counts per college
    from View_StudentAggregatedValues_APIDataset.
    Counts distinct students (by MAP Internal StudentID when available, else rows)
    where Military Credits > 0. Filters test and potential students.
    Returns empty dict if dataset unavailable.
    """
    ds = datasets.get("View_StudentAggregatedValues_APIDataset")
    if not ds:
        return {}

    rows   = ds["rows"]
    cm     = ds["col_map"]
    i_col  = cm.get("College", 0)
    i_milcr = cm.get("Military Credits", 16)
    i_pot  = cm.get("Potential Student", 18)
    i_test = cm.get("Test Student", 20)
    i_sid  = cm.get("MAP Internal StudentID", 15)

    # Track (college, student_id) to avoid double-counting
    seen = set()
    counts = {}
    for row in rows:
        college = (row[i_col] or "").strip()
        if not college or college in _TEST_COLLEGES:
            continue
        test_s = (row[i_test] or "").strip().lower()
        pot_s  = (row[i_pot]  or "").strip().lower()
        if test_s == "yes" or pot_s == "yes":
            continue
        mil_cr_raw = (row[i_milcr] or "").strip()
        if not mil_cr_raw or mil_cr_raw == "0":
            continue
        try:
            if float(mil_cr_raw) <= 0:
                continue
        except ValueError:
            continue

        sid = (row[i_sid] or "").strip()
        key = (college, sid) if sid else (college, id(row))
        if key not in seen:
            seen.add(key)
            counts[college] = counts.get(college, 0) + 1

    return counts


def _parse_exhibits(datasets):
    """Extract exhibit metrics from available datasets.
    Prefers View_ArticulatedMAPExhibits (has Collaborative Type column).
    Falls back to View_ArticulatedCollegeCourses if needed.
    """
    # ── Primary source: ArticulatedMAPExhibits (has Collaborative Type) ──
    ds = datasets.get("View_ArticulatedMAPExhibits_APIDataset")
    if ds:
        rows = ds["rows"]
        cm = ds["col_map"]

        i_college = cm.get("College", 0)
        i_exhibit = cm.get("ExhibitID", 1)
        i_title = cm.get("Exhibit Title", 2)
        i_artic = cm.get("Articulation College", 4)
        i_collab = cm.get("Collaborative Type", 7)
        i_cpl = cm.get("CPL Type Description", 13)

        exhibit_ids = set()
        originating_colleges = set()
        articulation_colleges = set()
        ccc_credit_recs = 0
        ccc_exhibit_ids = set()
        ccc_artic_colleges = set()
        cpl_type_counts = {}

        # Grouped-exhibit counts matching the EACR table: keyed on
        # (unified_title, issuing_agency, CPL Type, Collaborative Type) once the
        # credential identity layer (kb/unified_titles.json + kb/credentials.json)
        # is consulted. MAP frequently issues multiple ExhibitIDs for what is
        # conceptually one exhibit AND colleges enter the same credential under
        # several freehand titles — both forms of fragmentation collapse here.
        # See docs/exhibit_unification_vision.md §6.1; was previously keyed on
        # (raw title, CPL Type, Collab Type) before the EACR Phase 4 re-pivot.
        unified_lookup, issuer_pick = _load_credential_kb()
        exhibit_groups = set()
        ccc_exhibit_groups = set()

        for row in rows:
            college = (row[i_college] or "").strip()
            exhibit_id = (row[i_exhibit] or "").strip()
            raw_title = row[i_title] or ""
            title = raw_title.strip()
            artic_college = (row[i_artic] or "").strip()
            collab_type = (row[i_collab] or "").strip()
            cpl_type = (row[i_cpl] or "").strip()

            exhibit_ids.add(exhibit_id)
            if title and exhibit_id:
                ident = _classify_exhibit(raw_title, unified_lookup, issuer_pick)
                grp = (ident["unified_title"], ident["issuing_agency"],
                       cpl_type, collab_type)
                exhibit_groups.add(grp)
                if "CCC" in collab_type:
                    ccc_exhibit_groups.add(grp)
            if college:
                originating_colleges.add(college)
            if artic_college:
                articulation_colleges.add(artic_college)

            if "CCC" in collab_type:
                ccc_credit_recs += 1
                ccc_exhibit_ids.add(exhibit_id)
                ccc_artic_colleges.add(artic_college)

            if cpl_type:
                cpl_type_counts[cpl_type] = cpl_type_counts.get(cpl_type, 0) + 1

        local_credit_recs = len(rows) - ccc_credit_recs
        local_exhibit_groups = exhibit_groups - ccc_exhibit_groups

        return {
            "total_credit_recs": len(rows),
            "unique_exhibits": len(exhibit_groups),
            "unique_exhibits_raw_ids": len(exhibit_ids),
            "originating_colleges": len(originating_colleges),
            "articulation_colleges": len(articulation_colleges),
            "articulation_college_names": sorted(articulation_colleges),
            "ccc_collaborative": {
                "credit_recs": ccc_credit_recs,
                "unique_exhibits": len(ccc_exhibit_groups),
                "unique_exhibits_raw_ids": len(ccc_exhibit_ids),
                "adopting_colleges": len(ccc_artic_colleges),
                "college_names": sorted(ccc_artic_colleges),
            },
            "local": {
                "credit_recs": local_credit_recs,
                "unique_exhibits": len(local_exhibit_groups),
                "unique_exhibits_raw_ids": len(exhibit_ids - ccc_exhibit_ids),
            },
            "cpl_type_breakdown": cpl_type_counts,
        }

    # ── Fallback: ArticulatedCollegeCourses (no Collaborative Type) ──
    ds = datasets.get("View_ArticulatedCollegeCourses_APIDataset")
    if ds:
        rows = ds["rows"]
        cm = ds["col_map"]

        i_college = cm.get("College", 0)
        i_exhibit = cm.get("ExhibitID", 15)
        i_cpl = cm.get("CPL Type Description", 11)

        exhibit_ids = set()
        colleges = set()
        cpl_type_counts = {}

        for row in rows:
            college = (row[i_college] or "").strip()
            exhibit_id = (row[i_exhibit] or "").strip()
            cpl_type = (row[i_cpl] or "").strip()
            if exhibit_id:
                exhibit_ids.add(exhibit_id)
            if college:
                colleges.add(college)
            if cpl_type:
                cpl_type_counts[cpl_type] = cpl_type_counts.get(cpl_type, 0) + 1

        return {
            "total_credit_recs": len(rows),
            "unique_exhibits": len(exhibit_ids),
            "originating_colleges": len(colleges),
            "articulation_colleges": len(colleges),
            "articulation_college_names": sorted(colleges),
            "ccc_collaborative": {
                "credit_recs": 0,
                "unique_exhibits": 0,
                "adopting_colleges": 0,
                "college_names": [],
            },
            "local": {
                "credit_recs": len(rows),
                "unique_exhibits": len(exhibit_ids),
            },
            "cpl_type_breakdown": cpl_type_counts,
        }

    return None


def _parse_credit_distribution(datasets):
    """Parse View_CreditDistributionByCollege for per-college credit metrics.
    Returns a dict with system totals and per-college detail, or None.
    """
    ds = datasets.get("View_CreditDistributionByCollege_APIDataset")
    if not ds:
        return None

    rows = ds["rows"]
    cm = ds["col_map"]

    i_college = cm.get("College", 0)
    i_eligible = cm.get("Eligible Credits", 7)
    i_transcribed = cm.get("Transcribed Credits", 9)
    i_applied = cm.get("Applied Credits", 2)
    i_students = cm.get("Students Awarded", 8)

    total_eligible = 0
    total_transcribed = 0
    total_applied = 0
    total_students = 0
    per_college = []

    for row in rows:
        college = (row[i_college] or "").strip()
        eligible = float(row[i_eligible] or 0)
        transcribed = float(row[i_transcribed] or 0)
        applied = float(row[i_applied] or 0)
        students = int(float(row[i_students] or 0))

        total_eligible += eligible
        total_transcribed += transcribed
        total_applied += applied
        total_students += students

        per_college.append({
            "college": college,
            "eligible_credits": eligible,
            "transcribed_credits": transcribed,
            "applied_credits": applied,
            "students_awarded": students,
        })

    return {
        "colleges": len(per_college),
        "total_eligible_credits": total_eligible,
        "total_transcribed_credits": total_transcribed,
        "total_applied_credits": total_applied,
        "total_students_awarded": total_students,
        "per_college": sorted(per_college, key=lambda x: -x["eligible_credits"]),
    }


def _fmt_int(n):
    """Format an integer with comma separators."""
    return f"{n:,}"


def merge_exhibit_metrics(kpis, exhibit_data):
    """
    Merge exhibit metrics into headline KPIs.
    Updates credit_recommendations and adds MAP Exhibits, CCC Collaborative,
    and Articulating Colleges cards.  Optionally enriches with credit distribution data.
    """
    if not exhibit_data:
        return kpis

    ccc = exhibit_data["ccc_collaborative"]

    # ── 1. Replace/enhance CREDIT RECOMMENDATIONS KPI ──
    kpis["credit_recommendations"] = {
        "value": _fmt_int(exhibit_data["total_credit_recs"]),
        "label": "Credit Recommendations",
        "sub": "Source: MAP Custom Reporting Module",
        "breakdowns": [
            {"label": "CCC Collaborative", "value": _fmt_int(ccc["credit_recs"]),
             "note": "statewide faculty workgroups"},
            {"label": "Local", "value": _fmt_int(exhibit_data["local"]["credit_recs"]),
             "note": "individual college articulations"},
        ],
        "live": True,
    }

    # ── 2. MAP EXHIBITS KPI ──
    kpis["map_exhibits"] = {
        "value": _fmt_int(exhibit_data["unique_exhibits"]),
        "label": "MAP Exhibits",
        "sub": f"{_fmt_int(exhibit_data['originating_colleges'])} originating colleges",
        "breakdowns": [
            {"label": "CCC Collaborative", "value": _fmt_int(ccc["unique_exhibits"]),
             "note": "statewide exhibits"},
            {"label": "Local", "value": _fmt_int(exhibit_data["local"]["unique_exhibits"]),
             "note": "college-created exhibits"},
        ],
        "live": True,
    }

    # ── 3. CCC COLLABORATIVE ADOPTION KPI ──
    kpis["ccc_collaborative"] = {
        "value": _fmt_int(ccc["adopting_colleges"]),
        "label": "CCC Collaborative Adoption",
        "sub": "Colleges adopting statewide exhibits",
        "breakdowns": [
            {"label": "Collaborative Exhibits", "value": _fmt_int(ccc["unique_exhibits"])},
            {"label": "Collaborative Credit Recs", "value": _fmt_int(ccc["credit_recs"])},
        ],
        "live": True,
    }

    # ── 4. ARTICULATING COLLEGES KPI ──
    kpis["articulation_colleges"] = {
        "value": _fmt_int(exhibit_data["articulation_colleges"]),
        "label": "Articulating Colleges",
        "sub": "Colleges with MAP exhibits",
        "breakdowns": [
            {"label": "Originating Colleges", "value": _fmt_int(exhibit_data["originating_colleges"]),
             "note": "created exhibits"},
            {"label": "Adopting CCC Collaborative", "value": _fmt_int(ccc["adopting_colleges"]),
             "note": "statewide exhibits"},
        ],
        "live": True,
    }

    # ── CPL Type breakdown in footnote on Credit Recs card ──
    cpl_types = exhibit_data.get("cpl_type_breakdown", {})
    if cpl_types:
        sorted_types = sorted(cpl_types.items(), key=lambda x: -x[1])
        fn = [f"{name}: {_fmt_int(count)}" for name, count in sorted_types[:5]]
        kpis["credit_recommendations"]["footnote"] = fn

    # ── Enrich with credit distribution data if available ──
    cd = exhibit_data.get("credit_distribution")
    if cd:
        # Add Applied Credits breakdown to Transcribed Units card (if it exists)
        if "transcripted_units" in kpis:
            existing_bds = kpis["transcripted_units"].get("breakdowns", [])
            # Add applied credits as a breakdown if not already present
            has_applied = any("applied" in (b.get("label", "").lower()) for b in existing_bds)
            if not has_applied:
                existing_bds.append({
                    "label": "Applied Credits",
                    "value": _fmt_int(int(cd["total_applied_credits"])),
                    "note": "posted to student records",
                })
                kpis["transcripted_units"]["breakdowns"] = existing_bds

    kpis["_exhibit_updated"] = exhibit_data.get("generated_at", "")
    kpis["_exhibit_source"] = exhibit_data.get("source_file", "")
    kpis["_exhibit_datasets"] = exhibit_data.get("datasets_found", [])
    return kpis


# ── MAP Articulation Analysis Tables ─────────────────────────────────────

def _load_top_code_lookup():
    """Load MAP TOP Code → (CCC Discipline, CCC SW Sector) mapping from
    TOP_Code_Lookup.xlsx. Returns a dict where each value is itself a dict
    {"discipline": ..., "sector": ...}. Callers that only want the discipline
    can use the .get(code, {}).get("discipline", "Unknown") pattern, but the
    convenience wrapper _top_disc() handles the common case.
    Searches the same locations as the exhibit JSON."""
    for folder in _EXHIBIT_LOCATIONS:
        path = os.path.join(folder, "TOP_Code_Lookup.xlsx")
        if os.path.exists(path):
            break
    else:
        return {}
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["TOP Code Lookup"]
        lookup = {}
        # Column G (index 6) holds the CCC SW Sector classification.
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            code = str(row[0]).strip() if row[0] is not None else ""
            disc = str(row[2]).strip() if row[2] else "Unknown"
            sector = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            if code:
                lookup[code] = {"discipline": disc, "sector": sector}
        wb.close()
        return lookup
    except Exception as e:
        print(f"  WARNING: Could not load TOP_Code_Lookup.xlsx: {e}")
        return {}


def _top_disc(top_lookup, code, default="Not Mapped"):
    """Convenience: discipline name for a TOP code from _load_top_code_lookup()."""
    entry = top_lookup.get(code)
    if isinstance(entry, dict):
        return entry.get("discipline") or default
    # Backwards-compat: an older flat dict shape
    return entry or default


def _top_sector(top_lookup, code, default=""):
    """Convenience: SW sector name for a TOP code from _load_top_code_lookup()."""
    entry = top_lookup.get(code)
    if isinstance(entry, dict):
        return entry.get("sector") or default
    return default


# ── Credential identity layer (kb/unified_titles.json + kb/credentials.json) ──
# Loaded once per pipeline run and shared by _parse_exhibits() (MAP Exhibits
# KPI) + _build_statewide_adoption() (EACR table) so both group exhibits on
# the same key. See docs/exhibit_unification_vision.md §6.1.
_CREDENTIAL_KB_CACHE = None


def _load_credential_kb():
    """Return (unified_lookup, issuer_pick).

    unified_lookup: raw_title → {unified_title, confidence_title, quality_flag, …}
                    Keyed by both the exact KB key (un-stripped, as the classifier
                    stored it) AND the stripped variant — the EACR generators
                    historically `.strip()` raw titles, so the stripped key
                    catches whitespace-variant cache hits the classifier
                    persisted under their raw MAP form.
    issuer_pick:    unified_title → highest-confidence credential record.
                    Vision §6.1's per-row issuer disambiguation (ICC vs NFPA
                    Fire Inspector I) requires row-level context not present
                    in MAP today; until then, the deterministic-pick keeps
                    all raw rows for a unified_title on the same card.
    """
    global _CREDENTIAL_KB_CACHE
    if _CREDENTIAL_KB_CACHE is not None:
        return _CREDENTIAL_KB_CACHE

    ut_path = os.path.join(SCRIPT_DIR, "kb", "unified_titles.json")
    cr_path = os.path.join(SCRIPT_DIR, "kb", "credentials.json")
    if not (os.path.exists(ut_path) and os.path.exists(cr_path)):
        print("  WARNING: kb/unified_titles.json or kb/credentials.json missing; "
              "EACR will fall back to raw-title grouping.")
        _CREDENTIAL_KB_CACHE = ({}, {})
        return _CREDENTIAL_KB_CACHE

    try:
        with open(ut_path, encoding="utf-8") as f:
            ut_raw = json.load(f)
        with open(cr_path, encoding="utf-8") as f:
            cr_raw = json.load(f)
    except Exception as e:
        print(f"  WARNING: failed to load credential KB: {e}")
        _CREDENTIAL_KB_CACHE = ({}, {})
        return _CREDENTIAL_KB_CACHE

    unified_lookup = {}
    for raw_key, entry in ut_raw.items():
        unified_lookup[raw_key] = entry
        stripped = raw_key.strip()
        if stripped and stripped != raw_key and stripped not in unified_lookup:
            unified_lookup[stripped] = entry

    issuer_pick = {}
    for ut, records in cr_raw.items():
        if not records:
            continue
        best = sorted(
            records,
            key=lambda r: (-(r.get("confidence_issuer") or 0.0),
                           r.get("issuing_agency") or ""),
        )[0]
        issuer_pick[ut] = best

    print(f"  Credential KB loaded: {len(ut_raw):,} raw→unified entries, "
          f"{len(cr_raw):,} unified→issuer entries")
    _CREDENTIAL_KB_CACHE = (unified_lookup, issuer_pick)
    return _CREDENTIAL_KB_CACHE


def _classify_exhibit(raw_title, unified_lookup=None, issuer_pick=None):
    """Return a dict describing this raw_title's place in the credential layer.

    Returned keys:
      unified_title, issuing_agency, training_agency,
      confidence_title, confidence_issuer, quality_flag, is_classified.

    For unclassified raw titles (no KB entry), `unified_title` falls back to
    `raw_title.strip()` so the row still has a group key — this preserves
    EACR coverage even before the classifier has seen a title.
    """
    if unified_lookup is None or issuer_pick is None:
        unified_lookup, issuer_pick = _load_credential_kb()

    # Exact match (preserves whitespace as classifier stored it) → stripped fallback.
    entry = unified_lookup.get(raw_title)
    if entry is None:
        stripped = (raw_title or "").strip()
        if stripped:
            entry = unified_lookup.get(stripped)

    if entry is None:
        # Unclassified — keep raw title (stripped) as the group key.
        return {
            "unified_title": (raw_title or "").strip(),
            "issuing_agency": "",
            "training_agency": "",
            "confidence_title": 0.0,
            "confidence_issuer": 0.0,
            "quality_flag": "",
            "is_classified": False,
        }

    ut = entry.get("unified_title") or (raw_title or "").strip()
    conf_t = float(entry.get("confidence_title") or 0.0)
    qflag = entry.get("quality_flag") or ""

    issuer_rec = issuer_pick.get(ut)
    if issuer_rec:
        return {
            "unified_title": ut,
            "issuing_agency": issuer_rec.get("issuing_agency") or "",
            "training_agency": issuer_rec.get("training_agency") or "",
            "confidence_title": conf_t,
            "confidence_issuer": float(issuer_rec.get("confidence_issuer") or 0.0),
            "quality_flag": qflag,
            "is_classified": True,
        }

    # KB has unified_title but no credentials.json record yet.
    return {
        "unified_title": ut,
        "issuing_agency": "",
        "training_agency": "",
        "confidence_title": conf_t,
        "confidence_issuer": 0.0,
        "quality_flag": qflag,
        "is_classified": True,
    }



def build_exhibit_analysis_tables(exhibit_data):
    """
    Compute analysis tables from the combined CustomReport JSON.
    Uses View_ArticulatedMAPExhibits as the primary data source.
    Returns a dict of analysis tables ready for dashboard rendering.
    """
    if not exhibit_data or not EXHIBIT_FILE:
        return None

    try:
        with open(EXHIBIT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return None

    # Find the ArticulatedMAPExhibits dataset (preferred)
    ds = None
    for report in data:
        if report.get("viewName") == "View_ArticulatedMAPExhibits_APIDataset":
            ds = report
            break
    if not ds:
        return None

    rows = ds["columnValue"]
    cm = {c: i for i, c in enumerate(ds.get("columnName", []))}
    top_lookup = _load_top_code_lookup()

    i_college = cm.get("College", 0)
    i_exhibit = cm.get("ExhibitID", 1)
    i_title = cm.get("Exhibit Title", 2)
    i_artic = cm.get("Articulation College", 4)
    i_course = cm.get("Course", 5)
    i_collab = cm.get("Collaborative Type", 7)
    i_top = cm.get("TOP Code", 8)
    i_cid = cm.get("CID Number", 9)
    i_mol = cm.get("CPL Mode of Learning", 11)
    i_cpl = cm.get("CPL Type Description", 13)

    from collections import defaultdict

    # ── By College ──
    by_college = defaultdict(lambda: {"recs": 0, "exhibits": set(), "disciplines": set(),
                                      "ccc": 0, "industry": 0})
    # ── By Discipline ──
    by_disc = defaultdict(lambda: {"recs": 0, "exhibits": set(), "courses": set(),
                                   "colleges": set(), "ccc": 0})
    # ── By CPL Type ──
    by_cpl = defaultdict(lambda: {"recs": 0, "exhibits": set(), "colleges": set()})
    # ── By Mode of Learning ──
    by_mol = defaultdict(lambda: {"recs": 0, "exhibits": set(), "colleges": set()})
    # ── Top Exhibits ──
    by_exhibit_title = defaultdict(lambda: {"recs": 0, "courses": set(), "colleges": set(),
                                            "cpl_type": "", "disc": ""})
    # ── Collaborative Analysis ──
    by_collab = defaultdict(lambda: {"recs": 0, "exhibits": set(), "colleges": set(),
                                     "disciplines": set()})
    # ── Per-College Per-Discipline Detail (powers the College Activity discipline filter) ──
    by_college_disc = defaultdict(lambda: defaultdict(lambda: {"recs": 0, "exhibits": set()}))

    total_recs = len(rows)

    for row in rows:
        college = (row[i_college] or "").strip()
        eid = (row[i_exhibit] or "").strip()
        title = (row[i_title] or "").strip()
        artic = (row[i_artic] or "").strip()
        course = (row[i_course] or "").strip()
        collab = (row[i_collab] or "").strip()
        top_code = (row[i_top] or "").strip()
        mol = (row[i_mol] or "").strip() or "Unknown"
        cpl = (row[i_cpl] or "").strip() or "Other"

        disc = _top_disc(top_lookup, top_code)
        is_ccc = "CCC" in collab
        is_industry = cpl == "Industry Certification"

        # Collaborative category
        if is_ccc:
            collab_cat = "CCC Collaborative"
        elif collab in ("", "Other"):
            collab_cat = "Local"
        else:
            collab_cat = "Industry/Other"

        # By College
        if artic:
            c = by_college[artic]
            c["recs"] += 1
            c["exhibits"].add(eid)
            c["disciplines"].add(disc)
            if is_ccc: c["ccc"] += 1
            if is_industry: c["industry"] += 1

            # Per-college per-discipline
            cd = by_college_disc[artic][disc]
            cd["recs"] += 1
            cd["exhibits"].add(eid)

        # By Discipline
        d = by_disc[disc]
        d["recs"] += 1
        d["exhibits"].add(eid)
        d["courses"].add(course)
        d["colleges"].add(artic)
        if is_ccc: d["ccc"] += 1

        # By CPL Type
        t = by_cpl[cpl]
        t["recs"] += 1
        t["exhibits"].add(eid)
        t["colleges"].add(artic)

        # By Mode of Learning
        m = by_mol[mol]
        m["recs"] += 1
        m["exhibits"].add(eid)
        m["colleges"].add(artic)

        # Top Exhibits
        e = by_exhibit_title[title]
        e["recs"] += 1
        e["courses"].add(course)
        e["colleges"].add(artic)
        if not e["cpl_type"]: e["cpl_type"] = cpl
        if not e["disc"] or e["disc"] == "Not Mapped": e["disc"] = disc

        # Collaborative
        cc = by_collab[collab_cat]
        cc["recs"] += 1
        cc["exhibits"].add(eid)
        cc["colleges"].add(artic)
        cc["disciplines"].add(disc)

    # ── Serialize to JSON-friendly format ──
    def pct(n): return round(n / total_recs * 100, 1) if total_recs else 0

    tables = {
        "by_college": sorted([
            {"college": k, "credit_recs": v["recs"], "exhibits": len(v["exhibits"]),
             "disciplines": len(v["disciplines"]), "ccc_collaborative": v["ccc"],
             "industry_certs": v["industry"], "pct": pct(v["recs"])}
            for k, v in by_college.items()
        ], key=lambda x: -x["credit_recs"]),

        "by_discipline": sorted([
            {"discipline": k, "credit_recs": v["recs"], "exhibits": len(v["exhibits"]),
             "courses": len(v["courses"]), "colleges": len(v["colleges"]),
             "ccc_collaborative": v["ccc"], "pct": pct(v["recs"])}
            for k, v in by_disc.items()
        ], key=lambda x: -x["credit_recs"]),

        "by_cpl_type": sorted([
            {"cpl_type": k, "credit_recs": v["recs"], "exhibits": len(v["exhibits"]),
             "colleges": len(v["colleges"]), "pct": pct(v["recs"])}
            for k, v in by_cpl.items()
        ], key=lambda x: -x["credit_recs"]),

        "by_mode_of_learning": sorted([
            {"mode": k, "credit_recs": v["recs"], "exhibits": len(v["exhibits"]),
             "colleges": len(v["colleges"]), "pct": pct(v["recs"])}
            for k, v in by_mol.items()
        ], key=lambda x: -x["credit_recs"]),

        "top_exhibits": sorted([
            {"title": k, "credit_recs": v["recs"], "courses": len(v["courses"]),
             "colleges": len(v["colleges"]), "cpl_type": v["cpl_type"], "discipline": v["disc"]}
            for k, v in by_exhibit_title.items()
        ], key=lambda x: -x["credit_recs"])[:50],

        "collaborative_analysis": sorted([
            {"category": k, "credit_recs": v["recs"], "exhibits": len(v["exhibits"]),
             "colleges": len(v["colleges"]), "disciplines": len(v["disciplines"]),
             "pct": pct(v["recs"])}
            for k, v in by_collab.items()
        ], key=lambda x: -x["credit_recs"]),

        "college_discipline_detail": {
            college: {
                disc: {"recs": d["recs"], "exhibits": len(d["exhibits"])}
                for disc, d in disc_map.items()
            }
            for college, disc_map in by_college_disc.items()
        },

        "total_credit_recs": total_recs,
        "generated_at": exhibit_data.get("generated_at", ""),
    }

    # ── Statewide Exhibit Adoption Analysis ──
    # Cross-reference CCC Collaborative exhibits with College Courses to find potential adopters
    statewide_adoption = _build_statewide_adoption(data, rows, cm)
    if statewide_adoption:
        tables["statewide_adoption"] = statewide_adoption

    # ── Articulations by Unified Course (course-identity layer) ──
    # Collapses the same course across colleges into one row via the coci
    # identity layer; surfaces cross-college adoption leverage.
    articulations_by_course = _build_articulations_by_course()
    if articulations_by_course:
        tables["articulations_by_course"] = articulations_by_course

    print(f"  Built exhibit analysis tables: {len(tables['by_college'])} colleges, "
          f"{len(tables['by_discipline'])} disciplines, {len(tables['by_cpl_type'])} CPL types, "
          f"{len(tables['by_mode_of_learning'])} modes, {len(tables['top_exhibits'])} top exhibits"
          + (f", {len(statewide_adoption)} statewide exhibits" if statewide_adoption else ""))

    return tables


def _build_statewide_adoption(all_data, exhibit_rows, exhibit_cm):
    """
    For ALL exhibits (CCC Collaborative + Local), identify:
      - Colleges that have already adopted it
      - Colleges that could adopt it (matching programs by MAP TOP code via ProgramsofStudy)
      - Credit recommendations (course → credit text) for each exhibit

    Uses View_ProgramsofStudy for TOP-code-based potential-adopter matching
    (ProgramsofStudy carries the MAP integer TOP codes, same as exhibits).
    Returns a sorted list of exhibit dicts, or None.
    """
    from collections import defaultdict

    top_lookup = _load_top_code_lookup()  # MAP code → discipline name

    # ── Build MAP TOP code → colleges from ProgramsofStudy ──
    # ProgramsofStudy uses MAP integer TOP codes (same as exhibits)
    programs_ds = None
    for report in all_data:
        if report.get("viewName") == "View_ProgramsofStudy_APIDataset":
            programs_ds = report
            break

    top_to_colleges = defaultdict(set)
    if programs_ds:
        pcm = {c: i for i, c in enumerate(programs_ds.get("columnName", []))}
        for row in programs_ds["columnValue"]:
            college = (row[pcm.get("College", 0)] or "").strip()
            tc = (row[pcm.get("Top Code", 9)] or "").strip()
            if college and tc:
                top_to_colleges[tc].add(college)
        print(f"  ProgramsofStudy: {len(top_to_colleges)} TOP codes, "
              f"{len(set(c for cs in top_to_colleges.values() for c in cs))} colleges")

    # ── Also build CID → colleges from View_CollegeCourses for CID-based matching ──
    cid_to_colleges = defaultdict(set)
    for report in all_data:
        if report.get("viewName") == "View_CollegeCourses_APIDataset":
            ccm = {c: i for i, c in enumerate(report.get("columnName", []))}
            for row in report["columnValue"]:
                college = (row[ccm.get("College", 0)] or "").strip()
                cid = (row[ccm.get("CID Number", 1)] or "").strip()
                if college and cid:
                    cid_to_colleges[cid].add(college)
            break

    # ── Gather ALL exhibits (both CCC Collaborative and Local) ──
    i_collab = exhibit_cm.get("Collaborative Type", 7)
    i_eid = exhibit_cm.get("ExhibitID", 1)
    i_title = exhibit_cm.get("Exhibit Title", 2)
    i_artic = exhibit_cm.get("Articulation College", 4)
    i_course = exhibit_cm.get("Course", 5)
    i_credit = exhibit_cm.get("Credit Recommendation", 6)
    i_cid = exhibit_cm.get("CID Number", 9)
    i_top = exhibit_cm.get("TOP Code", 8)
    i_cpl = exhibit_cm.get("CPL Type Description", 13)

    # ── EACR Phase 4 re-pivot (2026-05-26): group on the credential identity ──
    # Group key is (unified_title, issuing_agency, CPL Type, Collaborative Type)
    # for raw titles classified into the credential layer; for unclassified
    # raw titles, we keep the stripped raw title as the unified key so coverage
    # is preserved. This collapses two forms of fragmentation at once:
    #   1. ID fragmentation — MAP issues separate ExhibitIDs per articulated
    #      course (was previously handled by the raw-title grouping).
    #   2. Title drift — colleges enter the same credential under multiple
    #      freehand titles ("Google IT Support Professional Certification" vs
    #      "Google IT Support Professional Certificate" vs "CMPET 315 …").
    # See docs/exhibit_unification_vision.md §6.1 + kb/eacr_dryrun/report.md
    # (the PR-C0 measurement that estimated 3,345 → ~2,406 cards). TOP Code is
    # excluded from the key because ~295 single-credential exhibits legitimately
    # span multiple TOP codes (Dental Board cert across TOPs 101/89/171).
    unified_lookup, issuer_pick = _load_credential_kb()

    all_exhibits = defaultdict(lambda: {
        "eids": set(),
        "adopters": set(),
        "cids": set(),
        "tops": set(),
        "raw_titles": set(),  # for the consumer's "also entered as…" disclosure
        "confidence_titles": [],  # per-row for modal
        "quality_flags": set(),  # any constituent flag rolls up
        "credit_recs": [],  # list of {course, credit} dicts (deduped)
        "training_agency": "",  # set on first classified row
        "confidence_issuer": 0.0,
        "is_classified": False,
    })

    for row in exhibit_rows:
        eid = (row[i_eid] or "").strip()
        raw_title = row[i_title] or ""
        title = raw_title.strip()
        if not eid or not title:
            continue
        cpl = (row[i_cpl] or "").strip()
        collab = (row[i_collab] or "").strip()

        ident = _classify_exhibit(raw_title, unified_lookup, issuer_pick)
        group_key = (
            ident["unified_title"],
            ident["issuing_agency"],
            cpl,
            collab,
        )
        e = all_exhibits[group_key]
        e["eids"].add(eid)
        e["raw_titles"].add(title)
        if ident["confidence_title"]:
            e["confidence_titles"].append(ident["confidence_title"])
        if ident["quality_flag"]:
            e["quality_flags"].add(ident["quality_flag"])
        if ident["is_classified"]:
            e["is_classified"] = True
            # First classified row sets training_agency + confidence_issuer
            # (all rows for the same group resolve to the same issuer record).
            if not e["training_agency"]:
                e["training_agency"] = ident["training_agency"]
            if not e["confidence_issuer"]:
                e["confidence_issuer"] = ident["confidence_issuer"]

        artic = (row[i_artic] or "").strip()
        if artic:
            e["adopters"].add(artic)
        cid = (row[i_cid] or "").strip()
        if cid:
            e["cids"].add(cid)
        top = (row[i_top] or "").strip()
        if top:
            e["tops"].add(top)
        # Collect credit recommendations (dedup across constituent raw rows).
        course = (row[i_course] or "").strip()
        credit = (row[i_credit] or "").strip()
        if course and credit:
            rec_key = (course, credit)
            if rec_key not in {(r["course"], r["credit"]) for r in e["credit_recs"]}:
                e["credit_recs"].append({"course": course, "credit": credit})

    # ── Compute potential + assemble output rows ──
    from collections import Counter as _Counter
    results = []
    for group_key, e in all_exhibits.items():
        unified_title, issuing_agency, cpl_type, collab_type = group_key
        adopters = e["adopters"]
        tops_sorted = sorted(e["tops"])
        map_top = tops_sorted[0] if tops_sorted else ""

        # Potential adopters via MAP TOP code (ProgramsofStudy) — union across all TOPs
        potential_top = set()
        for t in e["tops"]:
            potential_top |= top_to_colleges.get(t, set())

        # Potential adopters via CID (CollegeCourses)
        potential_cid = set()
        for cid in e["cids"]:
            potential_cid.update(cid_to_colleges.get(cid, set()))

        new_colleges = (potential_cid | potential_top) - adopters
        disc = _top_disc(top_lookup, map_top)

        # Career Cluster (CCC Strong Workforce sector). Groups that span multiple
        # TOP codes can in principle resolve to multiple sectors. Pick the most
        # common; ties broken alphabetically for stability.
        sector_counts = _Counter(s for s in (_top_sector(top_lookup, t) for t in e["tops"]) if s)
        if sector_counts:
            sector = sorted(sector_counts.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]
        else:
            sector = ""

        # Classify as Statewide (CCC Collaborative) or Local
        is_statewide = "CCC" in collab_type
        collab_label = "CCC Collaborative" if is_statewide else (collab_type or "Local")

        # Stable identifier for the merged group (concatenation of member MAP IDs).
        # Used by the UI as a row-selection / dedup key. Post Phase 4 the merged_id
        # may be larger than before (more raw exhibit IDs fold per card) — PR-D's
        # _EACR_FLAG keys keyed against the old merged_id are mapped via
        # kb/eacr_dryrun/alias_map.json (migration runs at PR-C2 land).
        merged_id = "|".join(sorted(e["eids"]))

        # Card "title" surfaces the unified credential name; raw_titles[] is the
        # underlying spelling-variant list for the "also entered as…" disclosure
        # PR-C2 will surface in the UI. For unclassified groups the unified_title
        # equals the (stripped) raw title, so the title field stays meaningful.
        conf_t_modal = (
            max(set(e["confidence_titles"]), key=e["confidence_titles"].count)
            if e["confidence_titles"] else 0.0
        )
        quality_flag = sorted(e["quality_flags"])[0] if e["quality_flags"] else ""

        results.append({
            "exhibit_id": merged_id,
            "exhibit_ids": sorted(e["eids"]),
            "title": unified_title,
            "unified_title": unified_title,
            "is_classified": e["is_classified"],
            "issuing_agency": issuing_agency,
            "training_agency": e["training_agency"],
            "confidence_title": conf_t_modal,
            "confidence_issuer": e["confidence_issuer"],
            "quality_flag": quality_flag,
            "raw_titles": sorted(e["raw_titles"]),
            "cpl_type": cpl_type,
            "discipline": disc,
            "sector": sector,
            "collaborative_type": collab_label,
            "adopters": len(adopters),
            "adopter_names": sorted(adopters),
            "potential": len(new_colleges),
            "potential_names": sorted(new_colleges),
            "total_addressable": len(adopters) + len(new_colleges),
            "credit_recs": e["credit_recs"],
        })

    # Sort: exhibits with most potential first, then by adopters descending
    results.sort(key=lambda x: (-x["potential"], -x["adopters"]))
    return results


def _build_articulations_by_course():
    """Group earned MAP articulations by UNIFIED COURSE IDENTITY (the coci
    course-identity layer) instead of raw MAP rows, so the same course taught at
    many colleges collapses to one row and cross-college adoption leverage is
    visible.

    Reads kb/coci_articulations.json — the staging layer that already resolves
    each earned articulation to a C-ID/CCN/M-ID identity (Course → (subject,
    number) → C-ID/CCN when the row carries a CID Number, else M-ID, with
    multi-M-ID rows disambiguated by the local Course Title). Returns a list of
    per-identity dicts (ranked by adoption leverage), or None if the file is
    absent.

    Over-merge guardrail: adoption leverage on identities flagged `over_merged`
    is NOT presented as actionable (the cluster may conflate distinct courses);
    such rows carry over_merged=True so the renderer can suppress/badge the
    number rather than suggest a bogus cross-college adoption.
    """
    from collections import Counter, defaultdict

    path = os.path.join(SCRIPT_DIR, "kb", "coci_articulations.json")
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            doc = json.load(f)
    except Exception:
        return None

    idents = doc.get("identities", {})
    records = doc.get("articulations", [])
    if not records:
        return None

    agg = defaultdict(lambda: {
        "earned": set(), "creds": Counter(), "credit_recs": Counter(),
        "titles": Counter(), "leverage": 0, "over_merged": False, "flagged": False,
    })
    for r in records:
        cid = r.get("course_id")
        if not cid:
            continue
        a = agg[cid]
        a["earned"].update(r.get("earned_by_colleges", []))
        if r.get("unified_title"):
            a["creds"][r["unified_title"]] += 1
        for cr in (r.get("credit_recommendations") or []):
            a["credit_recs"][cr] += 1
        for lc in (r.get("local_courses") or []):
            if lc.get("title"):
                a["titles"][lc["title"]] += 1
        a["leverage"] = max(a["leverage"], r.get("adoption_leverage_count", 0) or 0)
        if r.get("over_merged"):
            a["over_merged"] = True
        if r.get("quality_flag"):
            a["flagged"] = True

    out = []
    for cid, a in agg.items():
        ident = idents.get(cid, {})
        over = a["over_merged"] or bool(ident.get("over_merged"))
        title = ident.get("title") or (a["titles"].most_common(1)[0][0] if a["titles"] else "")
        cred = a["creds"].most_common(1)[0][0] if a["creds"] else ""
        n_cred = len(a["creds"])
        credit_rec = a["credit_recs"].most_common(1)[0][0] if a["credit_recs"] else ""
        out.append({
            "course_id": cid,
            "id_system": ident.get("identity_system", ""),
            "title": title,
            "discipline": ident.get("discipline") or "—",
            "credit_status": ident.get("credit_status") or "",
            "colleges_earned": len(a["earned"]),
            "credit_rec": credit_rec,
            "credential": cred,
            "credential_count": n_cred,
            # Leverage = peer colleges teaching the same identity that have NOT yet
            # earned this articulation. Zeroed-out for over-merged clusters so the
            # number is never read as an actionable adoption target.
            "leverage": 0 if over else a["leverage"],
            "over_merged": over,
            "flagged": a["flagged"],
        })
    # Clean, high-leverage identities first; over-merged rows sink to the bottom.
    out.sort(key=lambda x: (x["over_merged"], -x["leverage"], -x["colleges_earned"]))
    return out

def _write_analytics_xlsx_export(card_id, title, headers, rows, output_dir):
    """Write a single CPL Analytics table to <output_dir>/<card_id>.xlsx.
    rows is a list of lists (one row per data row); headers is a list of strings.
    The first sheet row is the headers with the navy fill used by the KPI cards.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None
    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit
    ws.append(headers)
    header_font = Font(bold=True, color="C9A84C", size=11)
    header_fill = PatternFill(start_color="0A2240", end_color="0A2240", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for row in rows:
        ws.append(row)
    # Auto-fit column widths (approximate)
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                cell_len = len(str(row[col_idx - 1] or ""))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 60)
    out_path = os.path.join(output_dir, f"{card_id}.xlsx")
    wb.save(out_path)
    return out_path


def export_credential_reference():
    """Build credential_reference_data.js — the lean payload consumed by the
    Credential Reference tab. Joins the credential-identity layer
    (kb/unified_titles.json + credentials.json) with the course-identity
    layer (coci_articulations.json + the minted/unified/singleton catalogs)
    so curators see, per credential identity:

      - the canonical credential name (unified_title) + issuer + trainer
      - the predominant discipline + TOP code across articulations
      - whether ANY articulation is CCC Collaborative (statewide badge)
      - per common-course identity (CCN-ID/C-ID/M-ID/Cluster) that articulates
        to this credential: identity + title + discipline + the local college
        course rows (subject/number/title) and earning colleges

    Replaces the tab's previous runtime fetch of unified_titles.json +
    coci_articulations.json (~3MB+); the baked payload is ~300-500KB and
    pre-joined so the tab is responsive on load.
    """
    from collections import Counter, defaultdict
    from datetime import datetime as _dt
    from datetime import timezone as _tz

    kdir = os.environ.get("UC_KB_DIR") or os.path.join(SCRIPT_DIR, "kb")
    odir = os.environ.get("UC_OUT_DIR") or SCRIPT_DIR
    out_js = os.path.join(odir, "credential_reference_data.js")

    def _load(name):
        p = os.path.join(kdir, name)
        return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else None

    ut_doc = _load("unified_titles.json")
    cred_doc = _load("credentials.json")
    if not ut_doc or not cred_doc:
        with open(out_js, "w", encoding="utf-8") as f:
            f.write("/* credential-identity KB files missing */\n"
                    "window.CPL_CREDENTIAL_REFERENCE = {unified_titles: []};\n")
        print("  Credential Reference: kb files absent — wrote empty global")
        return

    # Curator overrides from Supabase (synced daily via
    # kb/_apply_credential_review.py into kb/credential_review_overlay.json).
    # When a row has an override, the override value is baked into the row's
    # visible field AND the original AI baseline is preserved on a parallel
    # _original_* field, so the dashboard's runtime applyOverlay() can show
    # the "originally: X" hint correctly. unified_title_override is
    # recorded in the overlay but NOT applied here (display-only by PR-4
    # design; rename promotion is Cred-Ref PR-5b's job).
    review_doc = _load("credential_review_overlay.json") or {}
    review_overrides = (review_doc.get("overrides") or {})

    art_doc = _load("coci_articulations.json") or {}
    identities = art_doc.get("identities", {}) or {}
    articulations = art_doc.get("articulations", []) or []

    # Build a discipline/top/title lookup per course_id. identities[cid] is the
    # primary source but is sometimes empty ({}), so cascade to the broader
    # catalogs (minted_courses → singletons → unified_courses clusters).
    course_meta = {}
    for cid, rec in identities.items():
        if rec:
            course_meta[cid] = {
                "title": rec.get("title", "") or "",
                "discipline": rec.get("discipline", "") or "",
                "top": rec.get("top_code", "") or "",
                "sys": rec.get("identity_system", "") or "",
            }
    for name, top_key in (
        ("coci_minted_courses.json", "courses"),
        ("coci_minted_singletons.json", "courses"),
        ("coci_unified_courses.json", "clusters"),
    ):
        doc = _load(name) or {}
        for cid, rec in (doc.get(top_key) or {}).items():
            if cid in course_meta:
                continue
            course_meta[cid] = {
                "title": rec.get("common_title") or rec.get("title") or rec.get("unified_title") or "",
                "discipline": rec.get("discipline", "") or "",
                "top": rec.get("top_code", "") or "",
                "sys": rec.get("id_system") or rec.get("identity_system")
                       or ("Cluster" if top_key == "clusters" else "M-ID"),
            }
        del doc  # free the 28MB minted_courses promptly

    # Pre-aggregate raw-variant info from unified_titles.json
    raw_count_by_ut = Counter()
    quality_by_ut = {}
    confs_by_ut = defaultdict(list)
    raw_to_ut = {}  # raw_title → unified_title (for audit-tag rollup)
    for raw_title, ent in ut_doc.items():
        ut = ent.get("unified_title") or "(blank)"
        raw_count_by_ut[ut] += 1
        raw_to_ut[raw_title] = ut
        if ent.get("quality_flag"):
            quality_by_ut[ut] = ent["quality_flag"]
        confs_by_ut[ut].append(ent.get("confidence_title", 0))

    # Pre-aggregate audit-tag counts per unified_title — sum across raw variants.
    # Falls back to {} if the audit file isn't generated yet (first-run / clean repo).
    audit_doc = _load(os.path.join("exhibit_audit", "latest.json")) or {}
    audit_tags_by_ut = defaultdict(lambda: Counter())
    for c in (audit_doc.get("title_cards") or []):
        ut = raw_to_ut.get(c.get("raw_title"))
        if not ut:
            continue
        for t in (c.get("tags") or []):
            audit_tags_by_ut[ut][t] += 1

    # Group articulations by credential
    art_by_ut = defaultdict(list)
    for a in articulations:
        ut = a.get("unified_title")
        if ut:
            art_by_ut[ut].append(a)

    rows = []
    for ut, _n in raw_count_by_ut.items():
        creds = cred_doc.get(ut, [])
        primary = creds[0] if creds else {}
        confs = confs_by_ut[ut]
        rc = Counter(round(c, 2) for c in confs)
        conf_modal = rc.most_common(1)[0][0] if rc else 0

        ut_arts = art_by_ut.get(ut, [])
        disc_counter, top_counter = Counter(), Counter()
        statewide = False
        by_cid = defaultdict(list)
        for a in ut_arts:
            cid = a.get("course_id")
            if not cid:
                continue
            by_cid[cid].append(a)
            # coci_articulations.json normalises MAP's "CCC Collaborative" to
            # "CCC". Match either form so the badge survives a future kb-pipeline
            # change that re-introduces the long form.
            ct = a.get("collaborative_type") or ""
            if ct == "CCC" or ct == "CCC Collaborative":
                statewide = True
            m = course_meta.get(cid) or {}
            if m.get("discipline"):
                disc_counter[m["discipline"]] += 1
            if m.get("top"):
                top_counter[m["top"]] += 1

        disc_modal = disc_counter.most_common(1)[0][0] if disc_counter else ""
        top_modal = top_counter.most_common(1)[0][0] if top_counter else ""

        articulations_out = []
        for cid, recs in by_cid.items():
            m = course_meta.get(cid) or {}
            # Collapse all local_courses across recs; track colleges per local-course pair.
            local_set = {}
            for r in recs:
                cols = r.get("earned_by_colleges", []) or []
                for lc in (r.get("local_courses") or []):
                    key = (lc.get("subject", ""), lc.get("number", ""), lc.get("title", ""))
                    if key not in local_set:
                        local_set[key] = set()
                    local_set[key].update(cols)
            local_list = [
                {"subj": s, "num": n, "t": t, "colleges": sorted(cs)}
                for (s, n, t), cs in sorted(local_set.items())
            ]
            articulations_out.append({
                "cid": cid,
                "sys": m.get("sys", ""),
                "title": m.get("title", ""),
                "disc": m.get("discipline", ""),
                "top": m.get("top", ""),
                "local": local_list,
            })
        # Order common-course identities: official first (CCN-ID, C-ID), then M-ID / Cluster.
        sys_order = {"CCN-ID": 0, "C-ID": 1, "M-ID": 2, "Cluster": 3}
        articulations_out.sort(key=lambda x: (sys_order.get(x["sys"], 9), x["cid"]))

        audit_tags = dict(audit_tags_by_ut.get(ut, {}))
        issuer_val = primary.get("issuing_agency")
        trainer_val = primary.get("training_agency")
        qflag_val = quality_by_ut.get(ut)
        row = {
            "ut": ut,
            "raw_count": raw_count_by_ut[ut],
            "issuer": issuer_val,
            "trainer": trainer_val,
            "conf_title": conf_modal,
            "conf_issuer": primary.get("confidence_issuer", 0),
            "quality_flag": qflag_val,
            "disc_modal": disc_modal,
            "top_modal": top_modal,
            "statewide": statewide,
            "audit_tags": audit_tags,
            "audit_tag_total": sum(audit_tags.values()),
            "articulations": articulations_out,
            "n_articulation_lines": sum(len(a["local"]) for a in articulations_out),
        }
        # Apply curator overrides (Mode A: non-identity fields only). When a
        # field is overridden, the AI baseline gets preserved on _original_*
        # so the runtime applyOverlay() can show the "originally: X" hint.
        ov = review_overrides.get(ut)
        if ov:
            if "issuing_agency_override" in ov:
                row["_original_issuer"] = issuer_val
                row["issuer"] = ov["issuing_agency_override"] or None
            if "training_agency_override" in ov:
                row["_original_trainer"] = trainer_val
                row["trainer"] = ov["training_agency_override"] or None
            if "quality_flag_override" in ov:
                row["_original_quality_flag"] = qflag_val
                row["quality_flag"] = ov["quality_flag_override"] or None
            if ov.get("reviewed_by"):
                row["curated_by"] = ov["reviewed_by"]
            if ov.get("reviewed_at"):
                row["curated_at"] = ov["reviewed_at"]
        rows.append(row)

    rows.sort(key=lambda r: r["ut"].lower())

    # PR-3 grouping: include a 2-digit-TOP → category title map so the
    # Credential Reference tab can show "TOP 12 — Health" instead of just
    # "12" when grouping. Lifted from kb/discipline_canonical_subj4.json
    # (the same source the CSC tab uses for its TOP categories).
    top_categories = {}
    disc_seed = _load("discipline_canonical_subj4.json") or {}
    for _disc, _rec in (disc_seed.get("disciplines") or {}).items():
        k = _rec.get("top_category_2digit")
        t = _rec.get("top_category_title")
        if k and t and k not in top_categories:
            top_categories[k] = t

    payload = {
        "_generated_at": _dt.now(_tz.utc).isoformat(timespec="seconds"),
        "_generated_by": "excel_to_dashboard.py:export_credential_reference()",
        "_stats": {
            "unified_titles": len(rows),
            "articulated_titles": sum(1 for r in rows if r["articulations"]),
            "total_articulation_lines": sum(r["n_articulation_lines"] for r in rows),
            "statewide_titles": sum(1 for r in rows if r["statewide"]),
        },
        "top_categories": top_categories,
        "unified_titles": rows,
    }
    with open(out_js, "w", encoding="utf-8") as f:
        f.write("window.CPL_CREDENTIAL_REFERENCE = ")
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")
    sz = os.path.getsize(out_js)
    s = payload["_stats"]
    print(f"  Credential Reference: wrote {sz//1024} KB "
          f"({s['unified_titles']} unified titles · {s['articulated_titles']} articulated · "
          f"{s['total_articulation_lines']} local-course lines · {s['statewide_titles']} statewide)")


def export_unified_courses():
    """Build the Unified Courses tab data (window.CPL_UNIFIED_COURSES in
    unified_courses_data.js) + the full xlsx export, from the kb/coci_*.json
    staging files. Writes an empty global if the KB files are absent so the
    tab degrades gracefully. STAGING data — does not touch curated KB files."""
    from datetime import datetime as _dt
    # UC_KB_DIR / UC_OUT_DIR are a test seam (e.g. validating a re-mint preview);
    # unset in production, so the daily build reads kb/ and writes the repo root.
    kdir = os.environ.get("UC_KB_DIR") or os.path.join(SCRIPT_DIR, "kb")
    odir = os.environ.get("UC_OUT_DIR") or SCRIPT_DIR
    out_js = os.path.join(odir, "unified_courses_data.js")

    def _load(name):
        p = os.path.join(kdir, name)
        return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else None

    cat_doc = _load("coci_minted_courses.json")
    if not cat_doc:
        with open(out_js, "w", encoding="utf-8") as f:
            f.write("/* Unified Courses data unavailable (kb/coci_minted_courses.json missing) */\n"
                    "window.CPL_UNIFIED_COURSES = {rows: [], colleges: []};\n")
        print("  Unified Courses: kb files absent — wrote empty global")
        return

    cat = cat_doc["courses"]
    sg = (_load("coci_minted_singletons.json") or {}).get("courses", {})
    clusters = (_load("coci_unified_courses.json") or {}).get("clusters", {})
    art_doc = _load("coci_articulations.json") or {}
    art_ident = art_doc.get("identities", {})
    earned = {}
    for g in art_doc.get("articulations", []):
        earned.setdefault(g["course_id"], set()).update(g.get("earned_by_colleges", []))

    # Human curation overlay (git-canonical, synced from Supabase by
    # kb/_apply_curation.py). Applied on top of the AI drafts so curation is
    # regen-safe: a curated discipline overrides a blank, and the row is marked
    # reviewed so the daily regen never clobbers it.
    curation = (_load("coci_curation.json") or {}).get("curations", {})

    def disc_of(cid, base):
        c = curation.get(cid)
        return (c.get("discipline") or base) if c else base

    def reviewed_by_of(cid):
        return (curation.get(cid) or {}).get("reviewed_by")

    def reviewed_at_of(cid):
        return ((curation.get(cid) or {}).get("reviewed_at") or "")[:10]

    def _add_prov(row, cid, v):
        # Surface how a Generated discipline was inferred (subject_map vs
        # title_keyword) so reviewers can triage the riskier keyword fills.
        # Skip curated entries (human discipline, not an AI draft); the tab
        # renders those as Verified anyway. Lean: only emit when an AI source
        # exists, so blank/manual/anchor rows carry no extra keys.
        if not curation.get(cid):
            src = v.get("discipline_source")
            if src:
                row["dsrc"] = src
                conf = v.get("discipline_confidence")
                if conf is not None:
                    row["dconf"] = conf
        return row

    def cur_desc_of(cid):
        return (curation.get(cid) or {}).get("description")

    # Reviewer consolidations ("Generate unified course"): member course_id →
    # target unified id, from the curation overlay. Members are folded into a
    # synthesized Verified unified row keyed by the target id.
    merge_into, merge_members = {}, {}
    for _cid, _c in curation.items():
        _t = _c.get("merge_into")
        if _t:
            merge_into[_cid] = _t
            merge_members.setdefault(_t, []).append(_cid)

    colleges, col_idx = [], {}

    def cidx(name):
        if name not in col_idx:
            col_idx[name] = len(colleges); colleges.append(name)
        return col_idx[name]

    def rollup(mids):
        """(adopted, potential) college name lists rolled up across member M-IDs.
        adopted = colleges that earned >=1 articulation; potential = offering minus
        adopted. Empty unless at least one member actually carries an articulation."""
        off, ad, any_art = set(), set(), False
        for m in mids:
            ident = art_ident.get(m)
            if ident:
                any_art = True
                off |= set(ident.get("colleges_offering", []))
                ad |= set(earned.get(m, set()))
        return (sorted(ad), sorted(off - ad)) if any_art else ([], [])

    def flags_of(v, cid, use_spread=True):
        sp = v.get("subject_spread", 1) or 1
        return {
            "over_merged": bool(use_spread and sp >= 8),
            "credit_mixed": bool(v.get("credit_status_mixed")),
            "top_mixed": bool(v.get("top_code_mixed")),
            "ncc_mixed": bool(v.get("noncredit_category_mixed")),
            "reviewed": bool(v.get("reviewed_at") or curation.get(cid)),
        }

    rows = []
    for mid, v in cat.items():
        if mid in merge_into or mid in merge_members:
            continue
        ad, pot = rollup([mid])
        rows.append(_add_prov({"kind": "Course", "id": mid, "title": v.get("common_title"),
                     "disc": disc_of(mid, v.get("discipline")), "credit": v.get("credit_status"),
                     "units": v.get("typical_units"), "top": v.get("top_code"),
                     "subj": [v["subject"]] if v.get("subject") else [],
                     "members": v.get("corroboration_members"), "conf": v.get("confidence"),
                     "id_system": "M-ID", "locked": False,
                     "flags": flags_of(v, mid),
                     "reviewed_by": reviewed_by_of(mid), "reviewed_at": reviewed_at_of(mid),
                     "adopted": [cidx(c) for c in ad], "potential": [cidx(c) for c in pot]}, mid, v))
    for uid, v in clusters.items():
        if uid in merge_into or uid in merge_members:
            continue
        ad, pot = rollup(v.get("members", []))
        rows.append(_add_prov({"kind": "Cluster", "id": uid,
                     "title": v.get("synthesized_title") or v.get("canonical_title"),
                     "disc": disc_of(uid, v.get("discipline")), "credit": v.get("credit_status"),
                     "units": v.get("typical_units"), "top": v.get("top_code"),
                     "subj": v.get("subjects", []), "members": v.get("member_count"), "conf": None,
                     "title_variants": v.get("title_variants", []),
                     "id_system": "Cluster", "locked": False,
                     "flags": flags_of(v, uid, use_spread=False),
                     "reviewed_by": reviewed_by_of(uid), "reviewed_at": reviewed_at_of(uid),
                     "adopted": [cidx(c) for c in ad], "potential": [cidx(c) for c in pot]}, uid, v))

    # Curated common-course anchor (C-ID / CCN-ID / M-ID) — shown READ-ONLY:
    # curation is disabled for these (common_courses.json is firewalled). Also
    # makes the Source filter meaningful, since the COCI staging is all M-ID.
    cc = _load("common_courses.json") or {}
    seen = set(cat.keys()) | set(clusters.keys())
    curated_n = 0
    for ccid, v in cc.items():
        if ccid in seen or ccid in merge_into or ccid in merge_members:
            continue
        curated_n += 1
        ad, pot = rollup([ccid])
        rows.append({"kind": "Course", "id": ccid, "title": v.get("common_title"),
                     "disc": v.get("discipline"), "credit": None,
                     "units": v.get("typical_units"), "top": None,
                     "subj": [v["subject"]] if v.get("subject") else [],
                     "members": v.get("source_college_count"), "conf": v.get("confidence"),
                     "id_system": v.get("id_system"), "locked": True,
                     "flags": {"over_merged": False, "credit_mixed": False, "top_mixed": False,
                               "ncc_mixed": False, "reviewed": True},
                     "reviewed_by": v.get("reviewed_by"), "reviewed_at": (v.get("reviewed_at") or "")[:10],
                     "adopted": [cidx(c) for c in ad], "potential": [cidx(c) for c in pot]})

    # CCN-ID anchor (AB 1111 Common Course Numbers) — read-only/locked reference
    # rows from kb/reference/ccn_courses.json, mirroring the curated anchor above.
    # Gives CCN-ID a Source-filter value and a merge target. Skips any CCN already
    # present in the staging/anchor sets so a future CCN promotion doesn't double-emit.
    ccn_doc = _load(os.path.join("reference", "ccn_courses.json")) or {}
    ccn_courses = ccn_doc.get("courses", [])
    seen_ids = seen | set(cc.keys()) | set(merge_into) | set(merge_members)
    ccn_n = 0
    for c in ccn_courses:
        ccid = c.get("ccn")
        if not ccid or ccid in seen_ids:
            continue
        ccn_n += 1
        rows.append({"kind": "Course", "id": ccid, "title": c.get("title"),
                     "disc": None, "credit": None, "units": None, "top": None,
                     "subj": [c["subject"]] if c.get("subject") else [],
                     "members": None, "conf": None,
                     "id_system": "CCN-ID", "locked": True,
                     "flags": {"over_merged": False, "credit_mixed": False, "top_mixed": False,
                               "ncc_mixed": False, "reviewed": True},
                     "reviewed_by": None, "reviewed_at": "",
                     "adopted": [], "potential": []})

    # Reviewer-consolidated unified courses (from merge_into curations) — one
    # synthesized Verified row per target id, with its members folded in.
    def _member_v(m):
        return cat.get(m) or sg.get(m) or cc.get(m) or clusters.get(m) or {}

    def _member_title(m):
        v = _member_v(m)
        return v.get("common_title") or v.get("synthesized_title") or v.get("canonical_title")

    # ── Cluster field aggregation (formerly hardcoded to None — caused
    # UC-CUR-MPG029OM to show "—/—/—/—" for Credit/Units/TOP/Conf even though
    # all 3 members carried unanimous Credit + TOP 0949.00). Walk members,
    # aggregate per field: unanimous → that value; modal (≥⅔ agree) → that
    # value + mixed flag; varied → None. Confidence synthesized as
    # mean(member_conf) ÷ coherence_factor. Curation overrides win if set.
    # See kb/_row_audit.py for the audit-side mirror; this is the
    # render-side. Phase 1b — first half (rendering fix).
    def _agg_unanimous(values):
        """Return (value, mixed_flag) — value if all non-null agree, else (None|modal, True)."""
        vals = [v for v in values if v is not None]
        if not vals:
            return None, False
        from collections import Counter as _C
        c = _C(vals)
        top_val, top_n = c.most_common(1)[0]
        if len(c) == 1:
            return top_val, False
        if top_n / len(vals) >= 0.67:
            return top_val, True
        return None, True  # truly varied — withhold a single value

    def _synth_cluster_conf(member_recs):
        confs = [r.get("confidence") for r in member_recs if r.get("confidence") is not None]
        if not confs:
            return None
        mean = sum(confs) / len(confs)
        credits = [r.get("credit_status") for r in member_recs if r.get("credit_status")]
        tops    = [r.get("top_code") for r in member_recs if r.get("top_code")]
        coherent = (len(set(credits)) <= 1) and (len(set(tops)) <= 1)
        factor = 0.85 if coherent else 0.70
        return round(min(1.0, mean / max(factor, 0.01)), 3)

    for tgt, members in sorted(merge_members.items()):
        cur = curation.get(tgt, {})
        # If the target is itself an existing course (merge INTO an M-ID/C-ID),
        # include it as a member so its title/subject/articulations are folded in.
        tgt_v = _member_v(tgt)
        all_members = ([tgt] if tgt_v else []) + list(members)
        member_recs = [_member_v(m) for m in all_members]
        subs = sorted({mv.get("subject") for mv in member_recs if mv.get("subject")})
        variants = [t for t in (_member_title(m) for m in all_members) if t]
        ad, pot = rollup(all_members)  # only M-ID members contribute articulations

        c_credit, credit_mixed = _agg_unanimous([mv.get("credit_status") for mv in member_recs])
        c_top,    top_mixed    = _agg_unanimous([mv.get("top_code")      for mv in member_recs])
        c_units,  _units_mixed = _agg_unanimous([mv.get("typical_units") for mv in member_recs])
        c_conf                 = _synth_cluster_conf(member_recs)

        rows.append({"kind": "Cluster", "id": tgt,
                     "title": cur.get("unified_title") or _member_title(tgt) or (variants[0] if variants else tgt),
                     "disc": cur.get("discipline") or tgt_v.get("discipline"),
                     "credit": cur.get("credit_status") or c_credit,
                     "units":  cur.get("typical_units") if cur.get("typical_units") is not None else c_units,
                     "top":    cur.get("top_code")      or c_top,
                     "subj": subs, "members": len(all_members),
                     "conf": cur.get("confidence_override") if cur.get("confidence_override") is not None else c_conf,
                     "id_system": "Cluster", "locked": False, "title_variants": variants,
                     "flags": {"over_merged": False,
                               "credit_mixed": credit_mixed and not cur.get("credit_status"),
                               "top_mixed":    top_mixed    and not cur.get("top_code"),
                               "ncc_mixed": False, "reviewed": True},
                     "reviewed_by": cur.get("reviewed_by"), "reviewed_at": (cur.get("reviewed_at") or "")[:10],
                     "adopted": [cidx(c) for c in ad], "potential": [cidx(c) for c in pot]})

    # NOTE: unified_courses_data.js is written further down, AFTER the raw COCI
    # list is indexed, so each row can carry its matched official C-ID/CCN
    # (Phase A crosswalk surfacing). See "write unified_courses_data.js" below.

    # NOTE: the "Generate unified course" search index (unified_courses_index.js)
    # is written further down, AFTER Phase B consolidation, so it reflects the
    # final identity set (consumed M-IDs dropped, official-ID rows added).

    # ---- raw COCI course list — read ONCE, feeds descriptions + member rows ---
    # The per-college course list carries CatalogDescription / CIDNumber /
    # CommonCourseNumber per local course. We index it for two lazy artifacts:
    # description fallbacks (below) and the member-college rows (further down).
    raw_xlsx = os.path.join(kdir, "reference", "coci_course_list.xlsx")
    _have_raw = os.path.exists(raw_xlsx)

    def _nrm(x):
        return " ".join(str(x).split()).upper() if x is not None else ""

    mcolleges, mcol_idx = [], {}

    def _mc(name):
        if name not in mcol_idx:
            mcol_idx[name] = len(mcolleges); mcolleges.append(name)
        return mcol_idx[name]

    # join key -> [ {c, n, t, d, cid, ccn} ] (d = CatalogDescription, kept in
    # memory only; stripped from the members file output). One ent object is
    # shared across the pair/cid/ccn indexes it belongs to.
    cid_rows, ccn_rows, cn_rows = {}, {}, {}
    top_titles = {}  # TOP code -> program title, deduped (raw TopCode is "code: title")
    # CIDNumber / CommonCourseNumber cells sometimes carry a literal sentinel
    # instead of being blank — treat these as "no official id".
    _NULLISH = {"", "NULL", "N/A", "NA", "NONE", "NOT APPLICABLE", "NOT APPLICABLE.", "TBD", "-"}

    if _have_raw:
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(raw_xlsx, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rit = ws.iter_rows(values_only=True)
        next(rit)  # header
        for row in rit:
            college, subj, num, title = row[0], row[2], row[3], row[4]
            units, cid, ccn, desc, topc = row[5], row[9], row[11], row[10], row[8]
            ctrl = _nrm(row[1])  # CourseControlNumber — exact member-join key (re-mint)
            if not college:
                continue
            code = (f"{subj} {num}".strip() if subj is not None
                    else (str(num) if num is not None else ""))
            cidn, ccnn = _nrm(cid), _nrm(ccn)
            if cidn in _NULLISH:
                cidn = ""
            if ccnn in _NULLISH:
                ccnn = ""
            tcode = ""
            if topc and str(topc).strip():
                ts = str(topc).strip()
                tcode, _, ttitle = ts.partition(":")
                tcode = tcode.strip()
                if tcode and ttitle.strip():
                    top_titles.setdefault(tcode, ttitle.strip())
            ent = {"c": _mc(str(college)), "n": code, "t": title or "",
                   "d": str(desc) if (desc and str(desc).strip()) else "",
                   "u": units if isinstance(units, (int, float)) else None,
                   "p": tcode,
                   "cid": cidn, "ccn": ccnn}
            if ctrl:
                cn_rows.setdefault(ctrl, []).append(ent)
            if cidn:
                cid_rows.setdefault(cidn, []).append(ent)
            if ccnn:
                ccn_rows.setdefault(ccnn, []).append(ent)
        wb.close()

    memships = (_load("coci_minted_memberships.json") or {}).get("memberships", {})
    promotions = (_load("promotions.json") or {}).get("promotions", {})

    # Member-join is now EXACT (the re-mint): each membership member carries its
    # own College/CourseControlNumber, so a displayed identity's member college
    # courses are looked up by control number — no (subject, number) ambiguity and
    # no title-Jaccard filter. C-ID/CCN rows join on the official id.
    def _id_cid(mid):
        cv = cc.get(mid)
        return _nrm(cv["c_id"]) if cv and cv.get("c_id") else ""

    def _leaf_ids(r):
        """Constituent minted/singleton identity ids for a displayed row
        (expands clusters + reviewer merge members)."""
        leaves = []

        def add_leaf(i):
            clv = clusters.get(i)
            if clv:
                for m in clv.get("members", []):
                    add_leaf(m)
            else:
                leaves.append(i)

        for i in [r["id"]] + merge_members.get(r["id"], []):
            add_leaf(i)
        return leaves

    def _leaf_cns(i):
        if i in memships:
            return [_nrm(p.get("control_number")) for p in memships[i] if p.get("control_number")]
        sv = sg.get(i)
        if sv and sv.get("control_number"):
            return [_nrm(sv["control_number"])]
        return []

    # Candidate college-course ents for a displayed row — exact control-number join.
    def _row_candidates(r):
        rid, idsys = r["id"], r.get("id_system")
        if idsys == "CCN-ID":
            return ccn_rows.get(_nrm(rid), [])
        if idsys == "C-ID":
            return cid_rows.get(_id_cid(rid) or _nrm(rid), [])
        out = []
        for i in _leaf_ids(r):
            for cn in _leaf_cns(i):
                out += cn_rows.get(cn, [])
        return out

    # Member entries (deduped, description stripped) for a displayed row.
    def _row_ents(r):
        seen, out = set(), []
        for ent in _row_candidates(r):
            k = (ent["c"], ent["n"], ent["t"])
            if k in seen:
                continue
            seen.add(k); out.append({"c": ent["c"], "n": ent["n"], "t": ent["t"],
                                     "u": ent.get("u"), "p": ent.get("p") or "", "d": ent.get("d") or ""})
        return out

    # Representative raw CatalogDescription for a row (longest among candidates).
    def _row_rawdesc(r):
        best = None
        for ent in _row_candidates(r):
            d = ent.get("d")
            if d and (best is None or len(d) > len(best)):
                best = d
        return best

    # Phase A crosswalk surfacing — sourced from the kb promotions manifest
    # (computed control-number-exact in the re-mint), NOT by scanning member CIDs:
    # the new memberships are remnant-only and carry none. Aggregates the official
    # C-ID/CCN target(s) of the row's constituent minted identities: {"cid": "ACCT
    # 110"} single agreed, {"cid_conflict": [...]} when they disagree (over-merge
    # signal), {"ccn": "..."} for a CCN. No identity change — purely informational.
    def _row_official(r):
        if r.get("id_system") in ("C-ID", "CCN-ID"):
            return None  # already an official-ID identity
        cids, ccns = set(), set()
        for i in _leaf_ids(r):
            for oid in (promotions.get(i, {}).get("official_targets") or {}):
                if oid.startswith("CCN:"):
                    ccns.add(oid[4:])
                elif oid.startswith("C-ID:"):
                    cids.add(oid[5:])
        out = {}
        if len(ccns) == 1:
            out["ccn"] = next(iter(ccns))
        if len(cids) == 1:
            out["cid"] = next(iter(cids))
        elif len(cids) > 1:
            out["cid_conflict"] = sorted(cids)
        return out or None

    # ---- write unified_courses_data.js (deferred so rows carry matched IDs) --
    if _have_raw:
        for r in rows:
            m = _row_official(r)
            if m:
                r["match"] = m

    # ---- Phase B: consolidate-by-official-ID (inline, regen-safe) ------------
    # Decision 2026-05-22: group every minted/cluster row whose title-consistent
    # members agree on ONE official C-ID or CCN into a single official-identity
    # row — folding under an existing anchor when one exists, else synthesizing
    # the official row. Collapses the N:1 fragmentation (many minted M-IDs that
    # are really the same C-ID course). NEVER touches cid_conflict rows. No KB
    # mutation: the consolidation is recomputed each run and the underlying M-ID
    # keys are recorded on the row (`consolidated_from`) + registered in
    # merge_into/merge_members so the lazy member/detail joins fold correctly and
    # curation/articulation pointers survive. Only acts when it changes something
    # (an anchor to fold into, or >1 M-ID claiming one id); a lone M-ID with no
    # anchor keeps just its Phase A badge.
    if _have_raw:
        anchor_by_id = {r["id"]: r for r in rows
                        if r.get("locked") and r.get("id_system") in ("C-ID", "CCN-ID")}
        groups = {}
        for r in rows:
            if r.get("locked") or r.get("id_system") not in ("M-ID", "Cluster"):
                continue
            m = r.get("match") or {}
            if m.get("ccn"):
                oid, osys = m["ccn"], "CCN-ID"
            elif m.get("cid") and not m.get("cid_conflict"):
                oid, osys = m["cid"], "C-ID"
            else:
                continue
            groups.setdefault(oid, {"sys": osys, "members": []})["members"].append(r)

        consumed, synthesized, folds = set(), [], 0
        for oid, g in groups.items():
            mems = g["members"]
            anchor = anchor_by_id.get(oid)
            if not anchor and len(mems) < 2:
                continue
            if anchor:
                folds += 1
            mem_ids = [r["id"] for r in mems]
            for mid_ in mem_ids:
                consumed.add(mid_)
                merge_into[mid_] = oid
                merge_members.setdefault(oid, []).append(mid_)
            ad, pot = rollup(mem_ids)
            subj = sorted({s for r in mems for s in (r.get("subj") or [])}
                          | set((anchor or {}).get("subj") or []))
            variants = sorted({r.get("title") for r in mems if r.get("title")})
            dc = {}
            for r in mems:
                if r.get("disc"):
                    dc[r["disc"]] = dc.get(r["disc"], 0) + 1
            modal_disc = max(dc, key=dc.get) if dc else None
            target = anchor or {
                "kind": "Course", "id": oid, "id_system": g["sys"], "locked": False,
                "credit": None, "units": None, "top": None, "conf": None,
                "flags": {"over_merged": False, "credit_mixed": False,
                          "top_mixed": False, "ncc_mixed": False, "reviewed": False},
                "reviewed_by": None, "reviewed_at": "",
            }
            target["consolidated_from"] = mem_ids
            target["subj"] = subj
            target["members"] = len(mem_ids) + (1 if anchor else 0)
            tv = sorted(set(target.get("title_variants") or []) | set(variants))
            if tv:
                target["title_variants"] = tv
            if not target.get("title"):
                target["title"] = variants[0] if variants else oid
            if not target.get("disc") and modal_disc:
                target["disc"] = modal_disc
            target["adopted"] = sorted(set(target.get("adopted") or []) | {cidx(c) for c in ad})
            target["potential"] = sorted(set(target.get("potential") or []) | {cidx(c) for c in pot})
            target.pop("match", None)  # it IS the official id now
            if not anchor:
                synthesized.append(target)
        if consumed:
            rows = [r for r in rows if r["id"] not in consumed]
            rows.extend(synthesized)
            print(f"  Unified Courses: Phase B consolidated {len(consumed)} M-IDs into "
                  f"{len(synthesized)} new official-ID rows + {folds} anchor folds")

    # Compact all-course title index for the "Generate unified course" dialog —
    # a separate file the tab lazy-loads only when a curator opens it. Built from
    # the FINAL row set (post-Phase-B) plus stand-alone singletons, so consumed
    # M-IDs aren't offered as ghost merge targets and the new official-ID
    # identities are searchable. Each entry: [id, title, subject(s), kind, units].
    out_idx = os.path.join(odir, "unified_courses_index.js")
    idx = []
    for r in rows:
        k = r.get("id_system") if r.get("id_system") in ("C-ID", "CCN-ID") else r.get("kind")
        idx.append([r["id"], r.get("title"), ";".join(r.get("subj") or []), k, r.get("units")])
    for sid, v in sg.items():
        if sid not in merge_into:
            idx.append([sid, v.get("common_title"), v.get("subject"), "Stand-Alone", v.get("typical_units")])
    with open(out_idx, "w", encoding="utf-8") as f:
        f.write("/* Unified Courses search index — id,title,subject,kind. Lazy-loaded by the curation dialog. */\n"
                "window.CPL_UC_INDEX = " + json.dumps(idx, ensure_ascii=False, separators=(",", ":")) + ";\n")
    print(f"  Unified Courses: wrote {out_idx} ({len(idx)} index entries)")

    # ---- suggested-merge worklist (lazy) ------------------------------------
    # Precompute groups of courses that look like the SAME course, so a reviewer
    # can confirm merges in one click. Grouping key is a level-SAFE title
    # signature: parentheticals removed, articles dropped, roman numerals
    # normalized to digits, remaining tokens sorted — so "Japanese I"/"Japanese 1"
    # group but "Japanese I" and "Japanese II" do NOT. NEVER auto-applied — purely
    # a worklist surfaced in the tab; the curator confirms each group. (Generated
    # file — add to the daily workflow git-add list.)
    #
    # Two sections (V2):
    #  - groups  (identity-anchored) — every group has >=1 main-payload identity
    #    (M-ID/Cluster, not locked, not a cid_conflict over-merge) + >=2 members;
    #    orphan singletons whose signature matches an identity are attached.
    #    Confirming MERGES into that existing identity. Higher value, shown first.
    #  - singleton_groups (NEW in V2) — signatures with >=2 single-college
    #    (Stand-Alone) courses that match NO existing identity. Confirming MINTS a
    #    brand-new unified course from them. Each carries `same_college`: True when
    #    every member resolves to one college (likely intra-college variant ladders
    #    / credit-noncredit / language pairs, NOT cross-college duplicates) — these
    #    are flagged in the UI and ranked last so genuine cross-college candidates
    #    surface first. (Decision 2026-05-22: include but flag + rank last.)
    from collections import Counter as _Ctr
    _SUG_DROP = {"the", "of", "to", "and", "for", "with", "in", "a", "an", "on", "at", "as"}
    _SUG_ROMAN = {"i": "1", "ii": "2", "iii": "3", "iv": "4", "v": "5",
                  "vi": "6", "vii": "7", "viii": "8", "ix": "9", "x": "10"}

    def _sug_sig(t):
        t = re.sub(r"\([^)]*\)", " ", str(t or "").lower())
        t = re.sub(r"[^a-z0-9 ]+", " ", t)
        return " ".join(sorted(_SUG_ROMAN.get(w, w) for w in t.split() if w not in _SUG_DROP))

    def _sug_score(members):
        # cohesion: subject agreement (modal share) + units agreement + size.
        subc = _Ctr(x for m in members for x in str(m["s"] or "").split(";") if x)
        modal = (subc.most_common(1)[0][1] / len(members)) if subc else 0.0
        us = [m["u"] for m in members if m["u"] is not None]
        units_ok = (sum(1 for u in us if abs(u - us[0]) < 0.5) / len(members)) if us else 0.0
        return round(0.5 * modal + 0.3 * units_ok + 0.2 * min(len(members), 10) / 10.0, 3)

    sug = {}
    for r in rows:
        if r.get("locked") or r.get("id_system") not in ("M-ID", "Cluster"):
            continue
        if (r.get("match") or {}).get("cid_conflict"):
            continue
        s = _sug_sig(r.get("title"))
        if not s:
            continue
        sug.setdefault(s, {"main": [], "sing": []})["main"].append(
            {"id": r["id"], "t": r.get("title"), "s": ";".join(r.get("subj") or []),
             "u": r.get("units"), "k": r.get("id_system")})
    # Singletons: attach to an anchored signature when one exists; otherwise
    # bucket by signature as a singleton-only merge candidate (V2).
    sing_only = {}
    for sid, v in sg.items():
        if sid in merge_into:
            continue
        s = _sug_sig(v.get("common_title"))
        if not s:
            continue
        m = {"id": sid, "t": v.get("common_title"), "s": v.get("subject") or "",
             "u": v.get("typical_units"), "k": "Stand-Alone", "g": 1}
        g = sug.get(s)
        if g:  # attach only to an existing main-anchored signature
            g["sing"].append(m)
        else:
            sing_only.setdefault(s, []).append(m)

    sug_groups = []
    for s, g in sug.items():
        members = g["main"] + g["sing"]
        if not g["main"] or len(members) < 2:
            continue
        sug_groups.append({"sig": s, "n": len(members), "score": _sug_score(members),
                           "members": members})
    sug_groups.sort(key=lambda x: -x["score"])

    # Singleton-only groups (>=2 stand-alone members, no anchor). Compute a
    # same_college flag via the title-filtered forward join to the raw COCI list
    # (the same join the member-row build uses), so intra-college variant ladders
    # are flagged and demoted rather than presented as cross-college duplicates.
    def _sing_colleges(sid):
        if not _have_raw:
            return None
        return {ent["c"] for ent in _row_candidates({"id": sid, "id_system": "M-ID"})}

    singleton_groups = []
    for s, members in sing_only.items():
        if len(members) < 2:
            continue
        cols = set()
        for m in members:
            c = _sing_colleges(m["id"])
            if c:
                cols |= c
        # same_college True only when we positively resolved colleges and they
        # collapse to one; unknown (raw list absent / no match) -> not flagged.
        same_college = bool(cols) and len(cols) <= 1
        singleton_groups.append({"sig": s, "n": len(members), "score": _sug_score(members),
                                 "same_college": same_college, "members": members})
    # Cross-college candidates first (by cohesion); flagged same-college last.
    singleton_groups.sort(key=lambda x: (x["same_college"], -x["score"]))

    out_sug = os.path.join(odir, "unified_courses_suggestions.js")
    _sc_flagged = sum(1 for g in singleton_groups if g["same_college"])
    sug_payload = {"generated_at": _dt.now().strftime("%Y-%m-%d %H:%M"),
                   "count": len(sug_groups), "groups": sug_groups,
                   "singleton_count": len(singleton_groups),
                   "singleton_groups": singleton_groups}
    with open(out_sug, "w", encoding="utf-8") as f:
        f.write("/* Unified Courses suggested-merge worklist — lazy-loaded. groups = "
                "identity-anchored same-title merges; singleton_groups = NEW unified "
                "courses minted from single-college matches (same_college flag = likely "
                "intra-college variants). HUMAN-CONFIRMED in the tab, NEVER auto-applied. */\n"
                "window.CPL_UC_SUGGESTIONS = " + json.dumps(sug_payload, ensure_ascii=False, separators=(",", ":")) + ";\n")
    print(f"  Unified Courses: wrote {out_sug} ({len(sug_groups)} anchored + "
          f"{len(singleton_groups)} singleton-only groups [{_sc_flagged} same-college flagged])")

    mq = (_load(os.path.join("reference", "mq_disciplines.json")) or {}).get("disciplines", [])
    payload = {"generated_at": _dt.now().strftime("%Y-%m-%d %H:%M"), "beta": True,
               "colleges": colleges, "mq_disciplines": sorted(mq),
               "count_inbrowser": len(rows),
               "count_total": len(cat) + len(sg) + len(clusters) + curated_n + ccn_n,
               # What's already folded into git (kb/coci_curation.json at build
               # time) — the client diffs the live Supabase overlay against this
               # to count edits still awaiting the daily sync.
               "committed_curation": {cid: c.get("discipline") for cid, c in curation.items()},
               # Descriptions already folded into git (kb/coci_curation.json) — the
               # client diffs the live description overlay against this to count
               # description edits still awaiting the daily sync.
               "committed_descriptions": {cid: c["description"] for cid, c in curation.items() if c.get("description")},
               # TOP code -> program title (small, ~400 entries) so the main
               # list can show the TOP title on hover without the lazy member file.
               "topmap": top_titles,
               "export_path": "exports/unified_courses.xlsx", "rows": rows}
    with open(out_js, "w", encoding="utf-8") as f:
        f.write("/* Unified Courses (COCI identity layer) — auto-generated. AI-assisted STAGING. */\n"
                "window.CPL_UNIFIED_COURSES = " + json.dumps(payload, ensure_ascii=False) + ";\n")
    print(f"  Unified Courses: wrote {out_js} ({len(rows)} in-browser rows, {len(colleges)} colleges)")

    # ---- lazy course-description detail file --------------------------------
    # Descriptions live in a SEPARATE file the tab lazy-loads only when a reviewer
    # opens the row-details modal, keeping unified_courses_data.js lean. Precedence
    # per id: curated (kb_curation) > existing layer (minted/synthesized/C-ID) >
    # raw CatalogDescription fallback from the per-college list.
    coci_ref = (_load(os.path.join("reference", "coci_courses.json")) or {}).get("courses", {})
    base_desc = {}  # course_id -> {"d": text, "s": source}
    for mid, v in cat.items():
        if v.get("description"):
            base_desc[mid] = {"d": v["description"], "s": v.get("description_source") or ""}
    for uid, v in clusters.items():
        if v.get("synthesized_description"):
            base_desc[uid] = {"d": v["synthesized_description"], "s": "synthesized"}
    for ccid, v in cc.items():
        if v.get("description"):
            base_desc[ccid] = {"d": v["description"], "s": v.get("description_source") or ""}
    for ccid, v in coci_ref.items():
        if v.get("description") and ccid not in base_desc:
            base_desc[ccid] = {"d": v["description"], "s": v.get("id_system") or "COCI"}

    # Emit details for in-browser rows AND stand-alone courses (so their modal has
    # a description too). Iterate in stable order so output key order is stable.
    details, _filled_raw = {}, 0

    def _detail_for(r):
        nonlocal _filled_raw
        cid = r["id"]
        if cid in details:
            return
        cd = cur_desc_of(cid)
        if cd:
            details[cid] = {"d": cd, "s": "curated"}
        elif cid in base_desc:
            details[cid] = base_desc[cid]
        else:
            rd = _row_rawdesc(r)
            if rd:
                details[cid] = {"d": rd, "s": "local catalog (COCI)"}
                _filled_raw += 1

    for r in rows:
        _detail_for(r)
    for sid, v in sg.items():
        if sid in merge_into or sid in merge_members:
            continue
        _detail_for({"id": sid, "id_system": "M-ID"})
    out_det = os.path.join(odir, "unified_courses_details.js")
    with open(out_det, "w", encoding="utf-8") as f:
        f.write("/* Unified Courses descriptions — id -> {d:description, s:source}. "
                "Lazy-loaded by the row-details modal. Curated descriptions override the base. */\n"
                "window.CPL_UC_DETAILS = " + json.dumps(details, ensure_ascii=False, separators=(",", ":")) + ";\n")
    print(f"  Unified Courses: wrote {out_det} ({len(details)} descriptions, {_filled_raw} filled from raw list)")

    # ---- lazy stand-alone (singleton) rows ----------------------------------
    # The ~57k single-college courses are deliberately kept OUT of the main
    # payload (it would balloon to tens of MB). They live in their own file the
    # tab loads only when a reviewer selects the "Stand-Alone" kind filter, so
    # they become visible + curatable without slowing the default view.
    sa_rows = []
    for sid, v in sg.items():
        if sid in merge_into or sid in merge_members:
            continue
        sa_rows.append(_add_prov({"kind": "Stand-Alone", "id": sid, "title": v.get("common_title"),
                        "disc": disc_of(sid, v.get("discipline")), "credit": v.get("credit_status"),
                        "units": v.get("typical_units"), "top": v.get("top_code"),
                        "subj": [v["subject"]] if v.get("subject") else [],
                        "members": 1, "conf": v.get("confidence"),
                        "id_system": "M-ID", "locked": False,
                        "flags": flags_of(v, sid),
                        "reviewed_by": reviewed_by_of(sid), "reviewed_at": reviewed_at_of(sid),
                        "adopted": [], "potential": []}, sid, v))
    if _have_raw:
        for r in sa_rows:
            m = _row_official(r)
            if m:
                r["match"] = m
    out_sa = os.path.join(odir, "unified_courses_standalone.js")
    sa_payload = {"generated_at": _dt.now().strftime("%Y-%m-%d %H:%M"), "rows": sa_rows}
    with open(out_sa, "w", encoding="utf-8") as f:
        f.write("/* Unified Courses stand-alone (single-college) rows — lazy-loaded "
                "only when the 'Stand-Alone' kind filter is selected. */\n"
                "window.CPL_UC_STANDALONE = " + json.dumps(sa_payload, ensure_ascii=False, separators=(",", ":")) + ";\n")
    print(f"  Unified Courses: wrote {out_sa} ({len(sa_rows)} stand-alone rows)")

    # ---- lazy member-college rows (forward join to the raw COCI course list) -
    # For each displayed identity, list the member college courses that roll up
    # to it (forward join via _row_ents — see the shared index build above).
    # Forward-joining avoids guessing on ambiguous (subject, number) pairs: an
    # over-merged identity may surface a course under multiple (already-flagged)
    # cards rather than picking one. Lazy-loaded only when a row is expanded.
    if _have_raw:
        # members.js stays lean: per member college course we keep code/title +
        # units (u) + the self-describing TOP code (p, e.g. "0949.00: Automotive
        # Collision Repair"). The CatalogDescription is heavy (~56MB across all
        # member rows) so it is split into a SEPARATE on-demand file the tab loads
        # only when a curator asks to see member descriptions. mdesc[id] is a list
        # PARALLEL to members[id] (same order), each truncated to 500 chars.
        members, mdesc = {}, {}
        for r in rows + sa_rows:
            ents = _row_ents(r)
            if not ents:
                continue
            members[r["id"]] = [{"c": e["c"], "n": e["n"], "t": e["t"],
                                 "u": e.get("u"), "p": e.get("p") or ""} for e in ents]
            ds = [(e.get("d") or "")[:500] for e in ents]
            if any(ds):
                mdesc[r["id"]] = ds

        out_mem = os.path.join(odir, "unified_courses_members.js")
        mem_payload = {"generated_at": _dt.now().strftime("%Y-%m-%d %H:%M"),
                       "colleges": mcolleges, "members": members, "topmap": top_titles}
        with open(out_mem, "w", encoding="utf-8") as f:
            f.write("/* Unified Courses member-college rows — id -> [{c:collegeIdx, n:code, t:title, u:units, p:topcode}]. "
                    "topmap maps a TOP code -> program title. Lazy-loaded when a row is expanded. "
                    "colleges[] holds the names. Descriptions are in unified_courses_member_desc.js. */\n"
                    "window.CPL_UC_MEMBERS = " + json.dumps(mem_payload, ensure_ascii=False, separators=(",", ":")) + ";\n")
        total = sum(len(v) for v in members.values())
        print(f"  Unified Courses: wrote {out_mem} ({len(members)} identities, {total} member rows, {len(mcolleges)} colleges)")

        out_md = os.path.join(odir, "unified_courses_member_desc.js")
        md_payload = {"generated_at": _dt.now().strftime("%Y-%m-%d %H:%M"), "desc": mdesc}
        with open(out_md, "w", encoding="utf-8") as f:
            f.write("/* Unified Courses member-course descriptions — id -> [desc,...] PARALLEL to "
                    "unified_courses_members.js members[id] (each truncated to 500 chars). On-demand: "
                    "loaded only when a curator opens member descriptions. */\n"
                    "window.CPL_UC_MEMBER_DESC = " + json.dumps(md_payload, ensure_ascii=False, separators=(",", ":")) + ";\n")
        print(f"  Unified Courses: wrote {out_md} ({len(mdesc)} identities with member descriptions)")
    else:
        print("  Unified Courses: kb/reference/coci_course_list.xlsx absent — skipped member rows")

    # ---- full xlsx export (Course + Cluster + Singleton, incl. college name lists) ----
    headers = ["Kind", "ID", "Title", "Discipline", "Credit Status", "Units", "TOP Code",
               "Subject(s)", "Members", "Confidence", "Over-merged", "Credit mixed", "TOP mixed",
               "Noncredit mixed", "Reviewed", "Curated by", "Curated on", "Adopted (count)",
               "Adopted colleges", "Adoptable (count)", "Adoptable colleges"]
    xrows = []

    def xrow(kind, cid, title, disc, credit, units, top, subj, members, conf, fl, mids):
        ad, pot = rollup(mids)
        xrows.append([kind, cid, title, disc, credit, units, top, "; ".join(subj or []),
                      members, ("" if conf is None else conf),
                      *["Y" if fl[k] else "" for k in ("over_merged", "credit_mixed", "top_mixed", "ncc_mixed", "reviewed")],
                      reviewed_by_of(cid) or "", reviewed_at_of(cid),
                      len(ad), "; ".join(ad), len(pot), "; ".join(pot)])

    for mid, v in cat.items():
        xrow("Course", mid, v.get("common_title"), disc_of(mid, v.get("discipline")), v.get("credit_status"),
             v.get("typical_units"), v.get("top_code"), [v.get("subject")] if v.get("subject") else [],
             v.get("corroboration_members"), v.get("confidence"), flags_of(v, mid), [mid])
    for sid, v in sg.items():
        xrow("Stand-Alone", sid, v.get("common_title"), disc_of(sid, v.get("discipline")), v.get("credit_status"),
             v.get("typical_units"), v.get("top_code"), [v.get("subject")] if v.get("subject") else [],
             1, v.get("confidence"), flags_of(v, sid), [sid])
    for uid, v in clusters.items():
        xrow("Cluster", uid, v.get("synthesized_title") or v.get("canonical_title"), disc_of(uid, v.get("discipline")),
             v.get("credit_status"), v.get("typical_units"), v.get("top_code"), v.get("subjects", []),
             v.get("member_count"), None, flags_of(v, uid, use_spread=False), v.get("members", []))

    _write_analytics_xlsx_export("unified_courses", "Unified Courses", headers, xrows,
                                 os.path.join(odir, "exports"))
    print(f"  Unified Courses: wrote exports/unified_courses.xlsx ({len(xrows)} rows)")


def render_exhibit_analysis_html(tables, kpi_params=None, xlsx_export_dir=None):
    """
    Render exhibit analysis tables as scrollable HTML cards inside a
    collapsible "CPL Analytics" section that mirrors the KPI Metrics header.
    Per-table xlsx exports are written to xlsx_export_dir (relative to the
    dashboard) so the in-card Export buttons link to a real file.
    """
    if not tables:
        return ""

    def fmt(n):
        return f"{n:,}" if isinstance(n, int) else str(n)

    def pct_bar(pct_val):
        w = min(pct_val, 100)
        return (f'<div style="display:inline-block;width:60px;height:8px;background:rgba(255,255,255,0.1);'
                f'border-radius:4px;vertical-align:middle;margin-left:6px;">'
                f'<div style="width:{w}%;height:100%;background:#C9A84C;border-radius:4px;"></div></div>')

    # ── Card builder helper ──
    def table_card(card_id, title, subtitle, headers, rows_data, row_renderer,
                   footer=None, totals_row_html=None, xlsx_rows=None):
        """xlsx_rows: optional flat list-of-lists used to pre-generate an
        Excel export at xlsx_export_dir/<card_id>.xlsx. If provided AND
        xlsx_export_dir is set, a download button is rendered in the header.
        totals_row_html: optional <tr>...</tr> string appended to <tbody>."""
        header_cells = "".join(f'<th>{h}</th>' for h in headers)
        body_rows = "".join(row_renderer(r) for r in rows_data)
        if totals_row_html:
            body_rows += totals_row_html
        footer_html = (
            f'  <div class="exhibit-card-footer">{footer}</div>\n'
            if footer else ""
        )
        algo_html = render_algo_details(card_id, params=kpi_params)
        algo_block = f'  <div style="padding:0 1rem 0.8rem;">{algo_html}</div>\n' if algo_html else ""

        export_btn = ""
        if xlsx_rows is not None and xlsx_export_dir:
            xlsx_path = _write_analytics_xlsx_export(
                card_id, title, headers,
                xlsx_rows,
                os.path.join(os.path.dirname(os.path.abspath(__file__)), xlsx_export_dir),
            )
            if xlsx_path:
                rel = f"{xlsx_export_dir.rstrip('/')}/{card_id}.xlsx"
                export_btn = (
                    f'    <a class="analytics-export-btn" href="{rel}" download '
                    f'title="Download as Excel (.xlsx)">'
                    f'<span style="font-size:0.7rem;">&#11015;</span> Excel</a>\n'
                )

        return (
            f'<div class="exhibit-card" id="{card_id}">\n'
            f'  <div class="exhibit-card-header">\n'
            f'    <div class="exhibit-card-title-row">\n'
            f'      <div>\n'
            f'        <div class="exhibit-card-title">{title}</div>\n'
            f'        <div class="exhibit-card-subtitle">{subtitle}</div>\n'
            f'      </div>\n'
            + export_btn +
            f'    </div>\n'
            f'  </div>\n'
            f'  <div class="exhibit-card-body">\n'
            f'    <table class="exhibit-table">\n'
            f'      <thead><tr>{header_cells}</tr></thead>\n'
            f'      <tbody>{body_rows}</tbody>\n'
            f'    </table>\n'
            f'  </div>\n'
            + footer_html
            + algo_block
            + f'</div>\n'
        )

    def total_row(cells):
        """Render a <tr.exhibit-total-row> from a list of (text, css_class) tuples."""
        tds = "".join(f'<td class="{c}">{t}</td>' for t, c in cells)
        return f'<tr class="exhibit-total-row">{tds}</tr>\n'

    # ── 1. By College ──
    total_recs = tables['total_credit_recs']
    by_college_totals = {
        "credit_recs": sum(r["credit_recs"] for r in tables["by_college"]),
        "exhibits": sum(r["exhibits"] for r in tables["by_college"]),
        "ccc_collaborative": sum(r["ccc_collaborative"] for r in tables["by_college"]),
        "industry_certs": sum(r["industry_certs"] for r in tables["by_college"]),
    }
    college_card = table_card(
        "exhibit-by-college",
        "Credit Recommendations by College",
        f"{len(tables['by_college'])} articulating colleges | {fmt(total_recs)} total credit recs",
        ["#", "College", "Credit Recs", "Exhibits", "Disciplines", "CCC Collab", "Industry Certs", "%"],
        tables["by_college"],
        lambda r: (f'<tr><td>{tables["by_college"].index(r)+1}</td>'
                   f'<td class="exhibit-cell-name">{r["college"]}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["credit_recs"])}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["exhibits"])}</td>'
                   f'<td class="exhibit-cell-num">{r["disciplines"]}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["ccc_collaborative"])}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["industry_certs"])}</td>'
                   f'<td class="exhibit-cell-pct">{r["pct"]}%{pct_bar(r["pct"])}</td></tr>\n'),
        footer=f'% = each college\'s share of the {fmt(total_recs)} total credit recommendations statewide',
        totals_row_html=total_row([
            ("Total", ""),
            (f'{len(tables["by_college"])} colleges', "exhibit-cell-name"),
            (fmt(by_college_totals["credit_recs"]), "exhibit-cell-num"),
            (fmt(by_college_totals["exhibits"]), "exhibit-cell-num"),
            ("—", "exhibit-cell-num"),
            (fmt(by_college_totals["ccc_collaborative"]), "exhibit-cell-num"),
            (fmt(by_college_totals["industry_certs"]), "exhibit-cell-num"),
            ("100%", "exhibit-cell-pct"),
        ]),
        xlsx_rows=[
            [i + 1, r["college"], r["credit_recs"], r["exhibits"], r["disciplines"],
             r["ccc_collaborative"], r["industry_certs"], r["pct"]]
            for i, r in enumerate(tables["by_college"])
        ] + [["Total", f'{len(tables["by_college"])} colleges',
              by_college_totals["credit_recs"], by_college_totals["exhibits"], None,
              by_college_totals["ccc_collaborative"], by_college_totals["industry_certs"], 100]],
    )

    # ── 2. By Discipline ──
    by_disc_totals = {
        "credit_recs": sum(r["credit_recs"] for r in tables["by_discipline"]),
        "exhibits": sum(r["exhibits"] for r in tables["by_discipline"]),
        "courses": sum(r["courses"] for r in tables["by_discipline"]),
        "ccc_collaborative": sum(r["ccc_collaborative"] for r in tables["by_discipline"]),
    }
    disc_card = table_card(
        "exhibit-by-discipline",
        "Credit Recommendations by CCC Discipline",
        f"{len(tables['by_discipline'])} discipline areas",
        ["Discipline", "Credit Recs", "Exhibits", "Courses", "Colleges", "CCC Collab", "%"],
        tables["by_discipline"],
        lambda r: (f'<tr><td class="exhibit-cell-name">{r["discipline"]}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["credit_recs"])}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["exhibits"])}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["courses"])}</td>'
                   f'<td class="exhibit-cell-num">{r["colleges"]}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["ccc_collaborative"])}</td>'
                   f'<td class="exhibit-cell-pct">{r["pct"]}%{pct_bar(r["pct"])}</td></tr>\n'),
        totals_row_html=total_row([
            ("Total", "exhibit-cell-name"),
            (fmt(by_disc_totals["credit_recs"]), "exhibit-cell-num"),
            (fmt(by_disc_totals["exhibits"]), "exhibit-cell-num"),
            (fmt(by_disc_totals["courses"]), "exhibit-cell-num"),
            ("—", "exhibit-cell-num"),
            (fmt(by_disc_totals["ccc_collaborative"]), "exhibit-cell-num"),
            ("100%", "exhibit-cell-pct"),
        ]),
        xlsx_rows=[
            [r["discipline"], r["credit_recs"], r["exhibits"], r["courses"],
             r["colleges"], r["ccc_collaborative"], r["pct"]]
            for r in tables["by_discipline"]
        ] + [["Total", by_disc_totals["credit_recs"], by_disc_totals["exhibits"],
              by_disc_totals["courses"], None, by_disc_totals["ccc_collaborative"], 100]],
    )

    # ── 3. By CPL Type ──
    by_cpl_totals = {
        "credit_recs": sum(r["credit_recs"] for r in tables["by_cpl_type"]),
        "exhibits": sum(r["exhibits"] for r in tables["by_cpl_type"]),
    }
    cpl_card = table_card(
        "exhibit-by-cpl-type",
        "Credit Recommendations by CPL Type",
        f"{len(tables['by_cpl_type'])} CPL types",
        ["CPL Type", "Credit Recs", "Exhibits", "Colleges", "%"],
        tables["by_cpl_type"],
        lambda r: (f'<tr><td class="exhibit-cell-name">{r["cpl_type"]}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["credit_recs"])}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["exhibits"])}</td>'
                   f'<td class="exhibit-cell-num">{r["colleges"]}</td>'
                   f'<td class="exhibit-cell-pct">{r["pct"]}%{pct_bar(r["pct"])}</td></tr>\n'),
        totals_row_html=total_row([
            ("Total", "exhibit-cell-name"),
            (fmt(by_cpl_totals["credit_recs"]), "exhibit-cell-num"),
            (fmt(by_cpl_totals["exhibits"]), "exhibit-cell-num"),
            ("—", "exhibit-cell-num"),
            ("100%", "exhibit-cell-pct"),
        ]),
        xlsx_rows=[
            [r["cpl_type"], r["credit_recs"], r["exhibits"], r["colleges"], r["pct"]]
            for r in tables["by_cpl_type"]
        ] + [["Total", by_cpl_totals["credit_recs"], by_cpl_totals["exhibits"], None, 100]],
    )

    # ── 4. By Mode of Learning ──
    by_mol_totals = {
        "credit_recs": sum(r["credit_recs"] for r in tables["by_mode_of_learning"]),
        "exhibits": sum(r["exhibits"] for r in tables["by_mode_of_learning"]),
    }
    mol_card = table_card(
        "exhibit-by-mol",
        "Credit Recommendations by Mode of Learning",
        f"{len(tables['by_mode_of_learning'])} modes",
        ["Mode of Learning", "Credit Recs", "Exhibits", "Colleges", "%"],
        tables["by_mode_of_learning"],
        lambda r: (f'<tr><td class="exhibit-cell-name">{r["mode"]}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["credit_recs"])}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["exhibits"])}</td>'
                   f'<td class="exhibit-cell-num">{r["colleges"]}</td>'
                   f'<td class="exhibit-cell-pct">{r["pct"]}%{pct_bar(r["pct"])}</td></tr>\n'),
        totals_row_html=total_row([
            ("Total", "exhibit-cell-name"),
            (fmt(by_mol_totals["credit_recs"]), "exhibit-cell-num"),
            (fmt(by_mol_totals["exhibits"]), "exhibit-cell-num"),
            ("—", "exhibit-cell-num"),
            ("100%", "exhibit-cell-pct"),
        ]),
        xlsx_rows=[
            [r["mode"], r["credit_recs"], r["exhibits"], r["colleges"], r["pct"]]
            for r in tables["by_mode_of_learning"]
        ] + [["Total", by_mol_totals["credit_recs"], by_mol_totals["exhibits"], None, 100]],
    )

    # ── 5. Collaborative Analysis ──
    by_collab_totals = {
        "credit_recs": sum(r["credit_recs"] for r in tables["collaborative_analysis"]),
        "exhibits": sum(r["exhibits"] for r in tables["collaborative_analysis"]),
    }
    collab_card = table_card(
        "exhibit-collaborative",
        "CCC Collaborative vs. Local Exhibits",
        "Statewide faculty workgroup articulations vs. individual college articulations",
        ["Category", "Credit Recs", "Exhibits", "Colleges", "Disciplines", "%"],
        tables["collaborative_analysis"],
        lambda r: (f'<tr><td class="exhibit-cell-name">{r["category"]}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["credit_recs"])}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["exhibits"])}</td>'
                   f'<td class="exhibit-cell-num">{r["colleges"]}</td>'
                   f'<td class="exhibit-cell-num">{r["disciplines"]}</td>'
                   f'<td class="exhibit-cell-pct">{r["pct"]}%{pct_bar(r["pct"])}</td></tr>\n'),
        totals_row_html=total_row([
            ("Total", "exhibit-cell-name"),
            (fmt(by_collab_totals["credit_recs"]), "exhibit-cell-num"),
            (fmt(by_collab_totals["exhibits"]), "exhibit-cell-num"),
            ("—", "exhibit-cell-num"),
            ("—", "exhibit-cell-num"),
            ("100%", "exhibit-cell-pct"),
        ]),
        xlsx_rows=[
            [r["category"], r["credit_recs"], r["exhibits"], r["colleges"], r["disciplines"], r["pct"]]
            for r in tables["collaborative_analysis"]
        ] + [["Total", by_collab_totals["credit_recs"], by_collab_totals["exhibits"], None, None, 100]],
    )

    # ── 6. Top Exhibits (ranking — no Total row; Excel export included) ──
    top_card = table_card(
        "exhibit-top-50",
        "Top 50 Most-Articulated Exhibits",
        "Ranked by total credit recommendations across all colleges",
        ["#", "Exhibit Title", "Credit Recs", "Courses", "Colleges", "CPL Type", "Discipline"],
        tables["top_exhibits"],
        lambda r: (f'<tr><td>{tables["top_exhibits"].index(r)+1}</td>'
                   f'<td class="exhibit-cell-name">{r["title"]}</td>'
                   f'<td class="exhibit-cell-num">{fmt(r["credit_recs"])}</td>'
                   f'<td class="exhibit-cell-num">{r["courses"]}</td>'
                   f'<td class="exhibit-cell-num">{r["colleges"]}</td>'
                   f'<td>{r["cpl_type"]}</td>'
                   f'<td>{r["discipline"]}</td></tr>\n'),
        xlsx_rows=[
            [i + 1, r["title"], r["credit_recs"], r["courses"], r["colleges"],
             r["cpl_type"], r["discipline"]]
            for i, r in enumerate(tables["top_exhibits"])
        ],
    )

    # ── 7. Articulations by Unified Course (course-identity grouping) ──
    abc_card = ""
    abc_all = tables.get("articulations_by_course", [])
    if abc_all:
        abc_top = abc_all[:50]
        n_ident = len(abc_all)
        n_over = sum(1 for r in abc_all if r["over_merged"])
        n_multi = sum(1 for r in abc_all if r["colleges_earned"] > 1)

        def _abc_lev(r):
            if r["over_merged"]:
                return '<span class="exhibit-cell-num" title="Over-merged cluster — leverage withheld">&#9888; flagged</span>'
            return f'<span class="exhibit-cell-num">{fmt(r["leverage"])}</span>'

        def _abc_course(r):
            t = r["title"] or ""
            sub = f'<div style="font-size:0.72rem;opacity:0.7;">{r["course_id"]} &middot; {r["id_system"]}</div>' if t else f'{r["course_id"]} &middot; {r["id_system"]}'
            return (f'{t}{sub}' if t else sub)

        def _abc_cred(r):
            c = r["credential"] or "—"
            if r["credential_count"] > 1:
                c += f' <span style="opacity:0.6;">(+{r["credential_count"]-1})</span>'
            return c

        abc_card = table_card(
            "articulations-by-course",
            "Articulations by Unified Course",
            f"{fmt(n_ident)} unified course identities &middot; {fmt(n_multi)} earned at &gt;1 college &middot; top 50 by adoption leverage",
            ["#", "Unified Course", "Discipline", "Colleges Earned", "Credit Recommendation", "Credential", "Adoption Leverage"],
            abc_top,
            lambda r: (f'<tr><td>{abc_top.index(r)+1}</td>'
                       f'<td class="exhibit-cell-name">{_abc_course(r)}</td>'
                       f'<td>{r["discipline"]}</td>'
                       f'<td class="exhibit-cell-num">{fmt(r["colleges_earned"])}</td>'
                       f'<td>{r["credit_rec"] or "—"}</td>'
                       f'<td>{_abc_cred(r)}</td>'
                       f'<td>{_abc_lev(r)}</td></tr>\n'),
            footer=(f'Earned articulations grouped by unified course identity. '
                    f'Adoption leverage = peer colleges teaching the same course that have not yet earned it. '
                    f'{fmt(n_over)} over-merged identities show leverage as &#9888; flagged (withheld). '
                    f'Full set ({fmt(n_ident)} identities) in the Excel export.'),
            xlsx_rows=[
                [i + 1,
                 f'{r["title"]} ({r["course_id"]} · {r["id_system"]})' if r["title"] else f'{r["course_id"]} · {r["id_system"]}',
                 r["discipline"], r["colleges_earned"], r["credit_rec"],
                 r["credential"] + (f' (+{r["credential_count"]-1})' if r["credential_count"] > 1 else ""),
                 "over-merged (withheld)" if r["over_merged"] else r["leverage"]]
                for i, r in enumerate(abc_all)
            ],
        )

    # ── 8. Statewide Exhibit Adoption (dynamic — powered by statewide_data.js) ──
    statewide_card = ""
    statewide_data = tables.get("statewide_adoption", [])
    if statewide_data:
        statewide_card = '<div id="statewide-interactive-container" style="margin-top:2rem;"></div>\n'

    # ── Assemble section as a collapsible wrapper that mirrors the KPI Metrics chrome ──
    gen_at = tables.get("generated_at", "")
    section = (
        '<!-- ═══ CPL Analytics Section ═══ -->\n'
        '<div class="kpi-section-wrapper" id="cplAnalyticsWrapper">\n'
        '    <div class="kpi-section-header" onclick="(function(){var w=document.getElementById(\'cplAnalyticsWrapper\');w.classList.toggle(\'collapsed\');})()">\n'
        '        <span class="kpi-section-title">CPL Analytics</span>\n'
        f'        <span class="kpi-section-updated">Source: MAP Custom Reporting · Generated {gen_at}</span>\n'
        '        <span class="kpi-toggle-arrow">&#9650;</span>\n'
        '    </div>\n'
        '    <div class="cpl-analytics-body" id="exhibitAnalysisSection">\n'
        '        <div class="exhibit-cards-grid">\n'
        f'            {collab_card}\n'
        f'            {cpl_card}\n'
        f'            {mol_card}\n'
        f'            {disc_card}\n'
        f'            {college_card}\n'
        f'            {top_card}\n'
        f'            {abc_card}\n'
        '        </div>\n'
        + (f'        {statewide_card}\n' if statewide_card else '')
        + '    </div>\n'
        '</div>\n'
    )

    return section


# ── CSS for exhibit analysis cards ──
EXHIBIT_ANALYSIS_CSS = """
/* ═══ MAP Articulation Analysis Cards ═══ */
.cpl-analytics-body,
.activity-kpi-body {
    background-color: #ffffff;
    padding: 1.5rem 2rem 2rem;
    margin: 0;
}
.kpi-section-wrapper.collapsed .cpl-analytics-body,
.kpi-section-wrapper.collapsed .activity-kpi-body,
.kpi-section-wrapper.collapsed .workplan-projects-body { display: none; }
.activity-kpi-body .activity-kpi-section { margin-bottom: 0; }
.exhibit-cards-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(500px, 1fr));
    gap: 1.5rem;
    max-width: 1400px;
    margin: 0 auto;
}
.exhibit-card {
    background: #0A2240;
    border: 1px solid rgba(201,168,76,0.25);
    border-top: 4px solid #C9A84C;
    border-radius: 8px;
    overflow: hidden;
    box-shadow: 0 2px 6px rgba(0,0,0,0.15);
}
.exhibit-card-header {
    padding: 0.8rem 1rem 0.5rem;
    border-bottom: 1px solid rgba(255,255,255,0.08);
}
.exhibit-card-title-row {
    display: flex;
    align-items: flex-start;
    justify-content: space-between;
    gap: 0.6rem;
}
.exhibit-card-title {
    font-size: 0.95rem;
    font-weight: 700;
    color: #C9A84C;
}
.exhibit-card-subtitle {
    font-size: 0.7rem;
    color: rgba(255,255,255,0.55);
    margin-top: 2px;
}
.analytics-export-btn {
    display: inline-flex;
    align-items: center;
    gap: 0.25rem;
    background: transparent;
    color: #C9A84C;
    border: 1px solid rgba(201,168,76,0.5);
    padding: 4px 10px;
    border-radius: 4px;
    font-size: 0.7rem;
    font-weight: 600;
    text-decoration: none;
    white-space: nowrap;
    transition: background 0.15s, color 0.15s;
}
.analytics-export-btn:hover {
    background: #C9A84C;
    color: #0A2240;
}
.exhibit-total-row td {
    background: rgba(201,168,76,0.12);
    color: #C9A84C;
    font-weight: 700;
    border-top: 2px solid rgba(201,168,76,0.4);
    position: sticky;
    bottom: 0;
}
.exhibit-card-body {
    max-height: 420px;
    overflow: auto;
    scrollbar-width: thin;
    scrollbar-color: rgba(201,168,76,0.3) transparent;
}
.exhibit-card-footer {
    padding: 0.45rem 1rem;
    font-size: 0.64rem;
    color: rgba(255,255,255,0.38);
    border-top: 1px solid rgba(255,255,255,0.06);
    font-style: italic;
}
.exhibit-card-body::-webkit-scrollbar { width: 6px; height: 6px; }
.exhibit-card-body::-webkit-scrollbar-thumb { background: rgba(201,168,76,0.3); border-radius: 3px; }
.exhibit-card-body::-webkit-scrollbar-track { background: transparent; }
.exhibit-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.72rem;
    color: rgba(255,255,255,0.85);
}
.exhibit-table thead {
    position: sticky;
    top: 0;
    z-index: 2;
}
.exhibit-table th {
    background: rgba(10,34,64,0.98);
    color: #C9A84C;
    font-size: 0.65rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.03em;
    padding: 0.45rem 0.5rem;
    text-align: left;
    border-bottom: 1px solid rgba(201,168,76,0.3);
    white-space: nowrap;
}
.exhibit-table td {
    padding: 0.35rem 0.5rem;
    border-bottom: 1px solid rgba(255,255,255,0.04);
    vertical-align: top;
}
.exhibit-table tbody tr:hover {
    background: rgba(201,168,76,0.06);
}
.exhibit-cell-name {
    max-width: 280px;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
    font-weight: 600;
}
/* EACR Phase 4 PR-C2 — supporting elements around the unified-title card name. */
.sw-issuer-subtitle {
    font-size: 0.7rem;
    color: rgba(255,255,255,0.55);
    margin-top: 1px;
    font-style: italic;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    max-width: 280px;
}
.sw-title-badges {
    display: flex;
    flex-wrap: wrap;
    gap: 0.25rem;
    margin-top: 2px;
}
.sw-conf-badge, .sw-quality-badge {
    display: inline-block;
    font-size: 0.58rem;
    font-weight: 700;
    padding: 1px 6px;
    border-radius: 3px;
    letter-spacing: 0.02em;
}
.sw-conf-badge {
    background: rgba(201,168,76,0.18);
    color: #C9A84C;
    border: 1px solid rgba(201,168,76,0.4);
}
.sw-conf-low {
    background: rgba(201,80,80,0.18);
    color: #e8a3a3;
    border-color: rgba(201,80,80,0.4);
}
.sw-quality-badge {
    background: rgba(201,80,80,0.15);
    color: #e8a3a3;
    border: 1px solid rgba(201,80,80,0.3);
}
.sw-also-entered {
    margin-top: 3px;
}
.sw-also-entered > summary {
    font-size: 0.7rem;
    color: rgba(155,188,216,0.85);
    cursor: pointer;
    list-style: none;
    user-select: none;
}
.sw-also-entered > summary::before {
    content: "▸ ";
    font-size: 0.65rem;
    transition: transform 0.15s;
    display: inline-block;
}
.sw-also-entered[open] > summary::before {
    transform: rotate(90deg);
}
.sw-also-entered > summary:hover {
    color: #C9A84C;
}
.sw-raw-titles {
    list-style: none;
    margin: 4px 0 0;
    padding: 0 0 0 14px;
    font-size: 0.7rem;
    color: rgba(255,255,255,0.6);
}
.sw-raw-titles li {
    padding: 1px 0;
    border-left: 1px solid rgba(155,188,216,0.2);
    padding-left: 6px;
    margin-left: -8px;
    white-space: normal;
    line-height: 1.35;
}
.exhibit-cell-num {
    text-align: right;
    font-variant-numeric: tabular-nums;
    white-space: nowrap;
}
.exhibit-cell-pct {
    text-align: right;
    white-space: nowrap;
    color: rgba(255,255,255,0.6);
}
/* Full-width card (statewide adoption) */
.exhibit-card-full {
    grid-column: 1 / -1;
}
/* College name tags */
.sw-college-list {
    line-height: 1.8;
    max-width: 500px;
}
.sw-college {
    display: inline-block;
    font-size: 0.62rem;
    padding: 1px 6px;
    border-radius: 3px;
    margin: 1px 2px;
    white-space: nowrap;
}
.sw-adopted {
    background: rgba(42,125,79,0.25);
    color: #6fcf97;
    border: 1px solid rgba(42,125,79,0.4);
}
.sw-potential {
    background: rgba(201,168,76,0.12);
    color: #C9A84C;
    border: 1px solid rgba(201,168,76,0.25);
}
/* ═══ Interactive Statewide Card ═══ */
#statewide-interactive-container { grid-column: 1 / -1; }
.sw-interactive { background:rgba(10,34,64,0.85); border:1px solid rgba(201,168,76,0.25); border-radius:10px; overflow:hidden; }
.sw-toolbar { padding:0.8rem 1rem; display:flex; flex-wrap:wrap; gap:0.5rem; align-items:center; border-bottom:1px solid rgba(255,255,255,0.08); }
.sw-toolbar input[type=text] { flex:1 1 200px; padding:0.4rem 0.6rem; border:1px solid rgba(255,255,255,0.2); border-radius:5px; background:rgba(255,255,255,0.06); color:#fff; font-size:0.78rem; outline:none; }
.sw-toolbar input[type=text]:focus { border-color:#C9A84C; }
.sw-toolbar input[type=text]::placeholder { color:rgba(255,255,255,0.35); }
.sw-filter-group { position:relative; display:inline-block; }
.sw-filter-btn { padding:0.35rem 0.7rem; border:1px solid rgba(255,255,255,0.2); border-radius:5px; background:rgba(255,255,255,0.06); color:rgba(255,255,255,0.8); font-size:0.72rem; cursor:pointer; white-space:nowrap; }
.sw-filter-btn:hover, .sw-filter-btn.active { border-color:#C9A84C; color:#C9A84C; }
.sw-filter-dropdown { display:none; position:absolute; top:100%; left:0; z-index:100; min-width:220px; max-height:280px; overflow-y:auto; background:#0e2a4a; border:1px solid rgba(201,168,76,0.3); border-radius:6px; box-shadow:0 6px 20px rgba(0,0,0,0.5); margin-top:4px; }
.sw-filter-dropdown.open { display:block; }
.sw-filter-dropdown label { display:flex; align-items:center; gap:0.4rem; padding:0.3rem 0.6rem; font-size:0.7rem; color:rgba(255,255,255,0.8); cursor:pointer; }
.sw-filter-dropdown label:hover { background:rgba(201,168,76,0.1); }
.sw-filter-dropdown input[type=checkbox] { accent-color:#C9A84C; }
.sw-filter-search { width:calc(100% - 1rem); margin:0.4rem 0.5rem; padding:0.3rem 0.5rem; border:1px solid rgba(255,255,255,0.15); border-radius:4px; background:rgba(255,255,255,0.05); color:#fff; font-size:0.68rem; }
.sw-action-bar { padding:0.5rem 1rem; display:flex; flex-wrap:wrap; gap:0.5rem; align-items:center; border-bottom:1px solid rgba(255,255,255,0.06); }
.sw-action-btn { padding:0.35rem 0.9rem; border:1px solid rgba(201,168,76,0.4); border-radius:5px; background:rgba(201,168,76,0.1); color:#C9A84C; font-size:0.72rem; font-weight:600; cursor:pointer; transition:all 0.15s; }
.sw-action-btn:hover { background:rgba(201,168,76,0.25); }
.sw-action-btn.primary { background:#C9A84C; color:#0A2240; border-color:#C9A84C; }
.sw-action-btn.primary:hover { background:#d4b35c; }
.sw-count { font-size:0.7rem; color:rgba(255,255,255,0.5); margin-left:auto; }
.sw-table-wrap { max-height:600px; overflow:auto; scrollbar-width:thin; scrollbar-color:rgba(201,168,76,0.3) transparent; }
.sw-table-wrap::-webkit-scrollbar { width:6px; height:6px; }
.sw-table-wrap::-webkit-scrollbar-thumb { background:rgba(201,168,76,0.3); border-radius:3px; }
.sw-chk { accent-color:#C9A84C; cursor:pointer; }
.sw-row-selected { background:rgba(201,168,76,0.08) !important; }
/* Badges */
.sw-badge { display:inline-block; font-size:0.58rem; font-weight:700; padding:1px 5px; border-radius:3px; text-transform:uppercase; letter-spacing:0.03em; }
.sw-badge-ccc { background:rgba(42,125,79,0.25); color:#6fcf97; border:1px solid rgba(42,125,79,0.4); }
.sw-badge-local { background:rgba(100,149,237,0.15); color:#6495ed; border:1px solid rgba(100,149,237,0.3); }
/* Credit recs toggle */
.sw-recs-toggle { cursor:pointer; color:#C9A84C; font-weight:600; }
.sw-recs-toggle:hover { text-decoration:underline; }
/* Credit recs panel */
.sw-recs-panel { background:rgba(0,0,0,0.2); border:1px solid rgba(201,168,76,0.15); border-radius:6px; padding:0.4rem 0.6rem; max-height:200px; overflow-y:auto; }
.sw-recs-row td { padding:0 !important; border-bottom:none !important; }
/* Pagination */
.sw-pagination { display:flex; justify-content:center; align-items:center; gap:4px; padding:0.6rem 1rem; border-top:1px solid rgba(255,255,255,0.06); }
.sw-page-btn { padding:4px 10px; border:1px solid rgba(255,255,255,0.15); border-radius:4px; background:rgba(255,255,255,0.04); color:rgba(255,255,255,0.7); font-size:0.68rem; cursor:pointer; }
.sw-page-btn:hover:not([disabled]) { background:rgba(201,168,76,0.15); border-color:#C9A84C; color:#C9A84C; }
.sw-page-btn.active { background:rgba(201,168,76,0.25); border-color:#C9A84C; color:#C9A84C; font-weight:700; }
.sw-page-btn[disabled] { opacity:0.3; cursor:default; }
/* Show more link */
.sw-show-more { cursor:pointer; color:#C9A84C; font-size:0.62rem; font-weight:600; }
.sw-show-more:hover { text-decoration:underline; }
/* Inline credit recs */
.sw-credit-recs { margin-top:0.25rem; padding-top:0.2rem; border-top:1px solid rgba(255,255,255,0.05); }
.sw-rec-line { font-size:0.64rem; color:rgba(255,255,255,0.55); line-height:1.5; padding:1px 0; }
.sw-rec-course { font-size:0.58rem; color:rgba(201,168,76,0.5); }
"""


def read_workplan_goals(wb):
    """
    Read the 'Annual Workplan Goals' sheet.
    Each activity block has 3 rows: GOAL, CURRENT, STRETCH.
    Columns: B=Activity description, C=row type, D-H=years (2025-26 to 2029-30), I=TOTAL.
    Returns list of dicts with annual goal/stretch data per activity.
    CURRENT row values are ignored (placeholder data).
    """
    if "Annual Workplan Goals" not in wb.sheetnames:
        print("  No 'Annual Workplan Goals' sheet found — skipping")
        return []

    ws = wb["Annual Workplan Goals"]
    year_labels = ["2025-26", "2026-27", "2027-28", "2028-29", "2029-30"]
    activities = []
    r = 4  # first data row

    while r <= ws.max_row:
        activity_desc = ws.cell(row=r, column=2).value
        row_type = str(ws.cell(row=r, column=3).value or "").strip().upper()
        if not activity_desc and not row_type:
            r += 1
            continue
        if row_type != "GOAL":
            r += 1
            continue

        # Parse activity ID and name from the description cell
        desc_text = str(activity_desc).strip()
        lines = desc_text.split("\n")
        first_line = lines[0].strip()
        # Extract ID like "1.1", "3.1.1", "3.1.2" from start
        import re as _re
        id_match = _re.match(r'^(\d+\.\d+(?:\.\d+)?)\s+(.+)', first_line)
        if id_match:
            act_id = id_match.group(1)
            act_name = id_match.group(2)
        else:
            act_id = first_line[:3].strip()
            act_name = first_line

        # Read GOAL row values (columns D=4 through H=8, I=9 for total)
        goal_values = []
        for c in range(4, 9):
            v = ws.cell(row=r, column=c).value
            goal_values.append(v if v is not None else 0)
        goal_total = ws.cell(row=r, column=9).value or 0

        # Read STRETCH row (should be r+2, skip CURRENT at r+1)
        stretch_values = []
        stretch_total = 0
        stretch_row = r + 2
        if stretch_row <= ws.max_row:
            stype = str(ws.cell(row=stretch_row, column=3).value or "").strip().upper()
            if stype == "STRETCH":
                for c in range(4, 9):
                    v = ws.cell(row=stretch_row, column=c).value
                    stretch_values.append(v if v is not None else 0)
                stretch_total = ws.cell(row=stretch_row, column=9).value or 0

        # Determine if values are percentages (0-1 range)
        is_pct = all(isinstance(v, float) and 0 < v < 1 for v in goal_values if v)

        activities.append({
            "id": act_id,
            "name": act_name,
            "is_percentage": is_pct,
            "years": year_labels,
            "goal": goal_values,
            "goal_total": goal_total,
            "stretch": stretch_values if stretch_values else [0] * 5,
            "stretch_total": stretch_total,
        })

        r += 3  # skip GOAL, CURRENT, STRETCH rows

    print(f"  Read {len(activities)} activities from Annual Workplan Goals")
    return activities


def build_workplan_goals_from_projects(projects, live_data=None):
    """
    Derive the workplan-goals data structure directly from the Project List tab,
    so the 'Annual Workplan Goals' sheet is no longer required.

    Each project row already carries columns:
        kpi_goal_2526, kpi_stretch_2526, kpi_goal_2627, …, kpi_goal_2930, kpi_stretch_2930

    Returns two things as a tuple:
        workplan_goals – list of dicts compatible with render_workplan_goals_html()
        annual_goals   – list of dicts compatible with render_annual_goals_table_html()
    """
    import re as _re

    year_keys = [
        ("2025-26", "kpi_goal_2526", "kpi_stretch_2526"),
        ("2026-27", "kpi_goal_2627", "kpi_stretch_2627"),
        ("2027-28", "kpi_goal_2728", "kpi_stretch_2728"),
        ("2028-29", "kpi_goal_2829", "kpi_stretch_2829"),
        ("2029-30", "kpi_goal_2930", "kpi_stretch_2930"),
    ]

    # Core sub-activity IDs (same as build_activity_kpis)
    core_ids = [
        "1.1", "1.2", "1.3", "1.4",
        "2.1", "2.2", "2.3", "2.4",
        "3.1", "3.2", "3.3", "3.4", "3.5", "3.6",
        "4.1",  # aggregated from 4.1a-4.1d
        "4.2", "4.3", "4.4", "4.5",
    ]
    sprint_ids = ["4.1a", "4.1b", "4.1c", "4.1d"]

    activity_labels = {
        "1": "Activity 1: Build AI-Enhanced CPL Infrastructure",
        "2": "Activity 2: Faculty Workgroups & Credit Recommendations",
        "3": "Activity 3: Build CPL Data Infrastructure",
        "4": "Activity 4: Sprints, Projects, Partnerships & Scale",
    }

    proj_map = {p["id"]: p for p in projects}

    def _parse_num(val):
        """Parse a formatted number string back to a float."""
        if val is None or val == "":
            return 0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace(",", "").replace("+", "")
        if s.lower().endswith("k"):
            try:
                return float(s[:-1]) * 1000
            except ValueError:
                return 0
        try:
            return float(s)
        except ValueError:
            return 0

    # ── Build per-sub-activity rows ──
    workplan_goals = []   # for render_workplan_goals_html
    annual_goals = []     # for render_annual_goals_table_html

    for pid in core_ids:
        # For 4.1, aggregate sprint sub-projects
        if pid == "4.1":
            sprint_projs = [proj_map[sid] for sid in sprint_ids if sid in proj_map]
            if not sprint_projs:
                continue
            name = "Sprints (Veteran, Apprenticeship, Adoption, 29 Palms)"
            goal_values = []
            stretch_values = []
            goal_dict = {}
            stretch_dict = {}
            current_dict = {}
            for yr_label, g_key, s_key in year_keys:
                g_sum = sum(_parse_num(sp.get(g_key, 0)) for sp in sprint_projs)
                s_sum = sum(_parse_num(sp.get(s_key, 0)) for sp in sprint_projs)
                goal_values.append(g_sum)
                stretch_values.append(s_sum)
                goal_dict[yr_label] = g_sum
                stretch_dict[yr_label] = s_sum
                current_dict[yr_label] = 0
            # Star college count comes from the live CCCCO scrape (single source of truth).
            current_metric = int((live_data or {}).get("star_college_count", 0))
        else:
            p = proj_map.get(pid)
            if not p:
                continue
            name = p["name"]
            goal_values = []
            stretch_values = []
            goal_dict = {}
            stretch_dict = {}
            current_dict = {}
            for yr_label, g_key, s_key in year_keys:
                g_val = _parse_num(p.get(g_key, 0))
                s_val = _parse_num(p.get(s_key, 0))
                goal_values.append(g_val)
                stretch_values.append(s_val)
                goal_dict[yr_label] = g_val
                stretch_dict[yr_label] = s_val
                current_dict[yr_label] = 0
            current_metric = _parse_num(p.get("kpi_metric", 0))

        goal_total = sum(goal_values)
        stretch_total = sum(stretch_values)

        # Populate current 2025-26 with actual KPI metric
        current_dict["2025-26"] = current_metric
        current_dict["total"] = current_metric

        # Determine activity group
        act_num = pid.split(".")[0]
        act_label = activity_labels.get(act_num, f"Activity {act_num}")

        # Check if values are percentages
        is_pct = all(0 < v < 1 for v in goal_values if v)

        # workplan_goals format (for render_workplan_goals_html)
        workplan_goals.append({
            "id": pid,
            "name": name,
            "is_percentage": is_pct,
            "years": ["2025-26", "2026-27", "2027-28", "2028-29", "2029-30"],
            "goal": goal_values,
            "goal_total": goal_total,
            "stretch": stretch_values if stretch_values else [0] * 5,
            "stretch_total": stretch_total,
        })

        # annual_goals format (for render_annual_goals_table_html)
        goal_dict["total"] = goal_total
        stretch_dict["total"] = stretch_total
        annual_goals.append({
            "id": pid,
            "name": name,
            "activity": act_label,
            "goal": goal_dict,
            "current": current_dict,
            "stretch": stretch_dict,
        })

    print(f"  Built {len(workplan_goals)} workplan goal rows from Project List")
    return workplan_goals, annual_goals


def read_budget_plan(wb):
    """
    Read comprehensive budget and expenditure data from the "20260324 CPL Budget" tab.

    Structure:
    - Rows 6-11: Funding sources (name, category, type, source, yearly budgets, expense, total)
    - Rows 12-14: Funding subtotals (Operations, Special Projects, Local Funding)
    - Row 15: Grand total
    - Rows 19-52: Expenditure line items
    - Rows 53-56: Summary by type
    - Rows 60-67: Expense categories summary
    - Rows 70-72: Expense areas
    - Rows 75-81: Expenditure factors (FTE, costs, funding, COLA, indirect)
    - Rows 87-99: Personnel plan (positions with FTE and compensation)
    - Row 100: Personnel totals
    """
    # Find the budget sheet (search for "Budget" in name, but not exact "Budget Plan")
    ws = None
    for sheet_name in wb.sheetnames:
        if "budget" in sheet_name.lower() and sheet_name.lower() != "budget plan":
            ws = wb[sheet_name]
            break

    if not ws:
        # Fallback to old sheet name for backward compatibility
        if "Budget Plan" in wb.sheetnames:
            ws = wb["Budget Plan"]
        else:
            # Return empty structure if no budget sheet found
            return {
                "funding_sources": [],
                "funding_subtotals": {},
                "expenditures": [],
                "expense_categories": [],
                "expense_areas": [],
                "factors": {},
                "personnel": [],
                "personnel_totals": {},
                "year_labels": ["2025-26", "2026-27", "2027-28", "2028-29", "2029-30"],
                "grand_total": 0,
            }

    year_labels = ["2025-26", "2026-27", "2027-28", "2028-29", "2029-30"]

    def safe_num(val):
        """Safely convert cell value to number (0 if None or non-numeric)."""
        if val is None:
            return 0
        if isinstance(val, (int, float)):
            return val
        try:
            return float(str(val).replace(",", ""))
        except (ValueError, AttributeError):
            return 0

    # ── FUNDING SOURCES (rows 6-11) ──
    # Cols: B=name, C=category, D=type, E=source, F-K=yearly budgets (5 years), L=total, M=avg
    funding_sources = []
    for r in range(6, 12):
        name = ws.cell(row=r, column=2).value or ""
        if not name:
            break

        source_code = ws.cell(row=r, column=5).value or ""
        budget_by_year = []
        for c in range(6, 11):  # cols F-J = 5 years
            budget_by_year.append(safe_num(ws.cell(row=r, column=c).value))

        expense_2025 = safe_num(ws.cell(row=r, column=7).value)  # col G = expense
        total = safe_num(ws.cell(row=r, column=12).value)  # col L = total
        avg = safe_num(ws.cell(row=r, column=13).value)  # col M = avg

        funding_sources.append({
            "name": name,
            "source_code": source_code,
            "budget_by_year": budget_by_year,
            "expense_2025": expense_2025,
            "total": total,
            "avg": avg,
        })

    # ── FUNDING SUBTOTALS (rows 12-14) ──
    funding_subtotals = {}
    subtotal_labels = ["ops_support", "special_projects", "local_funding"]
    for idx, label in enumerate(subtotal_labels):
        r = 12 + idx
        budget_by_year = []
        for c in range(6, 11):
            budget_by_year.append(safe_num(ws.cell(row=r, column=c).value))
        total = safe_num(ws.cell(row=r, column=12).value)
        funding_subtotals[label] = {
            "budget_by_year": budget_by_year,
            "total": total,
        }

    # ── GRAND TOTAL (row 15, col L) ──
    grand_total = safe_num(ws.cell(row=15, column=12).value)

    # ── EXPENDITURE LINE ITEMS (rows 19-52) ──
    # Cols: B=description, C=category, D=type, E=source, F-K=yearly budgets, G=expense 2025, L=total
    expenditures = []
    for r in range(19, 53):
        description = ws.cell(row=r, column=2).value or ""
        if not description:
            continue

        category = ws.cell(row=r, column=3).value or ""
        exp_type = ws.cell(row=r, column=4).value or ""  # Operations/Special Projects/Local Funding
        source = ws.cell(row=r, column=5).value or ""

        budget_by_year = []
        for c in range(6, 11):  # cols F-J = 5 years
            budget_by_year.append(safe_num(ws.cell(row=r, column=c).value))

        expense_2025 = safe_num(ws.cell(row=r, column=7).value)  # col G = expense
        total = safe_num(ws.cell(row=r, column=12).value)  # col L = total

        expenditures.append({
            "description": description,
            "category": category,
            "type": exp_type,
            "source": source,
            "budget_by_year": budget_by_year,
            "expense_2025": expense_2025,
            "total": total,
        })

    # ── EXPENSE CATEGORIES (rows 60-67) ──
    # Cols: B=name, F-K=yearly budgets, G=expense, L=total, M=avg
    expense_categories = []
    for r in range(60, 68):
        name = ws.cell(row=r, column=2).value or ""
        if not name:
            continue

        budget_by_year = []
        for c in range(6, 11):
            budget_by_year.append(safe_num(ws.cell(row=r, column=c).value))

        expense_2025 = safe_num(ws.cell(row=r, column=7).value)
        total = safe_num(ws.cell(row=r, column=12).value)
        avg = safe_num(ws.cell(row=r, column=13).value)

        expense_categories.append({
            "name": name,
            "budget_by_year": budget_by_year,
            "expense_2025": expense_2025,
            "total": total,
            "avg": avg,
        })

    # ── EXPENSE AREAS (rows 70-72) ──
    expense_areas = []
    for r in range(70, 73):
        name = ws.cell(row=r, column=2).value or ""
        if not name:
            continue

        budget_by_year = []
        for c in range(6, 11):
            budget_by_year.append(safe_num(ws.cell(row=r, column=c).value))

        total = safe_num(ws.cell(row=r, column=12).value)

        expense_areas.append({
            "name": name,
            "budget_by_year": budget_by_year,
            "total": total,
        })

    # ── EXPENDITURE FACTORS (rows 75-81) ──
    factors = {}
    factor_rows = {
        75: "fte",
        76: "platform_maint",
        77: "platform_dev",
        78: "college_funding",
        79: "colleges_eligible",
        80: "cola_rate",
        81: "indirect_rate",
    }

    for r, key in factor_rows.items():
        # Values are in col F (2025-26), H-K (2026-27 through 2029-30)
        year_cols = [6, 8, 9, 10, 11]
        year_vals = [safe_num(ws.cell(row=r, column=c).value) for c in year_cols]
        non_zero = [v for v in year_vals if v != 0]
        factors[key] = sum(non_zero) / len(non_zero) if non_zero else 0

    # ── PERSONNEL PLAN (rows 87-99) ──
    # Cols: B=title, F-K=FTE by year (5 years), G=not used, L=total compensation
    personnel = []
    for r in range(87, 100):
        title = ws.cell(row=r, column=2).value or ""
        if not title:
            continue

        fte_by_year = []
        for c in range(6, 11):  # cols F-J = 5 years
            fte_by_year.append(safe_num(ws.cell(row=r, column=c).value))

        total_comp = safe_num(ws.cell(row=r, column=12).value)  # col L = total comp

        personnel.append({
            "title": title,
            "fte_by_year": fte_by_year,
            "total_comp": total_comp,
        })

    # ── PERSONNEL TOTALS (row 100) ──
    total_fte = safe_num(ws.cell(row=100, column=6).value)  # col F
    total_comp = safe_num(ws.cell(row=100, column=12).value)  # col L
    personnel_totals = {
        "total_fte": total_fte,
        "total_comp": total_comp,
    }

    return {
        "funding_sources": funding_sources,
        "funding_subtotals": funding_subtotals,
        "expenditures": expenditures,
        "expense_categories": expense_categories,
        "expense_areas": expense_areas,
        "factors": factors,
        "personnel": personnel,
        "personnel_totals": personnel_totals,
        "year_labels": year_labels,
        "grand_total": grand_total,
    }


def render_budget_html(budget):
    """
    Generate a complete HTML section for the budget dashboard.
    Replaces content between <!-- Budget Section --> and <!-- Vision 2030 Section -->.

    Includes:
    - Funding Overview with stacked bar chart and source table
    - Expenditure Summary (by category and area)
    - Expenditure Detail (collapsible accordion by type)
    - Personnel Plan with FTE and compensation
    """

    if not budget or not budget.get("funding_sources"):
        # Return minimal section if no budget data
        return '''<!-- Budget Section -->
        <div class="budget-section">
            <h2>CPL Budget & Expenditures</h2>
            <p style="color:#999;">Budget data not available</p>
        </div>
        <!-- End Budget Section -->'''

    # Extract data
    funding_sources = budget.get("funding_sources", [])
    expenditures = budget.get("expenditures", [])
    expense_categories = budget.get("expense_categories", [])
    expense_areas = budget.get("expense_areas", [])
    personnel = budget.get("personnel", [])
    personnel_totals = budget.get("personnel_totals", {})
    factors = budget.get("factors", {})
    year_labels = budget.get("year_labels", ["2025-26", "2026-27", "2027-28", "2028-29", "2029-30"])
    grand_total = budget.get("grand_total", 0)

    # Color scheme
    colors = ["#0A2240", "#163A5F", "#2A7D4F", "#C9A84C", "#9BBCD8"]

    # === SECTION 1: FUNDING OVERVIEW ===
    funding_html = '<div class="budget-funding-overview">\n'
    funding_html += f'        <h3>CPL 5-Year Funding Plan</h3>\n'
    funding_html += f'        <p style="font-size:1.3rem;font-weight:bold;color:#0A2240;margin-bottom:1.5rem;">{fmt_dollars(grand_total)} Total Allocation</p>\n'

    # Stacked bar chart (CSS-based)
    if sum(s.get("total", 0) for s in funding_sources) > 0:
        funding_html += '        <div style="margin-bottom:2rem;">\n'
        funding_html += '            <div style="display:flex;height:40px;border-radius:4px;overflow:hidden;box-shadow:0 2px 4px rgba(0,0,0,0.1);">\n'

        total_funding = sum(s.get("total", 0) for s in funding_sources)
        for idx, source in enumerate(funding_sources):
            source_total = source.get("total", 0)
            pct = (source_total / total_funding * 100) if total_funding > 0 else 0
            color = colors[idx % len(colors)]
            source_label = source.get("name", "")
            funding_html += f'                <div style="flex:{pct:.1f}%;background-color:{color};display:flex;align-items:center;justify-content:center;color:white;font-size:0.75rem;font-weight:bold;text-align:center;" title="{source_label}: {fmt_dollars(source_total)}">'
            if pct > 8:
                funding_html += f'{pct:.0f}%'
            funding_html += '</div>\n'

        funding_html += '            </div>\n'
        funding_html += '            <div style="display:flex;margin-top:0.5rem;font-size:0.8rem;gap:1.5rem;flex-wrap:wrap;">\n'
        for idx, source in enumerate(funding_sources):
            color = colors[idx % len(colors)]
            source_name = source.get("name", "")
            source_total = source.get("total", 0)
            funding_html += f'                <div><span style="display:inline-block;width:12px;height:12px;background-color:{color};border-radius:2px;margin-right:4px;vertical-align:middle;"></span>{source_name}: {fmt_dollars(source_total)}</div>\n'
        funding_html += '            </div>\n'
        funding_html += '        </div>\n'

    # Funding sources table
    funding_html += '        <table style="width:100%;border-collapse:collapse;margin-bottom:2rem;font-size:0.9rem;">\n'
    funding_html += '            <thead>\n'
    funding_html += '                <tr style="background-color:#f5f5f5;border-bottom:2px solid #0A2240;">\n'
    funding_html += '                    <th style="padding:8px;text-align:left;font-weight:bold;">Funding Source</th>\n'
    for year in year_labels:
        funding_html += f'                    <th style="padding:8px;text-align:right;font-weight:bold;">{year}</th>\n'
    funding_html += '                    <th style="padding:8px;text-align:right;font-weight:bold;">2025-26 Expense</th>\n'
    funding_html += '                    <th style="padding:8px;text-align:right;font-weight:bold;border-right:2px solid #0A2240;">Total</th>\n'
    funding_html += '                </tr>\n'
    funding_html += '            </thead>\n'
    funding_html += '            <tbody>\n'

    for source in funding_sources:
        source_name = source.get("name", "")
        budget_by_year = source.get("budget_by_year", [])
        expense_2025 = source.get("expense_2025", 0)
        total = source.get("total", 0)

        funding_html += f'                <tr style="border-bottom:1px solid #ddd;">\n'
        funding_html += f'                    <td style="padding:8px;font-weight:500;">{source_name}</td>\n'
        for budget in budget_by_year:
            funding_html += f'                    <td style="padding:8px;text-align:right;">{fmt_dollars(budget)}</td>\n'
        funding_html += f'                    <td style="padding:8px;text-align:right;color:#666;">{fmt_dollars(expense_2025)}</td>\n'
        funding_html += f'                    <td style="padding:8px;text-align:right;border-right:2px solid #0A2240;font-weight:bold;">{fmt_dollars(total)}</td>\n'
        funding_html += f'                </tr>\n'

    funding_html += '            </tbody>\n'
    funding_html += '        </table>\n'
    funding_html += '    </div>\n'

    # === SECTION 2: EXPENDITURE SUMMARY ===
    summary_html = '    <div class="budget-expenditure-summary" style="margin-top:2rem;">\n'
    summary_html += '        <h3>Expenditure Summary</h3>\n'
    summary_html += '        <div style="display:grid;grid-template-columns:1fr 1fr;gap:2rem;margin-bottom:2rem;">\n'

    # By Expense Category
    summary_html += '            <div>\n'
    summary_html += '                <h4 style="margin-bottom:1rem;">By Expense Category</h4>\n'
    summary_html += '                <table style="width:100%;border-collapse:collapse;font-size:0.85rem;">\n'
    summary_html += '                    <tbody>\n'

    for cat in expense_categories[:8]:  # Top 8 categories
        cat_name = cat.get("name", "")
        total = cat.get("total", 0)
        summary_html += f'                        <tr style="border-bottom:1px solid #eee;">\n'
        summary_html += f'                            <td style="padding:6px;text-align:left;">{cat_name}</td>\n'
        summary_html += f'                            <td style="padding:6px;text-align:right;font-weight:bold;">{fmt_dollars(total)}</td>\n'
        summary_html += f'                        </tr>\n'

    summary_html += '                    </tbody>\n'
    summary_html += '                </table>\n'
    summary_html += '            </div>\n'

    # By Expense Area
    summary_html += '            <div>\n'
    summary_html += '                <h4 style="margin-bottom:1rem;">By Expense Area</h4>\n'
    summary_html += '                <table style="width:100%;border-collapse:collapse;font-size:0.85rem;">\n'
    summary_html += '                    <tbody>\n'

    for area in expense_areas:
        area_name = area.get("name", "")
        total = area.get("total", 0)
        summary_html += f'                        <tr style="border-bottom:1px solid #eee;">\n'
        summary_html += f'                            <td style="padding:6px;text-align:left;">{area_name}</td>\n'
        summary_html += f'                            <td style="padding:6px;text-align:right;font-weight:bold;">{fmt_dollars(total)}</td>\n'
        summary_html += f'                        </tr>\n'

    summary_html += '                    </tbody>\n'
    summary_html += '                </table>\n'
    summary_html += '            </div>\n'
    summary_html += '        </div>\n'
    summary_html += '    </div>\n'

    # === SECTION 3: EXPENDITURE DETAIL (Accordion) ===
    detail_html = '    <div class="budget-expenditure-detail" style="margin-top:2rem;">\n'
    detail_html += '        <h3>Expenditure Detail</h3>\n'

    # Group expenditures by type
    exp_by_type = {}
    for exp in expenditures:
        exp_type = exp.get("type", "Other")
        if exp_type not in exp_by_type:
            exp_by_type[exp_type] = []
        exp_by_type[exp_type].append(exp)

    for exp_type in ["Operations & Support", "Special Projects", "Local Funding"]:
        if exp_type not in exp_by_type:
            continue

        items = exp_by_type[exp_type]
        type_total = sum(e.get("total", 0) for e in items)

        detail_html += f'        <details style="margin-bottom:1rem;border:1px solid #ddd;border-radius:4px;padding:1rem;">\n'
        detail_html += f'            <summary style="font-weight:bold;cursor:pointer;padding:0.5rem 0;user-select:none;">\n'
        detail_html += f'                {exp_type} <span style="float:right;color:#666;">{fmt_dollars(type_total)}</span>\n'
        detail_html += f'            </summary>\n'
        detail_html += f'            <div style="margin-top:1rem;overflow-x:auto;">\n'
        detail_html += f'                <table style="width:100%;border-collapse:collapse;font-size:0.8rem;">\n'
        detail_html += f'                    <thead>\n'
        detail_html += f'                        <tr style="background-color:#f9f9f9;border-bottom:1px solid #ddd;">\n'
        detail_html += f'                            <th style="padding:6px;text-align:left;">Description</th>\n'
        detail_html += f'                            <th style="padding:6px;text-align:left;">Category</th>\n'
        detail_html += f'                            <th style="padding:6px;text-align:left;">Source</th>\n'
        for year in year_labels:
            detail_html += f'                            <th style="padding:6px;text-align:right;font-size:0.75rem;">{year}</th>\n'
        detail_html += f'                            <th style="padding:6px;text-align:right;">Total</th>\n'
        detail_html += f'                        </tr>\n'
        detail_html += f'                    </thead>\n'
        detail_html += f'                    <tbody>\n'

        for item in items:
            desc = item.get("description", "")[:50]  # Truncate to 50 chars
            category = item.get("category", "")
            source = item.get("source", "")
            budget_by_year = item.get("budget_by_year", [])
            total = item.get("total", 0)

            detail_html += f'                        <tr style="border-bottom:1px solid #eee;">\n'
            detail_html += f'                            <td style="padding:6px;font-size:0.75rem;">{desc}</td>\n'
            detail_html += f'                            <td style="padding:6px;font-size:0.75rem;">{category}</td>\n'
            detail_html += f'                            <td style="padding:6px;font-size:0.75rem;">{source}</td>\n'
            for budget in budget_by_year:
                detail_html += f'                            <td style="padding:6px;text-align:right;font-size:0.75rem;">{fmt_dollars(budget)}</td>\n'
            detail_html += f'                            <td style="padding:6px;text-align:right;font-weight:bold;font-size:0.75rem;">{fmt_dollars(total)}</td>\n'
            detail_html += f'                        </tr>\n'

        detail_html += f'                    </tbody>\n'
        detail_html += f'                </table>\n'
        detail_html += f'            </div>\n'
        detail_html += f'        </details>\n'

    detail_html += '    </div>\n'

    # === SECTION 4: PERSONNEL PLAN ===
    personnel_html = '    <div class="budget-personnel-plan" style="margin-top:2rem;">\n'
    personnel_html += '        <h3>Personnel Plan</h3>\n'

    if personnel:
        personnel_html += '        <div style="overflow-x:auto;">\n'
        personnel_html += '            <table style="width:100%;border-collapse:collapse;margin-bottom:1.5rem;font-size:0.85rem;">\n'
        personnel_html += '                <thead>\n'
        personnel_html += '                    <tr style="background-color:#f5f5f5;border-bottom:2px solid #0A2240;">\n'
        personnel_html += '                        <th style="padding:8px;text-align:left;font-weight:bold;">Position Title</th>\n'

        for year in year_labels:
            personnel_html += f'                        <th style="padding:8px;text-align:right;font-weight:bold;">FTE {year}</th>\n'

        personnel_html += '                        <th style="padding:8px;text-align:right;font-weight:bold;border-right:2px solid #0A2240;">Total Compensation</th>\n'
        personnel_html += '                    </tr>\n'
        personnel_html += '                </thead>\n'
        personnel_html += '                <tbody>\n'

        for pos in personnel:
            title = pos.get("title", "")
            fte_by_year = pos.get("fte_by_year", [])
            total_comp = pos.get("total_comp", 0)

            personnel_html += f'                    <tr style="border-bottom:1px solid #ddd;">\n'
            personnel_html += f'                        <td style="padding:8px;">{title}</td>\n'

            for fte in fte_by_year:
                personnel_html += f'                        <td style="padding:8px;text-align:right;">{fte:.2f}</td>\n'

            personnel_html += f'                        <td style="padding:8px;text-align:right;border-right:2px solid #0A2240;font-weight:bold;">{fmt_dollars(total_comp)}</td>\n'
            personnel_html += f'                    </tr>\n'

        # Personnel totals
        total_fte = personnel_totals.get("total_fte", 0)
        total_comp = personnel_totals.get("total_comp", 0)

        personnel_html += '                    <tr style="background-color:#f9f9f9;border-top:2px solid #0A2240;border-bottom:2px solid #0A2240;font-weight:bold;">\n'
        personnel_html += f'                        <td style="padding:8px;">TOTAL</td>\n'
        # Calculate totals by year from individual positions
        for year_idx in range(5):
            year_fte = sum(p.get("fte_by_year", [])[year_idx] if year_idx < len(p.get("fte_by_year", [])) else 0 for p in personnel)
            personnel_html += f'                        <td style="padding:8px;text-align:right;">{year_fte:.2f}</td>\n'
        personnel_html += f'                        <td style="padding:8px;text-align:right;border-right:2px solid #0A2240;">{fmt_dollars(total_comp)}</td>\n'
        personnel_html += '                    </tr>\n'

        personnel_html += '                </tbody>\n'
        personnel_html += '            </table>\n'
        personnel_html += '        </div>\n'

    # Key factors as info cards
    if factors:
        personnel_html += '        <div style="display:grid;grid-template-columns:repeat(auto-fit, minmax(180px, 1fr));gap:1rem;margin-top:1.5rem;">\n'

        factor_display = [
            ("Avg. Annual FTE", factors.get("fte", 0), lambda x: f"{x:.1f}"),
            ("Avg. COLA Rate", factors.get("cola_rate", 0), lambda x: f"{x:.1%}"),
            ("Avg. Indirect Rate", factors.get("indirect_rate", 0), lambda x: f"{x:.1%}"),
            ("Avg. Platform Maint.", factors.get("platform_maint", 0), fmt_dollars),
        ]

        for label, value, formatter in factor_display:
            try:
                formatted = formatter(value)
            except:
                formatted = str(value)

            personnel_html += f'            <div style="background-color:#f5f5f5;padding:1rem;border-radius:4px;border-left:4px solid #0A2240;">\n'
            personnel_html += f'                <div style="font-size:0.75rem;color:#666;text-transform:uppercase;font-weight:bold;margin-bottom:0.5rem;">{label}</div>\n'
            personnel_html += f'                <div style="font-size:1.3rem;font-weight:bold;color:#0A2240;">{formatted}</div>\n'
            personnel_html += f'            </div>\n'

        personnel_html += '        </div>\n'

    personnel_html += '    </div>\n'

    # ── Combine all sections ──
    budget_section = f'''<!-- Budget Section -->
    <div class="budget-section" style="background-color:#fafafa;padding:2rem;border-radius:8px;margin-bottom:2rem;">
        <h2>CPL Budget & Expenditure Plan</h2>
{funding_html}
{summary_html}
{detail_html}
{personnel_html}
    </div>
    <!-- End Budget Section -->'''

    return budget_section


def sync_goals_to_project_list(excel_file, workplan_goals):
    """
    Write annual workplan goal targets back into the Project List sheet.
    Maps each workplan activity (1.1–4.5) to the matching project row,
    writing the 2025-26 value to KPI Target 2026 (col T) and the
    Total/2030 value to KPI Target 2030 (col U).
    Opens a separate writable workbook to avoid data_only conflicts.
    """
    from openpyxl import load_workbook as _lwb
    wb_w = _lwb(excel_file)  # writable (preserves formulas)
    ws = wb_w["Project List"]

    # Build lookup: project ID → row number
    pid_rows = {}
    for r in range(4, ws.max_row + 1):
        pid = ws.cell(row=r, column=COL_ID).value
        if pid:
            pid_rows[str(pid).strip()] = r

    # Build workplan goal lookup by activity ID
    wg_by_id = {}
    for wg in workplan_goals:
        wg_by_id[wg["id"]] = wg

    updated = 0
    for act_id, wg in wg_by_id.items():
        row = pid_rows.get(act_id)
        if not row:
            continue

        goal_vals = wg.get("goal", [])
        goal_total = wg.get("goal_total", 0)
        stretch_vals = wg.get("stretch", [])
        stretch_total = wg.get("stretch_total", 0)

        # 2025-26 = first annual value (index 0)
        t2026 = goal_vals[0] if goal_vals else None
        # 2030 = total (for cumulative data, this is the last/final target)
        t2030 = goal_total

        if t2026 is not None:
            ws.cell(row=row, column=COL_KPI_G2526, value=t2026)
        if t2030:
            ws.cell(row=row, column=COL_KPI_G2930, value=t2030)
        updated += 1

    if updated:
        wb_w.save(excel_file)
        print(f"  Synced {updated} annual goals from workplan → Project List (cols T/U)")
    else:
        print("  No matching goals to sync")

    wb_w.close()


def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: Cannot find {EXCEL_FILE}")
        sys.exit(1)

    # Create a backup before reading (prevents data loss from corruption)
    backup_file = EXCEL_FILE + ".bak"
    try:
        import shutil
        shutil.copy2(EXCEL_FILE, backup_file)
    except Exception:
        pass  # non-critical

    print(f"Reading: {EXCEL_FILE}")

    # Archive current Project List notes to Update Log tab
    archive_updates_to_log(EXCEL_FILE)

    # Read all data from the Project List (single source of truth)
    wb = load_workbook(EXCEL_FILE, data_only=True)
    project_config  = read_project_config(wb)
    print(f"  Project: {project_config['title']} ({project_config['project_id']})")

    # Scan attachments folder (SharePoint sync or local)
    ATTACHMENTS_DIR = os.path.join(
        os.path.expanduser("~"),
        "Riverside Community College District",
        "California MAP Initiative - Documents",
        "CPL Workplan Dashboard",
        "Attachments",
    )
    _LOCAL_ATTACHMENTS = os.path.join(SCRIPT_DIR, "Attachments")
    att_dir = ATTACHMENTS_DIR if os.path.isdir(ATTACHMENTS_DIR) else _LOCAL_ATTACHMENTS
    # Will scan after projects are read (need project list for subfolder creation)
    projects        = read_projects(wb)
    config_overrides = read_config_overrides(wb)
    update_log      = read_update_log(wb)
    budget          = read_budget_plan(wb)
    # Load live CCCCO scrape early so it can seed star_college_count (single
    # source of truth for the Veteran Sprint headline + sprint composite).
    live_data       = read_live_metrics()
    kpis            = compute_headline_kpis(projects, budget, config_overrides, live_data)
    activity_kpis   = build_activity_kpis(projects)

    # ── KPI tunable parameters (from KPI_Config sheet, with defaults) ──
    # Auto-create the sheet on first run so users have a place to edit
    # thresholds without touching the codebase.
    if ensure_kpi_config_sheet(wb):
        try:
            wb.save(EXCEL_FILE)
            print(f"  Created KPI_Config sheet in {os.path.basename(EXCEL_FILE)} with default parameters")
        except Exception as e:
            print(f"  Could not save KPI_Config sheet ({e}); using defaults")
    kpi_params = read_kpi_parameters(wb)

    # Build workplan goals & annual goals from the Project List tab
    # (no longer needs the old 'Annual Workplan Goals' sheet)
    workplan_goals, annual_goals = build_workplan_goals_from_projects(projects, live_data)

    # Auto-create attachment subfolders for new activities/projects
    new_folders = ensure_attachment_subfolders(att_dir, projects)
    if new_folders:
        print(f"  Created {new_folders} new attachment subfolder(s)")
    attachments = scan_attachments(att_dir)
    print(f"  Attachments: {attachments['total']} files, by activity: {attachments['by_activity']}")

    now             = _now_pt().strftime("%B %d, %Y at %-I:%M %p PT")

    # ── Build custom KPI display order from column W if present ──
    # Map project IDs to headline KPI keys
    pid_to_kpi_key = {
        "3.1": "cumulative_students",
        "3.2": "eligible_units",
        "2.1": "credit_recommendations",
        "3.3": "active_colleges",
        "4.1a": "veteran_sprint",
    }
    proj_map_for_order = {p["id"]: p for p in projects}
    kpi_order_pairs = []  # (order_val, kpi_key)
    for pid, kpi_key in pid_to_kpi_key.items():
        p = proj_map_for_order.get(pid)
        if p and p.get("kpi_order") is not None:
            kpi_order_pairs.append((p["kpi_order"], kpi_key))
    if kpi_order_pairs:
        # Sort by order value, then append any keys not explicitly ordered
        kpi_order_pairs.sort(key=lambda x: x[0])
        kpi_display_order = [k for _, k in kpi_order_pairs]
        default_keys = ['cumulative_students', 'eligible_units', 'transcripted_units',
                        'credit_recommendations', 'map_exhibits', 'ccc_collaborative',
                        'active_colleges', 'articulation_colleges',
                        'estimated_savings', 'veteran_sprint', 'twenty_year_impact']
        for dk in default_keys:
            if dk not in kpi_display_order:
                kpi_display_order.append(dk)
        print(f"  Custom KPI display order from column W: {kpi_display_order}")
    else:
        kpi_display_order = None  # use default

    # Merge live dashboard metrics (population breakdowns) if available.
    # live_data was loaded earlier so star_college_count can seed the Veteran
    # Sprint headline before compute_headline_kpis runs.
    if live_data:
        kpis = merge_live_metrics(kpis, live_data)
        print(f"Merged live metrics from {live_data.get('scraped_at', 'unknown')}")
    else:
        print("No live_metrics.json found — using Excel data only")

    # Merge MAP Exhibit metrics (credit recommendations, exhibits, CCC Collaborative)
    exhibit_data = read_exhibit_metrics()
    if exhibit_data:
        kpis = merge_exhibit_metrics(kpis, exhibit_data)
        ds_list = exhibit_data.get('datasets_found', [])
        print(f"Merged exhibit metrics from {exhibit_data.get('source_file', 'unknown')} "
              f"({len(ds_list)} datasets, "
              f"{exhibit_data['total_credit_recs']:,} credit recs, "
              f"{exhibit_data['unique_exhibits']:,} exhibits, "
              f"{exhibit_data['ccc_collaborative']['adopting_colleges']} CCC Collaborative colleges)")
    else:
        print("No exhibit data found — skipping exhibit KPIs")

    # Log daily snapshot and render trends + activity cards
    kpi_history = log_daily_snapshot(live_data, exhibit_data)
    trends_card_html   = render_kpi_history_card(kpi_history, kpi_params=kpi_params)
    college_last_activity    = exhibit_data.get("college_last_activity",    {}) if exhibit_data else {}
    college_military_students = exhibit_data.get("college_military_students", {}) if exhibit_data else {}

    # Build exhibit tables early — College Activity card needs per-college
    # exhibit counts and per-college/discipline detail to render correctly.
    exhibit_tables = build_exhibit_analysis_tables(exhibit_data) if exhibit_data else None

    activity_card_html = render_college_activity_card(
        live_data,
        last_activity=college_last_activity,
        military_students=college_military_students,
        exhibit_tables=exhibit_tables,
        kpi_params=kpi_params,
    )

    # ── Vision 2030 progress — derive from live data + config overrides ──
    # Goal 1: compute from cumulative students / 250,000 target
    try:
        _students_raw = str(kpis['cumulative_students']['value']).replace(",", "")
        _v2030_g1_progress = min(100, round(int(_students_raw) / 250000 * 100))
    except (ValueError, TypeError, KeyError):
        _v2030_g1_progress = 17  # safe fallback
    # Goals 2 & 3: read from config overrides (no live data source)
    _v2030_g2_progress = int(config_overrides.get("V2030_G2_PROGRESS", 55))
    _v2030_g3_progress = int(config_overrides.get("V2030_G3_PROGRESS", 60))
    # Status text: read from config overrides or use defaults
    _v2030_g2_current = str(config_overrides.get("V2030_G2_CURRENT",
        "MAP at 116 colleges; Dashboard live; Portal launching"))
    _v2030_g3_current = str(config_overrides.get("V2030_G3_CURRENT",
        "AB 123 chaptered; $20M allocated; 1,000+ trained"))

    data = {
        "last_updated": now,
        "data_as_of":   now,
        "kpis":         kpis,
        "activity_kpis": activity_kpis,
        "update_log":   update_log,
        "budget":       budget,
        "vision2030": {
            "actions": [
                {"id": "Action 1a", "desc": "Scale CPL opportunities with focus on military service, apprenticeships, and technical industry certifications."},
                {"id": "Action 5", "desc": "Provide flexible course scheduling and CPL opportunities to optimize working learners' abilities to reach educational goals."},
            ],
            "goals": [
                {
                    "id": "Goal 1",
                    "name": "Expand Equitable Access & Boost Student Success",
                    "target": "250,000 Californians by 2030",
                    "progress": _v2030_g1_progress,
                    "current": f"{kpis['cumulative_students']['value']} cumulative students",
                },
                {
                    "id": "Goal 2",
                    "name": "Build Unified, Interoperable, Student-Centered System",
                    "target": "CPL embedded in outreach, onboarding, advising",
                    "progress": _v2030_g2_progress,
                    "current": _v2030_g2_current,
                },
                {
                    "id": "Goal 3",
                    "name": "Sustainable Policies, Resources & Professional Learning",
                    "target": "Faculty-driven policies; ongoing PD; scalable tools",
                    "progress": _v2030_g3_progress,
                    "current": _v2030_g3_current,
                },
            ],
        },
        "projects": projects,
        "workplan_goals": workplan_goals,
    }

    js_content = (
        "/*\n"
        " * CPL Initiative Dashboard -- Live Data File\n"
        " * Auto-generated by excel_to_dashboard.py on " + now + "\n"
        " * Source: CPL_Initiative_Project_List.xlsx\n"
        " *\n"
        " * Includes:\n"
        " *   - 6 headline KPIs (derived from sub-activity data)\n"
        " *   - 19 activity-level KPI cards grouped by Activity 1-4\n"
        " *   - {KPI} token substitution in project text fields\n"
        " *   - Budget data from 5-year plan\n"
        " *   - Vision 2030 alignment\n"
        " */\n\n"
        "window.CPL_DATA = " + json.dumps(data, indent=2, ensure_ascii=False) + ";\n"
    )

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(js_content)

    # ── Write statewide exhibit data as separate JS for interactive card ──
    # exhibit_tables was built earlier (above the render_college_activity_card call)
    if exhibit_tables and exhibit_tables.get("statewide_adoption"):
        sw_js_path = os.path.join(SCRIPT_DIR, "statewide_data.js")
        sw_data = {
            "exhibits": exhibit_tables["statewide_adoption"],
            "analysis": {
                "by_college": exhibit_tables["by_college"],
                "by_discipline": exhibit_tables["by_discipline"],
                "by_cpl_type": exhibit_tables["by_cpl_type"],
                "by_mode_of_learning": exhibit_tables["by_mode_of_learning"],
                "collaborative_analysis": exhibit_tables["collaborative_analysis"],
                "top_exhibits": exhibit_tables["top_exhibits"],
            },
            "generated_at": exhibit_tables.get("generated_at", ""),
            "total_credit_recs": exhibit_tables.get("total_credit_recs", 0),
        }
        sw_js = ("/* Statewide Exhibit Adoption Data — auto-generated */\n"
                 "window.CPL_STATEWIDE = " + json.dumps(sw_data, indent=2, ensure_ascii=False) + ";\n")
        with open(sw_js_path, "w", encoding="utf-8") as f:
            f.write(sw_js)
        print(f"  Exported statewide data to {sw_js_path} ({len(exhibit_tables['statewide_adoption'])} exhibits)")

    # ── Build the Unified Courses tab data + xlsx export (COCI KB staging layer) ──
    try:
        export_unified_courses()
    except Exception as e:
        print(f"  Unified Courses export failed ({e}); tab will show empty")

    # ── Build the Credential Reference tab data (credential-identity layer +
    #    common-course join). Lean payload consumed by credential_reference.js. ──
    try:
        export_credential_reference()
    except Exception as e:
        print(f"  Credential Reference export failed ({e}); tab will fall back to runtime fetch")

    # ── Inject data inline AND render static HTML into the dashboard ──
    if os.path.exists(HTML_FILE):
        with open(HTML_FILE, "r", encoding="utf-8") as f:
            html = f.read()

        # ── Update dashboard title from project config ──
        proj_title = project_config.get("title", "CPL Initiative")
        proj_desc = project_config.get("description", "")
        attach_url = project_config.get("attachments_url", "")
        dash_title = f"{proj_title} &mdash; Project Dashboard"

        # Replace <title> tag
        import re
        html = re.sub(r'<title>[^<]*</title>', f'<title>{proj_title} — Project Dashboard</title>', html)
        # Replace <h1> in header
        html = re.sub(
            r'<h1>[^<]*</h1>',
            f'<h1>{proj_title} &mdash; Project Dashboard</h1>',
            html,
            count=1
        )
        # Remove any previously injected description/attachment block
        PROJ_INFO_START = '<!-- PROJ-INFO-START -->'
        PROJ_INFO_END = '<!-- PROJ-INFO-END -->'
        pi_start = html.find(PROJ_INFO_START)
        pi_end = html.find(PROJ_INFO_END)
        if pi_start != -1 and pi_end != -1:
            html = html[:pi_start] + html[pi_end + len(PROJ_INFO_END):]

        # Build description (collapsible, default collapsed) + See Attachments block
        proj_info_parts = [PROJ_INFO_START]
        if proj_desc:
            # Use full description in a collapsible <details> element
            escaped_desc = proj_desc.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            proj_info_parts.append(
                f'<details class="project-description" style="max-width:800px;margin:0.4rem auto 0;'
                f'cursor:pointer;">'
                f'<summary style="font-size:0.82rem;color:#9BBCD8;font-weight:600;'
                f'list-style:none;display:inline-flex;align-items:center;gap:0.3rem;">'
                f'<span class="desc-arrow" style="font-size:0.7rem;transition:transform 0.2s;">&#9654;</span>'
                f' Project Description</summary>'
                f'<div style="font-size:0.82rem;color:#ccc;line-height:1.5;'
                f'margin-top:0.4rem;padding:0.5rem 0.8rem;'
                f'border-left:2px solid rgba(155,188,216,0.3);text-align:left;">'
                f'{escaped_desc}</div></details>'
                f'<style>.project-description[open] .desc-arrow{{transform:rotate(90deg);}}</style>')
        att_count = attachments.get("total", 0)
        badge_html = (f' <span style="background:#C9A84C;color:#0A2240;font-size:0.65rem;'
                      f'font-weight:700;padding:1px 6px;border-radius:8px;margin-left:4px;">'
                      f'{att_count}</span>') if att_count > 0 else ''
        proj_info_parts.append(
            f'<div style="margin-top:0.5rem;">'
            f'<a href="#" class="attach-btn" target="_blank" '
            f'style="display:inline-flex;align-items:center;gap:0.3rem;'
            f'font-size:0.82rem;color:#fff;text-decoration:none;font-weight:600;'
            f'padding:6px 14px;border:1px solid rgba(255,255,255,0.3);border-radius:4px;'
            f'background:rgba(255,255,255,0.1);cursor:pointer;transition:background 0.2s;"'
            f' onmouseover="this.style.background=\'rgba(255,255,255,0.2)\'"'
            f' onmouseout="this.style.background=\'rgba(255,255,255,0.1)\'"'
            f' title="Open SharePoint folder — use Upload or drag &amp; drop to add files">'
            f'&#128206; See Attachments{badge_html}</a></div>')
        proj_info_parts.append(PROJ_INFO_END)
        proj_info_html = '\n        '.join(proj_info_parts)

        # Insert after the subtitle div
        subtitle_end = html.find('</div>', html.find('class="subtitle"'))
        if subtitle_end != -1:
            insert_pos = subtitle_end + len('</div>')
            html = html[:insert_pos] + '\n        ' + proj_info_html + html[insert_pos:]

        print(f"  Updated dashboard title: {proj_title}")

        # ── Inject attachments URL from Excel config into JS global ──
        att_script_marker = '<script src="dashboard_filters.js">'
        att_script_pos = html.find(att_script_marker)
        if att_script_pos != -1 and attach_url:
            escaped_url = attach_url.replace("'", "\\'")
            inject_js = f"<script>window.CPL_ATTACHMENTS_URL='{escaped_url}';</script>\n    "
            # Remove any previously injected CPL_ATTACHMENTS_URL script
            old_marker = "<script>window.CPL_ATTACHMENTS_URL="
            old_pos = html.find(old_marker)
            if old_pos != -1:
                old_end = html.find("</script>", old_pos) + len("</script>")
                html = html[:old_pos] + html[old_end:].lstrip()
                att_script_pos = html.find(att_script_marker)
            html = html[:att_script_pos] + inject_js + html[att_script_pos:]
            print(f"  Injected attachments URL from Excel config")

        # ── Inject CPL Knowledge Base excerpt for the College Custom Report generator ──
        try:
            kb_text = fetch_cpl_kb()
        except Exception as e:
            print(f"  KB fetch failed entirely ({e}); proceeding without window.CPL_KB")
            kb_text = ""
        # Remove any previously injected CPL_KB script
        kb_start_marker = "<script>window.CPL_KB="
        kb_old_pos = html.find(kb_start_marker)
        if kb_old_pos != -1:
            kb_old_end = html.find("</script>", kb_old_pos) + len("</script>")
            html = html[:kb_old_pos] + html[kb_old_end:].lstrip()
        if kb_text:
            kb_anchor = '<script src="dashboard_filters.js">'
            kb_anchor_pos = html.find(kb_anchor)
            if kb_anchor_pos != -1:
                kb_json = json.dumps(kb_text)  # safe JS string with proper escaping
                kb_inject = f"<script>window.CPL_KB={kb_json};</script>\n    "
                html = html[:kb_anchor_pos] + kb_inject + html[kb_anchor_pos:]
                print(f"  Injected CPL KB excerpt ({len(kb_text):,} chars) from {CPL_KB_REPO}")

        # ── Inject statewide interactive scripts (college_lookup, statewide_data, statewide_interactive) ──
        sw_scripts = [
            '<script src="docx.min.js"></script>',
            '<script src="college_lookup.js"></script>',
            '<script src="statewide_data.js"></script>',
            '<script src="statewide_interactive.js"></script>',
            '<script src="college_report_generator.js"></script>',
            '<script src="unified_courses_data.js"></script>',
            '<script src="unified_courses.js"></script>',
        ]
        # Remove any existing statewide script tags first to guarantee correct order
        for sw_tag in sw_scripts:
            html = html.replace('    ' + sw_tag + '\n', '')
        # Insert all in order before </body>
        body_end = html.rfind('</body>')
        if body_end == -1:
            body_end = len(html)
        block = ''.join('    ' + t + '\n' for t in sw_scripts)
        html = html[:body_end] + block + html[body_end:]
        print("  Ensured statewide interactive script tags present (correct order)")

        START_MARKER = "<!-- DATA-START"
        END_MARKER   = "<!-- DATA-END -->"
        i_start = html.find(START_MARKER)
        i_end   = html.find(END_MARKER)

        # ── Self-heal: if DATA-END is missing (file truncated), rebuild from DATA-START ──
        if i_start != -1 and i_end == -1:
            print("  DATA-END marker missing (file truncated) — rebuilding from DATA-START")
            html = html[:i_start]  # discard everything from DATA-START onwards
            # Re-add both markers with minimal content (will be replaced below)
            html += "<!-- DATA-START (auto-replaced by excel_to_dashboard.py — do not edit manually) -->\n"
            html += "    <!-- DATA-END -->\n</body>\n</html>\n"
            i_start = html.find(START_MARKER)
            i_end = html.find(END_MARKER)

        if i_start != -1 and i_end != -1:
            # ── Filter JS is now in external file dashboard_filters.js ──
            # (loaded via <script src="dashboard_filters.js"> in the HTML template)
            # No injection needed — just clean up any old FILTER-JS blocks
            FILTER_START = "<!-- FILTER-JS-START -->"
            FILTER_END   = "<!-- FILTER-JS-END -->"
            fj_start = html.find(FILTER_START)
            fj_end = html.find(FILTER_END)
            if fj_start != -1 and fj_end != -1:
                html = html[:fj_start] + html[fj_end + len(FILTER_END):]
                print("  Removed old inline FILTER-JS block (now external file)")

            # ── Inject data inline (DATA-START block — bottom of file, OK if truncated) ──
            inline_block = (
                "<!-- DATA-START (auto-replaced by excel_to_dashboard.py — do not edit manually) -->\n"
                "    <script>\n"
                "    window.CPL_DATA = " + json.dumps(data, indent=2, ensure_ascii=False) + ";\n"
                "    </script>\n"
                "    "
            )
            html = html[:i_start] + inline_block + html[i_end:]

            # ── Replace the KPI Summary Cards section with fully rendered static HTML ──
            # End-anchor is the outer "Workplan Activities & Projects" wrapper marker
            # (NOT <!-- Filter Bar -->), since Filter Bar now lives INSIDE that wrapper
            # and using it would wipe the wrapper opening on every run.
            outer_proj_marker = '<!-- ═══ Workplan Activities & Projects Section ═══ -->'
            kpi_section_start = html.find('<!-- KPI Summary Cards -->')
            kpi_section_end = html.find(outer_proj_marker)
            if kpi_section_start != -1 and kpi_section_end != -1:
                kpi_cards_html = render_kpi_section_html(kpis, kpi_display_order, kpi_params=kpi_params)
                # Format the live-scrape timestamp as Pacific Time for the section header.
                kpi_scraped_label = "—"
                _scraped_iso = (live_data or {}).get("scraped_at", "") if live_data else ""
                if _scraped_iso:
                    try:
                        _dt = datetime.fromisoformat(_scraped_iso.replace("Z", "+00:00"))
                        _pt = _dt.astimezone(_PT).replace(tzinfo=None)
                        kpi_scraped_label = _pt.strftime("%b %-d, %Y · %-I:%M %p PT")
                    except Exception:
                        kpi_scraped_label = _scraped_iso[:16].replace("T", " ") + " UTC"
                new_kpi_section = (
                    '<!-- KPI Summary Cards -->\n'
                    '    <div class="kpi-section-wrapper" id="kpiSectionWrapper">\n'
                    '        <div class="kpi-section-header" onclick="(function(){var w=document.getElementById(\'kpiSectionWrapper\');w.classList.toggle(\'collapsed\');})()"> \n'
                    '            <span class="kpi-section-title">KPI Metrics</span>\n'
                    f'            <span class="kpi-section-updated">Live data refreshed {kpi_scraped_label}</span>\n'
                    '            <span class="kpi-toggle-arrow">&#9650;</span>\n'
                    '        </div>\n'
                    '    <div class="kpi-section">\n'
                    + kpi_cards_html
                    + trends_card_html
                    + activity_card_html +
                    '    </div>\n'
                    '    </div>\n\n    '
                )
                html = html[:kpi_section_start] + new_kpi_section + html[kpi_section_end:]
                print("  Rendered static KPI cards with breakdowns")

            # ── Inject CPL Analytics section (collapsible — replaces old toggle button + cards) ──
            # exhibit_tables was built earlier during College Activity rendering.
            # Strip any pre-existing copy first so repeat runs don't accumulate.
            import re as _re
            html = _re.sub(
                r'<div style="text-align:center;"><button class="exhibit-toggle-btn".*?</button></div>\s*',
                '', html, flags=_re.DOTALL,
            )
            # Strip patterns end-anchor on the outer wrapper marker so that
            # the wrapper opening (and everything inside it) is preserved.
            html = _re.sub(
                r'<!-- ═══ MAP Articulation Analysis Section ═══ -->.*?(?=<!-- ═══ Workplan Activities & Projects Section ═══ -->)',
                '', html, flags=_re.DOTALL,
            )
            html = _re.sub(
                r'<!-- ═══ CPL Analytics Section ═══ -->.*?(?=<!-- ═══ Workplan Activities & Projects Section ═══ -->)',
                '', html, flags=_re.DOTALL,
            )
            if exhibit_tables:
                exhibit_html = render_exhibit_analysis_html(
                    exhibit_tables, kpi_params=kpi_params, xlsx_export_dir="exports",
                )
                # Insert BEFORE the outer wrapper marker so Analytics stays OUTSIDE
                # the Workplan Activities & Projects wrapper.
                outer_pos = html.find('<!-- ═══ Workplan Activities & Projects Section ═══ -->')
                if outer_pos != -1:
                    html = html[:outer_pos] + exhibit_html + '\n    ' + html[outer_pos:]
                    print(f"  Injected CPL Analytics section ({len(exhibit_tables['by_college'])} college cards, "
                          f"{len(exhibit_tables['top_exhibits'])} top exhibits, "
                          f"{len(exhibit_tables.get('articulations_by_course', []))} unified-course articulation rows, "
                          f"7 xlsx exports written to exports/)")

                # Inject CSS before closing </style> tag (idempotent — strips any
                # previously injected copies so repeat runs don't accumulate duplicates).
                # Both the current marker block AND a legacy "MAP Exhibit
                # Analysis Cards" block (renamed 2026-05-18) need to be
                # stripped so the static template never accumulates copies.
                import re as _re
                pattern = _re.compile(
                    r'\n?/\* ═══ MAP Articulation Analysis Cards ═══ \*/.*?\.sw-rec-course \{[^}]*\}\n?',
                    _re.DOTALL,
                )
                html = pattern.sub('', html)
                legacy_pattern = _re.compile(
                    r'\n?/\* ═══ MAP Exhibit Analysis Cards ═══ \*/.*?\.exhibit-toggle-btn:hover \{[^}]*\}\n?',
                    _re.DOTALL,
                )
                html = legacy_pattern.sub('', html)

                # Also strip any prior copy of ALGO_DETAILS_CSS so it doesn't
                # accumulate on repeat runs.
                ALGO_CSS_MARKER = '/* ═══ Collapsible Algorithm Descriptions ═══ */'
                if ALGO_CSS_MARKER in html:
                    import re as _re
                    algo_pattern = _re.compile(
                        r'\n?/\* ═══ Collapsible Algorithm Descriptions ═══ \*/.*?\.algo-details \.algo-meta \{[^}]*\}\n?',
                        _re.DOTALL,
                    )
                    html = algo_pattern.sub('', html)

                style_end = html.find('</style>')
                if style_end != -1:
                    html = (html[:style_end]
                            + '\n' + EXHIBIT_ANALYSIS_CSS
                            + '\n' + ALGO_DETAILS_CSS
                            + '\n' + html[style_end:])

            # ── Workplan Activity Metrics: relocate above the Filter Bar and
            # wrap in the same collapsible chrome used by KPI Metrics. ──
            # First, remove any existing copy (old position between Filter Bar
            # and Projects Grid, or a previously-wrapped copy above the Filter
            # Bar) so the section never duplicates.
            import re as _re
            # Strip the legacy in-grid placement.
            html = _re.sub(
                r'<div class="activity-kpi-section" id="activityKpiSection">.*?</div>\s*(?=<!-- Projects Grid -->)',
                '', html, flags=_re.DOTALL,
            )
            # Strip any prior wrapped copy (this generator's own previous output).
            html = _re.sub(
                r'<!-- ═══ Workplan Activity Metrics Section ═══ -->.*?(?=<!-- Filter Bar -->)',
                '', html, flags=_re.DOTALL,
            )
            act_html = render_activity_kpis_html(activity_kpis, annual_goals, update_log, attachments=attachments)
            # The "CPL Workplan Progress — Path to 2030" trend chart logically
            # belongs to this section, so we render it inside the same
            # collapsible body. It used to sit at the top of Projects Grid,
            # which meant it kept showing when Workplan Activity collapsed.
            _current_students = kpis.get("cumulative_students", {}).get("value", "43,321")
            _bd_list = kpis.get("cumulative_students", {}).get("breakdowns", [])
            _sub_pops = {}
            for _bd in _bd_list:
                _lbl = _bd.get("label", "").lower()
                if "military" in _lbl:
                    _sub_pops["military"] = _bd.get("value", "")
                elif "workforce" in _lbl or "other" in _lbl:
                    _sub_pops["workforce"] = _bd.get("value", "")
                elif "apprentice" in _lbl:
                    _sub_pops["apprentice"] = _bd.get("value", "")
            workplan_charts_html = render_workplan_charts_html(
                _current_students, _sub_pops, workplan_goals, config_overrides,
            )
            # Activity Metrics is now a subsection inside the outer
            # "Workplan Activities & Projects" wrapper (defined in the static
            # template), so it no longer has its own kpi-section-wrapper /
            # header — one collapse toggles activities, filters, and projects
            # together.
            new_act_section = (
                '<!-- ═══ Workplan Activity Metrics Section ═══ -->\n'
                '        <div class="activity-kpi-body">\n'
                '            <div class="activity-kpi-section" id="activityKpiSection">\n'
                + act_html +
                '            </div>\n'
                + workplan_charts_html +
                '        </div>\n\n        '
            )
            filter_marker = '<!-- Filter Bar -->'
            filter_pos = html.find(filter_marker)
            if filter_pos != -1:
                html = html[:filter_pos] + new_act_section + html[filter_pos:]
                print(f"  Rendered Workplan Activity Metrics ({sum(len(g['kpis']) for g in activity_kpis)} sub-activities, inside Workplan Activities & Projects)")

            # ── Inject the Annual Workplan Goals section before Projects Grid ──
            workplan_goals_html = render_workplan_goals_html(workplan_goals)
            if workplan_goals_html:
                # Insert before the Projects Grid marker
                wg_marker = '<!-- Projects Grid -->'
                wg_pos = html.find(wg_marker)
                if wg_pos != -1:
                    # Add a section marker for future updates
                    wrapped = ('<!-- Annual Workplan Goals -->\n'
                               + workplan_goals_html
                               + '        <!-- End Annual Workplan Goals -->\n\n        ')
                    # Check if there's already an Annual Workplan Goals section to replace
                    awg_start = html.find('<!-- Annual Workplan Goals -->')
                    awg_end = html.find('<!-- End Annual Workplan Goals -->')
                    if awg_start != -1 and awg_end != -1:
                        html = html[:awg_start] + wrapped + html[awg_end + len('<!-- End Annual Workplan Goals -->'):].lstrip('\n')
                        print(f"  Replaced Annual Workplan Goals section ({len(workplan_goals)} activities)")
                    else:
                        html = html[:wg_pos] + wrapped + html[wg_pos:]
                        print(f"  Injected Annual Workplan Goals section ({len(workplan_goals)} activities)")

            # ── Replace the Projects Grid with fully rendered static HTML ──
            # Use <!-- End Projects Grid --> as the inner-boundary marker so
            # the surrounding tab-pane wrappers (added in Phase D) survive
            # repeated runs. Fall back to <!-- Budget Section --> on a fresh
            # template that doesn't yet have the End marker.
            proj_grid_start = html.find('<!-- Projects Grid -->')
            proj_grid_end = html.find('<!-- End Projects Grid -->')
            if proj_grid_end != -1:
                proj_grid_end_consumes = proj_grid_end + len('<!-- End Projects Grid -->')
            else:
                proj_grid_end = html.find('<!-- Budget Section -->')
                proj_grid_end_consumes = proj_grid_end
            if proj_grid_start != -1 and proj_grid_end != -1:
                proj_cards_html = render_projects_grid_html(projects, update_log, attachments=attachments)
                project_count = len([p for p in projects if not p["id"].startswith("D.")])
                # Workplan Progress chart now lives inside the Workplan Activity
                # Metrics section (so it collapses with it). The Projects Grid
                # is just the project cards now.
                new_proj_section = (
                    '<!-- Projects Grid -->\n'
                    '        <h2 style="margin-bottom:1.5rem;">Projects <span id="projectCount" style="font-size:0.9rem;color:#888;">(' + str(project_count) + ')</span></h2>\n'
                    '        <div id="projectsGrid">\n'
                    + proj_cards_html +
                    '        </div>\n'
                    '        <!-- End Projects Grid -->\n'
                )
                html = html[:proj_grid_start] + new_proj_section + html[proj_grid_end_consumes:]
                print(f"  Rendered static project cards ({project_count} projects, grouped by Goal)")
                print("  Rendered Workplan Progress Chart inside Workplan Activity Metrics (3 trend lines)")

            # ── Teaser cards on the Dashboard tab linking to the other tabs ──
            # Replace the static placeholder; idempotent because the placeholder
            # is wrapped in delimiting comments and stays in the template.
            teaser_html = (
                '<!-- Teaser cards (auto-generated, do not edit manually) -->\n'
                '        <div class="tab-teaser-grid">\n'
                f'            <div class="tab-teaser-card">\n'
                f'                <h3>Annual Workplan Goals</h3>\n'
                f'                <p>Year-by-year targets and stretch goals tracked across {len(workplan_goals)} sub-activities, with current progress and pacing toward 2030.</p>\n'
                f'                <a class="tab-teaser-link" href="#workplan-goals">View Workplan Goals →</a>\n'
                f'            </div>\n'
                f'            <div class="tab-teaser-card">\n'
                f'                <h3>Budget</h3>\n'
                f'                <p>CPL 5-year funding plan, expenditure detail, personnel, and category roll-ups including AB 123 and ESS 25-82 allocations.</p>\n'
                f'                <a class="tab-teaser-link" href="#budget">View Budget →</a>\n'
                f'            </div>\n'
                f'            <div class="tab-teaser-card">\n'
                f'                <h3>Vision 2030 Alignment</h3>\n'
                f'                <p>CPL Initiative alignment with Vision 2030 Actions 1a &amp; 5, and the three CPL system goals — with live progress against each.</p>\n'
                f'                <a class="tab-teaser-link" href="#vision-2030">View Vision 2030 →</a>\n'
                f'            </div>\n'
                '        </div>\n'
                '        <!-- End Teaser cards -->\n'
            )
            # Replace either the bare placeholder (first run) or an existing teaser block (later runs).
            import re as _re
            placeholder_pattern = _re.compile(
                r'<!-- TEASER_CARDS_PLACEHOLDER[^>]*-->|'
                r'<!-- Teaser cards \(auto-generated, do not edit manually\) -->.*?<!-- End Teaser cards -->',
                _re.DOTALL,
            )
            html, n_sub = placeholder_pattern.subn(teaser_html.rstrip(), html, count=1)
            if n_sub:
                print("  Injected Dashboard tab teaser cards")

            # ── Render and inject the Budget section ──
            budget_section_start = html.find('<!-- Budget Section -->')
            budget_section_end = html.find('<!-- End Budget Section -->')
            budget_html = render_budget_html(budget)

            if budget_section_start != -1 and budget_section_end != -1:
                # Replace existing rendered budget section
                html = html[:budget_section_start] + budget_html + html[budget_section_end + len('<!-- End Budget Section -->'):]
                print("  Injected Budget section")
            elif budget_section_start != -1:
                # First run: replace old static budget through Vision 2030 marker
                v2030_marker = html.find('<!-- Vision 2030 Section -->', budget_section_start)
                if v2030_marker != -1:
                    html = html[:budget_section_start] + budget_html + '\n\n        ' + html[v2030_marker:]
                else:
                    html = html[:budget_section_start] + budget_html + html[budget_section_start:]
                print("  Injected Budget section")

            # ── Populate the Lead filter dropdown with unique leads ──
            leads = sorted(set(p.get("lead", "") for p in projects if p.get("lead") and not p["id"].startswith("D.")))
            lead_options = '<option value="">All</option>\n'
            for lead in leads:
                lead_options += f'                <option value="{lead}">{lead}</option>\n'
            html = re.sub(
                r'(<select id="filterLead">)\s*<option value="">All</option>\s*(</select>)',
                r'\1\n                ' + lead_options + r'            \2',
                html,
                flags=re.DOTALL
            )

            # ── Update the last-updated timestamp ──
            html = re.sub(
                r'(<div class="last-updated">)Last Updated:.*?(</div>)',
                r'\1Last Updated: ' + data["last_updated"] + r'\2',
                html
            )

            # ── Inject Refresh Data button below the last-updated div ──
            refresh_btn = (
                '<div style="text-align:center;margin-top:0.5rem;">'
                '<button id="refreshBtn" onclick="'
                'if(!confirm(\'⚠️ Manual Pipeline Run\\n\\n'
                'This will re-scrape live data from the MAP CPL Dashboard, '
                'regenerate all charts and KPIs, and overwrite today\\\'s deployed dashboard.\\n\\n'
                'The process typically takes 3-5 minutes.\\n\\n'
                'Continue?\'))return;'
                'var b=this;b.disabled=true;b.textContent=\'⏳ Pipeline triggered — updating...\';'
                'fetch(\'https://cpl-proxy.slee-548.workers.dev/trigger\','
                '{method:\'POST\',headers:{\'Content-Type\':\'application/json\'},'
                'body:JSON.stringify({secret:\'CPL_SCRAPE_2026\'})})'
                '.then(function(r){return r.json()})'
                '.then(function(d){'
                'if(d.success){b.textContent=\'✅ Pipeline running — page will refresh in 5 min\';'
                'setTimeout(function(){location.reload()},300000)}'
                'else{b.textContent=\'❌ \'+( d.error||\'Unknown error\');b.disabled=false}})'
                '.catch(function(e){b.textContent=\'❌ \'+e.message;b.disabled=false})" style="'
                'background:transparent;color:#C9A84C;border:1px solid #C9A84C;'
                'padding:6px 18px;font-weight:600;cursor:pointer;border-radius:4px;'
                'font-size:0.8rem;font-family:\'Source Sans 3\',Arial,sans-serif;'
                'transition:all 0.2s;" '
                'onmouseover="if(!this.disabled){this.style.background=\'#C9A84C\';this.style.color=\'#0A2240\'}" '
                'onmouseout="if(!this.disabled){this.style.background=\'transparent\';this.style.color=\'#C9A84C\'}">'
                '&#x21bb; Refresh Today&#39;s Data</button></div>'
            )
            # Remove any existing refresh button first, then insert after last-updated
            html = re.sub(r'<div style="text-align:center;margin-top:0\.5rem;">.*?Refresh.*?Data.*?</div>', '', html)
            html = html.replace(
                '<div class="last-updated">Last Updated: ' + data["last_updated"] + '</div>',
                '<div class="last-updated">Last Updated: ' + data["last_updated"] + '</div>\n        ' + refresh_btn
            )

            # ── Update footer date ──
            html = re.sub(
                r'(<span id="footerDate">).*?(</span>)',
                r'\1' + data["last_updated"] + r'\2',
                html
            )

            # ── Inject Annual Workplan Goals table ──
            # Replace in place between the existing markers so the section
            # stays inside whatever wrapper it's nested in (since Phase D
            # that's the workplan-goals tab pane). Falls back to inserting
            # before Vision 2030 Section only on a fresh template.
            if annual_goals:
                goals_table_html = render_annual_goals_table_html(annual_goals)
                ag_start = html.find('<!-- Annual Workplan Goals -->')
                ag_end = html.find('<!-- End Annual Workplan Goals -->')
                if ag_start != -1 and ag_end != -1:
                    wrapped = '<!-- Annual Workplan Goals -->\n' + goals_table_html + '        <!-- End Annual Workplan Goals -->'
                    html = html[:ag_start] + wrapped + html[ag_end + len('<!-- End Annual Workplan Goals -->'):]
                    print(f"  Rendered Annual Workplan Goals table ({len(annual_goals)} rows)")
                else:
                    # Fresh template: insert before Vision 2030 Section
                    v2030_insert = html.find('<!-- Vision 2030 Section -->')
                    if v2030_insert == -1:
                        v2030_insert = html.find('<!-- DATA-START')
                    if v2030_insert != -1:
                        wrapped = '<!-- Annual Workplan Goals -->\n' + goals_table_html + '        <!-- End Annual Workplan Goals -->\n\n        '
                        html = html[:v2030_insert] + wrapped + html[v2030_insert:]
                        print(f"  Rendered Annual Workplan Goals table ({len(annual_goals)} rows, first-run insert)")

            # ── Rebuild the Vision 2030 section with updated goal data ──
            # End boundary is <!-- End Vision 2030 Section --> (added Phase D)
            # so the surrounding tab-pane wrappers survive repeat runs.
            # Falls back to <!-- DATA-START --> on a fresh template.
            v2030_start = html.find('<!-- Vision 2030 Section -->')
            v2030_end = html.find('<!-- End Vision 2030 Section -->')
            if v2030_end != -1:
                v2030_end_consumes = v2030_end + len('<!-- End Vision 2030 Section -->')
            else:
                v2030_end = html.find('<!-- DATA-START')
                v2030_end_consumes = v2030_end
            data_start_marker = v2030_end  # back-compat var name kept for the next block
            if v2030_start != -1 and v2030_end != -1:
                # Use live KPI data for progress calculations
                recs_val = int(str(kpis.get("credit_recommendations", {}).get("value", "576")).replace(",", ""))
                students_val = int(str(kpis.get("cumulative_students", {}).get("value", "42620")).replace(",", ""))
                colleges_val = int(str(kpis.get("active_colleges", {}).get("value", "84")).replace(",", ""))
                recs_pct = round(recs_val / 1000 * 100, 1)
                students_pct = round(students_val / 250000 * 100, 1)

                new_v2030 = f'''<!-- Vision 2030 Section -->
        <div class="vision-section">
            <h2>Vision 2030 Alignment</h2>

            <div class="vision-grid">
                <div class="vision-card">
                    <h3>Vision 2030 Action 1a</h3>
                    <p>Develop infrastructure to ensure that multiple pathways to a baccalaureate degree are available to all students, including completion via the community college system, including expansion of CPL and credential recognition.</p>
                </div>

                <div class="vision-card">
                    <h3>Vision 2030 Action 5</h3>
                    <p>Advance equitable and transparent credentialing systems within and across the California higher education systems that recognize both traditional and competency-based learning.</p>
                </div>

                <div class="vision-card">
                    <h3 style="color:#C9A84C;">{CPL_GOALS["Goal 1"]["title"]}</h3>
                    <p style="font-size:0.85rem;">{CPL_GOALS["Goal 1"]["target"]}</p>
                    <ul style="font-size:0.82rem;color:#555;margin:0.3rem 0 0.5rem 1.2rem;padding:0;">'''
                for b in CPL_GOALS["Goal 1"]["bullets"]:
                    new_v2030 += f'\n                        <li style="margin-bottom:0.2rem;">{b}</li>'
                new_v2030 += f'''
                    </ul>
                    <div class="goal-progress">
                        <div class="goal-progress-bar">
                            <div class="goal-progress-fill" style="width:{recs_pct}%;"></div>
                        </div>
                        <span style="font-weight:600;">Credit Recs: {recs_pct}% ({fmt_number(recs_val)}/1,000) &nbsp;|&nbsp; Students: {students_pct}% ({fmt_number(students_val)}/250,000)</span>
                    </div>
                </div>

                <div class="vision-card">
                    <h3 style="color:#C9A84C;">{CPL_GOALS["Goal 2"]["title"]}</h3>
                    <p style="font-size:0.85rem;">{CPL_GOALS["Goal 2"]["target"]}</p>
                    <ul style="font-size:0.82rem;color:#555;margin:0.3rem 0 0.5rem 1.2rem;padding:0;">'''
                for b in CPL_GOALS["Goal 2"]["bullets"]:
                    new_v2030 += f'\n                        <li style="margin-bottom:0.2rem;">{b}</li>'
                new_v2030 += f'''
                    </ul>
                    <div class="goal-progress">
                        <div class="goal-progress-bar">
                            <div class="goal-progress-fill" style="width:{min(colleges_val / 116 * 100, 100):.1f}%;"></div>
                        </div>
                        <span style="font-weight:600;">Active Colleges: {colleges_val} of 116 ({colleges_val / 116 * 100:.0f}%)</span>
                    </div>
                </div>

                <div class="vision-card">
                    <h3 style="color:#C9A84C;">{CPL_GOALS["Goal 3"]["title"]}</h3>
                    <p style="font-size:0.85rem;">{CPL_GOALS["Goal 3"]["target"]}</p>
                    <ul style="font-size:0.82rem;color:#555;margin:0.3rem 0 0.5rem 1.2rem;padding:0;">'''
                for b in CPL_GOALS["Goal 3"]["bullets"]:
                    new_v2030 += f'\n                        <li style="margin-bottom:0.2rem;">{b}</li>'
                new_v2030 += '''
                    </ul>
                    <div class="goal-progress">
                        <div class="goal-progress-bar">
                            <div class="goal-progress-fill" style="width:65%;"></div>
                        </div>
                        <span style="font-weight:600;">1,000+ trained; $5M ongoing secured; Title 5 updates advancing</span>
                    </div>
                </div>
            </div>
        </div>

        '''
                # Append the End-marker so future runs can re-locate the boundary.
                new_v2030 = new_v2030.rstrip() + '\n        <!-- End Vision 2030 Section -->\n        '
                html = html[:v2030_start] + new_v2030 + html[v2030_end_consumes:]
                print("  Rebuilt Vision 2030 section with updated CPL Goal data")

            # ── Remove the old standalone applyFilters script block if present ──
            # (now injected via DATA-START block instead)
            old_filter_start = html.find('// ── Filter & Search for Project Cards ──')
            if old_filter_start != -1:
                # Find the <script> tag before it and </script> tag after
                script_open = html.rfind('<script>', 0, old_filter_start)
                script_close = html.find('</script>', old_filter_start)
                if script_open != -1 and script_close != -1:
                    html = html[:script_open] + html[script_close + len('</script>'):]
                    print("  Removed old standalone applyFilters script block")

            # ── Ensure proper closing tags (survives Cowork sync truncation) ──
            closing_tags = "\n</body>\n</html>\n"
            # Strip any existing closing tags to avoid duplicates
            html = html.rstrip()
            for tag in ['</html>', '</body>']:
                if html.endswith(tag):
                    html = html[:html.rfind(tag)].rstrip()
            html += closing_tags

            with open(HTML_FILE, "w", encoding="utf-8") as f:
                f.write(html)
            print(f"Injected data inline into {HTML_FILE}")
        else:
            print(f"WARNING: DATA-START/DATA-END markers not found in {HTML_FILE}")
            print("  The HTML will not update. Re-add the markers between <!-- DATA-START --> and <!-- DATA-END -->.")
    else:
        print(f"WARNING: {HTML_FILE} not found — skipping inline injection.")

    # ── Summary ──
    print(f"\nExported {len(projects)} projects to {OUTPUT_FILE}")
    print(f"Headline KPIs: Students={kpis['cumulative_students']['value']}, "
          f"Units={kpis['transcripted_units']['value']}, "
          f"Recs={kpis['credit_recommendations']['value']}, "
          f"Colleges={kpis['active_colleges']['value']}")
    for grp in activity_kpis:
        print(f"  {grp['activity_id']}: {len(grp['kpis'])} sub-activity KPIs")

    # Check for {KPI} tokens that weren't substituted (debugging aid)
    unresolved = [p["id"] for p in projects if "{KPI}" in p["desc"] or "{KPI}" in p["update"] or "{KPI}" in p["milestones"]]
    if unresolved:
        print(f"\nWARNING: {len(unresolved)} project(s) still contain unresolved {{KPI}} tokens: {unresolved}")
        print("  These projects have {KPI} in their text but no KPI Metric value in column R.")
    else:
        print("\nAll {KPI} tokens resolved successfully.")

    # ── Generate Word reports ──
    import subprocess
    report_script = os.path.join(SCRIPT_DIR, "generate_reports.js")
    if os.path.exists(report_script):
        try:
            result = subprocess.run(
                ["node", report_script],
                capture_output=True, text=True, timeout=60, cwd=SCRIPT_DIR
            )
            if result.returncode == 0:
                for line in result.stdout.strip().split("\n"):
                    print(line)
            else:
                print(f"  Report generation warning: {result.stderr.strip()}")
        except Exception as e:
            print(f"  (Skipping report generation: {e})")

    # Mirror to index.html so GitHub Pages (which serves the repo root)
    # stays in sync with CPL_Dashboard.html. The daily workflow does the same
    # cp step, but doing it here too means manual regenerations don't drift.
    try:
        import shutil
        shutil.copyfile(HTML_FILE, os.path.join(SCRIPT_DIR, "index.html"))
        print("Mirrored CPL_Dashboard.html -> index.html")
    except Exception as e:
        print(f"  Warning: could not mirror to index.html ({e})")

    print("Done! Refresh CPL_Dashboard.html in your browser to see changes.")


if __name__ == "__main__":
    main()
