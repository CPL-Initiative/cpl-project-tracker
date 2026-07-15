#!/usr/bin/env python3
"""Definitive rebuild — category from the consultant's CERTIFIED designations.

Finding: in the 2026-07-15 workbook the two tabs disagree in BOTH directions
(45.0199: Desc=Both/Cross=NotCTE→certified Both; 45.0702: Desc=NotCTE/Cross=Both
→certified Both). Neither tab is reliably correct, so the authority is the
consultant's certified green column in cip_cte_discrepancies_260715.xlsx.

Resolution per code:
  1. in the certified 244  -> certified value (authority)
  2. else Desc == Cross    -> that agreed value
  3. else (orphan/only one tab) -> whichever tab has it
  4. else uncertified conflict -> keep + REPORT (should be 0)
"""
import openpyxl, json, os
from collections import defaultdict, Counter

SCRATCH = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(SCRATCH, "cip_searchable_260715.xlsx")
CERT = os.path.join(SCRATCH, "..", "..", "..", "..")  # unused
CERT_FILE = "/root/.claude/uploads/ed89db7a-67e9-5a23-bd48-2f4f883c197d/29ea80c3-cip_cte_discrepancies_260715.xlsx"

CATMAP = {"CTE": "CTE", "Both": "Both", "Not CTE": "Non-CTE", "Noncredit CIP": "Noncredit",
          "Remove": "Retired", "Canadian CIP": "Reserved", "CIP No Longer in Use": "Retired", "": ""}
GOFORWARD = {"CTE", "Both", "Non-CTE", "Noncredit"}

def clean(v): return "" if v is None else str(v).strip()
def head(s): s = clean(s); return s.split(" - ", 1)[0].strip() if " - " in s else s
def tail(s): s = clean(s); return s.split(" - ", 1)[1].strip() if " - " in s else ""
def canon(code):
    code = clean(code)
    l, r = code.split(".", 1) if "." in code else (code, "")
    return l.zfill(2) + "." + r.ljust(4, "0")[:4]

# ── certified designations (green "Crosswalk (Online)" col) ──
cw = openpyxl.load_workbook(CERT_FILE, data_only=True).active
certified = {}
for r in cw.iter_rows(min_row=2, values_only=True):
    certified[canon(r[0])] = clean(r[4])   # col E = CTE Designation from Crosswalk (Online)

# ── workbook 260715 ──
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
desc = {}
for r in wb["CIP Descriptions"].iter_rows(min_row=2, values_only=True):
    code = canon(r[1])
    if clean(r[1]):
        ex = clean(r[13])
        if ex.lower().startswith("examples:"):
            ex = ex[len("examples:"):].strip(" -")
        desc[code] = {"t": clean(r[8]).rstrip(".") or tail(r[9]), "def": clean(r[11]),
                      "ex": ex, "fam": clean(r[0]), "act": clean(r[5]), "flag": clean(r[10])}
xw = defaultdict(set); fam_titles = {}
for r in wb["TOP-CIP Data"].iter_rows(min_row=2, values_only=True):
    cip = canon(r[11]) if clean(r[11]) else ""
    if clean(r[11]):
        xw[cip].add(clean(r[15]))
    fc, ft = head(r[14]), tail(r[14])
    if len(fc) == 2 and fc.isdigit() and ft:
        fam_titles[fc] = ft
# a couple of reserved families have no crosswalk title
fam_titles.setdefault("21", "Reserved (Canadian CIP)")
fam_titles.setdefault("55", "Reserved (Canadian CIP)")

rows, fams, uncertified = [], {}, []
cert_hits = 0
for code, d in desc.items():
    xs = xw.get(code)
    xflag = list(xs)[0] if (xs and len(xs) == 1) else ""
    dflag = d["flag"]
    if code in certified:
        src = certified[code]; cert_hits += 1
    elif xflag and dflag and xflag == dflag:
        src = dflag
    elif xflag and not dflag:
        src = xflag
    elif dflag and not xflag:
        src = dflag
    elif xflag and dflag and xflag != dflag:
        uncertified.append((code, dflag, xflag)); src = xflag  # provisional
    else:
        src = dflag or xflag
    famt = fam_titles.get(d["fam"], "")
    if d["fam"] and famt:
        fams[d["fam"]] = famt
    rows.append({"code": code, "t": d["t"], "cat": CATMAP.get(src, ""),
                 "fam": d["fam"], "def": d["def"], "ex": d["ex"], "act": d["act"]})

rows.sort(key=lambda r: r["code"])

# C-ID / CCN presence — rolled up TOP->CIP.
#   C-ID = Course Identification Numbering (statewide transfer-model articulation)
#   CCN  = Common Course Numbering (AB 1111, the new statewide common numbering)
# A CIP is flagged if any TOP that maps to it has courses carrying a C-ID or CCN.
# It's a FLOOR signal (course-level identifiers present), not full transferability.
try:
    cwb = openpyxl.load_workbook("/home/user/cpl-project-tracker/kb/reference/coci_course_list.xlsx",
                                 read_only=True, data_only=True)
    it = cwb.active.iter_rows(values_only=True)
    ch = {h: i for i, h in enumerate(next(it))}
    ti, ci, ni = ch.get("TopCode"), ch.get("CIDNumber"), ch.get("CommonCourseNumber")
    top_flag = {}
    for row in it:
        raw = clean(row[ti]) if ti is not None else ""
        if not raw:
            continue
        tc = raw.split(":", 1)[0].strip() if ":" in raw else raw
        f = top_flag.setdefault(tc, {"cid": False, "ccn": False})
        if ci is not None and clean(row[ci]):
            f["cid"] = True
        if ni is not None and clean(row[ni]):
            f["ccn"] = True
    craw = open("/home/user/cpl-project-tracker/cip_crosswalk_data.js", encoding="utf-8").read()
    CD = json.loads(craw[craw.index("window.CIP_CROSSWALK =") + len("window.CIP_CROSSWALK ="):].strip().rstrip(";"))
    xset = set()
    for p in CD["pairs"]:
        tc, cc = p[0], p[1]
        f = top_flag.get(tc) if (tc and cc) else None
        if f and (f["cid"] or f["ccn"]):
            xset.add(cc)
    for r in rows:
        if r["code"] in xset:
            r["x"] = 1
    print("C-ID/CCN-flagged rows:", sum(1 for r in rows if r.get("x")),
          "| TOPs with C-ID:", sum(1 for v in top_flag.values() if v["cid"]),
          "| TOPs with CCN:", sum(1 for v in top_flag.values() if v["ccn"]))
except Exception as ex:  # noqa
    print("C-ID/CCN join skipped:", ex)

open(os.path.join(SCRATCH, "cip_proto_data.json"), "w", encoding="utf-8").write(
    json.dumps({"fams": fams, "rows": rows, "_src": "CIP_Searchable_260715 + certified"},
               ensure_ascii=False, separators=(",", ":")))

byc = {r["code"]: r for r in rows}
print("total:", len(rows), "| certified hits:", cert_hits, "of", len(certified))
print("category counts:", dict(Counter(r["cat"] for r in rows)))
print("go-forward:", sum(1 for r in rows if r["cat"] in GOFORWARD), "| families:", len(fams))
print("UNCERTIFIED conflicts (desc!=cross, not in certified list):", len(uncertified))
for u in uncertified[:20]: print("   ", u)
for c in ["45.0199", "45.0702", "32.0107", "01.0000"]:
    r = byc.get(c); print(f"  {c}: {r['cat']!r}  {r['t'][:44]!r}" if r else f"  {c}: MISSING")
